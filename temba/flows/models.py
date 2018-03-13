# -*- coding: utf-8 -*-
from __future__ import absolute_import, division, print_function, unicode_literals

import iso8601
import itertools
import json
import logging
import numbers
import phonenumbers
import regex
import six
import time
import traceback
import six.moves.urllib_request as urllib2

from array import array
from collections import OrderedDict, defaultdict
from datetime import timedelta, datetime
from decimal import Decimal
from django.conf import settings
from django.core.cache import cache
from django.core.files.storage import default_storage
from django.core.files.temp import NamedTemporaryFile
from django.core.urlresolvers import reverse
from django.contrib.auth.models import User, Group
from django.contrib.postgres.fields import ArrayField
from django.db import models, connection as db_connection
from django.db.models import Q, Count, QuerySet, Sum, Max, Prefetch
from django.utils import timezone
from django.utils.functional import cached_property
from django.utils.translation import ugettext_lazy as _, ungettext_lazy as _n
from django.utils.html import escape
from django_redis import get_redis_connection
from enum import Enum
from openpyxl import Workbook
from six.moves import range
from smartmin.models import SmartModel
from temba.airtime.models import AirtimeTransfer
from temba.assets.models import register_asset_store
from temba.contacts.models import Contact, ContactGroup, ContactField, ContactURN, URN, TEL_SCHEME, NEW_CONTACT_VARIABLE
from temba.channels.models import Channel, ChannelSession
from temba.locations.models import AdminBoundary
from temba.msgs.models import Broadcast, Msg, FLOW, INBOX, INCOMING, QUEUED, FAILED, INITIALIZING, HANDLED, Label
from temba.msgs.models import PENDING, DELIVERED, USSD as MSG_TYPE_USSD, OUTGOING, UnreachableException
from temba.orgs.models import Org, Language, get_current_export_version
from temba.utils import analytics, chunk_list, on_transaction_commit, goflow
from temba.utils.dates import get_datetime_format, str_to_datetime, datetime_to_str, json_date_to_datetime
from temba.utils.email import is_valid_address
from temba.utils.export import BaseExportTask, BaseExportAssetStore
from temba.utils.expressions import ContactFieldCollector
from temba.utils.models import SquashableModel, TembaModel, RequireUpdateFieldsMixin, generate_uuid, JSONAsTextField
from temba.utils.queues import push_task
from temba.utils.text import slugify_with
from temba.values.models import Value
from temba_expressions.utils import tokenize
from uuid import uuid4


logger = logging.getLogger(__name__)

FLOW_DEFAULT_EXPIRES_AFTER = 60 * 12
START_FLOW_BATCH_SIZE = 500


class FlowException(Exception):
    pass


class FlowInvalidCycleException(FlowException):
    def __init__(self, node_uuids):
        self.node_uuids = node_uuids


class FlowUserConflictException(FlowException):
    def __init__(self, other_user, last_saved_on):
        self.other_user = other_user
        self.last_saved_on = last_saved_on


class FlowVersionConflictException(FlowException):
    def __init__(self, rejected_version):
        self.rejected_version = rejected_version


FLOW_LOCK_TTL = 60  # 1 minute
FLOW_LOCK_KEY = 'org:%d:lock:flow:%d:%s'

FLOW_PROP_CACHE_KEY = 'org:%d:cache:flow:%d:%s'
FLOW_PROP_CACHE_TTL = 24 * 60 * 60 * 7  # 1 week

UNREAD_FLOW_RESPONSES = 'unread_flow_responses'


class FlowLock(Enum):
    """
    Locks that are flow specific
    """
    participation = 1
    definition = 3


class FlowPropsCache(Enum):
    """
    Properties of a flow that we cache
    """
    terminal_nodes = 1
    category_nodes = 2


@six.python_2_unicode_compatible
class FlowSession(models.Model):
    """
    A contact's session with the flow engine
    """
    STATUS_WAITING = 'W'
    STATUS_COMPLETED = 'C'
    STATUS_INTERRUPTED = 'I'
    STATUS_EXPIRED = 'X'
    STATUS_FAILED = 'F'

    STATUS_CHOICES = ((STATUS_WAITING, "Waiting"),
                      (STATUS_COMPLETED, "Completed"),
                      (STATUS_INTERRUPTED, "Interrupted"),
                      (STATUS_EXPIRED, "Expired"),
                      (STATUS_FAILED, "Failed"))

    GOFLOW_STATUSES = {'waiting': STATUS_WAITING, 'completed': STATUS_COMPLETED, 'errored': STATUS_FAILED}

    org = models.ForeignKey(Org, help_text="The organization this session belongs to")

    contact = models.ForeignKey('contacts.Contact', help_text="The contact that this session is with")

    connection = models.OneToOneField('channels.ChannelSession', null=True, related_name='session',
                                      help_text=_("The channel connection used for flow sessions over IVR or USSD"))

    status = models.CharField(max_length=1, choices=STATUS_CHOICES, null=True, help_text="The status of this session")

    responded = models.BooleanField(default=False, help_text='Whether the contact has responded in this session')

    output = JSONAsTextField(null=True, default=dict)

    @classmethod
    def create(cls, contact, connection):
        return cls.objects.create(org=contact.org, contact=contact, connection=connection)

    @classmethod
    def get_waiting(cls, contact):
        return cls.objects.filter(contact=contact, status=FlowSession.STATUS_WAITING).first()

    @classmethod
    def interrupt_waiting(cls, contacts):
        cls.objects.filter(contact__in=contacts, status=FlowSession.STATUS_WAITING).update(status=FlowSession.STATUS_INTERRUPTED)

    @classmethod
    def bulk_start(cls, contacts, flow, parent_run_summary=None, msg_in=None, params=None):
        """
        Starts a contact in the given flow
        """
        cls.interrupt_waiting(contacts)

        Contact.bulk_cache_initialize(flow.org, contacts)

        client = goflow.get_client()

        asset_timestamp = int(time.time() * 1000000)

        runs = []
        for contact in contacts:
            # build request to flow server
            request = client.request_builder(asset_timestamp).asset_server(flow.org)

            if settings.TESTING:
                # TODO find a way to run an assets server during testing?
                request.include_all(flow.org)

            # only include message if it's a real message
            if msg_in and msg_in.created_on:
                request.add_msg_received(msg_in)

            try:
                if parent_run_summary:
                    output = request.start_by_flow_action(flow.org, contact, flow, parent_run_summary)
                else:
                    output = request.start_manual(flow.org, contact, flow, params)

            except goflow.FlowServerException:
                continue

            status = FlowSession.GOFLOW_STATUSES[output.session['status']]

            # create our session
            session = cls.objects.create(
                org=contact.org,
                contact=contact,
                status=status,
                output=output.session,
                responded=bool(msg_in and msg_in.created_on)
            )

            contact_runs = session.sync_runs(output, msg_in)
            runs.append(contact_runs[0])

        return runs

    def resume(self, msg_in=None, expired_child_run=None):
        """
        Resumes an existing flow session
        """
        # find the run that was waiting for input
        waiting_run = None
        for run in self.output['runs']:
            if run['status'] == 'waiting':
                waiting_run = run
                break

        if not waiting_run:  # pragma: no cover
            raise ValueError("Can't resume a session with no waiting run")

        client = goflow.get_client()

        # build request to flow server
        asset_timestamp = int(time.time() * 1000000)
        request = client.request_builder(asset_timestamp).asset_server(self.org)

        if settings.TESTING:
            # TODO find a way to run an assets server during testing?
            request.include_all(self.org)

        # only include message if it's a real message
        if msg_in and msg_in.created_on:
            request.add_msg_received(msg_in)
        if expired_child_run:  # pragma: needs cover
            request.add_run_expired(expired_child_run)

        # TODO determine if contact or environment has changed
        # request = request.add_contact_changed(self.contact)
        # request = request.add_environment_changed(self.org)

        try:
            new_output = request.resume(self.output)

            # update our output
            self.output = new_output.session
            self.responded = bool(msg_in and msg_in.created_on)
            self.status = FlowSession.GOFLOW_STATUSES[new_output.session['status']]
            self.save(update_fields=('output', 'responded', 'status'))

            # update our session
            self.sync_runs(new_output, msg_in, waiting_run)

        except goflow.FlowServerException:
            # something has gone wrong so this session is over
            self.status = FlowSession.STATUS_FAILED
            self.save(update_fields=('status',))

            self.runs.update(is_active=False, exited_on=timezone.now(), exit_type=FlowRun.EXIT_TYPE_COMPLETED)

        return True, []

    def sync_runs(self, output, msg_in, prev_waiting_run=None):
        """
        Update our runs with the session output
        """
        # make a map of steps to runs
        step_to_run = {}
        for run in output.session['runs']:
            for step in run['path']:
                if step['uuid']:
                    step_to_run[step['uuid']] = run['uuid']

        run_receiving_input = prev_waiting_run or output.session['runs'][0]

        # update each of our runs
        runs = []
        msgs_to_send = []
        first_run = None
        for run in output.session['runs']:
            run_log = [event for event in output.log if event.step_uuid and step_to_run[event.step_uuid] == run['uuid']]

            wait = output.session['wait'] if run['status'] == "waiting" else None

            # currently outgoing messages can only have response_to set if sent from same run
            run_input = msg_in if run['uuid'] == run_receiving_input['uuid'] else None

            run, msgs = FlowRun.create_or_update_from_goflow(self, self.contact, run, run_log, wait, run_input)
            runs.append(run)
            msgs_to_send += msgs

            if not first_run:
                first_run = run

        # if we're no longer active and we're in a simulation, create an action log to show we've left the flow
        if self.contact.is_test and not self.status == FlowSession.STATUS_WAITING:  # pragma: no cover
            ActionLog.create(first_run, '%s has exited this flow' % self.contact.get_display(self.org, short=True))

        # trigger message sending
        if msgs_to_send:
            msgs_to_send = sorted(msgs_to_send, key=lambda m: m.created_on)
            self.org.trigger_send(msgs_to_send)

        return runs

    def __str__(self):  # pragma: no cover
        return six.text_type(self.contact)


@six.python_2_unicode_compatible
class Flow(TembaModel):
    UUID = 'uuid'
    ENTRY = 'entry'
    RULE_SETS = 'rule_sets'
    ACTION_SETS = 'action_sets'
    RULES = 'rules'
    CONFIG = 'config'
    ACTIONS = 'actions'
    DESTINATION = 'destination'
    EXIT_UUID = 'exit_uuid'
    LABEL = 'label'
    WEBHOOK_URL = 'webhook'
    WEBHOOK_ACTION = 'webhook_action'
    FINISHED_KEY = 'finished_key'
    RULESET_TYPE = 'ruleset_type'
    OPERAND = 'operand'
    METADATA = 'metadata'

    BASE_LANGUAGE = 'base_language'
    SAVED_BY = 'saved_by'
    VERSION = 'version'

    CONTACT_CREATION = 'contact_creation'
    CONTACT_PER_RUN = 'run'
    CONTACT_PER_LOGIN = 'login'

    SAVED_ON = 'saved_on'
    NAME = 'name'
    REVISION = 'revision'
    FLOW_TYPE = 'flow_type'
    ID = 'id'
    EXPIRES = 'expires'

    X = 'x'
    Y = 'y'

    FLOW = 'F'
    MESSAGE = 'M'
    VOICE = 'V'
    SURVEY = 'S'
    USSD = 'U'

    RULES_ENTRY = 'R'
    ACTIONS_ENTRY = 'A'

    FLOW_TYPES = ((FLOW, _("Message flow")),
                  (MESSAGE, _("Single Message Flow")),
                  (VOICE, _("Phone call flow")),
                  (SURVEY, _("Android Survey")),
                  (USSD, _("USSD flow")))

    ENTRY_TYPES = ((RULES_ENTRY, "Rules"),
                   (ACTIONS_ENTRY, "Actions"))

    START_MSG_FLOW_BATCH = 'start_msg_flow_batch'

    VERSIONS = [
        "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "10.1", "10.2", "10.3", "10.4", "11.0", "11.1", "11.2", "11.3"
    ]

    name = models.CharField(max_length=64,
                            help_text=_("The name for this flow"))

    labels = models.ManyToManyField('FlowLabel', related_name='flows', verbose_name=_("Labels"), blank=True,
                                    help_text=_("Any labels on this flow"))

    org = models.ForeignKey(Org, related_name='flows')

    entry_uuid = models.CharField(null=True, max_length=36, unique=True)

    entry_type = models.CharField(max_length=1, null=True, choices=ENTRY_TYPES,
                                  help_text=_("The type of node this flow starts with"))

    is_archived = models.BooleanField(default=False,
                                      help_text=_("Whether this flow is archived"))

    flow_type = models.CharField(max_length=1, choices=FLOW_TYPES, default=FLOW,
                                 help_text=_("The type of this flow"))

    metadata = JSONAsTextField(null=True, blank=True, default=dict,
                               help_text=_("Any extra metadata attached to this flow, strictly used by the user interface."))

    expires_after_minutes = models.IntegerField(default=FLOW_DEFAULT_EXPIRES_AFTER,
                                                help_text=_("Minutes of inactivity that will cause expiration from flow"))

    ignore_triggers = models.BooleanField(default=False,
                                          help_text=_("Ignore keyword triggers while in this flow"))

    saved_on = models.DateTimeField(auto_now_add=True,
                                    help_text=_("When this item was saved"))

    saved_by = models.ForeignKey(User, related_name="flow_saves",
                                 help_text=_("The user which last saved this flow"))

    base_language = models.CharField(max_length=4, null=True, blank=True,
                                     help_text=_('The primary language for editing this flow'),
                                     default='base')

    version_number = models.CharField(default=get_current_export_version, max_length=8,
                                      help_text=_("The flow version this definition is in"))

    flow_dependencies = models.ManyToManyField('Flow', related_name='dependent_flows', verbose_name=("Flow Dependencies"), blank=True,
                                               help_text=_("Any flows this flow uses"))

    group_dependencies = models.ManyToManyField(ContactGroup, related_name='dependent_flows', verbose_name=_("Group Dependencies"), blank=True,
                                                help_text=_("Any groups this flow uses"))

    field_dependencies = models.ManyToManyField(ContactField, related_name='dependent_flows', verbose_name=_(''), blank=True,
                                                help_text=('Any fields this flow depends on'))

    flow_server_enabled = models.BooleanField(default=False, help_text=_('Run this flow using the flow server'))

    @classmethod
    def create(cls, org, user, name, flow_type=FLOW, expires_after_minutes=FLOW_DEFAULT_EXPIRES_AFTER, base_language=None):
        flow = Flow.objects.create(org=org, name=name, flow_type=flow_type,
                                   expires_after_minutes=expires_after_minutes, base_language=base_language,
                                   saved_by=user, created_by=user, modified_by=user)

        analytics.track(user.username, 'nyaruka.flow_created', dict(name=name))
        return flow

    @classmethod
    def create_single_message(cls, org, user, message, base_language):
        """
        Creates a special 'single message' flow
        """
        name = 'Single Message (%s)' % six.text_type(uuid4())
        flow = Flow.create(org, user, name, flow_type=Flow.MESSAGE)
        flow.update_single_message_flow(message, base_language)
        return flow

    @classmethod
    def label_to_slug(cls, label):
        return regex.sub(r'[^a-z0-9]+', '_', label.lower(), regex.V0)

    @classmethod
    def create_join_group(cls, org, user, group, response=None, start_flow=None):
        """
        Creates a special 'join group' flow
        """
        base_language = org.primary_language.iso_code if org.primary_language else 'base'

        name = Flow.get_unique_name(org, 'Join %s' % group.name)
        flow = Flow.create(org, user, name, base_language=base_language)
        flow.version_number = '11.2'
        flow.save(update_fields=('version_number',))

        entry_uuid = six.text_type(uuid4())
        definition = {
            'version': flow.version_number,
            'entry': entry_uuid,
            'base_language': base_language,
            'rule_sets': [],
            'action_sets': [
                {
                    'x': 100, 'y': 0,
                    'uuid': entry_uuid,
                    'exit_uuid': str(uuid4()),
                    'actions': [
                        {
                            'uuid': str(uuid4()),
                            'type': 'add_group',
                            'group': {'uuid': group.uuid, 'name': group.name}
                        },
                        {
                            'uuid': str(uuid4()),
                            'type': 'save',
                            'field': 'name',
                            'label': 'Contact Name',
                            'value': '@(PROPER(REMOVE_FIRST_WORD(step.value)))'
                        }
                    ]
                }
            ]
        }

        if response:
            definition['action_sets'][0]['actions'].append({
                'uuid': str(uuid4()),
                'type': 'reply',
                'msg': {base_language: response}
            })

        if start_flow:
            definition['action_sets'][0]['actions'].append({
                'uuid': str(uuid4()),
                'type': 'flow',
                'flow': {'uuid': start_flow.uuid, 'name': start_flow.name}
            })

        flow.update(FlowRevision.migrate_definition(definition, flow))
        return flow

    def use_flow_server(self):
        """
        For now we manually switch flows to using the flow server, though in the settings we can override this for some
        flow types.
        """
        if not settings.FLOW_SERVER_URL:  # pragma: no cover
            return False

        if settings.FLOW_SERVER_FORCE and self.flow_type in (self.MESSAGE, self.FLOW):
            return True

        return self.flow_server_enabled

    @classmethod
    def is_before_version(cls, to_check, version):
        version_str = six.text_type(to_check)
        version = six.text_type(version)
        for ver in Flow.VERSIONS:
            if ver == version_str and version != ver:
                return True
            elif version == ver:
                return False
        return False

    @classmethod
    def import_flows(cls, exported_json, org, user, same_site=False):
        """
        Import flows from our flow export file
        """
        created_flows = []
        flow_uuid_map = dict()

        # create all the flow containers first
        for flow_spec in exported_json['flows']:

            FlowRevision.validate_flow_definition(flow_spec)

            flow_type = flow_spec.get('flow_type', Flow.FLOW)
            name = flow_spec['metadata']['name'][:64].strip()

            flow = None

            # Don't create our campaign message flows, we'll do that later
            # this check is only needed up to version 3 of exports
            if flow_type != Flow.MESSAGE:
                # check if we can find that flow by id first
                if same_site:
                    flow = Flow.objects.filter(org=org, is_active=True, uuid=flow_spec['metadata']['uuid']).first()
                    if flow:  # pragma: needs cover
                        expires_minutes = flow_spec['metadata'].get('expires', FLOW_DEFAULT_EXPIRES_AFTER)
                        if flow_type == Flow.VOICE:
                            expires_minutes = min([expires_minutes, 15])

                        flow.expires_after_minutes = expires_minutes
                        flow.name = Flow.get_unique_name(org, name, ignore=flow)
                        flow.save(update_fields=['name', 'expires_after_minutes'])

                # if it's not of our world, let's try by name
                if not flow:
                    flow = Flow.objects.filter(org=org, is_active=True, name=name).first()

                # if there isn't one already, create a new flow
                if not flow:
                    expires_minutes = flow_spec['metadata'].get('expires', FLOW_DEFAULT_EXPIRES_AFTER)
                    if flow_type == Flow.VOICE:
                        expires_minutes = min([expires_minutes, 15])

                    flow = Flow.create(org, user, Flow.get_unique_name(org, name), flow_type=flow_type,
                                       expires_after_minutes=expires_minutes)

                created_flows.append(dict(flow=flow, flow_spec=flow_spec))

                if 'uuid' in flow_spec['metadata']:
                    flow_uuid_map[flow_spec['metadata']['uuid']] = flow.uuid

        # now let's update our flow definitions with any referenced flows
        def remap_flow(element):
            # first map our id accordingly
            if element['uuid'] in flow_uuid_map:
                element['uuid'] = flow_uuid_map[element['uuid']]

            existing_flow = Flow.objects.filter(uuid=element['uuid'], org=org, is_active=True).first()
            if not existing_flow:
                existing_flow = Flow.objects.filter(org=org, name=element['name'], is_active=True).first()
                if existing_flow:
                    element['uuid'] = existing_flow.uuid

        for created in created_flows:
            for ruleset in created['flow_spec'][Flow.RULE_SETS]:
                if ruleset['ruleset_type'] == RuleSet.TYPE_SUBFLOW:
                    remap_flow(ruleset['config']['flow'])

            for actionset in created['flow_spec'][Flow.ACTION_SETS]:
                for action in actionset['actions']:
                    if action['type'] in ['flow', 'trigger-flow']:
                        remap_flow(action['flow'])
            remap_flow(created['flow_spec']['metadata'])
            created['flow'].import_definition(created['flow_spec'])

        # remap our flow ids according to how they were resolved
        if 'campaigns' in exported_json:
            for campaign in exported_json['campaigns']:
                for event in campaign['events']:
                    if 'flow' in event:
                        flow_uuid = event['flow']['uuid']
                        if flow_uuid in flow_uuid_map:
                            event['flow']['uuid'] = flow_uuid_map[flow_uuid]

        if 'triggers' in exported_json:
            for trigger in exported_json['triggers']:
                if 'flow' in trigger:
                    flow_uuid = trigger['flow']['uuid']
                    if flow_uuid in flow_uuid_map:
                        trigger['flow']['uuid'] = flow_uuid_map[flow_uuid]

        return exported_json

    @classmethod
    def copy(cls, flow, user):
        copy = Flow.create(flow.org, user, "Copy of %s" % flow.name[:55], flow_type=flow.flow_type)

        # grab the json of our original
        flow_json = flow.as_json()

        copy.import_definition(flow_json)

        # copy our expiration as well
        copy.expires_after_minutes = flow.expires_after_minutes
        copy.save()

        return copy

    @classmethod
    def get_node(cls, flow, uuid, destination_type):

        if not uuid or not destination_type:
            return None

        if destination_type == FlowStep.TYPE_RULE_SET:
            return RuleSet.get(flow, uuid)
        else:
            return ActionSet.get(flow, uuid)

    @classmethod
    def handle_call(cls, call, text=None, saved_media_url=None, hangup=False, resume=False):
        run = FlowRun.objects.filter(connection=call, is_active=True).select_related('org').order_by('-created_on').first()

        # what we will send back
        voice_response = call.channel.generate_ivr_response()

        if run is None:  # pragma: no cover
            voice_response.hangup()
            return voice_response

        flow = run.flow

        # make sure we have the latest version
        flow.ensure_current_version()

        run.voice_response = voice_response

        # make sure our test contact is handled by simulation
        if call.contact.is_test:
            Contact.set_simulation(True)

        # create a message to hold our inbound message
        from temba.msgs.models import IVR
        if text or saved_media_url:

            # we don't have text for media, so lets use the media value there too
            if saved_media_url and ':' in saved_media_url:
                text = saved_media_url.partition(':')[2]

            msg = Msg.create_incoming(call.channel, six.text_type(call.contact_urn),
                                      text, status=PENDING, msg_type=IVR,
                                      attachments=[saved_media_url] if saved_media_url else None,
                                      connection=run.connection)
        else:
            msg = Msg(org=call.org, contact=call.contact, text='', id=0)

        # find out where we last left off
        step = run.steps.all().order_by('-arrived_on').first()

        # if we are just starting the flow, create our first step
        if not step:
            # lookup our entry node
            destination = ActionSet.objects.filter(flow=run.flow, uuid=flow.entry_uuid).first()
            if not destination:
                destination = RuleSet.objects.filter(flow=run.flow, uuid=flow.entry_uuid).first()

            # and add our first step for our run
            if destination:
                step = flow.add_step(run, destination, [])

        # go and actually handle wherever we are in the flow
        destination = Flow.get_node(run.flow, step.step_uuid, step.step_type)
        (handled, msgs) = Flow.handle_destination(destination, step, run, msg, user_input=text is not None, resume_parent_run=resume)

        # if we stopped needing user input (likely), then wrap our response accordingly
        voice_response = Flow.wrap_voice_response_with_input(call, run, voice_response)

        # if we handled it, mark it so
        if handled and msg.id:
            Msg.mark_handled(msg)

        # if we didn't handle it, this is a good time to hangup
        if not handled or hangup:
            voice_response.hangup()
            run.set_completed(final_step=step)

        return voice_response

    @classmethod
    def wrap_voice_response_with_input(cls, call, run, voice_response):
        """ Finds where we are in the flow and wraps our voice_response with whatever comes next """
        step = run.steps.all().order_by('-pk').first()
        destination = Flow.get_node(run.flow, step.step_uuid, step.step_type)
        if isinstance(destination, RuleSet):
            response = call.channel.generate_ivr_response()
            callback = 'https://%s%s' % (run.org.get_brand_domain(), reverse('ivr.ivrcall_handle', args=[call.pk]))
            gather = destination.get_voice_input(response, action=callback)

            # recordings have to be tacked on last
            if destination.ruleset_type == RuleSet.TYPE_WAIT_RECORDING:
                voice_response.record(action=callback)

            elif destination.ruleset_type == RuleSet.TYPE_SUBFLOW:
                voice_response.redirect(url=callback)

            elif gather and hasattr(gather, 'document'):  # voicexml case
                gather.join(voice_response)

                voice_response = response

            elif gather:  # TwiML case
                # nest all of our previous verbs in our gather
                for verb in voice_response.verbs:
                    gather.append(verb)

                voice_response = response

                # append a redirect at the end in case the user sends #
                voice_response.redirect(url=callback + "?empty=1")

        return voice_response

    @classmethod
    def get_unique_name(cls, org, base_name, ignore=None):
        """
        Generates a unique flow name based on the given base name
        """
        name = base_name[:64].strip()

        count = 2
        while True:
            flows = Flow.objects.filter(name=name, org=org, is_active=True)
            if ignore:  # pragma: needs cover
                flows = flows.exclude(pk=ignore.pk)

            if not flows.exists():
                break

            name = '%s %d' % (base_name[:59].strip(), count)
            count += 1

        return name

    @classmethod
    def should_close_connection(cls, run, current_destination, next_destination):
        if run.flow.flow_type == Flow.USSD:
            # this might be our last node that sends msg
            if not next_destination:
                return True
            else:
                if next_destination.is_messaging:
                    return False
                else:
                    return Flow.should_close_connection_graph(next_destination)
        else:
            return False

    @classmethod
    def should_close_connection_graph(cls, start_node):
        # modified DFS that is looking for nodes with messaging capabilities
        if start_node.get_step_type() == FlowStep.TYPE_RULE_SET:
            # keep rules only that have destination
            rules = [rule for rule in start_node.get_rules() if rule.destination]
            if not rules:
                return True
            else:
                for rule in rules:
                    next_node = Flow.get_node(start_node.flow, rule.destination, rule.destination_type)
                    if next_node.is_messaging:
                        return False
                    else:
                        if Flow.should_close_connection_graph(next_node):
                            continue
                        else:
                            return False
                return True
        elif start_node.get_step_type() == FlowStep.TYPE_ACTION_SET:
            if start_node.destination:
                next_node = Flow.get_node(start_node.flow, start_node.destination, start_node.destination_type)
                if next_node.is_messaging:
                    return False
                else:
                    return Flow.should_close_connection_graph(next_node)
            else:
                return True

    @classmethod
    def find_and_handle(cls, msg, started_flows=None, voice_response=None,
                        triggered_start=False, resume_parent_run=False,
                        resume_after_timeout=False, user_input=True, trigger_send=True, continue_parent=True):

        if started_flows is None:
            started_flows = []

        # resume via flow server if we have a waiting session for that
        session = FlowSession.get_waiting(contact=msg.contact)
        if session:
            return session.resume(msg_in=msg)

        steps = FlowStep.get_active_steps_for_contact(msg.contact, step_type=FlowStep.TYPE_RULE_SET)
        for step in steps:
            flow = step.run.flow
            flow.ensure_current_version()
            destination = Flow.get_node(flow, step.step_uuid, step.step_type)

            # this node doesn't exist anymore, mark it as left so they leave the flow
            if not destination:  # pragma: no cover
                step.run.set_completed(final_step=step)
                Msg.mark_handled(msg)
                return True, []

            (handled, msgs) = Flow.handle_destination(destination, step, step.run, msg, started_flows,
                                                      user_input=user_input, triggered_start=triggered_start,
                                                      resume_parent_run=resume_parent_run,
                                                      resume_after_timeout=resume_after_timeout, trigger_send=trigger_send,
                                                      continue_parent=continue_parent)

            if handled:
                return True, msgs

        return False, []

    @classmethod
    def handle_destination(cls, destination, step, run, msg,
                           started_flows=None, is_test_contact=False, user_input=False,
                           triggered_start=False, trigger_send=True, resume_parent_run=False, resume_after_timeout=False, continue_parent=True):

        if started_flows is None:
            started_flows = []

        def add_to_path(path, uuid):
            if uuid in path:
                path.append(uuid)
                raise FlowException("Flow cycle detected at runtime: %s" % path)
            path.append(uuid)

        start_time = time.time()
        path = []
        msgs = []

        # lookup our next destination
        handled = False

        while destination:
            result = {"handled": False}

            if destination.get_step_type() == FlowStep.TYPE_RULE_SET:
                should_pause = False

                # check if we need to stop
                if destination.is_pause():
                    should_pause = True

                if (user_input or resume_after_timeout) or not should_pause:
                    result = Flow.handle_ruleset(destination, step, run, msg, started_flows, resume_parent_run,
                                                 resume_after_timeout)
                    add_to_path(path, destination.uuid)

                    # add any messages generated by this ruleset (ussd and subflow)
                    msgs += result.get('msgs', [])

                    # USSD check for session end
                    if not result.get('interrupted') and \
                            Flow.should_close_connection(run, destination, result.get('destination')):

                        end_message = Msg.create_outgoing(msg.org, get_flow_user(msg.org), msg.contact, '',
                                                          channel=msg.channel,
                                                          connection=msg.connection, response_to=msg if msg.id else None)

                        end_message.connection.mark_ending()
                        msgs.append(end_message)
                        ActionLog.create(run, _("USSD Session was marked to end"))

                # USSD ruleset has extra functionality to send out messages.
                elif destination.is_ussd():
                    result = Flow.handle_ussd_ruleset_action(destination, step, run, msg)

                    msgs += result.get('msgs', [])

                # if we used this input, then mark our user input as used
                if should_pause:
                    user_input = False

                    # once we handle user input, reset our path
                    path = []

            elif destination.get_step_type() == FlowStep.TYPE_ACTION_SET:
                result = Flow.handle_actionset(destination, step, run, msg, started_flows, is_test_contact)
                add_to_path(path, destination.uuid)

                # USSD check for session end
                if Flow.should_close_connection(run, destination, result.get('destination')):
                    for msg in result['msgs']:
                        msg.connection.mark_ending()
                        ActionLog.create(run, _("USSD Session was marked to end"))

                # add any generated messages to be sent at once
                msgs += result.get('msgs', [])

            # if this is a triggered start, we only consider user input on the first step, so clear it now
            if triggered_start:
                user_input = False

            # pull out our current state from the result
            step = result.get('step')

            # lookup our next destination
            destination = result.get('destination', None)

            # if any one of our destinations handled us, consider it handled
            if result.get('handled', False):
                handled = True

            resume_parent_run = False
            resume_after_timeout = False

        # if we have a parent to continue, do so
        if getattr(run, 'continue_parent', False) and continue_parent:
            msgs += FlowRun.continue_parent_flow_run(run, trigger_send=False, continue_parent=True)

        if handled:
            analytics.gauge('temba.flow_execution', time.time() - start_time)

        # send any messages generated
        if msgs and trigger_send:
            msgs.sort(key=lambda message: message.created_on)
            Msg.objects.filter(id__in=[m.id for m in msgs]).exclude(status=DELIVERED).update(status=PENDING)
            run.flow.org.trigger_send(msgs)

        return handled, msgs

    @classmethod
    def handle_actionset(cls, actionset, step, run, msg, started_flows, is_test_contact=False):

        # not found, escape out, but we still handled this message, user is now out of the flow
        if not actionset:  # pragma: no cover
            run.set_completed(final_step=step)
            return dict(handled=True, destination=None, destination_type=None)

        # actually execute all the actions in our actionset
        msgs = actionset.execute_actions(run, msg, started_flows)
        run.add_messages(msgs, step=step)

        # and onto the destination
        destination = Flow.get_node(actionset.flow, actionset.destination, actionset.destination_type)
        if destination:
            step = run.flow.add_step(run, destination, previous_step=step, exit_uuid=actionset.exit_uuid)
        else:
            run.set_completed(final_step=step)
            step = None

        return dict(handled=True, destination=destination, step=step, msgs=msgs)

    @classmethod
    def handle_ruleset(cls, ruleset, step, run, msg, started_flows, resume_parent_run=False, resume_after_timeout=False):
        msgs = []

        if ruleset.is_ussd() and run.connection_interrupted:
            rule, value = ruleset.find_interrupt_rule(step, run, msg)
            if not rule:
                run.set_interrupted(final_step=step)
                return dict(handled=True, destination=None, destination_type=None, interrupted=True)
        else:
            if ruleset.ruleset_type == RuleSet.TYPE_SUBFLOW:
                if not resume_parent_run:
                    flow_uuid = ruleset.config.get('flow').get('uuid')
                    flow = Flow.objects.filter(org=run.org, uuid=flow_uuid).first()
                    flow.org = run.org
                    message_context = run.flow.build_expressions_context(run.contact, msg, run=run)

                    # our extra will be the current flow variables
                    extra = message_context.get('extra', {})
                    extra['flow'] = message_context.get('flow', {})

                    if msg.id:
                        run.add_messages([msg], step=step)
                        run.update_expiration(timezone.now())

                    if flow:
                        child_runs = flow.start([], [run.contact], started_flows=started_flows,
                                                restart_participants=True, extra=extra,
                                                parent_run=run, interrupt=False)

                        if child_runs:
                            child_run = child_runs[0]
                            msgs += child_run.start_msgs
                            continue_parent = getattr(child_run, 'continue_parent', False)
                        else:  # pragma: no cover
                            continue_parent = False

                        # it's possible that one of our children interrupted us with a start flow action
                        run.refresh_from_db(fields=('is_active',))
                        if continue_parent and run.is_active:
                            started_flows.remove(flow.id)
                        else:
                            return dict(handled=True, destination=None, destination_type=None, msgs=msgs)

            # find a matching rule
            rule, value = ruleset.find_matching_rule(step, run, msg, resume_after_timeout=resume_after_timeout)

        flow = ruleset.flow

        # add the message to our step
        if msg.id:
            run.add_messages([msg], step=step)
            run.update_expiration(timezone.now())

        if ruleset.ruleset_type in RuleSet.TYPE_MEDIA and msg.attachments:
            # store the media path as the value
            value = msg.attachments[0].split(':', 1)[1]

        step.save_rule_match(rule, value)
        ruleset.save_run_value(run, rule, value, msg.text)

        # output the new value if in the simulator
        if run.contact.is_test:
            if run.connection_interrupted:  # pragma: no cover
                ActionLog.create(run, _("@flow.%s has been interrupted") % (Flow.label_to_slug(ruleset.label)))
            else:
                ActionLog.create(run, _("Saved '%s' as @flow.%s") % (value, Flow.label_to_slug(ruleset.label)))

        # no destination for our rule?  we are done, though we did handle this message, user is now out of the flow
        if not rule.destination:
            if run.connection_interrupted:
                # run was interrupted and interrupt state not handled (not connected)
                run.set_interrupted(final_step=step)
                return dict(handled=True, destination=None, destination_type=None, interrupted=True, msgs=msgs)
            else:
                run.set_completed(final_step=step)
                return dict(handled=True, destination=None, destination_type=None, msgs=msgs)

        # Create the step for our destination
        destination = Flow.get_node(flow, rule.destination, rule.destination_type)
        if destination:
            step = flow.add_step(run, destination, exit_uuid=rule.uuid, previous_step=step)

        return dict(handled=True, destination=destination, step=step, msgs=msgs)

    @classmethod
    def handle_ussd_ruleset_action(cls, ruleset, step, run, msg):
        action = UssdAction.from_ruleset(ruleset, run)
        context = run.flow.build_expressions_context(run.contact, msg)
        msgs = action.execute(run, context, ruleset.uuid, msg)

        run.add_messages(msgs, step=step)

        return dict(handled=True, destination=None, step=step, msgs=msgs)

    @classmethod
    def apply_action_label(cls, user, flows, label, add):  # pragma: needs cover
        return label.toggle_label(flows, add)

    @classmethod
    def apply_action_archive(cls, user, flows):
        changed = []

        for flow in flows:

            # don't archive flows that belong to campaigns
            from temba.campaigns.models import CampaignEvent
            if not CampaignEvent.objects.filter(flow=flow, campaign__org=user.get_org(), campaign__is_archived=False).exists():
                flow.archive()
                changed.append(flow.pk)

        return changed

    @classmethod
    def apply_action_restore(cls, user, flows):
        changed = []
        for flow in flows:
            try:
                flow.restore()
                changed.append(flow.pk)
            except FlowException:  # pragma: no cover
                pass
        return changed

    @classmethod
    def get_versions_before(cls, version_number):
        versions = []
        version_str = six.text_type(version_number)
        for ver in Flow.VERSIONS:
            if version_str != ver:
                versions.append(ver)
            else:
                break
        return versions

    @classmethod
    def get_versions_after(cls, version_number):
        versions = []
        version_str = six.text_type(version_number)
        for ver in reversed(Flow.VERSIONS):
            if version_str != ver:
                versions.insert(0, ver)
            else:
                break
        return versions

    def as_select2(self):
        return dict(id=self.uuid, text=self.name)

    def release(self):
        """
        Releases this flow, marking it inactive. We remove all flow runs, steps and values in a background process.
        We keep FlowRevisions and FlowStarts however.
        """
        from .tasks import deactivate_flow_runs_task

        self.is_active = False
        self.save()

        # release any campaign events that depend on this flow
        from temba.campaigns.models import CampaignEvent
        for event in CampaignEvent.objects.filter(flow=self, is_active=True):
            event.release()

        # release any triggers that depend on this flow
        from temba.triggers.models import Trigger
        for trigger in Trigger.objects.filter(flow=self, is_active=True):
            trigger.release()

        self.group_dependencies.clear()
        self.flow_dependencies.clear()
        self.field_dependencies.clear()

        # deactivate our runs in the background
        on_transaction_commit(lambda: deactivate_flow_runs_task.delay(self.id))

    def get_category_counts(self, deleted_nodes=True):

        actives = self.rule_sets.all().values('uuid', 'label').order_by('y', 'x')

        uuids = [active['uuid'] for active in actives]
        keys = [Flow.label_to_slug(active['label']) for active in actives]
        counts = FlowCategoryCount.objects.filter(flow_id=self.id)

        # always filter by active keys
        counts = counts.filter(result_key__in=keys)

        # filter by active nodes if we aren't including deleted nodes
        if not deleted_nodes:
            counts = counts.filter(node_uuid__in=uuids)
        counts = counts.values('result_key', 'category_name').annotate(count=Sum('count'), result_name=Max('result_name'))

        results = {}
        for count in counts:
            key = count['result_key']
            result = results.get(key, {})
            if 'name' not in result:
                if count['category_name'] == 'All Responses':
                    continue
                result['key'] = key
                result['name'] = count['result_name']
                result['categories'] = [dict(name=count['category_name'], count=count['count'])]
                result['total'] = count['count']
            else:
                result['categories'].append(dict(name=count['category_name'], count=count['count']))
                result['total'] += count['count']
            results[count['result_key']] = result

        for k, v in six.iteritems(results):
            for cat in results[k]['categories']:
                if (results[k]['total']):
                    cat['pct'] = float(cat['count']) / float(results[k]['total'])
                else:
                    cat['pct'] = 0

        # order counts by their place on the flow
        result_list = []
        for active in actives:
            key = Flow.label_to_slug(active['label'])
            result = results.get(key)
            if result:
                result_list.append(result)

        return dict(counts=result_list)

    def deactivate_runs(self):
        """
        Exits all flow runs, values and steps for a flow. For now, intentionally leave our values
        and steps since those are not long for this world.
        """

        # grab the ids of all our active runs
        run_ids = self.runs.filter(is_active=True).values_list('id', flat=True)

        # batch this for 1,000 runs at a time so we don't grab locks for too long
        for id_batch in chunk_list(run_ids, 1000):
            now = timezone.now()
            runs = FlowRun.objects.filter(id__in=id_batch)
            runs.update(is_active=False, exited_on=now, exit_type=FlowRun.EXIT_TYPE_INTERRUPTED, modified_on=now)

        # clear all our cached stats
        self.clear_props_cache()

    def clear_props_cache(self):
        r = get_redis_connection()
        keys = [self.get_props_cache_key(c) for c in FlowPropsCache.__members__.values()]
        r.delete(*keys)

    def get_props_cache_key(self, kind):
        return FLOW_PROP_CACHE_KEY % (self.org_id, self.pk, kind.name)

    def lock_on(self, lock, qualifier=None, lock_ttl=None):
        """
        Creates the requested type of flow-level lock
        """
        r = get_redis_connection()
        lock_key = FLOW_LOCK_KEY % (self.org_id, self.pk, lock.name)
        if qualifier:  # pragma: needs cover
            lock_key += (":%s" % qualifier)

        if not lock_ttl:
            lock_ttl = FLOW_LOCK_TTL

        return r.lock(lock_key, lock_ttl)

    def get_node_counts(self, simulation):
        """
        Gets the number of contacts at each node in the flow. For simulator mode this manual counts steps by test
        contacts as these are not pre-calculated.
        """
        if not simulation:
            return FlowNodeCount.get_totals(self)

        # count unique values of current_node_uuid for active runs for test contacts
        totals = (
            self.runs.filter(contact__is_test=True, is_active=True)
            .values('current_node_uuid')
            .annotate(total=Count('current_node_uuid'))
        )

        return {six.text_type(t['current_node_uuid']): t['total'] for t in totals if t['total']}

    def get_segment_counts(self, simulation):
        """
        Gets the number of contacts to have taken each flow segment. For simulator mode this manual counts steps by test
        contacts as these are not pre-calculated.
        """
        if not simulation:
            return FlowPathCount.get_totals(self)

        simulator_runs = self.runs.filter(contact__is_test=True)
        path_counts = defaultdict(int)

        for run in simulator_runs:
            prev_step = None
            for step in run.path:
                if prev_step:
                    exit_uuid = prev_step['exit_uuid']
                    node_uuid = step['node_uuid']
                    path_counts['%s:%s' % (exit_uuid, node_uuid)] += 1

                prev_step = step

        return path_counts

    def get_activity(self, simulation=False):
        """
        Get the activity summary for a flow as a tuple of the number of active runs
        at each step and a map of the previous visits
        """
        return self.get_node_counts(simulation), self.get_segment_counts(simulation)

    def is_starting(self):
        """
        Returns whether this flow has active flow starts
        """
        return self.starts.filter(status__in=(FlowStart.STATUS_STARTING, FlowStart.STATUS_PENDING)).exists()

    def get_localized_text(self, text_translations, contact=None, default_text=''):
        """
        Given a language dict and a preferred language, return the best possible text match
        :param text_translations: The text in all supported languages, or string (which will just return immediately)
        :param contact: the contact we are interacting with
        :param default_text: What to use if all else fails
        :return: the localized text
        """
        org_languages = self.org.get_language_codes()

        # We return according to the following precedence:
        #   1) Contact's language (if it's a valid org language)
        #   2) Org Primary Language
        #   3) Flow Base Language
        #   4) Default Text
        preferred_languages = []

        if contact and contact.language and contact.language in org_languages:
            preferred_languages.append(contact.language)

        if self.org.primary_language:
            preferred_languages.append(self.org.primary_language.iso_code)

        preferred_languages.append(self.base_language)

        return Language.get_localized_text(text_translations, preferred_languages, default_text)

    def import_definition(self, flow_json):
        """
        Allows setting the definition for a flow from another definition.  All uuid's will be
        remmaped accordingly.
        """
        # uuid mappings
        uuid_map = dict()

        def copy_recording(url, path):
            if not url:
                return None

            try:  # pragma: needs cover
                url = "https://%s/%s" % (settings.AWS_BUCKET_DOMAIN, url)
                temp = NamedTemporaryFile(delete=True)
                temp.write(urllib2.urlopen(url).read())
                temp.flush()
                return default_storage.save(path, temp)
            except Exception:  # pragma: needs cover
                # its okay if its no longer there, we'll remove the recording
                return None

        def remap_uuid(json, attribute):
            if attribute in json and json[attribute]:
                uuid = json[attribute]
                new_uuid = uuid_map.get(uuid, None)
                if not new_uuid:
                    new_uuid = str(uuid4())
                    uuid_map[uuid] = new_uuid

                json[attribute] = new_uuid

        remap_uuid(flow_json, 'entry')
        for actionset in flow_json[Flow.ACTION_SETS]:
            remap_uuid(actionset, 'uuid')
            remap_uuid(actionset, 'exit_uuid')
            remap_uuid(actionset, 'destination')

            # for all of our recordings, pull them down and remap
            for action in actionset['actions']:
                if 'recording' in action:
                    # if its a localized
                    if isinstance(action['recording'], dict):
                        for lang, url in six.iteritems(action['recording']):
                            path = copy_recording(url, 'recordings/%d/%d/steps/%s.wav' % (self.org.pk, self.pk, action['uuid']))
                            action['recording'][lang] = path
                    else:
                        path = copy_recording(action['recording'], 'recordings/%d/%d/steps/%s.wav' % (self.org.pk, self.pk, action['uuid']))
                        action['recording'] = path

        for ruleset in flow_json[Flow.RULE_SETS]:
            remap_uuid(ruleset, 'uuid')
            for rule in ruleset.get('rules', []):
                remap_uuid(rule, 'uuid')
                remap_uuid(rule, 'destination')

        # now update with our remapped values
        self.update(flow_json)
        return self

    def archive(self):
        self.is_archived = True
        self.save(update_fields=['is_archived'])

        from .tasks import interrupt_flow_runs_task
        interrupt_flow_runs_task.delay(self.id)

        # archive our triggers as well
        from temba.triggers.models import Trigger
        Trigger.objects.filter(flow=self).update(is_archived=True)

    def restore(self):
        if self.flow_type == Flow.VOICE:  # pragma: needs cover
            if not self.org.supports_ivr():
                raise FlowException("%s requires a Twilio number")

        self.is_archived = False
        self.save(update_fields=['is_archived'])

    def update_single_message_flow(self, translations, base_language):
        if base_language not in translations:  # pragma: no cover
            raise ValueError("Must include translation for base language")

        self.flow_type = Flow.MESSAGE
        self.base_language = base_language
        self.version_number = '10.4'
        self.save(update_fields=('name', 'flow_type', 'base_language', 'version_number'))

        entry_uuid = str(uuid4())
        definition = {
            'version': self.version_number,
            'entry': entry_uuid,
            'base_language': base_language,
            'rule_sets': [],
            'action_sets': [
                {
                    'x': 100, 'y': 0,
                    'uuid': entry_uuid,
                    'exit_uuid': str(uuid4()),
                    'actions': [
                        {'uuid': str(uuid4()), 'type': 'reply', 'msg': translations}
                    ]
                }
            ]
        }

        self.update(FlowRevision.migrate_definition(definition, self))

    def get_steps(self):
        return FlowStep.objects.filter(run__flow=self)

    def get_run_stats(self):
        totals_by_exit = FlowRunCount.get_totals(self)
        total_runs = sum(totals_by_exit.values())

        return {
            'total': total_runs,
            'active': totals_by_exit[FlowRun.STATE_ACTIVE],
            'completed': totals_by_exit[FlowRun.EXIT_TYPE_COMPLETED],
            'expired': totals_by_exit[FlowRun.EXIT_TYPE_EXPIRED],
            'interrupted': totals_by_exit[FlowRun.EXIT_TYPE_INTERRUPTED],
            'completion': int(totals_by_exit[FlowRun.EXIT_TYPE_COMPLETED] * 100 // total_runs) if total_runs else 0
        }

    def build_expressions_context(self, contact, msg, run=None):
        contact_context = contact.build_expressions_context() if contact else dict()

        # our default value
        channel_context = None

        # add our message context
        if msg:
            message_context = msg.build_expressions_context()

            # some fake channel deets for simulation
            if msg.contact.is_test:
                channel_context = Channel.SIMULATOR_CONTEXT
            elif msg.channel:
                channel_context = msg.channel.build_expressions_context()
        else:
            message_context = dict(__default__='')

        # If we still don't know our channel and have a contact, derive the right channel to use
        if not channel_context and contact:
            _contact, contact_urn = Msg.resolve_recipient(self.org, self.created_by, contact, None)

            # only populate channel if this contact can actually be reached (ie, has a URN)
            if contact_urn:
                channel = contact.cached_send_channel(contact_urn=contact_urn)
                if channel:
                    channel_context = channel.build_expressions_context()

        if not run:
            run = self.runs.filter(contact=contact).order_by('-created_on').first()

        if run:
            run.org = self.org
            run.contact = contact

            run_context = run.fields
            flow_context = run.build_expressions_context(contact_context)
        else:
            run_context = {}
            flow_context = {}

        context = dict(flow=flow_context, channel=channel_context, step=message_context, extra=run_context)

        # if we have parent or child contexts, add them in too
        if run:
            run.contact = contact

            if run.parent:
                run.parent.flow.org = self.org
                if run.parent.contact_id == run.contact_id:
                    run.parent.contact = run.contact

                run.parent.org = self.org
                context['parent'] = run.parent.build_expressions_context()

            # see if we spawned any children and add them too
            child_run = run.cached_child
            if child_run:
                child_run.org = self.org
                child_run.contact = run.contact
                context['child'] = child_run.build_expressions_context()

        if contact:
            context['contact'] = contact_context

        return context

    def async_start(self, user, groups, contacts, restart_participants=False, include_active=True):
        """
        Causes us to schedule a flow to start in a background thread.
        """
        from .tasks import start_flow_task

        # create a flow start object
        flow_start = FlowStart.objects.create(flow=self,
                                              restart_participants=restart_participants,
                                              include_active=include_active,
                                              created_by=user, modified_by=user)

        contact_ids = [c.id for c in contacts]
        flow_start.contacts.add(*contact_ids)

        group_ids = [g.id for g in groups]
        flow_start.groups.add(*group_ids)

        on_transaction_commit(lambda: start_flow_task.delay(flow_start.pk))

    def start(self, groups, contacts, restart_participants=False, started_flows=None,
              start_msg=None, extra=None, flow_start=None, parent_run=None, interrupt=True, connection=None, include_active=True):
        """
        Starts a flow for the passed in groups and contacts.
        """
        # build up querysets of our groups for memory efficiency
        if isinstance(groups, QuerySet):  # pragma: no cover
            group_qs = groups
        else:
            group_qs = ContactGroup.all_groups.filter(id__in=[g.id for g in groups])

        # build up querysets of our contacts for memory efficiency
        if isinstance(contacts, QuerySet):  # pragma: no cover
            contact_qs = contacts
        else:
            contact_qs = Contact.objects.filter(id__in=[c.id for c in contacts])

        self.ensure_current_version()

        if started_flows is None:
            started_flows = []

        # prevents infinite loops
        if self.pk in started_flows:  # pragma: needs cover
            return []

        # add this flow to our list of started flows
        started_flows.append(self.pk)

        if not self.entry_uuid:  # pragma: needs cover
            return []

        if start_msg and start_msg.id:
            start_msg.msg_type = FLOW
            start_msg.save(update_fields=['msg_type'])

        all_contact_ids = Contact.all().filter(Q(all_groups__in=group_qs) | Q(pk__in=contact_qs))
        all_contact_ids = all_contact_ids.only('is_test').order_by('pk').values_list('pk', flat=True).distinct('pk')
        if not restart_participants:
            # exclude anybody who has already participated in the flow
            already_started = set(self.runs.all().values_list('contact_id', flat=True))
            all_contact_ids = [contact_id for contact_id in all_contact_ids if contact_id not in already_started]

        if not include_active:
            # exclude anybody who has an active flow run
            already_active = set(FlowRun.objects.filter(is_active=True, org=self.org).values_list('contact_id', flat=True))
            all_contact_ids = [contact_id for contact_id in all_contact_ids if contact_id not in already_active]

        # if we have a parent run, find any parents/grandparents that are active, we'll keep these active
        ancestor_ids = []
        ancestor = parent_run
        while ancestor:
            # we don't consider it an ancestor if it's not current in our start list
            if ancestor.contact.id not in all_contact_ids:
                break
            ancestor_ids.append(ancestor.id)
            ancestor = ancestor.parent

        # for the contacts that will be started, exit any existing flow runs
        for contact_batch in chunk_list(all_contact_ids, 1000):
            active_runs = FlowRun.objects.filter(is_active=True, contact__pk__in=contact_batch).exclude(id__in=ancestor_ids)
            FlowRun.bulk_exit(active_runs, FlowRun.EXIT_TYPE_INTERRUPTED)

        # if we are interrupting parent flow runs, mark them as completed
        if ancestor_ids and interrupt:
            ancestor_runs = FlowRun.objects.filter(id__in=ancestor_ids)
            FlowRun.bulk_exit(ancestor_runs, FlowRun.EXIT_TYPE_COMPLETED)

        contact_count = len(all_contact_ids)

        # update our total flow count on our flow start so we can keep track of when it is finished
        if flow_start:
            flow_start.contact_count = contact_count
            flow_start.save(update_fields=['contact_count'])

        # if there are no contacts to start this flow, then update our status and exit this flow
        if contact_count == 0:
            if flow_start:
                flow_start.update_status()
            return []

        if self.flow_type == Flow.VOICE:
            return self.start_call_flow(all_contact_ids, start_msg=start_msg,
                                        extra=extra, flow_start=flow_start, parent_run=parent_run)

        elif self.flow_type == Flow.USSD:
            return self.start_ussd_flow(all_contact_ids, start_msg=start_msg,
                                        extra=extra, flow_start=flow_start, parent_run=parent_run, connection=connection)
        else:
            return self.start_msg_flow(all_contact_ids,
                                       started_flows=started_flows, start_msg=start_msg,
                                       extra=extra, flow_start=flow_start, parent_run=parent_run)

    def start_ussd_flow(self, all_contact_ids, start_msg=None, extra=None, flow_start=None, parent_run=None, connection=None):
        from temba.ussd.models import USSDSession

        runs = []
        msgs = []

        channel = self.org.get_ussd_channel()

        if not channel or Channel.ROLE_USSD not in channel.role:
            return runs

        for contact_id in all_contact_ids:
            contact = Contact.objects.filter(pk=contact_id, org=self.org).first()
            run = FlowRun.create(self, contact, start=flow_start, parent=parent_run)
            if extra:  # pragma: needs cover
                run.update_fields(extra)

            if run.contact.is_test:  # pragma: no cover
                ActionLog.create(run, '%s has entered the "%s" flow' % (run.contact.get_display(self.org, short=True), run.flow.name))

            # [USSD PUSH] we have to create an outgoing connection for the recipient
            if not connection:
                contact = Contact.objects.filter(pk=contact_id, org=self.org).first()
                contact_urn = contact.get_urn(TEL_SCHEME)
                channel = self.org.get_ussd_channel(contact_urn=contact_urn)

                connection = USSDSession.objects.create(channel=channel, contact=contact, contact_urn=contact_urn,
                                                        org=self.org, direction=USSDSession.USSD_PUSH,
                                                        started_on=timezone.now(), status=USSDSession.INITIATED)

            run.session = connection.get_session()
            run.connection = connection
            run.save(update_fields=['session', 'connection'])

            # if we were started by other connection, save that off
            if parent_run and parent_run.connection:  # pragma: needs cover
                connection.parent = parent_run.connection
                connection.save()
            else:
                entry_rule = RuleSet.objects.filter(flow=self, uuid=self.entry_uuid).first()

                step = self.add_step(run, entry_rule, is_start=True, arrived_on=timezone.now())
                if entry_rule.is_ussd():
                    handled, step_msgs = Flow.handle_destination(entry_rule, step, run, start_msg, trigger_send=False, continue_parent=False)

                    # add these messages as ones that are ready to send
                    for msg in step_msgs:
                        msgs.append(msg)

            run.start_msgs = [start_msg]

            runs.append(run)

        # trigger our messages to be sent
        if msgs and not parent_run:
            # then send them off
            msgs.sort(key=lambda message: (message.contact_id, message.created_on))
            Msg.objects.filter(id__in=[m.id for m in msgs]).update(status=PENDING)

            # trigger a sync
            self.org.trigger_send(msgs)

        if flow_start:  # pragma: needs cover
            flow_start.update_status()

        if start_msg:
            Msg.mark_handled(start_msg)

        return runs

    def start_call_flow(self, all_contact_ids, start_msg=None, extra=None, flow_start=None, parent_run=None):
        from temba.ivr.models import IVRCall
        runs = []
        channel = self.org.get_call_channel()

        if not channel or Channel.ROLE_CALL not in channel.role:  # pragma: needs cover
            return runs

        for contact_id in all_contact_ids:
            contact = Contact.objects.filter(pk=contact_id, org=channel.org).first()
            contact_urn = contact.get_urn(TEL_SCHEME)
            channel = self.org.get_call_channel(contact_urn=contact_urn)

            # can't reach this contact, move on
            if not contact or not contact_urn or not channel:  # pragma: no cover
                continue

            run = FlowRun.create(self, contact, start=flow_start, parent=parent_run)
            if extra:  # pragma: needs cover
                run.update_fields(extra)

            # create our call objects
            if parent_run and parent_run.connection:
                call = parent_run.connection
                session = parent_run.session
            else:
                call = IVRCall.create_outgoing(channel, contact, contact_urn, self.created_by)
                session = FlowSession.create(contact, connection=call)

            # save away our created call
            run.session = session
            run.connection = call
            run.save(update_fields=['connection'])

            if not parent_run or not parent_run.connection:
                # trigger the call to start (in the background)
                IVRCall.objects.get(id=call.id).start_call()

            # no start msgs in call flows but we want the variable there
            run.start_msgs = []

            runs.append(run)

        if flow_start:  # pragma: needs cover
            flow_start.update_status()

        return runs

    def start_msg_flow(self, all_contact_ids, started_flows=None, start_msg=None, extra=None,
                       flow_start=None, parent_run=None):

        # only use flowserver if flow supports it, message is an actual message, and parent wasn't run in old engine
        if self.use_flow_server() and not (start_msg and not start_msg.contact_urn) and not parent_run:
            contacts = Contact.objects.filter(id__in=all_contact_ids).order_by('id')
            runs = FlowSession.bulk_start(contacts, self, msg_in=start_msg, params=extra)
            if flow_start:
                flow_start.runs.add(*runs)
                flow_start.update_status()

            return runs

        start_msg_id = start_msg.id if start_msg else None
        flow_start_id = flow_start.id if flow_start else None

        if started_flows is None:
            started_flows = []

        # for each send action, we need to create a broadcast, we'll group our created messages under these
        broadcasts = []

        if len(all_contact_ids) > 1:

            # create the broadcast for this flow
            send_actions = self.get_entry_send_actions()

            for send_action in send_actions:
                # check that we either have text or media, available for the base language
                if (send_action.msg and send_action.msg.get(self.base_language)) or (send_action.media and send_action.media.get(self.base_language)):

                    broadcast = Broadcast.create(self.org, self.created_by, send_action.msg, [],
                                                 media=send_action.media,
                                                 base_language=self.base_language,
                                                 send_all=send_action.send_all,
                                                 quick_replies=send_action.quick_replies)
                    broadcast.update_contacts(all_contact_ids)

                    # manually set our broadcast status to QUEUED, our sub processes will send things off for us
                    broadcast.status = QUEUED
                    broadcast.save(update_fields=['status'])

                    # add it to the list of broadcasts in this flow start
                    broadcasts.append(broadcast)

        # if there are fewer contacts than our batch size, do it immediately
        if len(all_contact_ids) < START_FLOW_BATCH_SIZE:
            return self.start_msg_flow_batch(all_contact_ids, broadcasts=broadcasts, started_flows=started_flows,
                                             start_msg=start_msg, extra=extra, flow_start=flow_start,
                                             parent_run=parent_run)

        # otherwise, create batches instead
        else:
            # for all our contacts, build up start sms batches
            task_context = dict(contacts=[], flow=self.pk, flow_start=flow_start_id,
                                started_flows=started_flows, broadcasts=[b.id for b in broadcasts], start_msg=start_msg_id, extra=extra)

            batch_contacts = task_context['contacts']
            for contact_id in all_contact_ids:
                batch_contacts.append(contact_id)

                if len(batch_contacts) >= START_FLOW_BATCH_SIZE:
                    print("Starting flow '%s' for batch of %d contacts" % (self.name, len(task_context['contacts'])))
                    push_task(self.org, 'flows', Flow.START_MSG_FLOW_BATCH, task_context)
                    batch_contacts = []
                    task_context['contacts'] = batch_contacts

            if batch_contacts:
                print("Starting flow '%s' for batch of %d contacts" % (self.name, len(task_context['contacts'])))
                push_task(self.org, 'flows', Flow.START_MSG_FLOW_BATCH, task_context)

            return []

    def start_msg_flow_batch(self, batch_contact_ids, broadcasts, started_flows, start_msg=None,
                             extra=None, flow_start=None, parent_run=None):

        batch_contacts = Contact.objects.filter(id__in=batch_contact_ids)
        Contact.bulk_cache_initialize(self.org, batch_contacts)
        contact_map = {c.id: c for c in batch_contacts}

        simulation = len(batch_contacts) == 1 and batch_contacts[0].is_test

        # these fields are the initial state for our flow run
        run_fields = {}  # this should be the default value of the FlowRun.fields
        if extra:
            # we keep more values in @extra for new flow runs because we might be passing the state
            (normalized_fields, count) = FlowRun.normalize_fields(extra, settings.FLOWRUN_FIELDS_SIZE * 4)
            run_fields = normalized_fields

        # create all our flow runs for this set of contacts at once
        batch = []
        now = timezone.now()

        for contact_id in batch_contact_ids:
            contact = contact_map[contact_id]
            run = FlowRun.create(self, contact, fields=run_fields, start=flow_start, created_on=now,
                                 parent=parent_run, db_insert=False, responded=start_msg is not None)
            batch.append(run)

        runs = FlowRun.objects.bulk_create(batch)

        # build a map of contact to flow run
        run_map = dict()
        for run in runs:
            run.flow = self
            run.org = self.org

            run_map[run.contact_id] = run
            if run.contact.is_test:
                ActionLog.create(run, '%s has entered the "%s" flow' % (run.contact.get_display(self.org, short=True), run.flow.name))

        # update our expiration date on our runs, we do this by calculating it on one run then updating all others
        run.update_expiration(timezone.now())

        # if we have more than one run, update the others to the same expiration
        if len(run_map) > 1:
            FlowRun.objects.filter(id__in=[r.id for r in runs]).update(expires_on=run.expires_on, modified_on=timezone.now())

        # if we have some broadcasts to optimize for
        message_map = dict()
        if broadcasts:
            # create our expressions context
            expressions_context_base = self.build_expressions_context(None, start_msg)
            if extra:
                expressions_context_base['extra'] = extra

            # and add each contact and message to each broadcast
            for broadcast in broadcasts:
                broadcast.org = self.org
                # provide the broadcast with a partial recipient list
                partial_recipients = list(), batch_contacts

                # create the sms messages
                created_on = timezone.now()
                broadcast.send(expressions_context=expressions_context_base, trigger_send=False,
                               response_to=start_msg, status=INITIALIZING, msg_type=FLOW, created_on=created_on,
                               partial_recipients=partial_recipients, run_map=run_map)

                # map all the messages we just created back to our contact
                for msg in Msg.objects.filter(broadcast=broadcast, created_on=created_on).select_related('channel'):
                    msg.broadcast = broadcast
                    if msg.contact_id not in message_map:
                        message_map[msg.contact_id] = [msg]
                    else:  # pragma: needs cover
                        message_map[msg.contact_id].append(msg)

        # now execute our actual flow steps
        (entry_actions, entry_rules) = (None, None)
        if self.entry_type == Flow.ACTIONS_ENTRY:
            entry_actions = ActionSet.objects.filter(uuid=self.entry_uuid).first()
            if entry_actions:
                entry_actions.flow = self

        elif self.entry_type == Flow.RULES_ENTRY:
            entry_rules = RuleSet.objects.filter(flow=self, uuid=self.entry_uuid).first()
            if entry_rules:
                entry_rules.flow = self

        msgs = []
        optimize_sending_action = len(broadcasts) > 0

        for run in runs:
            contact = run.contact

            # each contact maintains its own list of started flows
            started_flows_by_contact = list(started_flows)
            run_msgs = message_map.get(contact.id, [])
            arrived_on = timezone.now()

            try:
                if entry_actions:
                    run_msgs += entry_actions.execute_actions(run, start_msg, started_flows_by_contact,
                                                              skip_leading_reply_actions=not optimize_sending_action)

                    step = self.add_step(run, entry_actions, run_msgs, is_start=True, arrived_on=arrived_on)

                    # and onto the destination
                    if entry_actions.destination:
                        destination = Flow.get_node(entry_actions.flow,
                                                    entry_actions.destination,
                                                    entry_actions.destination_type)

                        next_step = self.add_step(run, destination, previous_step=step, exit_uuid=entry_actions.exit_uuid)

                        msg = Msg(org=self.org, contact=contact, text='', id=0)
                        handled, step_msgs = Flow.handle_destination(destination, next_step, run, msg, started_flows_by_contact,
                                                                     is_test_contact=simulation, trigger_send=False, continue_parent=False)
                        run_msgs += step_msgs

                    else:
                        run.set_completed(final_step=step)

                elif entry_rules:
                    step = self.add_step(run, entry_rules, run_msgs, is_start=True, arrived_on=arrived_on)

                    # if we have a start message, go and handle the rule
                    if start_msg:
                        Flow.find_and_handle(start_msg, started_flows_by_contact, triggered_start=True)

                    # if we didn't get an incoming message, see if we need to evaluate it passively
                    elif not entry_rules.is_pause():
                        # create an empty placeholder message
                        msg = Msg(org=self.org, contact=contact, text='', id=0)
                        handled, step_msgs = Flow.handle_destination(entry_rules, step, run, msg, started_flows_by_contact, trigger_send=False, continue_parent=False)
                        run_msgs += step_msgs

                if start_msg:
                    run.add_messages([start_msg], step=step)

                # set the msgs that were sent by this run so that any caller can deal with them
                run.start_msgs = run_msgs

                # add these messages as ones that are ready to send
                for msg in run_msgs:
                    msgs.append(msg)

            except Exception:
                logger.error('Failed starting flow %d for contact %d' % (self.id, contact.id), exc_info=1, extra={'stack': True})

                # mark this flow as interrupted
                run.set_interrupted()

                # mark our messages as failed
                Msg.objects.filter(id__in=[m.id for m in run_msgs]).update(status=FAILED)

                # remove our msgs from our parent's concerns
                run.start_msgs = []

        # trigger our messages to be sent
        if msgs and not parent_run:
            # then send them off
            msgs.sort(key=lambda message: (message.contact_id, message.created_on))
            Msg.objects.filter(id__in=[m.id for m in msgs]).update(status=PENDING)

            # trigger a sync
            self.org.trigger_send(msgs)

        # if we have a flow start, check whether we are complete
        if flow_start:
            flow_start.update_status()

        return runs

    def add_step(self, run, node, msgs=None, exit_uuid=None, is_start=False, previous_step=None, arrived_on=None):
        if msgs is None:
            msgs = []

        if not arrived_on:
            arrived_on = timezone.now()

        if previous_step:
            previous_step.left_on = arrived_on
            previous_step.rule_uuid = exit_uuid
            previous_step.next_uuid = node.uuid
            previous_step.save(update_fields=('left_on', 'rule_uuid', 'next_uuid'))

        # update our timeouts
        timeout = node.get_timeout() if isinstance(node, RuleSet) else None
        run.update_timeout(arrived_on, timeout)

        if not is_start:
            # mark any other states for this contact as evaluated, contacts can only be in one place at time
            self.get_steps().filter(run=run, left_on=None).update(left_on=arrived_on, next_uuid=node.uuid,
                                                                  rule_uuid=exit_uuid)

        # then add our new step and associate it with our message
        step = FlowStep.objects.create(run=run, contact=run.contact, step_type=node.get_step_type(),
                                       step_uuid=node.uuid, arrived_on=arrived_on)

        # for each message, associate it with this step and set the label on it
        run.add_messages(msgs, step=step)

        path = run.path

        # complete previous step
        if path and exit_uuid:
            path[-1][FlowRun.PATH_EXIT_UUID] = exit_uuid

        # create new step
        path.append({FlowRun.PATH_NODE_UUID: node.uuid, FlowRun.PATH_ARRIVED_ON: arrived_on.isoformat()})

        # trim path to ensure it can't grow indefinitely
        if len(path) > FlowRun.PATH_MAX_STEPS:
            path = path[len(path) - FlowRun.PATH_MAX_STEPS:]

        run.path = path
        run.current_node_uuid = path[-1][FlowRun.PATH_NODE_UUID]
        run.save(update_fields=('path', 'current_node_uuid'))

        return step

    def get_entry_send_actions(self):
        """
        Returns all the entry actions (the first actions in a flow) that are reply actions. This is used
        for grouping all our outgoing messages into a single Broadcast.
        """
        if not self.entry_uuid or self.entry_type != Flow.ACTIONS_ENTRY:
            return []

        # get our entry actions
        entry_actions = ActionSet.objects.filter(uuid=self.entry_uuid).first()
        send_actions = []

        if entry_actions:
            actions = entry_actions.get_actions()

            for action in actions:
                # if this isn't a reply action, bail, they might be modifying the contact
                if not isinstance(action, ReplyAction):
                    break

                send_actions.append(action)

        return send_actions

    def get_dependencies(self, flow_map=None):
        from temba.contacts.models import ContactGroup

        # need to make sure we have the latest version to inspect dependencies
        self.ensure_current_version()

        dependencies = set()

        # find all the flows we reference, note this won't include archived flows
        for action_set in self.action_sets.all():
            for action in action_set.get_actions():
                if hasattr(action, 'flow'):
                    dependencies.add(action.flow)
                if hasattr(action, 'groups'):
                    for group in action.groups:
                        if isinstance(group, ContactGroup):
                            dependencies.add(group)

        for ruleset in self.rule_sets.all():
            if ruleset.ruleset_type == RuleSet.TYPE_SUBFLOW:
                flow_uuid = ruleset.config['flow']['uuid']
                flow = flow_map.get(flow_uuid) if flow_map else Flow.objects.filter(uuid=flow_uuid).first()
                if flow:
                    dependencies.add(flow)

        return dependencies

    def as_json(self, expand_contacts=False):
        """
        Returns the JSON definition for this flow.

          expand_contacts:
            Add names for contacts and groups that are just ids. This is useful for human readable
            situations such as the flow editor.

        """
        flow = dict()

        if self.entry_uuid:
            flow[Flow.ENTRY] = self.entry_uuid
        else:
            flow[Flow.ENTRY] = None

        actionsets = []
        for actionset in ActionSet.objects.filter(flow=self).order_by('pk'):
            actionsets.append(actionset.as_json())

        def lookup_action_contacts(action, contacts, groups):

            if 'contact' in action:  # pragma: needs cover
                contacts.append(action['contact']['uuid'])

            if 'contacts' in action:
                for contact in action['contacts']:
                    contacts.append(contact['uuid'])

            if 'group' in action:  # pragma: needs cover
                g = action['group']
                if isinstance(g, dict):
                    if 'uuid' in g:
                        groups.append(g['uuid'])

            if 'groups' in action:
                for group in action['groups']:
                    if isinstance(group, dict):
                        if 'uuid' in group:
                            groups.append(group['uuid'])

        def replace_action_contacts(action, contacts, groups):

            if 'contact' in action:  # pragma: needs cover
                contact = contacts.get(action['contact']['uuid'], None)
                if contact:
                    action['contact'] = contact.as_json()

            if 'contacts' in action:
                expanded_contacts = []
                for contact in action['contacts']:
                    contact = contacts.get(contact['uuid'], None)
                    if contact:
                        expanded_contacts.append(contact.as_json())

                action['contacts'] = expanded_contacts

            if 'group' in action:  # pragma: needs cover
                # variable substitution
                group = action['group']
                if isinstance(group, dict):
                    if 'uuid' in group:
                        group = groups.get(group['uuid'], None)
                        if group:
                            action['group'] = dict(uuid=group.uuid, name=group.name)

            if 'groups' in action:
                expanded_groups = []
                for group in action['groups']:

                    # variable substitution
                    if not isinstance(group, dict):
                        expanded_groups.append(group)
                    else:
                        group_instance = groups.get(group['uuid'], None)
                        if group_instance:
                            expanded_groups.append(dict(uuid=group_instance.uuid, name=group_instance.name))
                        else:
                            expanded_groups.append(group)

                action['groups'] = expanded_groups

        if expand_contacts:
            groups = []
            contacts = []

            for actionset in actionsets:
                for action in actionset['actions']:
                    lookup_action_contacts(action, contacts, groups)

            # load them all
            contacts = dict((_.uuid, _) for _ in Contact.all().filter(org=self.org, uuid__in=contacts))
            groups = dict((_.uuid, _) for _ in ContactGroup.user_groups.filter(org=self.org, uuid__in=groups))

            # and replace them
            for actionset in actionsets:
                for action in actionset['actions']:
                    replace_action_contacts(action, contacts, groups)

        flow[Flow.ACTION_SETS] = actionsets

        # add in our rulesets
        rulesets = []
        for ruleset in RuleSet.objects.filter(flow=self).order_by('pk'):
            rulesets.append(ruleset.as_json())
        flow[Flow.RULE_SETS] = rulesets

        # required flow running details
        flow[Flow.BASE_LANGUAGE] = self.base_language
        flow[Flow.FLOW_TYPE] = self.flow_type
        flow[Flow.VERSION] = get_current_export_version()
        flow[Flow.METADATA] = self.get_metadata()
        return flow

    def get_metadata(self):

        metadata = dict()
        if self.metadata:
            metadata = self.metadata

        revision = self.revisions.all().order_by('-revision').first()

        last_saved = self.saved_on
        if self.saved_by == get_flow_user(self.org):
            last_saved = self.modified_on

        metadata[Flow.NAME] = self.name
        metadata[Flow.SAVED_ON] = datetime_to_str(last_saved)
        metadata[Flow.REVISION] = revision.revision if revision else 1
        metadata[Flow.UUID] = self.uuid
        metadata[Flow.EXPIRES] = self.expires_after_minutes

        return metadata

    @classmethod
    def detect_invalid_cycles(cls, json_dict):
        """
        Checks for invalid cycles in our flow
        :param json_dict: our flow definition
        :return: invalid cycle path as list of uuids if found, otherwise empty list
        """

        # Adapted from a blog post by Guido:
        # http://neopythonic.blogspot.com/2009/01/detecting-cycles-in-directed-graph.html

        # Maintain path as a a depth-first path in the implicit tree;
        # path is represented as an OrderedDict of {node: [child,...]} pairs.

        nodes = list()
        node_map = {}

        for ruleset in json_dict.get(Flow.RULE_SETS, []):
            nodes.append(ruleset.get('uuid'))
            node_map[ruleset.get('uuid')] = ruleset

        for actionset in json_dict.get(Flow.ACTION_SETS, []):
            nodes.append(actionset.get('uuid'))
            node_map[actionset.get('uuid')] = actionset

        def get_destinations(uuid):
            node = node_map.get(uuid)

            if not node:  # pragma: needs cover
                return []

            rules = node.get('rules', [])
            destinations = []
            if rules:

                if node.get('ruleset_type', None) in RuleSet.TYPE_WAIT:
                    return []

                for rule in rules:
                    if rule.get('destination'):
                        destinations.append(rule.get('destination'))

            elif node.get('destination'):
                destinations.append(node.get('destination'))
            return destinations

        while nodes:
            root = nodes.pop()
            path = OrderedDict({root: get_destinations(root)})
            while path:
                # children at the fringe of the tree
                children = path[next(reversed(path))]
                while children:
                    child = children.pop()

                    # found a loop
                    if child in path:
                        pathlist = list(path)
                        return pathlist[pathlist.index(child):] + [child]

                    # new path
                    if child in nodes:
                        path[child] = get_destinations(child)
                        nodes.remove(child)
                        break
                else:
                    # no more children; pop back up a level
                    path.popitem()
        return None

    def ensure_current_version(self, min_version=None):
        """
        Makes sure the flow is at the current version. If it isn't it will
        migrate the definition forward updating the flow accordingly.
        """
        to_version = min_version or get_current_export_version()

        if Flow.is_before_version(self.version_number, to_version):
            with self.lock_on(FlowLock.definition):
                revision = self.revisions.all().order_by('-revision').all().first()
                if revision:
                    json_flow = revision.get_definition_json()
                else:  # pragma: needs cover
                    json_flow = self.as_json()

                self.update(json_flow, user=get_flow_user(self.org))
                self.refresh_from_db()

    def update(self, json_dict, user=None, force=False):
        """
        Updates a definition for a flow and returns the new revision
        """

        def get_step_type(dest, rulesets, actionsets):
            if dest:
                if rulesets.get(dest, None):
                    return FlowStep.TYPE_RULE_SET
                if actionsets.get(dest, None):
                    return FlowStep.TYPE_ACTION_SET
            return None

        cycle_node_uuids = Flow.detect_invalid_cycles(json_dict)
        if cycle_node_uuids:
            raise FlowInvalidCycleException(cycle_node_uuids)

        try:
            # make sure the flow version hasn't changed out from under us
            if json_dict.get(Flow.VERSION) != get_current_export_version():
                raise FlowVersionConflictException(json_dict.get(Flow.VERSION))

            flow_user = get_flow_user(self.org)
            # check whether the flow has changed since this flow was last saved
            if user and not force:
                saved_on = json_dict.get(Flow.METADATA, {}).get(Flow.SAVED_ON, None)
                org = user.get_org()

                # check our last save if we aren't the system flow user
                if user != flow_user:
                    migrated = self.saved_by == flow_user
                    last_save = self.saved_on

                    # use modified on if it was a migration
                    if migrated:
                        last_save = self.modified_on

                    if not saved_on or str_to_datetime(saved_on, org.timezone) < last_save:
                        raise FlowUserConflictException(self.saved_by, last_save)

            top_y = 0
            top_uuid = None

            # load all existing objects into dicts by uuid
            existing_actionsets = {actionset.uuid: actionset for actionset in self.action_sets.all()}
            existing_rulesets = {ruleset.uuid: ruleset for ruleset in self.rule_sets.all()}

            # set of uuids which we've seen, we use this set to remove objects no longer used in this flow
            seen = set()
            destinations = set()

            # our steps in our current update submission
            current_actionsets = {}
            current_rulesets = {}

            # parse our actions
            for actionset in json_dict.get(Flow.ACTION_SETS, []):

                uuid = actionset.get(Flow.UUID)

                # validate our actions, normalizing them as JSON after reading them
                actions = [_.as_json() for _ in Action.from_json_array(self.org, actionset.get(Flow.ACTIONS))]

                if actions:
                    current_actionsets[uuid] = actions

            for ruleset in json_dict.get(Flow.RULE_SETS, []):
                uuid = ruleset.get(Flow.UUID)
                current_rulesets[uuid] = ruleset
                seen.add(uuid)

            # create all our rule sets
            for ruleset in json_dict.get(Flow.RULE_SETS, []):

                uuid = ruleset.get(Flow.UUID)
                rules = ruleset.get(Flow.RULES)
                label = ruleset.get(Flow.LABEL, None)
                operand = ruleset.get(Flow.OPERAND, None)
                finished_key = ruleset.get(Flow.FINISHED_KEY)
                ruleset_type = ruleset.get(Flow.RULESET_TYPE)
                config = ruleset.get(Flow.CONFIG)

                if not config:
                    config = dict()

                # cap our lengths
                if label:
                    label = label[:64]

                if operand:
                    operand = operand[:128]

                (x, y) = (ruleset.get(Flow.X), ruleset.get(Flow.Y))

                if not top_uuid or y < top_y:
                    top_y = y
                    top_uuid = uuid

                # parse our rules, this will materialize any necessary dependencies
                parsed_rules = []
                rule_objects = Rule.from_json_array(self.org, rules)
                for r in rule_objects:
                    parsed_rules.append(r.as_json())
                rules = parsed_rules

                for rule in rules:
                    if 'destination' in rule:
                        # if the destination was excluded for not having any actions
                        # remove the connection for our rule too
                        if rule['destination'] not in current_actionsets and rule['destination'] not in seen:
                            rule['destination'] = None
                        else:
                            destination_uuid = rule.get('destination', None)
                            destinations.add(destination_uuid)

                            # determine what kind of destination we are pointing to
                            rule['destination_type'] = get_step_type(destination_uuid,
                                                                     current_rulesets, current_actionsets)

                            # print "Setting destination [%s] type to: %s" % (destination_uuid, rule['destination_type'])

                existing = existing_rulesets.get(uuid, None)

                if existing:
                    existing.label = ruleset.get(Flow.LABEL, None)
                    existing.rules = rules
                    existing.operand = operand
                    existing.label = label
                    existing.finished_key = finished_key
                    existing.ruleset_type = ruleset_type
                    existing.config = config
                    (existing.x, existing.y) = (x, y)
                    existing.save()
                else:

                    existing = RuleSet.objects.create(flow=self,
                                                      uuid=uuid,
                                                      label=label,
                                                      rules=rules,
                                                      finished_key=finished_key,
                                                      ruleset_type=ruleset_type,
                                                      operand=operand,
                                                      config=config,
                                                      x=x, y=y)

                existing_rulesets[uuid] = existing

                # update our value type based on our new rules
                existing.value_type = existing.get_value_type()
                RuleSet.objects.filter(pk=existing.pk).update(value_type=existing.value_type)

            # now work through our action sets
            for actionset in json_dict.get(Flow.ACTION_SETS, []):
                uuid = actionset.get(Flow.UUID)
                exit_uuid = actionset.get(Flow.EXIT_UUID)

                # skip actionsets without any actions. This happens when there are no valid
                # actions in an actionset such as when deleted groups or flows are the only actions
                if uuid not in current_actionsets:
                    continue

                actions = current_actionsets[uuid]
                seen.add(uuid)

                (x, y) = (actionset.get(Flow.X), actionset.get(Flow.Y))

                if not top_uuid or y < top_y:
                    top_y = y
                    top_uuid = uuid

                existing = existing_actionsets.get(uuid, None)

                # lookup our destination
                destination_uuid = actionset.get('destination')
                destination_type = get_step_type(destination_uuid, current_rulesets, current_actionsets)

                if destination_uuid:
                    if not destination_type:
                        destination_uuid = None

                # only create actionsets if there are actions
                if actions:
                    if existing:
                        # print "Updating %s to point to %s" % (unicode(actions), destination_uuid)
                        existing.destination = destination_uuid
                        existing.destination_type = destination_type
                        existing.exit_uuid = exit_uuid
                        existing.actions = actions
                        (existing.x, existing.y) = (x, y)
                        existing.save()
                    else:
                        existing = ActionSet.objects.create(flow=self,
                                                            uuid=uuid,
                                                            destination=destination_uuid,
                                                            destination_type=destination_type,
                                                            exit_uuid=exit_uuid,
                                                            actions=actions,
                                                            x=x, y=y)

                        existing_actionsets[uuid] = existing

            existing_actionsets_to_delete = set()
            seen_existing_actionsets = {}

            # now work through all our objects once more, making sure all uuids map appropriately
            for uuid, actionset in existing_actionsets.items():
                if uuid not in seen:
                    existing_actionsets_to_delete.add(uuid)
                else:
                    seen_existing_actionsets[uuid] = actionset

            # delete actionset which are not seen
            ActionSet.objects.filter(uuid__in=existing_actionsets_to_delete).delete()

            existing_actionsets = seen_existing_actionsets

            existing_rulesets_to_delete = set()
            seen_existing_rulesets = {}

            for uuid, ruleset in existing_rulesets.items():
                if uuid not in seen:
                    existing_rulesets_to_delete.add(uuid)

                    # instead of deleting it, make it a phantom ruleset until we do away with values_value
                    ruleset.flow = None
                    ruleset.uuid = str(uuid4())
                    ruleset.save(update_fields=('flow', 'uuid'))
                else:
                    seen_existing_rulesets[uuid] = ruleset

            existing_rulesets = seen_existing_rulesets

            # make sure all destinations are present though
            for destination in destinations:
                if destination not in existing_rulesets and destination not in existing_actionsets:  # pragma: needs cover
                    raise FlowException("Invalid destination: '%s', no matching actionset or ruleset" % destination)

            entry = json_dict.get('entry', None)

            # check if we are pointing to a destination that is no longer valid
            if entry not in existing_rulesets and entry not in existing_actionsets:
                entry = None

            if not entry and top_uuid:
                entry = top_uuid

            # set our entry
            if entry in existing_actionsets:
                self.entry_uuid = entry
                self.entry_type = Flow.ACTIONS_ENTRY
            elif entry in existing_rulesets:
                self.entry_uuid = entry
                self.entry_type = Flow.RULES_ENTRY

            # if we have a base language, set that
            self.base_language = json_dict.get('base_language', None)

            # set our metadata
            self.metadata = None
            if Flow.METADATA in json_dict:
                self.metadata = json_dict[Flow.METADATA]

            if user:
                self.saved_by = user

            # if it's our migration user, don't update saved on
            if user and user != flow_user:
                self.saved_on = timezone.now()

            self.version_number = get_current_export_version()
            self.save()

            # clear property cache
            self.clear_props_cache()

            # create a version of our flow for posterity
            if user is None:
                user = self.created_by

            # last version
            revision_num = 1
            last_revision = self.revisions.order_by('-revision').first()
            if last_revision:
                revision_num = last_revision.revision + 1

            # create a new version
            revision = self.revisions.create(definition=json_dict,
                                             created_by=user,
                                             modified_by=user,
                                             spec_version=get_current_export_version(),
                                             revision=revision_num)

            self.update_dependencies()

        except FlowUserConflictException as e:
            raise e
        except Exception as e:
            # user will see an error in the editor but log exception so we know we got something to fix
            logger.exception(six.text_type(e))
            traceback.print_exc()
            raise e

        return revision

    def update_dependencies(self):

        # if we are an older version, induce a system rev which will update our dependencies
        if Flow.is_before_version(self.version_number, get_current_export_version()):
            self.ensure_current_version()
            return

        # otherwise, go about updating our dependencies assuming a current flow
        groups = set()
        flows = set()
        collector = ContactFieldCollector()

        # find any references in our actions
        fields = set()
        for actionset in self.action_sets.all():
            for action in actionset.get_actions():
                if action.TYPE in (AddToGroupAction.TYPE, DeleteFromGroupAction.TYPE):
                    # iterate over them so we can type crack to ignore expression strings :(
                    for group in action.groups:
                        if isinstance(group, ContactGroup):
                            groups.add(group)
                        else:
                            # group names can be an expression
                            fields.update(collector.get_contact_fields(group))

                if action.TYPE == TriggerFlowAction.TYPE:
                    flows.add(action.flow)
                    for recipient in action.variables:
                        fields.update(collector.get_contact_fields(recipient))

                if action.TYPE in ('reply', 'send', 'say'):
                    for lang, msg in six.iteritems(action.msg):
                        fields.update(collector.get_contact_fields(msg))

                    if hasattr(action, 'media'):
                        for lang, text in six.iteritems(action.media):
                            fields.update(collector.get_contact_fields(text))

                    if hasattr(action, 'variables'):
                        for recipient in action.variables:
                            fields.update(collector.get_contact_fields(recipient))

                if action.TYPE == 'email':
                    fields.update(collector.get_contact_fields(action.subject))
                    fields.update(collector.get_contact_fields(action.message))

                if action.TYPE == 'save':
                    fields.add(action.field)
                    fields.update(collector.get_contact_fields(action.value))

                # voice recordings
                if action.TYPE == 'play':
                    fields.update(collector.get_contact_fields(action.url))

        # find references in our rulesets
        for ruleset in self.rule_sets.all():
            if ruleset.ruleset_type == RuleSet.TYPE_SUBFLOW:
                flow_uuid = ruleset.config.get('flow').get('uuid')
                flow = Flow.objects.filter(org=self.org, uuid=flow_uuid).first()
                if flow:
                    flows.add(flow)
            elif ruleset.ruleset_type == RuleSet.TYPE_WEBHOOK:
                webhook_url = ruleset.config.get('webhook')
                fields.update(collector.get_contact_fields(webhook_url))
            else:
                # check our operand for expressions
                fields.update(collector.get_contact_fields(ruleset.operand))

                # check all the rules and their localizations
                rules = ruleset.get_rules()

                for rule in rules:
                    if hasattr(rule.test, 'test'):
                        if type(rule.test.test) == dict:
                            for lang, text in six.iteritems(rule.test.test):
                                fields.update(collector.get_contact_fields(text))
                        # voice rules are not localized
                        elif isinstance(rule.test.test, six.string_types):
                            fields.update(collector.get_contact_fields(rule.test.test))
                    if isinstance(rule.test, InGroupTest):
                        groups.add(rule.test.group)

        if len(fields):
            existing = ContactField.objects.filter(org=self.org, key__in=fields).values_list('key')

            # create any field that doesn't already exist
            for field in fields:
                if ContactField.is_valid_key(field) and field not in existing:
                    # reverse slug to get a reasonable label
                    label = ' '.join([word.capitalize() for word in field.split('_')])
                    ContactField.get_or_create(self.org, self.modified_by, field, label)

        fields = ContactField.objects.filter(org=self.org, key__in=fields)

        self.group_dependencies.clear()
        self.group_dependencies.add(*groups)

        self.flow_dependencies.clear()
        self.flow_dependencies.add(*flows)

        self.field_dependencies.clear()
        self.field_dependencies.add(*fields)

    def __str__(self):
        return self.name

    class Meta:
        ordering = ('-modified_on',)


class FlowRun(RequireUpdateFieldsMixin, models.Model):
    STATE_ACTIVE = 'A'

    EXIT_TYPE_COMPLETED = 'C'
    EXIT_TYPE_INTERRUPTED = 'I'
    EXIT_TYPE_EXPIRED = 'E'
    EXIT_TYPE_CHOICES = ((EXIT_TYPE_COMPLETED, _("Completed")),
                         (EXIT_TYPE_INTERRUPTED, _("Interrupted")),
                         (EXIT_TYPE_EXPIRED, _("Expired")))

    INVALID_EXTRA_KEY_CHARS = regex.compile(r'[^a-zA-Z0-9_]')

    RESULT_NAME = 'name'
    RESULT_NODE_UUID = 'node_uuid'
    RESULT_CATEGORY = 'category'
    RESULT_CATEGORY_LOCALIZED = 'category_localized'
    RESULT_VALUE = 'value'
    RESULT_INPUT = 'input'
    RESULT_CREATED_ON = 'created_on'

    PATH_NODE_UUID = 'node_uuid'
    PATH_ARRIVED_ON = 'arrived_on'
    PATH_EXIT_UUID = 'exit_uuid'
    PATH_MAX_STEPS = 100

    uuid = models.UUIDField(unique=True, default=uuid4)

    org = models.ForeignKey(Org, related_name='runs', db_index=False)

    flow = models.ForeignKey(Flow, related_name='runs')

    contact = models.ForeignKey(Contact, related_name='runs')

    session = models.ForeignKey(FlowSession, related_name='runs', null=True,
                                help_text=_("The session that handled this flow run, only for voice flows"))

    connection = models.ForeignKey('channels.ChannelSession', related_name='runs', null=True, blank=True,
                                   help_text=_("The session that handled this flow run, only for voice flows"))

    is_active = models.BooleanField(default=True,
                                    help_text=_("Whether this flow run is currently active"))

    fields = JSONAsTextField(blank=True, null=True, object_pairs_hook=OrderedDict, default=dict,
                             help_text=_("A JSON representation of any custom flow values the user has saved away"))

    created_on = models.DateTimeField(default=timezone.now,
                                      help_text=_("When this flow run was created"))

    modified_on = models.DateTimeField(auto_now=True,
                                       help_text=_("When this flow run was last updated"))

    exited_on = models.DateTimeField(null=True,
                                     help_text=_("When the contact exited this flow run"))

    exit_type = models.CharField(null=True, max_length=1, choices=EXIT_TYPE_CHOICES,
                                 help_text=_("Why the contact exited this flow run"))

    expires_on = models.DateTimeField(null=True,
                                      help_text=_("When this flow run will expire"))

    timeout_on = models.DateTimeField(null=True,
                                      help_text=_("When this flow will next time out (if any)"))

    responded = models.BooleanField(default=False, help_text='Whether contact has responded in this run')

    start = models.ForeignKey('flows.FlowStart', null=True, blank=True, related_name='runs',
                              help_text=_("The FlowStart objects that started this run"))

    submitted_by = models.ForeignKey(settings.AUTH_USER_MODEL, null=True, db_index=False,
                                     help_text="The user which submitted this flow run")

    parent = models.ForeignKey('flows.FlowRun', null=True, help_text=_("The parent run that triggered us"))

    results = JSONAsTextField(null=True, default=dict,
                              help_text=_("The results collected during this flow run in JSON format"))

    path = JSONAsTextField(null=True, default=list,
                           help_text=_("The path taken during this flow run in JSON format"))

    message_ids = ArrayField(base_field=models.BigIntegerField(), null=True,
                             help_text=_("The IDs of messages associated with this run"))

    current_node_uuid = models.UUIDField(null=True,
                                         help_text=_("The current node location of this run in the flow"))

    @cached_property
    def cached_child(self):
        child = FlowRun.objects.filter(parent=self).order_by('-created_on').select_related('flow').first()
        if child:
            child.org = self.org
            child.flow.org = self.org
            child.contact = self.contact
        return child

    def clear_cached_child(self):
        if 'cached_child' in self.__dict__:
            del self.__dict__['cached_child']

    @classmethod
    def create_or_update_from_goflow(cls, session, contact, run_output, run_log, wait, msg_in):
        """
        Creates or updates a flow run from the given output returned from goflow
        """
        uuid = run_output['uuid']
        results = run_output.get('results', {})
        is_active = run_output['status'] in ('active', 'waiting')
        is_error = run_output['status'] == 'errored'
        created_on = iso8601.parse_date(run_output['created_on'])
        exited_on = iso8601.parse_date(run_output['exited_on']) if run_output['exited_on'] else None
        expires_on = iso8601.parse_date(run_output['expires_on']) if run_output['expires_on'] else None
        timeout_on = iso8601.parse_date(wait['timeout_on']) if wait and wait.get('timeout_on') else None

        # does this run already exist?
        existing = cls.objects.filter(org=contact.org, contact=contact, uuid=uuid).select_related('flow').first()

        if not is_active:
            exit_type = cls.EXIT_TYPE_INTERRUPTED if is_error else cls.EXIT_TYPE_COMPLETED
        else:
            exit_type = None

        # we store a simplified version of the path
        path = []
        for s in run_output['path']:
            step = {
                FlowRun.PATH_NODE_UUID: s['node_uuid'],
                FlowRun.PATH_ARRIVED_ON: s['arrived_on']
            }
            if 'exit_uuid' in s:
                step[FlowRun.PATH_EXIT_UUID] = s['exit_uuid']
            path.append(step)
        current_node_uuid = path[-1][FlowRun.PATH_NODE_UUID]

        if existing:
            existing.path = path
            existing.current_node_uuid = current_node_uuid
            existing.results = results
            existing.expires_on = expires_on
            existing.timeout_on = timeout_on
            existing.modified_on = timezone.now()
            existing.exited_on = exited_on
            existing.exit_type = exit_type
            existing.responded |= bool(msg_in)
            existing.is_active = is_active
            existing.save(update_fields=('path', 'current_node_uuid', 'results', 'expires_on', 'modified_on', 'exited_on', 'exit_type', 'responded', 'is_active'))
            run = existing

            msgs_to_send, run_messages = run.apply_events(run_log, msg_in)

        else:
            # use our now for created_on/modified_on as this needs to be as close as possible to database time for API
            # run syncing to work
            now = timezone.now()

            flow = Flow.objects.get(org=contact.org, uuid=run_output['flow_uuid'])

            parent_uuid = run_output.get('parent_uuid')
            parent = cls.objects.get(org=contact.org, uuid=parent_uuid) if parent_uuid else None

            run = cls.objects.create(org=contact.org, uuid=uuid,
                                     flow=flow, contact=contact,
                                     parent=parent,
                                     path=path,
                                     current_node_uuid=current_node_uuid,
                                     results=results,
                                     session=session,
                                     responded=bool(msg_in),
                                     is_active=is_active,
                                     exited_on=exited_on, exit_type=exit_type,
                                     expires_on=expires_on, timeout_on=timeout_on,
                                     created_on=now, modified_on=now)

            # old simulation needs an action log showing we entered the flow
            if contact.is_test:  # pragma: no cover
                log_text = '%s has entered the "%s" flow' % (contact.get_display(contact.org, short=True), flow.name)
                ActionLog.create(run, log_text, created_on=created_on)

            msgs_to_send, run_messages = run.apply_events(run_log, msg_in)

        # attach any messages to this run
        if not run.message_ids:
            run.message_ids = []
        if msg_in:
            run.message_ids.append(msg_in.id)
        for m in run_messages:
            run.message_ids.append(m.id)
        if run.message_ids:
            run.save(update_fields=('message_ids',))

        return run, msgs_to_send

    def apply_events(self, run_log, msg_in=None):
        all_msgs_to_send = []
        all_run_messages = []

        for entry in run_log:
            # print("⚡ %s %s" % (entry.event['type'], json.dumps({k: v for k, v in six.iteritems(entry.event) if k != 'type'})))

            if entry.event['type'] == 'broadcast_created':
                msgs_to_send, run_messages = self.apply_broadcast_created(entry.event, msg_in)

                all_msgs_to_send += msgs_to_send
                all_run_messages += run_messages
            else:
                apply_func = getattr(self, 'apply_%s' % entry.event['type'], None)
                if apply_func:
                    apply_func(entry.event, msg_in)

        return all_msgs_to_send, all_run_messages

    def apply_broadcast_created(self, event, msg):
        """
        An outgoing broadcast being sent (not necessarily this contact)
        """
        urns = event.get('urns', [])
        contact_refs = event.get('contacts', [])
        group_refs = event.get('groups', [])
        created_on = iso8601.parse_date(event['created_on'])
        user = get_flow_user(self.org)

        # convert attachment URLs to absolute URLs
        attachments = []
        for attachment in event.get('attachments', []):
            media_type, media_url = attachment.split(':', 1)

            if not media_url.startswith('http://') and not media_url.startswith('https://'):
                media_url = "https://%s/%s" % (settings.AWS_BUCKET_DOMAIN, media_url)

            attachments.append("%s:%s" % (media_type, media_url))

        urns, contacts = self._resolve_urns_and_contacts(user, urns, contact_refs, group_refs)

        msgs_to_send = []
        run_messages = []

        for recipient in itertools.chain(urns, contacts):
            high_priority = (recipient == self.contact and self.session.responded)
            channel = msg.channel if msg and msg.contact == recipient else None

            try:
                msg_out = Msg.create_outgoing(
                    self.org, user, recipient,
                    text=event['text'],
                    attachments=attachments,
                    quick_replies=event.get('quick_replies', []),
                    channel=channel,
                    high_priority=high_priority,
                    created_on=created_on,
                    response_to=msg if msg and msg.id else None
                )
                msgs_to_send.append(msg_out)
                if msg_out.contact == self.contact:
                    run_messages.append(msg_out)

            except UnreachableException:  # pragma: no cover
                continue

        return msgs_to_send, run_messages

    def apply_email_created(self, event, msg):
        """
        Email being sent out
        """
        from .tasks import send_email_action_task

        subject, body, addresses = event['subject'], event['body'], event['addresses']

        if not self.contact.is_test:
            on_transaction_commit(lambda: send_email_action_task.delay(self.org_id, addresses, subject, body))
        else:  # pragma: no cover
            quoted_addresses = ['"%s"' % elt for elt in addresses]
            ActionLog.info(self, _("\"%s\" would be sent to %s") % (event['body'], ", ".join(quoted_addresses)))

    def apply_contact_property_changed(self, event, msg):
        """
        Name or language being updated
        """
        prop = event['property']
        value = event['value']
        update_fields = []

        if prop == "language":
            self.contact.language = value or None
            update_fields.append('language')
        elif prop == "name":
            self.contact.name = value
            update_fields.append("name")
        else:  # pragma: no cover
            raise ValueError("Unknown property to update on contact: %s" % prop)

        self.contact.save(update_fields=update_fields)

        if self.contact.is_test:  # pragma: no cover
            ActionLog.create(self, _("Updated %s to '%s'") % (prop, value))

    def apply_contact_urn_added(self, event, msg):
        """
        New URN being added to the contact
        """
        user = get_flow_user(self.org)
        urns = [six.text_type(urn) for urn in self.contact.urns.order_by('-priority')]
        urns.append(event['urn'])

        # don't really update URNs on test contacts
        if self.contact.is_test:
            scheme, path, display = URN.to_parts(event['urn'])
            ActionLog.info(self, _("Added %s as @contact.%s - skipped in simulator" % (path, scheme)))
        else:
            self.contact.update_urns(user, urns)

    def apply_contact_field_changed(self, event, msg):
        """
        Properties of this contact being updated
        """
        user = get_flow_user(self.org)
        field = ContactField.objects.get(org=self.org, key=event['field']['key'])
        value = event['value']

        self.contact.set_field(user, field.key, value)

        if self.contact.is_test:
            ActionLog.create(self, _("Updated %s to '%s'") % (field.label, event['value']))

    def apply_run_result_changed(self, event, msg):
        # flow results are actually saved in create_or_update_from_goflow
        if self.contact.is_test:
            ActionLog.create(self, _("Saved '%s' as @flow.%s") % (event['value'], slugify_with(event['name'])),
                             created_on=iso8601.parse_date(event['created_on']))

    def apply_input_labels_added(self, event, msg):
        if not msg:  # pragma: no cover
            return

        for label_ref in event['labels']:
            label = Label.label_objects.get(org=self.org, uuid=label_ref['uuid'])
            if not self.contact.is_test:
                label.toggle_label([msg], True)
            else:  # pragma: no cover
                ActionLog.info(self, _("Added %s label to msg '%s'") % (label.name, msg.text))

    def apply_contact_groups_added(self, event, msg):
        """
        This contact being added to one or more groups
        """
        user = get_flow_user(self.org)
        for group_ref in event['groups']:
            group = ContactGroup.get_user_groups(self.org).get(uuid=group_ref['uuid'])
            if not self.contact.is_stopped and not self.contact.is_blocked:
                group.update_contacts(user, [self.contact], add=True)

            if self.contact.is_test:  # pragma: no cover
                ActionLog.info(self, _("Added %s to %s") % (self.contact.name, group.name))

    def apply_contact_groups_removed(self, event, msg):
        """
        This contact being removed from one or more groups
        """
        user = get_flow_user(self.org)
        for group_ref in event['groups']:
            group = ContactGroup.get_user_groups(self.org).get(uuid=group_ref['uuid'])
            if not self.contact.is_stopped and not self.contact.is_blocked:
                group.update_contacts(user, [self.contact], add=False)

            if self.contact.is_test:  # pragma: no cover
                ActionLog.info(self, _("Removed %s from %s") % (self.contact.name, group.name))

    def apply_session_triggered(self, event, msg):
        """
        New sessions being started for other contacts
        """
        urns = event.get('urns', [])
        flow_ref = event['flow']
        contact_refs = event.get('contacts', [])
        group_refs = event.get('groups', [])
        parent_run_summary = event['run']
        user = get_flow_user(self.org)

        flow = self.org.flows.filter(is_active=True, is_archived=False, uuid=flow_ref['uuid']).first()
        if flow:
            urns, contacts = self._resolve_urns_and_contacts(user, urns, contact_refs, group_refs)

            # extract only unique contacts
            contacts = set(contacts)
            for urn in urns:
                contacts.add(urn.contact)  # pragma: needs cover

            FlowSession.bulk_start(contacts, flow, parent_run_summary=parent_run_summary)

        else:  # pragma: no cover
            raise ValueError("No such flow with UUID %s" % flow_ref['uuid'])

    def apply_error(self, event, msg):
        """
        The flowserver recorded a non-fatal error event
        """
        # only interested in errors for turning them into simulator messages
        if not self.contact.is_test:  # pragma: no cover
            return

        error = event['text']

        if error.startswith("invalid URN: "):
            invalid_urn = error[14:-1]
            try:
                URN.to_parts(invalid_urn)
                ActionLog.info(self, _("Contact not updated, invalid connection for contact (%s)") % invalid_urn)
            except ValueError:
                ActionLog.info(self, _("Contact not updated, missing connection for contact"))

    def _resolve_urns_and_contacts(self, user, urn_strs, contact_refs, group_refs):
        """
        Helper function for send_msg and start_session events which include lists of urns, contacts and groups
        """
        contacts = list(Contact.objects.filter(org=self.org, is_active=True, is_stopped=False, is_blocked=False,
                                               uuid__in=[c['uuid'] for c in contact_refs]))
        groups = list(ContactGroup.user_groups.filter(org=self.org, is_active=True,
                                                      uuid__in=[g['uuid'] for g in group_refs]))
        urns = []
        for urn_str in urn_strs:
            contact, urn = Contact.get_or_create(self.org, urn_str, user=user)
            urns.append(urn)

        # resolve group contacts
        contacts = set(contacts)
        for group in groups:
            for contact in group.contacts.all():
                contacts.add(contact)

        return urns, contacts

    @classmethod
    def create(cls, flow, contact, start=None, session=None, connection=None, fields=None,
               created_on=None, db_insert=True, submitted_by=None, parent=None, responded=False):

        args = dict(org_id=flow.org_id, flow=flow, contact=contact, start=start,
                    session=session, connection=connection, fields=fields, submitted_by=submitted_by, parent=parent, responded=responded)

        if created_on:
            args['created_on'] = created_on

        if parent:
            parent.clear_cached_child()

        if db_insert:
            run = FlowRun.objects.create(**args)
        else:
            run = FlowRun(**args)

        run.contact = contact
        return run

    def build_expressions_context(self, contact_context=None):
        """
        Builds the @flow expression context for this run
        """
        def result_wrapper(res):
            """
            Wraps a result, lets us do a nice representation of both @flow.foo and @flow.foo.text
            """
            return {
                '__default__': res[FlowRun.RESULT_VALUE],
                'text': res.get(FlowRun.RESULT_INPUT),
                'time': res[FlowRun.RESULT_CREATED_ON],
                'category': res.get(FlowRun.RESULT_CATEGORY_LOCALIZED, res[FlowRun.RESULT_CATEGORY]),
                'value': res[FlowRun.RESULT_VALUE]
            }

        context = {}
        default_lines = []

        for key, result in six.iteritems(self.results):
            context[key] = result_wrapper(result)
            default_lines.append("%s: %s" % (result[FlowRun.RESULT_NAME], result[FlowRun.RESULT_VALUE]))

        context['__default__'] = "\n".join(default_lines)

        # if we don't have a contact context, build one
        if not contact_context:
            self.contact.org = self.org
            contact_context = self.contact.build_expressions_context()

        context['contact'] = contact_context

        return context

    @property
    def connection_interrupted(self):
        return self.connection and self.connection.status == ChannelSession.INTERRUPTED

    @classmethod
    def normalize_field_key(cls, key):
        return FlowRun.INVALID_EXTRA_KEY_CHARS.sub('_', key)[:255]

    @classmethod
    def normalize_fields(cls, fields, max_values=None, count=-1):
        """
        Turns an arbitrary dictionary into a dictionary containing only string keys and values
        """
        if max_values is None:
            max_values = settings.FLOWRUN_FIELDS_SIZE

        if isinstance(fields, six.string_types):
            return fields[:Value.MAX_VALUE_LEN], count + 1

        elif isinstance(fields, numbers.Number) or isinstance(fields, bool):
            return fields, count + 1

        elif isinstance(fields, dict):
            count += 1
            field_dict = OrderedDict()
            for (k, v) in fields.items():
                (field_dict[FlowRun.normalize_field_key(k)], count) = FlowRun.normalize_fields(v, max_values, count)

                if count >= max_values:
                    break

            return field_dict, count

        elif isinstance(fields, list):
            count += 1
            list_dict = OrderedDict()
            for (i, v) in enumerate(fields):
                (list_dict[str(i)], count) = FlowRun.normalize_fields(v, max_values, count)

                if count >= max_values:  # pragma: needs cover
                    break

            return list_dict, count

        elif fields is None:
            return "", count + 1
        else:  # pragma: no cover
            raise ValueError("Unsupported type %s in extra" % six.text_type(type(fields)))

    @classmethod
    def bulk_exit(cls, runs, exit_type):
        """
        Exits (expires, interrupts) runs in bulk
        """
        # when expiring phone calls, we want to issue hangups
        connection_runs = runs.exclude(connection=None)
        for run in connection_runs:
            connection = run.connection.get()

            # have our session close itself
            if exit_type == FlowRun.EXIT_TYPE_EXPIRED:
                connection.close()

        run_ids = list(runs.values_list('id', flat=True))

        from .tasks import continue_parent_flows

        # batch this for 1,000 runs at a time so we don't grab locks for too long
        for id_batch in chunk_list(run_ids, 1000):
            now = timezone.now()

            # mark all steps in these runs as having been left
            FlowStep.objects.filter(run__id__in=id_batch, left_on=None).update(left_on=now)

            runs = FlowRun.objects.filter(id__in=id_batch)
            runs.update(is_active=False, exited_on=now, exit_type=exit_type, modified_on=now)

            # continue the parent flows to continue async
            on_transaction_commit(lambda: continue_parent_flows.delay(id_batch))

    def add_messages(self, msgs, step=None):
        """
        Associates the given messages with this run
        """
        if self.message_ids is None:
            self.message_ids = []

        needs_update = False

        for msg in msgs:
            # no-op for no msg or mock msgs
            if not msg or not msg.id:
                continue

            if msg.id not in self.message_ids:
                self.message_ids.append(msg.id)
                needs_update = True

            if step:
                step.messages.add(msg)

                # if this msg is part of a broadcast, save that on our flowstep so we can later purge the msg
                if msg.broadcast:
                    step.broadcasts.add(msg.broadcast)

            # incoming non-IVR messages won't have a type yet so update that
            if not msg.msg_type or msg.msg_type == INBOX:
                msg.msg_type = FLOW
                msg.save(update_fields=['msg_type'])

            # if message is from contact, mark run as responded
            if msg.direction == INCOMING:
                if not self.responded:
                    self.responded = True
                    needs_update = True

        if needs_update:
            self.save(update_fields=('responded', 'message_ids'))

    def get_message_ids(self):
        """
        Gets all the messages associated with this run
        """
        return self.message_ids or []

    def get_messages(self):
        """
        Gets all the messages associated with this run
        """
        return Msg.objects.filter(id__in=self.message_ids) if self.message_ids else Msg.objects.none()

    def get_last_msg(self, direction=INCOMING):
        """
        Returns the last incoming msg on this run
        :param direction: the direction of the message to fetch, default INCOMING
        """
        return self.get_messages().filter(direction=direction).order_by('-created_on').first()

    @classmethod
    def continue_parent_flow_runs(cls, runs):
        """
        Hands flow control back to our parent run if we have one
        """
        runs = (
            runs.filter(parent__flow__is_active=True, parent__flow__is_archived=False, parent__is_active=True)
            .select_related('parent__flow')
        )
        for run in runs:
            cls.continue_parent_flow_run(run)

    @classmethod
    def continue_parent_flow_run(cls, run, trigger_send=True, continue_parent=True):

        # TODO: Remove this in favor of responded on session
        if run.responded and not run.parent.responded:
            run.parent.responded = True
            run.parent.save(update_fields=['responded'])

        msgs = []

        steps = run.parent.steps.filter(left_on=None, step_type=FlowStep.TYPE_RULE_SET)
        step = steps.select_related('run', 'run__flow', 'run__contact', 'run__flow__org').first()

        if step:
            # if our child was interrupted, so shall we be
            if run.exit_type == FlowRun.EXIT_TYPE_INTERRUPTED and run.contact.id == step.run.contact.id:
                FlowRun.bulk_exit(FlowRun.objects.filter(id=step.run.id), FlowRun.EXIT_TYPE_INTERRUPTED)
                return

            # resume via goflow if this run is using the new engine
            if run.parent.session and run.parent.session.output:  # pragma: needs cover
                session = FlowSession.objects.get(id=run.parent.session.id)
                return session.resume(expired_child_run=run)

            ruleset = RuleSet.objects.filter(uuid=step.step_uuid, ruleset_type=RuleSet.TYPE_SUBFLOW,
                                             flow__org=step.run.org).exclude(flow=None).first()
            if ruleset:
                # use the last incoming message on this step
                msg = step.messages.filter(direction=INCOMING).order_by('-created_on').first()

                # if we are routing back to the parent before a msg was sent, we need a placeholder
                if not msg:
                    msg = Msg()
                    msg.text = ''
                    msg.org = run.org
                    msg.contact = run.contact

                # finally, trigger our parent flow
                (handled, msgs) = Flow.find_and_handle(msg, user_input=False, started_flows=[run.flow, run.parent.flow],
                                                       resume_parent_run=True, trigger_send=trigger_send, continue_parent=continue_parent)

        return msgs

    def get_session_responded(self):
        """
        TODO: Replace with Session.responded when it exists
        """
        current_run = self
        while current_run and current_run.contact_id == self.contact_id:
            if current_run.responded:
                return True
            current_run = current_run.parent

        return False

    def is_ivr(self):
        """
        If this run is over an IVR connection
        """
        return self.connection and self.connection.is_ivr()

    def keep_active_on_exit(self):
        """
        If our run should be completed when we leave the last node
        """
        # we let parent runs over ivr get closed by the provider
        return self.is_ivr() and not self.parent and not self.connection.is_done()

    def resume_after_timeout(self, expired_timeout):
        """
        Resumes a flow that is at a ruleset that has timed out
        """
        last_step = FlowStep.get_active_steps_for_contact(self.contact).first()

        # this timeout is invalid, clear it
        if not last_step or last_step.run != self:
            self.timeout_on = None
            self.save(update_fields=('timeout_on', 'modified_on'))
            return

        node = last_step.get_node()

        # only continue if we are at a ruleset with a timeout
        if isinstance(node, RuleSet) and timezone.now() > self.timeout_on > last_step.arrived_on:
            timeout = node.get_timeout()

            # if our current node doesn't have a timeout, but our timeout is still right, then the ruleset
            # has changed out from under us and no longer has a timeout, clear our run's timeout_on
            if not timeout and abs(expired_timeout - self.timeout_on) < timedelta(milliseconds=1):
                self.timeout_on = None
                self.save(update_fields=('timeout_on', 'modified_on'))

            # this is a valid timeout, deal with it
            else:
                # get the last outgoing msg for this contact
                msg = self.get_last_msg(OUTGOING)

                # if the last outgoing message wasn't assigned a credit, then clear our timeout
                if msg and msg.topup_id is None:
                    self.timeout_on = None
                    self.save(update_fields=['timeout_on', 'modified_on'])

                # check that our last outgoing msg was sent and our timeout is in the past, otherwise reschedule
                elif msg and (not msg.sent_on or timezone.now() < msg.sent_on + timedelta(minutes=timeout) - timedelta(seconds=5)):
                    self.update_timeout(msg.sent_on if msg.sent_on else timezone.now(), timeout)

                # look good, lets resume this run
                else:
                    msg = self.get_last_msg(INCOMING)
                    if not msg:
                        msg = Msg()
                        msg.text = ''
                        msg.org = self.org
                        msg.contact = self.contact
                    Flow.find_and_handle(msg, resume_after_timeout=True)

    def release(self):
        """
        Permanently deletes this flow run
        """
        # remove each of our steps. we do this one at a time
        # so we can decrement the activity properly
        for step in self.steps.all():
            step.release()

        # lastly delete ourselves
        self.delete()

    def set_completed(self, final_step=None, completed_on=None):
        """
        Mark a run as complete
        """
        if self.contact.is_test:
            ActionLog.create(self, _('%s has exited this flow') % self.contact.get_display(self.flow.org, short=True))

        now = timezone.now()

        if not completed_on:
            completed_on = now

        # mark that we left this step
        if final_step:
            final_step.left_on = completed_on
            final_step.save(update_fields=['left_on'])

        # mark this flow as inactive
        if not self.keep_active_on_exit():
            self.exit_type = FlowRun.EXIT_TYPE_COMPLETED
            self.exited_on = completed_on
            self.is_active = False
            self.save(update_fields=('exit_type', 'exited_on', 'modified_on', 'is_active'))

        if hasattr(self, 'voice_response') and self.parent and self.parent.is_active:
            callback = 'https://%s%s' % (self.org.get_brand_domain(), reverse('ivr.ivrcall_handle', args=[self.connection.pk]))
            self.voice_response.redirect(url=callback + '?resume=1')
        else:
            # if we have a parent to continue
            if self.parent:
                # mark it for continuation
                self.continue_parent = True

    def set_interrupted(self, final_step=None):
        """
        Mark run as interrupted
        """
        if self.contact.is_test:  # pragma: needs cover
            ActionLog.create(self, _('%s has interrupted this flow') % self.contact.get_display(self.flow.org, short=True))

        now = timezone.now()

        if final_step:
            final_step.left_on = now
            final_step.save(update_fields=['left_on'])

        # mark this flow as inactive
        self.exit_type = FlowRun.EXIT_TYPE_INTERRUPTED
        self.exited_on = now
        self.is_active = False
        self.save(update_fields=('exit_type', 'exited_on', 'modified_on', 'is_active'))

    def update_timeout(self, now, minutes):
        """
        Updates our timeout for our run, either clearing it or setting it appropriately
        """
        if not minutes and self.timeout_on:
            self.timeout_on = None
            self.save(update_fields=['timeout_on', 'modified_on'])
        elif minutes:
            self.timeout_on = now + timedelta(minutes=minutes)
            self.save(update_fields=['timeout_on', 'modified_on'])

    def update_expiration(self, point_in_time=None):
        """
        Set our expiration according to the flow settings
        """
        if self.flow.expires_after_minutes:
            now = timezone.now()
            if not point_in_time:
                point_in_time = now
            self.expires_on = point_in_time + timedelta(minutes=self.flow.expires_after_minutes)

            # save our updated fields
            self.save(update_fields=['expires_on', 'modified_on'])

            # if it's in the past, just expire us now
            if self.expires_on < now:
                self.expire()

        # parent should always have a later expiration than the children
        if self.parent:
            self.parent.update_expiration(self.expires_on)

    def expire(self):
        self.bulk_exit(FlowRun.objects.filter(id=self.id), FlowRun.EXIT_TYPE_EXPIRED)

    @classmethod
    def exit_all_for_contacts(cls, contacts, exit_type):
        contact_runs = cls.objects.filter(is_active=True, contact__in=contacts)
        cls.bulk_exit(contact_runs, exit_type)

    def update_fields(self, field_map):
        # validate our field
        (field_map, count) = FlowRun.normalize_fields(field_map)
        if not self.fields:
            self.fields = field_map
        else:
            existing_map = self.fields
            existing_map.update(field_map)
            self.fields = existing_map

        self.save(update_fields=['fields'])

    def is_completed(self):
        return self.exit_type == FlowRun.EXIT_TYPE_COMPLETED

    def is_interrupted(self):
        return self.exit_type == FlowRun.EXIT_TYPE_INTERRUPTED

    def create_outgoing_ivr(self, text, recording_url, connection, response_to=None):

        # create a Msg object to track what happened
        from temba.msgs.models import DELIVERED, IVR

        attachments = None
        if recording_url:
            attachments = ['%s/x-wav:%s' % (Msg.MEDIA_AUDIO, recording_url)]

        msg = Msg.create_outgoing(self.flow.org, self.flow.created_by, self.contact, text, channel=self.connection.channel,
                                  response_to=response_to, attachments=attachments,
                                  status=DELIVERED, msg_type=IVR, connection=connection)

        # play a recording or read some text
        if msg:
            if recording_url:
                self.voice_response.play(url=recording_url)
            else:
                self.voice_response.say(text)

        return msg

    @classmethod
    def serialize_value(cls, value):
        """
        Utility method to give the serialized value for the passed in value
        """
        if value is None:
            return None

        if isinstance(value, datetime):
            return value.isoformat()
        elif isinstance(value, AdminBoundary):
            return value.path
        else:
            return six.text_type(value)

    def save_run_result(self, name, node_uuid, category, category_localized, raw_value, raw_input):
        # slug our name
        key = Flow.label_to_slug(name)

        # create our result dict
        results = self.results
        results[key] = {
            FlowRun.RESULT_NAME: name,
            FlowRun.RESULT_NODE_UUID: node_uuid,
            FlowRun.RESULT_CATEGORY: category,
            FlowRun.RESULT_VALUE: FlowRun.serialize_value(raw_value),
            FlowRun.RESULT_INPUT: raw_input,
            FlowRun.RESULT_CREATED_ON: timezone.now().isoformat(),
        }

        # if we have a different localized name for our category, save it as well
        if category != category_localized:
            results[key][FlowRun.RESULT_CATEGORY_LOCALIZED] = category_localized

        self.results = results
        self.modified_on = timezone.now()
        self.save(update_fields=['results', 'modified_on'])

    def __str__(self):
        return "FlowRun: %s Flow: %s\n%s" % (self.uuid, self.flow.uuid, json.dumps(self.results, indent=2))


@six.python_2_unicode_compatible
class FlowStep(models.Model):
    """
    A contact's visit to a node in a flow (rule set or action set)
    """
    TYPE_RULE_SET = 'R'
    TYPE_ACTION_SET = 'A'
    STEP_TYPE_CHOICES = ((TYPE_RULE_SET, "RuleSet"),
                         (TYPE_ACTION_SET, "ActionSet"))

    run = models.ForeignKey(FlowRun, related_name='steps')

    contact = models.ForeignKey(Contact, related_name='flow_steps')

    step_type = models.CharField(max_length=1, choices=STEP_TYPE_CHOICES, help_text=_("What type of node was visited"))

    step_uuid = models.CharField(max_length=36,
                                 help_text=_("The UUID of the ActionSet or RuleSet for this step"))

    rule_uuid = models.CharField(max_length=36, null=True,
                                 help_text=_("For uuid of the rule that matched on this ruleset, null on ActionSets"))

    rule_category = models.CharField(max_length=36, null=True,
                                     help_text=_("The category label that matched on this ruleset, null on ActionSets"))

    rule_value = models.TextField(null=True,
                                  help_text=_("The value that was matched in our category for this ruleset, null on ActionSets"))

    rule_decimal_value = models.DecimalField(max_digits=36, decimal_places=8, null=True,
                                             help_text=_("The decimal value that was matched in our category for this ruleset, null on ActionSets or if a non numeric rule was matched"))

    next_uuid = models.CharField(max_length=36, null=True,
                                 help_text=_("The uuid of the next step type we took"))

    arrived_on = models.DateTimeField(help_text=_("When the user arrived at this step in the flow"))

    left_on = models.DateTimeField(null=True,
                                   help_text=_("When the user left this step in the flow"))

    messages = models.ManyToManyField(Msg, related_name='steps',
                                      help_text=_("Any messages that are associated with this step (either sent or received)"))

    broadcasts = models.ManyToManyField(Broadcast, related_name='steps',
                                        help_text=_("Any broadcasts that are associated with this step (only sent)"))

    @classmethod
    def from_json(cls, json_obj, flow, run):
        """
        Creates a new flow step from the given Surveyor step JSON
        """
        node = json_obj['node']
        arrived_on = json_date_to_datetime(json_obj['arrived_on'])

        # find the previous step
        prev_step = cls.objects.filter(run=run).order_by('-left_on').first()

        # figure out which exit was taken by that step
        exit_uuid = None
        if prev_step:
            if prev_step.step_type == cls.TYPE_RULE_SET:
                exit_uuid = prev_step.rule_uuid
            else:
                prev_node = prev_step.get_node()
                if prev_node:
                    exit_uuid = prev_node.exit_uuid

        # generate the messages for this step
        msgs = []
        if node.is_ruleset():
            incoming = None
            if node.is_pause():
                # if a msg was sent to this ruleset, create it
                if json_obj['rule']:

                    media = None
                    if 'media' in json_obj['rule']:
                        media = json_obj['rule']['media']
                        (media_type, url) = media.split(':', 1)

                        # store the non-typed url in the value and text
                        json_obj['rule']['value'] = url
                        json_obj['rule']['text'] = url

                    # if we received a message
                    incoming = Msg.create_incoming(org=run.org, contact=run.contact, text=json_obj['rule']['text'],
                                                   attachments=[media] if media else None,
                                                   msg_type=FLOW, status=HANDLED, date=arrived_on,
                                                   channel=None, urn=None)
            else:  # pragma: needs cover
                incoming = Msg.objects.filter(org=run.org, direction=INCOMING, steps__run=run).order_by('-pk').first()

            if incoming:
                msgs.append(incoming)
        else:
            actions = Action.from_json_array(flow.org, json_obj['actions'])

            last_incoming = Msg.objects.filter(org=run.org, direction=INCOMING, steps__run=run).order_by('-pk').first()

            for action in actions:
                context = flow.build_expressions_context(run.contact, last_incoming)
                msgs += action.execute(run, context, node.uuid, msg=last_incoming, offline_on=arrived_on)

        step = flow.add_step(run, node, msgs=msgs, previous_step=prev_step, arrived_on=arrived_on, exit_uuid=exit_uuid)

        # if a rule was picked on this ruleset
        if node.is_ruleset() and json_obj['rule']:
            rule_uuid = json_obj['rule']['uuid']
            rule_value = json_obj['rule']['value']

            # update the value if we have an existing ruleset
            ruleset = RuleSet.objects.filter(flow=flow, uuid=node.uuid).first()
            if ruleset:
                rule = None
                for r in ruleset.get_rules():
                    if r.uuid == rule_uuid:
                        rule = r
                        break

                if not rule:
                    # the user updated the rules try to match the new rules
                    msg = Msg(org=run.org, contact=run.contact, text=json_obj['rule']['text'], id=0)
                    rule, value = ruleset.find_matching_rule(step, run, msg)

                    if not rule:
                        raise ValueError("No such rule with UUID %s" % rule_uuid)

                    rule_uuid = rule.uuid
                    rule_value = value

                ruleset.save_run_value(run, rule, rule_value, json_obj['rule']['text'])

            # update our step with our rule details
            step.rule_uuid = rule_uuid
            step.rule_value = rule_value
            step.save(update_fields=('rule_uuid', 'rule_value'))

        return step

    @classmethod
    def get_active_steps_for_contact(cls, contact, step_type=None):

        steps = FlowStep.objects.filter(run__is_active=True, run__flow__is_active=True, run__contact=contact,
                                        left_on=None)

        # don't consider voice steps, those are interactive
        steps = steps.exclude(run__flow__flow_type=Flow.VOICE)

        # real contacts don't deal with archived flows
        if not contact.is_test:
            steps = steps.filter(run__flow__is_archived=False)

        if step_type:
            steps = steps.filter(step_type=step_type)

        steps = steps.order_by('-pk')

        # optimize lookups
        return steps.select_related('run', 'run__flow', 'run__contact', 'run__flow__org', 'run__connection')

    def release(self):
        self.delete()

    def save_rule_match(self, rule, value):
        self.rule_uuid = rule.uuid

        if value is None:
            value = ''

        # format our rule value appropriately
        if isinstance(value, datetime):
            (date_format, time_format) = get_datetime_format(self.run.flow.org.get_dayfirst())
            self.rule_value = datetime_to_str(value, tz=self.run.flow.org.timezone, format=time_format, ms=False)
        else:
            self.rule_value = six.text_type(value)[:Msg.MAX_TEXT_LEN]

        self.save(update_fields=('rule_uuid', 'rule_value'))

    def get_node(self):
        """
        Returns the node (i.e. a RuleSet or ActionSet) associated with this step
        """
        if self.step_type == FlowStep.TYPE_RULE_SET:
            return RuleSet.objects.filter(uuid=self.step_uuid).first()
        else:  # pragma: needs cover
            return ActionSet.objects.filter(uuid=self.step_uuid).first()

    def __str__(self):
        return "%s - %s:%s" % (self.run.contact, self.step_type, self.step_uuid)


@six.python_2_unicode_compatible
class RuleSet(models.Model):
    TYPE_WAIT_MESSAGE = 'wait_message'

    # Ussd
    TYPE_WAIT_USSD_MENU = 'wait_menu'
    TYPE_WAIT_USSD = 'wait_ussd'

    # Calls
    TYPE_WAIT_RECORDING = 'wait_recording'
    TYPE_WAIT_DIGIT = 'wait_digit'
    TYPE_WAIT_DIGITS = 'wait_digits'

    # Surveys
    TYPE_WAIT_PHOTO = 'wait_photo'
    TYPE_WAIT_VIDEO = 'wait_video'
    TYPE_WAIT_AUDIO = 'wait_audio'
    TYPE_WAIT_GPS = 'wait_gps'

    TYPE_AIRTIME = 'airtime'
    TYPE_WEBHOOK = 'webhook'
    TYPE_RESTHOOK = 'resthook'
    TYPE_FLOW_FIELD = 'flow_field'
    TYPE_FORM_FIELD = 'form_field'
    TYPE_CONTACT_FIELD = 'contact_field'
    TYPE_EXPRESSION = 'expression'
    TYPE_RANDOM = 'random'
    TYPE_SUBFLOW = 'subflow'

    CONFIG_WEBHOOK = 'webhook'
    CONFIG_WEBHOOK_ACTION = 'webhook_action'
    CONFIG_WEBHOOK_HEADERS = 'webhook_headers'
    CONFIG_RESTHOOK = 'resthook'

    TYPE_MEDIA = (TYPE_WAIT_PHOTO, TYPE_WAIT_GPS, TYPE_WAIT_VIDEO, TYPE_WAIT_AUDIO, TYPE_WAIT_RECORDING)

    TYPE_WAIT = (TYPE_WAIT_MESSAGE, TYPE_WAIT_RECORDING, TYPE_WAIT_DIGIT, TYPE_WAIT_DIGITS, TYPE_WAIT_USSD_MENU,
                 TYPE_WAIT_USSD, TYPE_WAIT_PHOTO, TYPE_WAIT_VIDEO, TYPE_WAIT_AUDIO, TYPE_WAIT_GPS)

    TYPE_USSD = (TYPE_WAIT_USSD_MENU, TYPE_WAIT_USSD)

    TYPE_CHOICES = ((TYPE_WAIT_MESSAGE, "Wait for message"),
                    (TYPE_WAIT_USSD_MENU, "Wait for USSD menu"),
                    (TYPE_WAIT_USSD, "Wait for USSD message"),
                    (TYPE_WAIT_RECORDING, "Wait for recording"),
                    (TYPE_WAIT_DIGIT, "Wait for digit"),
                    (TYPE_WAIT_DIGITS, "Wait for digits"),
                    (TYPE_SUBFLOW, "Subflow"),
                    (TYPE_WEBHOOK, "Webhook"),
                    (TYPE_RESTHOOK, "Resthook"),
                    (TYPE_AIRTIME, "Transfer Airtime"),
                    (TYPE_FORM_FIELD, "Split by message form"),
                    (TYPE_CONTACT_FIELD, "Split on contact field"),
                    (TYPE_EXPRESSION, "Split by expression"),
                    (TYPE_RANDOM, "Split Randomly"))

    uuid = models.CharField(max_length=36, unique=True)

    flow = models.ForeignKey(Flow, related_name='rule_sets', null=True)

    label = models.CharField(max_length=64, null=True, blank=True,
                             help_text=_("The label for this field"))

    operand = models.CharField(max_length=128, null=True, blank=True,
                               help_text=_("The value that rules will be run against, if None defaults to @step.value"))

    webhook_url = models.URLField(null=True, blank=True, max_length=255,
                                  help_text=_("The URL that will be called with the user's response before we run our rules"))

    webhook_action = models.CharField(null=True, blank=True, max_length=8, default='POST',
                                      help_text=_('How the webhook should be executed'))

    rules = JSONAsTextField(help_text=_("The JSON encoded actions for this action set"), default=list)

    finished_key = models.CharField(max_length=1, null=True, blank=True,
                                    help_text="During IVR, this is the key to indicate we are done waiting")

    value_type = models.CharField(max_length=1, choices=Value.TYPE_CHOICES, default=Value.TYPE_TEXT,
                                  help_text="The type of value this ruleset saves")

    ruleset_type = models.CharField(max_length=16, choices=TYPE_CHOICES, null=True,
                                    help_text="The type of ruleset")

    response_type = models.CharField(max_length=1, help_text="The type of response that is being saved")

    config = JSONAsTextField(null=True, verbose_name=_("Ruleset Configuration"), default=dict,
                             help_text=_("RuleSet type specific configuration"))

    x = models.IntegerField()
    y = models.IntegerField()

    created_on = models.DateTimeField(auto_now_add=True, help_text=_("When this ruleset was originally created"))
    modified_on = models.DateTimeField(auto_now=True, help_text=_("When this ruleset was last modified"))

    @classmethod
    def get(cls, flow, uuid):
        return RuleSet.objects.filter(flow=flow, uuid=uuid).select_related('flow', 'flow__org').first()

    @property
    def is_messaging(self):
        return self.ruleset_type in (self.TYPE_USSD + (self.TYPE_WAIT_MESSAGE,))

    @classmethod
    def contains_step(cls, text):  # pragma: needs cover

        # remove any padding
        if text:
            text = text.strip()

        # match @step.value or @(step.value)
        return text and text[0] == '@' and 'step' in text

    def get_value_type(self):
        """
        Determines the value type that this ruleset will generate.
        """
        # we keep track of specialized rule types we see
        value_type = None

        for rule in self.get_rules():
            if isinstance(rule.test, TrueTest):
                continue

            rule_type = None

            if isinstance(rule.test, NumericTest):
                rule_type = Value.TYPE_DECIMAL

            elif isinstance(rule.test, DateTest):
                rule_type = Value.TYPE_DATETIME

            elif isinstance(rule.test, HasStateTest):
                rule_type = Value.TYPE_STATE

            elif isinstance(rule.test, HasDistrictTest):
                rule_type = Value.TYPE_DISTRICT

            elif isinstance(rule.test, HasWardTest):
                rule_type = Value.TYPE_WARD

            # this either isn't one of our value types or we have more than one type in this ruleset
            if not rule_type or (value_type and rule_type != value_type):
                return Value.TYPE_TEXT

            value_type = rule_type

        return value_type if value_type else Value.TYPE_TEXT

    def get_voice_input(self, voice_response, action=None):

        # recordings aren't wrapped input they get tacked on at the end
        if self.ruleset_type in [RuleSet.TYPE_WAIT_RECORDING, RuleSet.TYPE_SUBFLOW]:
            return voice_response
        elif self.ruleset_type == RuleSet.TYPE_WAIT_DIGITS:
            return voice_response.gather(finishOnKey=self.finished_key, timeout=120, action=action)
        else:
            # otherwise we assume it's single digit entry
            return voice_response.gather(numDigits=1, timeout=120, action=action)

    def is_pause(self):
        return self.ruleset_type in RuleSet.TYPE_WAIT

    def is_ussd(self):
        return self.ruleset_type in RuleSet.TYPE_USSD

    def get_timeout(self):
        for rule in self.get_rules():
            if isinstance(rule.test, TimeoutTest):
                return rule.test.minutes

        return None

    def find_matching_rule(self, step, run, msg, resume_after_timeout=False):
        orig_text = None
        if msg:
            orig_text = msg.text

        msg.contact = run.contact
        context = run.flow.build_expressions_context(run.contact, msg, run=run)

        if resume_after_timeout:
            for rule in self.get_rules():
                if isinstance(rule.test, TimeoutTest):
                    (result, value) = rule.matches(run, msg, context, orig_text)
                    if result > 0:
                        return rule, value

        elif self.ruleset_type in [RuleSet.TYPE_WEBHOOK, RuleSet.TYPE_RESTHOOK]:
            header = {}

            # figure out which URLs will be called
            if self.ruleset_type == RuleSet.TYPE_WEBHOOK:
                resthook = None
                urls = [self.config[RuleSet.CONFIG_WEBHOOK]]
                action = self.config[RuleSet.CONFIG_WEBHOOK_ACTION]

                if RuleSet.CONFIG_WEBHOOK_HEADERS in self.config:
                    headers = self.config[RuleSet.CONFIG_WEBHOOK_HEADERS]
                    for item in headers:
                        header[item.get('name')] = item.get('value')

            elif self.ruleset_type == RuleSet.TYPE_RESTHOOK:
                from temba.api.models import Resthook

                # look up the rest hook
                resthook_slug = self.config[RuleSet.CONFIG_RESTHOOK]
                resthook = Resthook.get_or_create(run.org, resthook_slug, run.flow.created_by)
                urls = resthook.get_subscriber_urls()

                # no urls? use None, as our empty case
                if not urls:
                    urls = [None]

                action = 'POST'

            # by default we are a failure (there are no resthooks for example)
            status_code = None
            body = ""

            for url in urls:
                from temba.api.models import WebHookEvent

                (value, errors) = Msg.evaluate_template(url, context, org=run.flow.org, url_encode=True)
                result = WebHookEvent.trigger_flow_webhook(run, value, self.uuid, msg, action,
                                                           resthook=resthook,
                                                           headers=header)

                # we haven't recorded any status yet, do so
                if not status_code:
                    status_code = result.status_code
                    body = result.body

                # our subscriber is no longer interested, remove this URL as a subscriber
                if result.status_code == 410:
                    resthook.remove_subscriber(url, run.flow.created_by)

                # if this is a success and we haven't ever succeeded, set our code and body
                elif 200 <= result.status_code < 300 and not (200 <= status_code < 300):  # pragma: needs cover
                    status_code = result.status_code
                    body = result.body

                # this was an empty URL, treat it as success regardless
                if url is None:
                    status_code = 200
                    body = _("No subscribers to this event")

            # default to a status code of 418 if we made no calls
            if not status_code:  # pragma: needs cover
                status_code = 418

            # find our matching rule, we pass in the status from our calls
            for rule in self.get_rules():
                (result, value) = rule.matches(run, msg, context, str(status_code))
                if result > 0:
                    return rule, body

        else:
            # if it's a form field, construct an expression accordingly
            if self.ruleset_type == RuleSet.TYPE_FORM_FIELD:
                delim = self.config.get('field_delimiter', ' ')
                self.operand = '@(FIELD(%s, %d, "%s"))' % (
                    self.operand[1:], self.config.get('field_index', 0) + 1, delim
                )

            # if we have a custom operand, figure that out
            text = None
            if self.operand:
                (text, errors) = Msg.evaluate_template(self.operand, context, org=run.flow.org)
            elif msg:
                text = msg.text

            if self.ruleset_type == RuleSet.TYPE_AIRTIME:

                # flow simulation will always simulate a suceessful airtime transfer
                # without saving the object in the DB
                if run.contact.is_test:
                    from temba.flows.models import ActionLog
                    log_txt = "Simulate Complete airtime transfer"
                    ActionLog.create(run, log_txt, safe=True)

                    airtime = AirtimeTransfer(status=AirtimeTransfer.SUCCESS)
                else:
                    airtime = AirtimeTransfer.trigger_airtime_event(self.flow.org, self, run.contact, msg)

                # rebuild our context again, the webhook may have populated something
                context = run.flow.build_expressions_context(run.contact, msg)

                # airtime test evaluate against the status of the airtime
                text = airtime.status

            try:
                rules = self.get_rules()
                for rule in rules:
                    (result, value) = rule.matches(run, msg, context, text)
                    if result:
                        # treat category as the base category
                        return rule, value
            finally:
                if msg:
                    msg.text = orig_text

        return None, None  # pragma: no cover

    def find_interrupt_rule(self, step, run, msg):
        rules = self.get_rules()
        for rule in rules:
            result, value = rule.matches(run, msg, {}, "")

            if result and value == "interrupted_status":
                return rule, value
        return None, None

    def save_run_value(self, run, rule, raw_value, raw_input):
        run.save_run_result(
            name=self.label,
            node_uuid=self.uuid,
            category=rule.get_category_name(run.flow.base_language),
            category_localized=rule.get_category_name(run.flow.base_language, run.contact.language),
            raw_value=raw_value,
            raw_input=raw_input
        )

    def get_step_type(self):
        return FlowStep.TYPE_RULE_SET

    def get_rules_dict(self):
        return self.rules

    def get_rules(self):
        return Rule.from_json_array(self.flow.org, self.rules)

    def get_rule_uuids(self):  # pragma: needs cover
        return [rule['uuid'] for rule in self.rules]

    def set_rules(self, rules):
        rules_dict = []
        for rule in rules:
            rules_dict.append(rule.as_json())

        self.rules = rules_dict

    def as_json(self):
        return dict(uuid=self.uuid, x=self.x, y=self.y, label=self.label, rules=self.get_rules_dict(),
                    finished_key=self.finished_key, ruleset_type=self.ruleset_type, response_type=self.response_type,
                    operand=self.operand, config=self.config)

    def __str__(self):  # pragma: no cover
        if self.label:
            return "RuleSet: %s - %s" % (self.uuid, self.label)
        else:
            return "RuleSet: %s" % (self.uuid,)


@six.python_2_unicode_compatible
class ActionSet(models.Model):
    uuid = models.CharField(max_length=36, unique=True)
    flow = models.ForeignKey(Flow, related_name='action_sets')

    destination = models.CharField(max_length=36, null=True)
    destination_type = models.CharField(max_length=1, choices=FlowStep.STEP_TYPE_CHOICES, null=True)

    exit_uuid = models.CharField(max_length=36, null=True)  # needed for migrating to new engine

    actions = JSONAsTextField(help_text=_("The JSON encoded actions for this action set"), default=dict)

    x = models.IntegerField()
    y = models.IntegerField()

    created_on = models.DateTimeField(auto_now_add=True, help_text=_("When this action was originally created"))
    modified_on = models.DateTimeField(auto_now=True, help_text=_("When this action was last modified"))

    @classmethod
    def get(cls, flow, uuid):
        return ActionSet.objects.filter(flow=flow, uuid=uuid).select_related('flow', 'flow__org').first()

    @property
    def is_messaging(self):
        actions = self.get_actions()
        for action in actions:
            if isinstance(action, (EndUssdAction, ReplyAction, SendAction)):
                return True
        return False

    def get_step_type(self):
        return FlowStep.TYPE_ACTION_SET

    def execute_actions(self, run, msg, started_flows, skip_leading_reply_actions=True):
        actions = self.get_actions()
        msgs = []

        run.contact.org = run.org
        context = run.flow.build_expressions_context(run.contact, msg, run=run)

        seen_other_action = False
        for a, action in enumerate(actions):
            if not isinstance(action, ReplyAction):
                seen_other_action = True

            # to optimize large flow starts, leading reply actions are handled as a single broadcast so don't repeat
            # them here
            if not skip_leading_reply_actions and isinstance(action, ReplyAction) and not seen_other_action:
                continue

            if isinstance(action, StartFlowAction):
                if action.flow.pk in started_flows:
                    pass
                else:
                    msgs += action.execute(run, context, self.uuid, msg, started_flows)

                    # reload our contact and reassign it to our run, it may have been changed deep down in our child flow
                    run.contact = Contact.objects.get(pk=run.contact.pk)

            else:
                msgs += action.execute(run, context, self.uuid, msg)

                # actions modify the run.contact, update the msg contact in case they did so
                if msg:
                    msg.contact = run.contact

            # if there are more actions, rebuild the parts of the context that may have changed
            if a < len(actions) - 1:
                context['contact'] = run.contact.build_expressions_context()
                context['extra'] = run.fields

        return msgs

    def get_actions(self):
        return Action.from_json_array(self.flow.org, self.actions)

    def as_json(self):
        return dict(uuid=self.uuid, x=self.x, y=self.y, destination=self.destination,
                    actions=self.actions, exit_uuid=self.exit_uuid)

    def __str__(self):  # pragma: no cover
        return "ActionSet: %s" % (self.uuid,)


class FlowRevision(SmartModel):
    """
    JSON definitions for previous flow revisions
    """
    flow = models.ForeignKey(Flow, related_name='revisions')

    definition = JSONAsTextField(help_text=_("The JSON flow definition"), default=dict)

    spec_version = models.CharField(default=get_current_export_version, max_length=8,
                                    help_text=_("The flow version this definition is in"))

    revision = models.IntegerField(null=True, help_text=_("Revision number for this definition"))

    @classmethod
    def validate_flow_definition(cls, flow_spec):

        non_localized_error = _('Malformed flow, encountered non-localized definition')

        # should always have a base_language
        if 'base_language' not in flow_spec or not flow_spec['base_language']:
            raise ValueError(non_localized_error)

        # language should match values in definition
        base_language = flow_spec['base_language']

        def validate_localization(lang_dict):

            # must be a dict
            if not isinstance(lang_dict, dict):
                raise ValueError(non_localized_error)

            # and contain the base_language
            if base_language not in lang_dict:  # pragma: needs cover
                raise ValueError(non_localized_error)

        for actionset in flow_spec['action_sets']:
            for action in actionset['actions']:
                if 'msg' in action and action['type'] != 'email':
                    validate_localization(action['msg'])

        for ruleset in flow_spec['rule_sets']:
            for rule in ruleset['rules']:
                validate_localization(rule['category'])

    @classmethod
    def migrate_export(cls, org, exported_json, same_site, version, to_version=None):
        from temba.flows import flow_migrations

        if not to_version:
            to_version = get_current_export_version()

        for version in Flow.get_versions_after(version):
            version_slug = version.replace(".", "_")
            migrate_fn = getattr(flow_migrations, 'migrate_export_to_version_%s' % version_slug, None)

            if migrate_fn:
                exported_json = migrate_fn(exported_json, org, same_site)

                # update the version of migrated flows
                flows = []
                for sub_flow in exported_json.get('flows', []):
                    sub_flow[Flow.VERSION] = version
                    flows.append(sub_flow)

                exported_json['flows'] = flows

            else:
                migrate_fn = getattr(flow_migrations, 'migrate_to_version_%s' % version_slug, None)
                if migrate_fn:
                    flows = []
                    for json_flow in exported_json.get('flows', []):
                        json_flow = migrate_fn(json_flow, None)

                        flows.append(json_flow)

                    exported_json['flows'] = flows

            # update each flow's version number
            for json_flow in exported_json.get('flows', []):
                json_flow[Flow.VERSION] = version

            if version == to_version:
                break

        return exported_json

    @classmethod
    def migrate_definition(cls, json_flow, flow, to_version=None):
        from temba.flows import flow_migrations

        if not to_version:
            to_version = get_current_export_version()

        for version in Flow.get_versions_after(json_flow.get(Flow.VERSION)):
            version_slug = version.replace(".", "_")
            migrate_fn = getattr(flow_migrations, 'migrate_to_version_%s' % version_slug, None)

            if migrate_fn:
                json_flow = migrate_fn(json_flow, flow)
                json_flow[Flow.VERSION] = version

            if version == to_version:
                break

        return json_flow

    def get_definition_json(self):

        definition = self.definition

        # if it's previous to version 6, wrap the definition to
        # mirror our exports for those versions
        if Flow.is_before_version(self.spec_version, "6"):
            definition = dict(definition=self.definition, flow_type=self.flow.flow_type,
                              expires=self.flow.expires_after_minutes, id=self.flow.pk,
                              revision=self.revision, uuid=self.flow.uuid)

        # make sure old revisions migrate properly
        definition[Flow.VERSION] = self.spec_version

        # migrate our definition if necessary
        if self.spec_version != get_current_export_version():
            definition = FlowRevision.migrate_definition(definition, self.flow)
        return definition

    def as_json(self, include_definition=False):

        name = self.created_by.get_full_name()
        return dict(user=dict(email=self.created_by.email, name=name),
                    created_on=datetime_to_str(self.created_on),
                    id=self.pk,
                    version=self.spec_version,
                    revision=self.revision)


class FlowCategoryCount(SquashableModel):
    """
    Maintains counts for categories across all possible results in a flow
    """
    SQUASH_OVER = ('flow_id', 'node_uuid', 'result_key', 'result_name', 'category_name')

    flow = models.ForeignKey(Flow, related_name='category_counts', help_text="The flow the result belongs to")
    node_uuid = models.UUIDField(db_index=True)
    result_key = models.CharField(max_length=128, help_text="The sluggified key for the result")
    result_name = models.CharField(max_length=128, help_text="The result the category belongs to")
    category_name = models.CharField(max_length=128, help_text="The category name for a result")
    count = models.IntegerField(default=0)

    @classmethod
    def get_squash_query(cls, distinct_set):
        sql = """
        WITH removed as (
          DELETE FROM %(table)s WHERE "id" IN (
            SELECT "id" FROM %(table)s
              WHERE "flow_id" = %%s AND "node_uuid" = %%s AND "result_key" = %%s AND "result_name" = %%s AND "category_name" = %%s
              LIMIT 10000
          ) RETURNING "count"
        )
        INSERT INTO %(table)s("flow_id", "node_uuid", "result_key", "result_name", "category_name", "count", "is_squashed")
        VALUES (%%s, %%s, %%s, %%s, %%s, GREATEST(0, (SELECT SUM("count") FROM removed)), TRUE);
        """ % {'table': cls._meta.db_table}

        params = (distinct_set.flow_id, distinct_set.node_uuid, distinct_set.result_key, distinct_set.result_name, distinct_set.category_name) * 2
        return sql, params

    def __str__(self):
        return "%s: %s" % (self.category_name, self.count)


@six.python_2_unicode_compatible
class FlowPathCount(SquashableModel):
    """
    Maintains hourly counts of flow paths
    """
    SQUASH_OVER = ('flow_id', 'from_uuid', 'to_uuid', 'period')

    flow = models.ForeignKey(Flow, related_name='activity', help_text=_("The flow where the activity occurred"))
    from_uuid = models.UUIDField(help_text=_("Which flow node they came from"))
    to_uuid = models.UUIDField(help_text=_("Which flow node they went to"))
    period = models.DateTimeField(help_text=_("When the activity occured with hourly precision"))
    count = models.IntegerField(default=0)

    @classmethod
    def get_squash_query(cls, distinct_set):
        sql = """
        WITH removed as (
            DELETE FROM %(table)s WHERE "flow_id" = %%s AND "from_uuid" = %%s AND "to_uuid" = %%s AND "period" = date_trunc('hour', %%s) RETURNING "count"
        )
        INSERT INTO %(table)s("flow_id", "from_uuid", "to_uuid", "period", "count", "is_squashed")
        VALUES (%%s, %%s, %%s, date_trunc('hour', %%s), GREATEST(0, (SELECT SUM("count") FROM removed)), TRUE);
        """ % {'table': cls._meta.db_table}

        params = (distinct_set.flow_id, distinct_set.from_uuid, distinct_set.to_uuid, distinct_set.period) * 2
        return sql, params

    @classmethod
    def get_totals(cls, flow):
        counts = cls.objects.filter(flow=flow)
        totals = list(counts.values_list('from_uuid', 'to_uuid').annotate(replies=Sum('count')))
        return {'%s:%s' % (t[0], t[1]): t[2] for t in totals}

    def __str__(self):  # pragma: no cover
        return "FlowPathCount(%d) %s:%s %s count: %d" % (self.flow_id, self.from_uuid, self.to_uuid, self.period, self.count)

    class Meta:
        index_together = ['flow', 'from_uuid', 'to_uuid', 'period']


class FlowPathRecentRun(models.Model):
    """
    Maintains recent runs for a flow path segment
    """
    PRUNE_TO = 5
    LAST_PRUNED_KEY = 'last_recentrun_pruned'

    id = models.BigAutoField(auto_created=True, primary_key=True, verbose_name='ID')

    from_uuid = models.UUIDField(help_text=_("Which flow node they came from"))
    to_uuid = models.UUIDField(help_text=_("Which flow node they went to"))
    run = models.ForeignKey(FlowRun, related_name='recent_runs')
    visited_on = models.DateTimeField(help_text=_("When the run visited this path segment"), default=timezone.now)

    @classmethod
    def get_recent(cls, exit_uuids, to_uuid, limit=PRUNE_TO):
        """
        Gets the recent runs for the given flow segments
        """
        recent = (
            cls.objects.filter(from_uuid__in=exit_uuids, to_uuid=to_uuid)
            .select_related('run')
            .order_by('-visited_on')
        )
        if limit:
            recent = recent[:limit]

        # batch fetch all the messages for these runs
        message_ids = set()
        for r in recent:
            message_ids.update(r.run.get_message_ids())
        msgs = {m.id: m for m in Msg.objects.filter(id__in=message_ids).only('id', 'text', 'created_on')}

        results = []
        for r in recent:
            # find the most recent message in the run when this visit happened
            msg = None
            for msg_id in reversed(r.run.get_message_ids()):
                msg = msgs.get(msg_id)
                if msg and msg.created_on < r.visited_on:
                    break

            results.append({
                'run': r.run,
                'text': msg.text if msg else "",
                'visited_on': r.visited_on
            })

        return results

    @classmethod
    def prune(cls):
        """
        Removes old recent run records leaving only PRUNE_TO most recent for each segment
        """
        last_id = cache.get(cls.LAST_PRUNED_KEY, -1)

        newest = cls.objects.order_by('-id').values('id').first()
        newest_id = newest['id'] if newest else -1

        sql = """
            DELETE FROM %(table)s WHERE id IN (
              SELECT id FROM (
                  SELECT
                    r.id,
                    dense_rank() OVER (PARTITION BY from_uuid, to_uuid ORDER BY visited_on DESC) AS pos
                  FROM %(table)s r
                  WHERE (from_uuid, to_uuid) IN (
                    -- get the unique segments added to since last prune
                    SELECT DISTINCT from_uuid, to_uuid FROM %(table)s WHERE id > %(last_id)d
                  )
              ) s WHERE s.pos > %(limit)d
            )""" % {'table': cls._meta.db_table, 'last_id': last_id, 'limit': cls.PRUNE_TO}

        cursor = db_connection.cursor()
        cursor.execute(sql)

        cache.set(cls.LAST_PRUNED_KEY, newest_id)

        return cursor.rowcount  # number of deleted entries

    class Meta:
        indexes = [
            models.Index(fields=['from_uuid', 'to_uuid', '-visited_on'])
        ]


class FlowNodeCount(SquashableModel):
    """
    Maintains counts of unique contacts at each flow node.
    """
    SQUASH_OVER = ('node_uuid',)

    flow = models.ForeignKey(Flow)
    node_uuid = models.UUIDField(db_index=True)
    count = models.IntegerField(default=0)

    @classmethod
    def get_squash_query(cls, distinct_set):
        sql = """
        WITH removed as (
            DELETE FROM %(table)s WHERE "node_uuid" = %%s RETURNING "count"
        )
        INSERT INTO %(table)s("flow_id", "node_uuid", "count", "is_squashed")
        VALUES (%%s, %%s, GREATEST(0, (SELECT SUM("count") FROM removed)), TRUE);
        """ % {'table': cls._meta.db_table}

        return sql, (distinct_set.node_uuid, distinct_set.flow_id, distinct_set.node_uuid)

    @classmethod
    def get_totals(cls, flow):
        totals = list(cls.objects.filter(flow=flow).values_list('node_uuid').annotate(replies=Sum('count')))
        return {six.text_type(t[0]): t[1] for t in totals if t[1]}


@six.python_2_unicode_compatible
class FlowRunCount(SquashableModel):
    """
    Maintains counts of different states of exit types of flow runs on a flow. These are calculated
    via triggers on the database.
    """
    SQUASH_OVER = ('flow_id', 'exit_type')

    flow = models.ForeignKey(Flow, related_name='counts')
    exit_type = models.CharField(null=True, max_length=1, choices=FlowRun.EXIT_TYPE_CHOICES)
    count = models.IntegerField(default=0)

    @classmethod
    def get_squash_query(cls, distinct_set):
        if distinct_set.exit_type:
            sql = """
            WITH removed as (
                DELETE FROM %(table)s WHERE "flow_id" = %%s AND "exit_type" = %%s RETURNING "count"
            )
            INSERT INTO %(table)s("flow_id", "exit_type", "count", "is_squashed")
            VALUES (%%s, %%s, GREATEST(0, (SELECT SUM("count") FROM removed)), TRUE);
            """ % {'table': cls._meta.db_table}

            params = (distinct_set.flow_id, distinct_set.exit_type) * 2
        else:
            sql = """
            WITH removed as (
                DELETE FROM %(table)s WHERE "flow_id" = %%s AND "exit_type" IS NULL RETURNING "count"
            )
            INSERT INTO %(table)s("flow_id", "exit_type", "count", "is_squashed")
            VALUES (%%s, NULL, GREATEST(0, (SELECT SUM("count") FROM removed)), TRUE);
            """ % {'table': cls._meta.db_table}

            params = (distinct_set.flow_id,) * 2

        return sql, params

    @classmethod
    def get_totals(cls, flow):
        totals = list(cls.objects.filter(flow=flow).values_list('exit_type').annotate(replies=Sum('count')))
        totals = {t[0]: t[1] for t in totals}

        # for convenience, ensure dict contains all possible states
        all_states = (None, FlowRun.EXIT_TYPE_COMPLETED, FlowRun.EXIT_TYPE_EXPIRED, FlowRun.EXIT_TYPE_INTERRUPTED)
        totals = {s: totals.get(s, 0) for s in all_states}

        # we record active runs as exit_type=None but replace with actual constant for clarity
        totals[FlowRun.STATE_ACTIVE] = totals[None]
        del totals[None]

        return totals

    def __str__(self):  # pragma: needs cover
        return "RunCount[%d:%s:%d]" % (self.flow_id, self.exit_type, self.count)

    class Meta:
        index_together = ('flow', 'exit_type')


class ExportFlowResultsTask(BaseExportTask):
    """
    Container for managing our export requests
    """
    analytics_key = 'flowresult_export'
    email_subject = "Your results export is ready"
    email_template = 'flows/email/flow_export_download'

    INCLUDE_RUNS = 'include_runs'
    INCLUDE_MSGS = 'include_msgs'
    CONTACT_FIELDS = 'contact_fields'
    RESPONDED_ONLY = 'responded_only'
    EXTRA_URNS = 'extra_urns'

    flows = models.ManyToManyField(Flow, related_name='exports', help_text=_("The flows to export"))

    config = JSONAsTextField(null=True, default=dict,
                             help_text=_("Any configuration options for this flow export"))

    @classmethod
    def create(cls, org, user, flows, contact_fields, responded_only, include_runs, include_msgs, extra_urns):
        config = {ExportFlowResultsTask.INCLUDE_RUNS: include_runs,
                  ExportFlowResultsTask.INCLUDE_MSGS: include_msgs,
                  ExportFlowResultsTask.CONTACT_FIELDS: [c.id for c in contact_fields],
                  ExportFlowResultsTask.RESPONDED_ONLY: responded_only,
                  ExportFlowResultsTask.EXTRA_URNS: extra_urns}

        export = cls.objects.create(org=org, created_by=user, modified_by=user, config=config)
        for flow in flows:
            export.flows.add(flow)

        return export

    def get_email_context(self, branding):
        context = super(ExportFlowResultsTask, self).get_email_context(branding)
        context['flows'] = self.flows.all()
        return context

    def _get_runs_columns(self, extra_urn_columns, contact_fields, result_nodes, show_submitted_by=False, show_time=False):
        columns = []

        if show_submitted_by:
            columns.append(("Submitted By", self.WIDTH_MEDIUM))

        columns.append(("Contact UUID", self.WIDTH_MEDIUM))
        columns.append(("ID" if self.org.is_anon else "URN", self.WIDTH_SMALL))

        for extra_urn in extra_urn_columns:
            columns.append((extra_urn['label'], self.WIDTH_SMALL))

        columns.append(("Name", self.WIDTH_MEDIUM))
        columns.append(("Groups", self.WIDTH_MEDIUM))

        for cf in contact_fields:
            columns.append((cf.label, self.WIDTH_MEDIUM))

        if show_time:
            columns.append(("Started", self.WIDTH_MEDIUM))
            columns.append(("Exited", self.WIDTH_MEDIUM))

        for node in result_nodes:
            columns.append(("%s (Category) - %s" % (node.label, node.flow.name), self.WIDTH_SMALL))
            columns.append(("%s (Value) - %s" % (node.label, node.flow.name), self.WIDTH_SMALL))
            columns.append(("%s (Text) - %s" % (node.label, node.flow.name), self.WIDTH_SMALL))

        return columns

    def _add_runs_sheet(self, book, columns):
        name = "Runs (%d)" % (book.num_runs_sheets + 1) if book.num_runs_sheets > 0 else "Runs"
        sheet = book.create_sheet(name, index=book.num_runs_sheets)
        book.num_runs_sheets += 1

        self.set_sheet_column_widths(sheet, [c[1] for c in columns])
        self.append_row(sheet, [c[0] for c in columns])
        return sheet

    def _add_contacts_sheet(self, book, columns):
        name = "Contacts (%d)" % (book.num_contacts_sheets + 1) if book.num_contacts_sheets > 0 else "Contacts"
        sheet = book.create_sheet(name, index=(book.num_runs_sheets + book.num_contacts_sheets))
        book.num_contacts_sheets += 1

        self.set_sheet_column_widths(sheet, [c[1] for c in columns])
        self.append_row(sheet, [c[0] for c in columns])
        return sheet

    def _add_msgs_sheet(self, book):
        name = "Messages (%d)" % (book.num_msgs_sheets + 1) if book.num_msgs_sheets > 0 else "Messages"
        index = book.num_runs_sheets + book.num_contacts_sheets + book.num_msgs_sheets
        sheet = book.create_sheet(name, index)
        book.num_msgs_sheets += 1

        if self.org.is_anon:
            headers = ["Contact UUID", "ID", "Name", "Date", "Direction", "Message", "Channel"]
        else:
            headers = ["Contact UUID", "URN", "Name", "Date", "Direction", "Message", "Channel"]

        self.set_sheet_column_widths(sheet, [
            self.WIDTH_MEDIUM, self.WIDTH_SMALL, self.WIDTH_MEDIUM, self.WIDTH_MEDIUM, self.WIDTH_SMALL,
            self.WIDTH_LARGE, self.WIDTH_MEDIUM
        ])
        self.append_row(sheet, headers)
        return sheet

    def _get_contact_groups_display(self, contact):
        group_names = []
        for group in contact.all_groups.all():
            if group.group_type == ContactGroup.TYPE_USER_DEFINED:
                group_names.append(group.name)

        group_names.sort()
        return ", ".join(group_names)

    def _get_messages_for_runs(self, runs):
        """
        Batch fetches messages for the given runs and returns a dict of runs to messages
        """
        message_ids = set()
        for r in runs:
            message_ids.update(r.get_message_ids())

        messages = (
            Msg.objects.filter(id__in=message_ids, visibility=Msg.VISIBILITY_VISIBLE)
            .select_related('contact_urn')
            .prefetch_related('channel')
            .order_by('created_on')
        )

        msgs_by_id = {m.id: m for m in messages}

        msgs_by_run = defaultdict(list)
        for run in runs:
            for msg_id in run.get_message_ids():
                msg = msgs_by_id.get(msg_id)
                if msg:
                    msgs_by_run[run].append(msg)

        return msgs_by_run

    def write_export(self):
        config = self.config
        include_runs = config.get(ExportFlowResultsTask.INCLUDE_RUNS, False)
        include_msgs = config.get(ExportFlowResultsTask.INCLUDE_MSGS, False)
        responded_only = config.get(ExportFlowResultsTask.RESPONDED_ONLY, True)
        contact_field_ids = config.get(ExportFlowResultsTask.CONTACT_FIELDS, [])
        extra_urns = config.get(ExportFlowResultsTask.EXTRA_URNS, [])

        contact_fields = [cf for cf in self.org.cached_contact_fields if cf.id in contact_field_ids]

        # get all result saving nodes across all flows being exported
        show_submitted_by = False
        result_nodes = []
        flows = list(self.flows.filter(is_active=True).prefetch_related(
            Prefetch('rule_sets', RuleSet.objects.exclude(label=None).order_by('y', 'id'))
        ))
        for flow in flows:
            for node in flow.rule_sets.all():
                node.flow = flow
                result_nodes.append(node)

            if flow.flow_type == Flow.SURVEY:
                show_submitted_by = True

        extra_urn_columns = []
        if not self.org.is_anon:
            for extra_urn in extra_urns:
                label = ContactURN.EXPORT_FIELDS.get(extra_urn, dict()).get('label', '')
                extra_urn_columns.append(dict(label=label, scheme=extra_urn))

        runs_columns = self._get_runs_columns(extra_urn_columns, contact_fields, result_nodes, show_submitted_by=show_submitted_by, show_time=True)
        contacts_columns = self._get_runs_columns(extra_urn_columns, contact_fields, result_nodes)

        book = Workbook(write_only=True)
        book.num_runs_sheets = 0
        book.num_contacts_sheets = 0
        book.num_msgs_sheets = 0

        merged_result_values = [None, None, None] * len(result_nodes)

        current_contact = None
        current_contact_values = None

        # the current sheets
        runs_sheet = self._add_runs_sheet(book, runs_columns) if include_runs else None
        contacts_sheet = self._add_contacts_sheet(book, contacts_columns)
        msgs_sheet = None

        # grab the ids of the runs we're going to be exporting..
        runs = FlowRun.objects.filter(flow__in=flows).order_by('contact', 'id')
        if responded_only:
            runs = runs.filter(responded=True)
        run_ids = array(str('l'), runs.values_list('id', flat=True))

        # for tracking performance
        runs_exported = 0
        start = time.time()

        for id_batch in chunk_list(run_ids, 1000):
            run_batch = (
                FlowRun.objects.filter(id__in=id_batch, contact__is_test=False)
                .select_related('contact')
                .order_by('contact', 'id')
            )

            msgs_by_run = self._get_messages_for_runs(run_batch)

            for run in run_batch:
                # is this a new contact?
                if run.contact != current_contact:
                    if not contacts_sheet or contacts_sheet._max_row >= self.MAX_EXCEL_ROWS:  # pragma: no cover
                        contacts_sheet = self._add_contacts_sheet(book, contacts_columns)

                    if current_contact:
                        merged_sheet_row = []
                        merged_sheet_row += current_contact_values
                        merged_sheet_row += merged_result_values

                        self.append_row(contacts_sheet, merged_sheet_row)

                        merged_result_values = [None, None, None] * len(result_nodes)

                    current_contact = run.contact

                    # generate current contact info columns
                    current_contact_values = [
                        run.contact.uuid,
                        run.contact.id if self.org.is_anon else run.contact.get_urn_display(org=self.org, formatted=False)
                    ]

                    for extra_urn_column in extra_urn_columns:
                        urn_display = run.contact.get_urn_display(org=self.org, formatted=False, scheme=extra_urn_column['scheme'])
                        current_contact_values.append(urn_display)

                    current_contact_values.append(self.prepare_value(run.contact.name))
                    current_contact_values.append(self._get_contact_groups_display(run.contact))

                    for cf in contact_fields:
                        field_value = Contact.get_field_display_for_value(cf, run.contact.get_field(cf.key.lower()), self.org)
                        current_contact_values.append(self.prepare_value(field_value))

                # get this run's results by node UUID
                results_by_node = {result[FlowRun.RESULT_NODE_UUID]: result for result in run.results.values()}

                result_values = []
                for n, node in enumerate(result_nodes):
                    node_result = results_by_node.get(node.uuid, {})
                    node_category = node_result.get(FlowRun.RESULT_CATEGORY, "")
                    node_value = node_result.get(FlowRun.RESULT_VALUE, "")
                    node_input = node_result.get(FlowRun.RESULT_INPUT, "")
                    result_values += [node_category, node_value, node_input]

                    if node_result:
                        merged_result_values[n * 3 + 0] = node_category
                        merged_result_values[n * 3 + 1] = node_value
                        merged_result_values[n * 3 + 2] = node_input

                if include_runs:
                    if not runs_sheet or runs_sheet._max_row >= self.MAX_EXCEL_ROWS:  # pragma: no cover
                        runs_sheet = self._add_runs_sheet(book, runs_columns)

                    # build the whole row
                    runs_sheet_row = []

                    if show_submitted_by:
                        runs_sheet_row.append(run.submitted_by.username if run.submitted_by else '')

                    runs_sheet_row += current_contact_values
                    runs_sheet_row += [run.created_on, run.exited_on]
                    runs_sheet_row += result_values

                    self.append_row(runs_sheet, runs_sheet_row)

                # write out any message associated with this run
                if include_msgs:
                    msgs_sheet = self._write_run_messages(book, msgs_sheet, run, msgs_by_run)

                runs_exported += 1
                if runs_exported % 10000 == 0:  # pragma: needs cover
                    print("Result export of for org #%d - %d%% complete in %0.2fs" %
                          (self.org.id, runs_exported * 100 // len(run_ids), time.time() - start))

        if current_contact:
            merged_sheet_row = []
            merged_sheet_row += current_contact_values
            merged_sheet_row += merged_result_values
            self.append_row(contacts_sheet, merged_sheet_row)

        temp = NamedTemporaryFile(delete=True)
        book.save(temp)
        temp.flush()
        return temp, 'xlsx'

    def _write_run_messages(self, book, msgs_sheet, run, msgs_by_run):
        """
        Writes out any messages associated with the given run
        """
        for msg in msgs_by_run.get(run, []):
            if not msgs_sheet or msgs_sheet._max_row >= self.MAX_EXCEL_ROWS:
                msgs_sheet = self._add_msgs_sheet(book)

            if self.org.is_anon:
                urn_display = run.contact.id
            else:
                urn_display = msg.contact_urn.get_display(org=self.org, formatted=False) if msg.contact_urn else ''

            self.append_row(msgs_sheet, [
                run.contact.uuid,
                urn_display,
                self.prepare_value(run.contact.name),
                msg.created_on,
                "IN" if msg.direction == INCOMING else "OUT",
                msg.text,
                msg.channel.name if msg.channel else ''
            ])

        return msgs_sheet


@register_asset_store
class ResultsExportAssetStore(BaseExportAssetStore):
    model = ExportFlowResultsTask
    key = 'results_export'
    directory = 'results_exports'
    permission = 'flows.flow_export_results'
    extensions = ('xlsx',)


@six.python_2_unicode_compatible
class ActionLog(models.Model):
    """
    Log of an event that occurred whilst executing a flow in the simulator
    """
    LEVEL_INFO = 'I'
    LEVEL_WARN = 'W'
    LEVEL_ERROR = 'E'
    LEVEL_CHOICES = ((LEVEL_INFO, _("Info")), (LEVEL_WARN, _("Warning")), (LEVEL_ERROR, _("Error")))

    run = models.ForeignKey(FlowRun, related_name='logs')

    text = models.TextField(help_text=_("Log event text"))

    level = models.CharField(max_length=1, choices=LEVEL_CHOICES, default=LEVEL_INFO, help_text=_("Log event level"))

    created_on = models.DateTimeField(help_text=_("When this log event occurred"))

    @classmethod
    def create(cls, run, text, level=LEVEL_INFO, safe=False, created_on=None):
        if not safe:
            text = escape(text)

        text = text.replace('\n', "<br/>")

        if not created_on:
            created_on = timezone.now()

        try:
            return ActionLog.objects.create(run=run, text=text, level=level, created_on=created_on)
        except Exception:  # pragma: no cover
            return None  # it's possible our test run can be deleted out from under us

    @classmethod
    def info(cls, run, text, safe=False):
        return cls.create(run, text, cls.LEVEL_INFO, safe)

    @classmethod
    def warn(cls, run, text, safe=False):
        return cls.create(run, text, cls.LEVEL_WARN, safe)

    @classmethod
    def error(cls, run, text, safe=False):
        return cls.create(run, text, cls.LEVEL_ERROR, safe)

    def as_json(self):
        return dict(id=self.id,
                    direction="O",
                    level=self.level,
                    text=self.text,
                    created_on=self.created_on.strftime('%x %X'),
                    model="log")

    def simulator_json(self):
        return self.as_json()

    def __str__(self):  # pragma: needs cover
        return self.text


@six.python_2_unicode_compatible
class FlowStart(SmartModel):
    STATUS_PENDING = 'P'
    STATUS_STARTING = 'S'
    STATUS_COMPLETE = 'C'
    STATUS_FAILED = 'F'

    STATUS_CHOICES = ((STATUS_PENDING, "Pending"),
                      (STATUS_STARTING, "Starting"),
                      (STATUS_COMPLETE, "Complete"),
                      (STATUS_FAILED, "Failed"))

    uuid = models.UUIDField(unique=True, default=uuid4)

    flow = models.ForeignKey(Flow, related_name='starts', help_text=_("The flow that is being started"))

    groups = models.ManyToManyField(ContactGroup, help_text=_("Groups that will start the flow"))

    contacts = models.ManyToManyField(Contact, help_text=_("Contacts that will start the flow"))

    restart_participants = models.BooleanField(default=True,
                                               help_text=_("Whether to restart any participants already in this flow"))

    include_active = models.BooleanField(default=True,
                                         help_text=_("Include contacts currently active in flows"))

    contact_count = models.IntegerField(default=0,
                                        help_text=_("How many unique contacts were started down the flow"))

    status = models.CharField(max_length=1, default=STATUS_PENDING, choices=STATUS_CHOICES,
                              help_text=_("The status of this flow start"))

    extra = JSONAsTextField(null=True, default=dict,
                            help_text=_("Any extra parameters to pass to the flow start (json)"))

    @classmethod
    def create(cls, flow, user, groups=None, contacts=None, restart_participants=True, extra=None, include_active=True):
        if contacts is None:  # pragma: needs cover
            contacts = []

        if groups is None:  # pragma: needs cover
            groups = []

        start = FlowStart.objects.create(flow=flow,
                                         restart_participants=restart_participants,
                                         include_active=include_active,
                                         extra=extra,
                                         created_by=user, modified_by=user)

        for contact in contacts:
            start.contacts.add(contact)

        for group in groups:
            start.groups.add(group)

        return start

    def async_start(self):
        from temba.flows.tasks import start_flow_task
        on_transaction_commit(lambda: start_flow_task.delay(self.id))

    def start(self):
        self.status = FlowStart.STATUS_STARTING
        self.save(update_fields=['status'])

        try:
            groups = [g for g in self.groups.all()]
            contacts = [c for c in self.contacts.all().only('is_test')]

            # load up our extra if any
            extra = self.extra if self.extra else None

            return self.flow.start(groups, contacts, flow_start=self, extra=extra,
                                   restart_participants=self.restart_participants, include_active=self.include_active)

        except Exception as e:  # pragma: no cover
            import traceback
            traceback.print_exc()

            self.status = FlowStart.STATUS_FAILED
            self.save(update_fields=['status'])
            raise e

    def update_status(self):
        # only update our status to complete if we have started as many runs as our total contact count
        if self.runs.count() == self.contact_count:
            self.status = FlowStart.STATUS_COMPLETE
            self.save(update_fields=['status'])

    def __str__(self):  # pragma: no cover
        return "FlowStart %d (Flow %d)" % (self.id, self.flow_id)


@six.python_2_unicode_compatible
class FlowLabel(models.Model):
    org = models.ForeignKey(Org)

    uuid = models.CharField(max_length=36, unique=True, db_index=True, default=generate_uuid,
                            verbose_name=_("Unique Identifier"), help_text=_("The unique identifier for this label"))
    name = models.CharField(max_length=64, verbose_name=_("Name"),
                            help_text=_("The name of this flow label"))
    parent = models.ForeignKey('FlowLabel', verbose_name=_("Parent"), null=True, related_name="children")

    def get_flows_count(self):
        """
        Returns the count of flows tagged with this label or one of its children
        """
        return self.get_flows().count()

    def get_flows(self):
        return Flow.objects.filter(Q(labels=self) | Q(labels__parent=self)).filter(is_active=True, is_archived=False).distinct()

    @classmethod
    def create_unique(cls, base, org, parent=None):

        base = base.strip()

        # truncate if necessary
        if len(base) > 32:
            base = base[:32]

        # find the next available label by appending numbers
        count = 2
        while FlowLabel.objects.filter(name=base, org=org, parent=parent):
            # make room for the number
            if len(base) >= 32:
                base = base[:30]
            last = str(count - 1)
            if base.endswith(last):
                base = base[:-len(last)]
            base = "%s %d" % (base.strip(), count)
            count += 1

        return FlowLabel.objects.create(name=base, org=org, parent=parent)

    def toggle_label(self, flows, add):
        changed = []

        for flow in flows:
            # if we are adding the flow label and this flow doesnt have it, add it
            if add:
                if not flow.labels.filter(pk=self.pk):
                    flow.labels.add(self)
                    changed.append(flow.pk)

            # otherwise, remove it if not already present
            else:
                if flow.labels.filter(pk=self.pk):
                    flow.labels.remove(self)
                    changed.append(flow.pk)

        return changed

    def __str__(self):
        if self.parent:
            return "%s > %s" % (self.parent, self.name)
        return self.name

    class Meta:
        unique_together = ('name', 'parent', 'org')


__flow_users = None


def clear_flow_users():
    global __flow_users
    __flow_users = None


def get_flow_user(org):
    global __flow_users
    if not __flow_users:
        __flow_users = {}

    branding = org.get_branding()
    username = '%s_flow' % branding['slug']
    flow_user = __flow_users.get(username)

    # not cached, let's look it up
    if not flow_user:
        email = branding['support_email']
        flow_user = User.objects.filter(username=username).first()
        if flow_user:  # pragma: needs cover
            __flow_users[username] = flow_user
        else:
            # doesn't exist for this brand, create it
            flow_user = User.objects.create_user(username, email, first_name='System Update')
            flow_user.groups.add(Group.objects.get(name='Service Users'))
            __flow_users[username] = flow_user

    return flow_user


class Action(object):
    """
    Base class for actions that can be added to an action set and executed during a flow run
    """
    TYPE = 'type'
    UUID = 'uuid'

    __action_mapping = None

    def __init__(self, uuid):
        self.uuid = uuid if uuid else str(uuid4())

    @classmethod
    def from_json(cls, org, json_obj):
        if not cls.__action_mapping:
            cls.__action_mapping = {
                ReplyAction.TYPE: ReplyAction,
                SendAction.TYPE: SendAction,
                AddToGroupAction.TYPE: AddToGroupAction,
                DeleteFromGroupAction.TYPE: DeleteFromGroupAction,
                AddLabelAction.TYPE: AddLabelAction,
                EmailAction.TYPE: EmailAction,
                WebhookAction.TYPE: WebhookAction,
                SaveToContactAction.TYPE: SaveToContactAction,
                SetLanguageAction.TYPE: SetLanguageAction,
                SetChannelAction.TYPE: SetChannelAction,
                StartFlowAction.TYPE: StartFlowAction,
                SayAction.TYPE: SayAction,
                PlayAction.TYPE: PlayAction,
                TriggerFlowAction.TYPE: TriggerFlowAction,
                EndUssdAction.TYPE: EndUssdAction,
            }

        action_type = json_obj.get(cls.TYPE)
        if not action_type:  # pragma: no cover
            raise FlowException("Action definition missing 'type' attribute: %s" % json_obj)

        if action_type not in cls.__action_mapping:  # pragma: no cover
            raise FlowException("Unknown action type '%s' in definition: '%s'" % (action_type, json_obj))

        return cls.__action_mapping[action_type].from_json(org, json_obj)

    @classmethod
    def from_json_array(cls, org, json_arr):
        actions = []
        for inner in json_arr:
            action = Action.from_json(org, inner)
            if action:
                actions.append(action)
        return actions


class EmailAction(Action):
    """
    Sends an email to someone
    """
    TYPE = 'email'
    EMAILS = 'emails'
    SUBJECT = 'subject'
    MESSAGE = 'msg'

    def __init__(self, uuid, emails, subject, message):
        super(EmailAction, self).__init__(uuid)

        if not emails:
            raise FlowException("Email actions require at least one recipient")

        self.emails = emails
        self.subject = subject
        self.message = message

    @classmethod
    def from_json(cls, org, json_obj):
        emails = json_obj.get(EmailAction.EMAILS)
        message = json_obj.get(EmailAction.MESSAGE)
        subject = json_obj.get(EmailAction.SUBJECT)
        return cls(json_obj.get(cls.UUID), emails, subject, message)

    def as_json(self):
        return dict(type=self.TYPE, uuid=self.uuid, emails=self.emails, subject=self.subject, msg=self.message)

    def execute(self, run, context, actionset_uuid, msg, offline_on=None):
        from .tasks import send_email_action_task

        # build our message from our flow variables
        (message, errors) = Msg.evaluate_template(self.message, context, org=run.flow.org)
        (subject, errors) = Msg.evaluate_template(self.subject, context, org=run.flow.org)

        # make sure the subject is single line; replace '\t\n\r\f\v' to ' '
        subject = regex.sub('\s+', ' ', subject, regex.V0)

        valid_addresses = []
        invalid_addresses = []
        for email in self.emails:
            if email.startswith('@'):
                # a valid email will contain @ so this is very likely to generate evaluation errors
                (address, errors) = Msg.evaluate_template(email, context, org=run.flow.org)
            else:
                address = email

            address = address.strip()

            if is_valid_address(address):
                valid_addresses.append(address)
            else:
                invalid_addresses.append(address)

        if not run.contact.is_test:
            if valid_addresses:
                on_transaction_commit(lambda: send_email_action_task.delay(run.flow.org.id, valid_addresses, subject, message))
        else:
            if valid_addresses:
                valid_addresses = ['"%s"' % elt for elt in valid_addresses]
                ActionLog.info(run, _("\"%s\" would be sent to %s") % (message, ", ".join(valid_addresses)))
            if invalid_addresses:
                invalid_addresses = ['"%s"' % elt for elt in invalid_addresses]
                ActionLog.warn(run, _("Some email address appear to be invalid: %s") % ", ".join(invalid_addresses))
        return []


class WebhookAction(Action):
    """
    Forwards the steps in this flow to the webhook (if any)
    """
    TYPE = 'api'
    ACTION = 'action'

    def __init__(self, uuid, webhook, action='POST', webhook_headers=None):
        super(WebhookAction, self).__init__(uuid)

        self.webhook = webhook
        self.action = action
        self.webhook_headers = webhook_headers

    @classmethod
    def from_json(cls, org, json_obj):
        return cls(json_obj.get(cls.UUID),
                   json_obj.get('webhook', org.get_webhook_url()),
                   json_obj.get('action', 'POST'),
                   json_obj.get('webhook_headers', []))

    def as_json(self):
        return dict(type=self.TYPE, uuid=self.uuid, webhook=self.webhook, action=self.action,
                    webhook_headers=self.webhook_headers)

    def execute(self, run, context, actionset_uuid, msg, offline_on=None):
        from temba.api.models import WebHookEvent

        (value, errors) = Msg.evaluate_template(self.webhook, context, org=run.flow.org, url_encode=True)

        if errors:
            ActionLog.warn(run, _("URL appears to contain errors: %s") % ", ".join(errors))

        headers = {}
        if self.webhook_headers:
            for item in self.webhook_headers:
                headers[item.get('name')] = item.get('value')

        WebHookEvent.trigger_flow_webhook(run, value, actionset_uuid, msg, self.action, headers=headers)
        return []


class AddToGroupAction(Action):
    """
    Adds the user to a group
    """
    TYPE = 'add_group'
    GROUP = 'group'
    GROUPS = 'groups'

    def __init__(self, uuid, groups):
        super(AddToGroupAction, self).__init__(uuid)

        self.groups = groups

    @classmethod
    def from_json(cls, org, json_obj):
        return cls(json_obj.get(cls.UUID), cls.get_groups(org, json_obj))

    @classmethod
    def get_groups(cls, org, json_obj):

        # for backwards compatibility
        group_data = json_obj.get(AddToGroupAction.GROUP, None)
        if not group_data:
            group_data = json_obj.get(AddToGroupAction.GROUPS)
        else:
            group_data = [group_data]

        groups = []

        for g in group_data:
            if isinstance(g, dict):
                group_uuid = g.get('uuid', None)
                group_name = g.get('name')

                group = ContactGroup.get_or_create(org, org.created_by, group_name, group_uuid)
                groups.append(group)
            else:
                if g and g[0] == '@':
                    groups.append(g)
                else:  # pragma: needs cover
                    group = ContactGroup.get_user_group(org, g)
                    if group:
                        groups.append(group)
                    else:
                        groups.append(ContactGroup.create_static(org, org.get_user(), g))
        return groups

    def as_json(self):
        groups = []
        for g in self.groups:
            if isinstance(g, ContactGroup):
                groups.append(dict(uuid=g.uuid, name=g.name))
            else:
                groups.append(g)

        return dict(type=self.get_type(), uuid=self.uuid, groups=groups)

    def get_type(self):
        return AddToGroupAction.TYPE

    def execute(self, run, context, actionset_uuid, msg, offline_on=None):
        contact = run.contact
        add = AddToGroupAction.TYPE == self.get_type()
        user = get_flow_user(run.org)

        if contact:
            for group in self.groups:
                if not isinstance(group, ContactGroup):
                    (value, errors) = Msg.evaluate_template(group, context, org=run.flow.org)
                    group = None

                    if not errors:
                        group = ContactGroup.get_user_group(contact.org, value)
                        if not group:
                            ActionLog.error(run, _("Unable to find group with name '%s'") % value)

                    else:  # pragma: needs cover
                        ActionLog.error(run, _("Group name could not be evaluated: %s") % ', '.join(errors))

                if group:
                    # TODO should become a failure (because it should be impossible) and not just a simulator error
                    if group.is_dynamic:
                        # report to sentry
                        logger.error("Attempt to add/remove contacts on dynamic group '%s' [%d] "
                                     "in flow '%s' [%d] for org '%s' [%d]"
                                     % (group.name, group.pk, run.flow.name, run.flow.pk, run.org.name, run.org.pk))
                        if run.contact.is_test:
                            if add:
                                ActionLog.error(run, _("%s is a dynamic group which we can't add contacts to") % group.name)
                            else:  # pragma: needs cover
                                ActionLog.error(run, _("%s is a dynamic group which we can't remove contacts from") % group.name)
                        continue

                    group.org = run.org
                    group.update_contacts(user, [contact], add)
                    if run.contact.is_test:
                        if add:
                            ActionLog.info(run, _("Added %s to %s") % (run.contact.name, group.name))
                        else:
                            ActionLog.info(run, _("Removed %s from %s") % (run.contact.name, group.name))
        return []


class DeleteFromGroupAction(AddToGroupAction):
    """
    Removes the user from a group
    """
    TYPE = 'del_group'

    def get_type(self):
        return DeleteFromGroupAction.TYPE

    def as_json(self):
        groups = []
        for g in self.groups:
            if isinstance(g, ContactGroup):
                groups.append(dict(uuid=g.uuid, name=g.name))
            else:
                groups.append(g)

        return dict(type=self.get_type(), uuid=self.uuid, groups=groups)

    @classmethod
    def from_json(cls, org, json_obj):
        return cls(json_obj.get(cls.UUID), cls.get_groups(org, json_obj))

    def execute(self, run, context, actionset, msg, offline_on=None):
        if len(self.groups) == 0:
            contact = run.contact
            user = get_flow_user(run.org)
            if contact:
                # remove from all active and inactive user-defined, static groups
                for group in ContactGroup.user_groups.filter(org=contact.org,
                                                             group_type=ContactGroup.TYPE_USER_DEFINED,
                                                             query__isnull=True):
                    group.update_contacts(user, [contact], False)
                    if run.contact.is_test:  # pragma: needs cover
                        ActionLog.info(run, _("Removed %s from %s") % (run.contact.name, group.name))
            return []
        return AddToGroupAction.execute(self, run, context, actionset, msg, offline_on)


class AddLabelAction(Action):
    """
    Add a label to the incoming message
    """
    TYPE = 'add_label'
    LABELS = 'labels'

    def __init__(self, uuid, labels):
        super(AddLabelAction, self).__init__(uuid)

        self.labels = labels

    @classmethod
    def from_json(cls, org, json_obj):
        labels_data = json_obj.get(cls.LABELS)

        labels = []
        for label_data in labels_data:
            if isinstance(label_data, dict):
                label_uuid = label_data.get('uuid', None)
                label_name = label_data.get('name')

                if label_uuid and Label.label_objects.filter(org=org, uuid=label_uuid).first():
                    label = Label.label_objects.filter(org=org, uuid=label_uuid).first()
                    if label:
                        labels.append(label)
                else:
                    labels.append(Label.get_or_create(org, org.get_user(), label_name))

            elif isinstance(label_data, six.string_types):
                if label_data and label_data[0] == '@':
                    # label name is a variable substitution
                    labels.append(label_data)
                else:  # pragma: needs cover
                    labels.append(Label.get_or_create(org, org.get_user(), label_data))
            else:  # pragma: needs cover
                raise ValueError("Label data must be a dict or string")

        return cls(json_obj.get(cls.UUID), labels)

    def as_json(self):
        labels = []
        for action_label in self.labels:
            if isinstance(action_label, Label):
                labels.append(dict(uuid=action_label.uuid, name=action_label.name))
            else:
                labels.append(action_label)

        return dict(type=self.get_type(), uuid=self.uuid, labels=labels)

    def get_type(self):
        return AddLabelAction.TYPE

    def execute(self, run, context, actionset_uuid, msg, offline_on=None):
        for label in self.labels:
            if not isinstance(label, Label):
                contact = run.contact
                (value, errors) = Msg.evaluate_template(label, context, org=run.flow.org)

                if not errors:
                    label = Label.label_objects.filter(org=contact.org, name__iexact=value.strip()).first()
                    if not label:
                        ActionLog.error(run, _("Unable to find label with name '%s'") % value.strip())

                else:  # pragma: needs cover
                    label = None
                    ActionLog.error(run, _("Label name could not be evaluated: %s") % ', '.join(errors))

            if label and msg and msg.pk:
                if run.contact.is_test:  # pragma: needs cover
                    # don't really add labels to simulator messages
                    ActionLog.info(run, _("Added %s label to msg '%s'") % (label.name, msg.text))
                else:
                    label.toggle_label([msg], True)
        return []


class SayAction(Action):
    """
    Voice action for reading some text to a user
    """
    TYPE = 'say'
    MESSAGE = 'msg'
    RECORDING = 'recording'

    def __init__(self, uuid, msg, recording):
        super(SayAction, self).__init__(uuid)

        self.msg = msg
        self.recording = recording

    @classmethod
    def from_json(cls, org, json_obj):
        return cls(json_obj.get(cls.UUID), json_obj.get(cls.MESSAGE), json_obj.get(cls.RECORDING))

    def as_json(self):
        return dict(type=self.TYPE, uuid=self.uuid, msg=self.msg, recording=self.recording)

    def execute(self, run, context, actionset_uuid, event, offline_on=None):

        media_url = None
        if self.recording:

            # localize our recording
            recording = run.flow.get_localized_text(self.recording, run.contact)

            # if we have a localized recording, create the url
            if recording:  # pragma: needs cover
                media_url = "https://%s/%s" % (settings.AWS_BUCKET_DOMAIN, recording)

        # localize the text for our message, need this either way for logging
        message = run.flow.get_localized_text(self.msg, run.contact)
        (message, errors) = Msg.evaluate_template(message, context)

        msg = run.create_outgoing_ivr(message, media_url, run.connection)

        if msg:
            if run.contact.is_test:
                if media_url:  # pragma: needs cover
                    ActionLog.create(run, _('Played recorded message for "%s"') % message)
                else:
                    ActionLog.create(run, _('Read message "%s"') % message)
            return [msg]
        else:  # pragma: needs cover
            # no message, possibly failed loop detection
            run.voice_response.say(_("Sorry, an invalid flow has been detected. Good bye."))
            return []


class PlayAction(Action):
    """
    Voice action for reading some text to a user
    """
    TYPE = 'play'
    URL = 'url'

    def __init__(self, uuid, url):
        super(PlayAction, self).__init__(uuid)

        self.url = url

    @classmethod
    def from_json(cls, org, json_obj):
        return cls(json_obj.get(cls.UUID), json_obj.get(cls.URL))

    def as_json(self):
        return dict(type=self.TYPE, uuid=self.uuid, url=self.url)

    def execute(self, run, context, actionset_uuid, event, offline_on=None):
        (recording_url, errors) = Msg.evaluate_template(self.url, context)
        msg = run.create_outgoing_ivr(_('Played contact recording'), recording_url, run.connection)

        if msg:
            if run.contact.is_test:
                log_txt = _('Played recording at "%s"') % recording_url
                ActionLog.create(run, log_txt)
            return [msg]
        else:  # pragma: needs cover
            # no message, possibly failed loop detection
            run.voice_response.say(_("Sorry, an invalid flow has been detected. Good bye."))
            return []


class ReplyAction(Action):
    """
    Simple action for sending back a message
    """
    TYPE = 'reply'
    MESSAGE = 'msg'
    MSG_TYPE = None
    MEDIA = 'media'
    SEND_ALL = 'send_all'
    QUICK_REPLIES = 'quick_replies'

    def __init__(self, uuid, msg=None, media=None, quick_replies=None, send_all=False):
        super(ReplyAction, self).__init__(uuid)

        self.msg = msg
        self.media = media if media else {}
        self.send_all = send_all
        self.quick_replies = quick_replies if quick_replies else []

    @classmethod
    def from_json(cls, org, json_obj):
        # assert we have some kind of message in this reply
        msg = json_obj.get(cls.MESSAGE)
        if isinstance(msg, dict):
            if not msg:
                raise FlowException("Invalid reply action, empty message dict")

            if not any([v for v in msg.values()]):
                raise FlowException("Invalid reply action, missing at least one message")
        elif not msg:
            raise FlowException("Invalid reply action, no message")

        return cls(json_obj.get(cls.UUID), msg=json_obj.get(cls.MESSAGE), media=json_obj.get(cls.MEDIA, None),
                   quick_replies=json_obj.get(cls.QUICK_REPLIES), send_all=json_obj.get(cls.SEND_ALL, False))

    def as_json(self):
        return dict(type=self.TYPE, uuid=self.uuid, msg=self.msg, media=self.media, quick_replies=self.quick_replies,
                    send_all=self.send_all)

    @staticmethod
    def get_translated_quick_replies(metadata, run):
        """
        Gets the appropriate metadata translation for the given contact
        """
        language_metadata = []
        preferred_languages = [run.contact.language, run.flow.base_language]
        for item in metadata:
            text = Language.get_localized_text(text_translations=item, preferred_languages=preferred_languages)
            language_metadata.append(text)

        return language_metadata

    def execute(self, run, context, actionset_uuid, msg, offline_on=None):
        replies = []

        if self.msg or self.media:
            user = get_flow_user(run.org)

            text = ''
            if self.msg:
                text = run.flow.get_localized_text(self.msg, run.contact)

            quick_replies = []
            if self.quick_replies:
                quick_replies = ReplyAction.get_translated_quick_replies(self.quick_replies, run)

            attachments = None
            if self.media:
                # localize our media attachment
                media_type, media_url = run.flow.get_localized_text(self.media, run.contact).split(':', 1)

                # if we have a localized media, create the url
                if media_url and len(media_type.split('/')) > 1:
                    attachments = ["%s:https://%s/%s" % (media_type, settings.AWS_BUCKET_DOMAIN, media_url)]
                else:
                    attachments = ["%s:%s" % (media_type, media_url)]

            if offline_on:
                context = None
                created_on = offline_on
            else:
                created_on = None

            if msg and msg.id:
                replies = msg.reply(text, user, trigger_send=False, expressions_context=context,
                                    connection=run.connection, msg_type=self.MSG_TYPE, quick_replies=quick_replies,
                                    attachments=attachments, send_all=self.send_all, created_on=created_on)
            else:
                # if our run has been responded to or any of our parent runs have
                # been responded to consider us interactive with high priority
                high_priority = run.get_session_responded()
                replies = run.contact.send(text, user, trigger_send=False, expressions_context=context,
                                           connection=run.connection, msg_type=self.MSG_TYPE, attachments=attachments,
                                           quick_replies=quick_replies, created_on=created_on, all_urns=self.send_all,
                                           high_priority=high_priority)
        return replies


class EndUssdAction(ReplyAction):
    """
    Reply action that ends a USSD session gracefully with a message
    """
    TYPE = 'end_ussd'
    MSG_TYPE = MSG_TYPE_USSD


class UssdAction(ReplyAction):
    """
    USSD action to send outgoing USSD messages
    Created from a USSD ruleset
    It builds localised text with localised USSD menu support
    """
    TYPE = 'ussd'
    MESSAGE = 'ussd_message'
    TYPE_WAIT_USSD_MENU = 'wait_menu'
    TYPE_WAIT_USSD = 'wait_ussd'
    MSG_TYPE = MSG_TYPE_USSD

    def __init__(self, uuid=None, msg=None, base_language=None, languages=None, primary_language=None):
        super(UssdAction, self).__init__(uuid, msg)

        self.languages = languages
        if msg and base_language and primary_language:
            self.base_language = base_language if base_language in msg else primary_language
        else:
            self.base_language = None

    @classmethod
    def from_ruleset(cls, ruleset, run):
        if ruleset and hasattr(ruleset, 'config') and ruleset.config != {}:
            # initial message, menu obj
            rules = ruleset.rules
            msg = ruleset.config.get(cls.MESSAGE, '')
            org = run.flow.org

            # TODO: this will be arbitrary unless UI is changed to maintain consistent uuids
            uuid = ruleset.config.get(cls.UUID, six.text_type(uuid4()))

            # define languages
            base_language = run.flow.base_language
            org_languages = {l.iso_code for l in org.languages.all()}
            primary_language = getattr(getattr(org, 'primary_language', None), 'iso_code', None)

            # initialize UssdAction
            ussd_action = cls(uuid=uuid, msg=msg, base_language=base_language, languages=org_languages,
                              primary_language=primary_language)

            ussd_action.substitute_missing_languages()

            if ruleset.ruleset_type == cls.TYPE_WAIT_USSD_MENU:
                ussd_action.add_menu_to_msg(rules)

            return ussd_action
        else:
            return cls()

    def substitute_missing_languages(self):
        # if there is a translation missing fill it with the base language
        for language in self.languages:
            if language not in self.msg:
                self.msg[language] = self.msg.get(self.base_language)

    def get_menu_label(self, label, language):
        if language not in label:
            return str(label.get(self.base_language))
        else:
            return str(label[language])

    def add_menu_to_msg(self, rules):
        # start with a new line
        self.msg = {language: localised_msg + '\n' for language, localised_msg in six.iteritems(self.msg)}

        # add menu to the msg
        for rule in rules:
            if rule.get('label'):  # filter "other" and "interrupted"
                self.msg = {language: localised_msg + ": ".join(
                    (str(rule['test']['test']), self.get_menu_label(rule['label'], language),)) + '\n' for language, localised_msg in six.iteritems(self.msg)}


class VariableContactAction(Action):
    """
    Base action that resolves variables into contacts. Used for actions that take
    SendAction, TriggerAction, etc
    """
    CONTACTS = 'contacts'
    GROUPS = 'groups'
    VARIABLES = 'variables'
    PHONE = 'phone'
    PATH = 'path'
    SCHEME = 'scheme'
    URNS = 'urns'
    NAME = 'name'
    ID = 'id'

    def __init__(self, uuid, groups, contacts, variables):
        super(VariableContactAction, self).__init__(uuid)

        self.groups = groups
        self.contacts = contacts
        self.variables = variables

    @classmethod
    def parse_groups(cls, org, json_obj):
        # we actually instantiate our contacts here
        groups = []
        for group_data in json_obj.get(VariableContactAction.GROUPS):
            group_uuid = group_data.get(VariableContactAction.UUID, None)
            group_name = group_data.get(VariableContactAction.NAME)

            # flows from when true deletion was allowed need this
            if not group_name:
                group_name = 'Missing'

            group = ContactGroup.get_or_create(org, org.get_user(), group_name, group_uuid)
            groups.append(group)

        return groups

    @classmethod
    def parse_contacts(cls, org, json_obj):
        contacts = []
        for contact in json_obj.get(VariableContactAction.CONTACTS):
            name = contact.get(VariableContactAction.NAME, None)
            phone = contact.get(VariableContactAction.PHONE, None)
            contact_uuid = contact.get(VariableContactAction.UUID, None)

            urns = []
            for urn in contact.get(VariableContactAction.URNS, []):
                scheme = urn.get(VariableContactAction.SCHEME)
                path = urn.get(VariableContactAction.PATH)

                if scheme and path:
                    urns.append(URN.from_parts(scheme, path))

            contact = Contact.objects.filter(uuid=contact_uuid, org=org).first()

            if not contact and (phone or urn):
                if phone:  # pragma: needs cover
                    contact = Contact.get_or_create_by_urns(org, org.created_by, name=None, urns=[URN.from_tel(phone)])
                elif urns:
                    contact = Contact.get_or_create_by_urns(org, org.created_by, name=None, urns=urns)

                # if they dont have a name use the one in our action
                if name and not contact.name:  # pragma: needs cover
                    contact.name = name
                    contact.save(update_fields=['name'])

            if contact:
                contacts.append(contact)

        return contacts

    @classmethod
    def parse_variables(cls, org, json_obj):
        variables = []
        if VariableContactAction.VARIABLES in json_obj:
            variables = list(_.get(VariableContactAction.ID) for _ in json_obj.get(VariableContactAction.VARIABLES))
        return variables

    def build_groups_and_contacts(self, run, msg):
        expressions_context = run.flow.build_expressions_context(run.contact, msg, run=run)
        contacts = list(self.contacts)
        groups = list(self.groups)

        # see if we've got groups or contacts
        for variable in self.variables:
            # this is a marker for a new contact
            if variable == NEW_CONTACT_VARIABLE:
                # if this is a test contact, stuff a fake contact in for logging purposes
                if run.contact.is_test:  # pragma: needs cover
                    contacts.append(Contact(pk=-1))

                # otherwise, really create the contact
                else:
                    contacts.append(Contact.get_or_create_by_urns(run.org, get_flow_user(run.org), name=None, urns=()))

            # other type of variable, perform our substitution
            else:
                (variable, errors) = Msg.evaluate_template(variable, expressions_context, org=run.flow.org)

                # Check for possible contact uuid and use its contact
                contact_variable_by_uuid = Contact.objects.filter(uuid=variable, org=run.flow.org).first()
                if contact_variable_by_uuid:
                    contacts.append(contact_variable_by_uuid)
                    continue

                variable_group = ContactGroup.get_user_group(run.flow.org, name=variable)
                if variable_group:  # pragma: needs cover
                    groups.append(variable_group)
                else:
                    country = run.flow.org.get_country_code()
                    (number, valid) = URN.normalize_number(variable, country)
                    if number and valid:
                        contact, contact_urn = Contact.get_or_create(run.org, URN.from_tel(number), user=get_flow_user(run.org))
                        contacts.append(contact)

        return groups, contacts


class TriggerFlowAction(VariableContactAction):
    """
    Action that starts a set of contacts down another flow
    """
    TYPE = 'trigger-flow'

    def __init__(self, uuid, flow, groups, contacts, variables):
        super(TriggerFlowAction, self).__init__(uuid, groups, contacts, variables)

        self.flow = flow

    @classmethod
    def from_json(cls, org, json_obj):
        flow_json = json_obj.get('flow')
        flow_uuid = flow_json.get('uuid')
        flow = Flow.objects.filter(org=org, is_active=True, is_archived=False, uuid=flow_uuid).first()

        # it is possible our flow got deleted
        if not flow:
            return None

        groups = VariableContactAction.parse_groups(org, json_obj)
        contacts = VariableContactAction.parse_contacts(org, json_obj)
        variables = VariableContactAction.parse_variables(org, json_obj)

        return cls(json_obj.get(cls.UUID), flow, groups, contacts, variables)

    def as_json(self):
        contact_ids = [dict(uuid=_.uuid, name=_.name) for _ in self.contacts]
        group_ids = [dict(uuid=_.uuid, name=_.name) for _ in self.groups]
        variables = [dict(id=_) for _ in self.variables]

        return dict(type=self.TYPE, uuid=self.uuid, flow=dict(uuid=self.flow.uuid, name=self.flow.name),
                    contacts=contact_ids, groups=group_ids, variables=variables)

    def execute(self, run, context, actionset_uuid, msg, offline_on=None):
        if self.flow:
            (groups, contacts) = self.build_groups_and_contacts(run, msg)
            # start our contacts down the flow
            if not run.contact.is_test:
                # our extra will be our flow variables in our message context
                extra = context.get('extra', dict())
                child_runs = self.flow.start(groups, contacts, restart_participants=True, started_flows=[run.flow.pk],
                                             extra=extra, parent_run=run)

                # build up all the msgs that where sent by our flow
                msgs = []
                for run in child_runs:
                    msgs += run.start_msgs

                return msgs
            else:  # pragma: needs cover
                unique_contacts = set()
                for contact in contacts:
                    unique_contacts.add(contact.pk)

                for group in groups:
                    for contact in group.contacts.all():
                        unique_contacts.add(contact.pk)

                self.logger(run, self.flow, len(unique_contacts))

            return []  # pragma: needs cover
        else:  # pragma: no cover
            return []

    def logger(self, run, flow, contact_count):  # pragma: needs cover
        log_txt = _("Added %d contact(s) to '%s' flow") % (contact_count, flow.name)
        log = ActionLog.create(run, log_txt)
        return log


class SetLanguageAction(Action):
    """
    Action that sets the language for a contact
    """
    TYPE = 'lang'
    LANG = 'lang'
    NAME = 'name'

    def __init__(self, uuid, lang, name):
        super(SetLanguageAction, self).__init__(uuid)

        self.lang = lang
        self.name = name

    @classmethod
    def from_json(cls, org, json_obj):
        return cls(json_obj.get(cls.UUID), json_obj.get(cls.LANG), json_obj.get(cls.NAME))

    def as_json(self):
        return dict(type=self.TYPE, uuid=self.uuid, lang=self.lang, name=self.name)

    def execute(self, run, context, actionset_uuid, msg, offline_on=None):

        if len(self.lang) != 3:
            run.contact.language = None
        else:
            run.contact.language = self.lang

        run.contact.save(update_fields=['language'])
        self.logger(run)
        return []

    def logger(self, run):  # pragma: needs cover
        # only log for test contact
        if not run.contact.is_test:
            return False

        log_txt = _("Setting language to %s") % self.name
        log = ActionLog.create(run, log_txt)
        return log


class StartFlowAction(Action):
    """
    Action that starts the contact into another flow
    """
    TYPE = 'flow'
    FLOW = 'flow'
    NAME = 'name'

    def __init__(self, uuid, flow):
        super(StartFlowAction, self).__init__(uuid)

        self.flow = flow

    @classmethod
    def from_json(cls, org, json_obj):
        flow_obj = json_obj.get(cls.FLOW)
        flow_uuid = flow_obj.get('uuid')

        flow = Flow.objects.filter(org=org, is_active=True, is_archived=False, uuid=flow_uuid).first()

        # it is possible our flow got deleted
        if not flow:
            return None
        else:
            return cls(json_obj.get(cls.UUID), flow)

    def as_json(self):
        return dict(type=self.TYPE, uuid=self.uuid, flow=dict(uuid=self.flow.uuid, name=self.flow.name))

    def execute(self, run, context, actionset_uuid, msg, started_flows, offline_on=None):
        msgs = []

        # our extra will be our flow variables in our message context
        extra = context.get('extra', dict())

        # if they are both flow runs, just redirect the call
        if run.flow.flow_type == Flow.VOICE and self.flow.flow_type == Flow.VOICE:
            new_run = self.flow.start([], [run.contact], started_flows=started_flows,
                                      restart_participants=True, extra=extra, parent_run=run)[0]
            url = "https://%s%s" % (new_run.org.get_brand_domain(), reverse('ivr.ivrcall_handle', args=[new_run.connection.pk]))
            run.voice_response.redirect(url)
        else:
            child_runs = self.flow.start([], [run.contact], started_flows=started_flows, restart_participants=True,
                                         extra=extra, parent_run=run)
            for run in child_runs:
                msgs += run.start_msgs

        self.logger(run)
        return msgs

    def logger(self, run):  # pragma: needs cover
        # only log for test contact
        if not run.contact.is_test:
            return False

        log_txt = _("Starting other flow %s") % self.flow.name

        log = ActionLog.create(run, log_txt)

        return log


class SaveToContactAction(Action):
    """
    Action to save a variable substitution to a field on a contact
    """
    TYPE = 'save'
    FIELD = 'field'
    LABEL = 'label'
    VALUE = 'value'

    def __init__(self, uuid, label, field, value):
        super(SaveToContactAction, self).__init__(uuid)

        self.label = label
        self.field = field
        self.value = value

    @classmethod
    def get_label(cls, org, field, label=None):

        # make sure this field exists
        if field == 'name':
            label = 'Contact Name'
        elif field == 'first_name':
            label = 'First Name'
        elif field == 'tel_e164':
            label = 'Phone Number'
        elif field in ContactURN.CONTEXT_KEYS_TO_SCHEME.keys():
            label = six.text_type(ContactURN.CONTEXT_KEYS_TO_LABEL[field])
        else:
            contact_field = ContactField.objects.filter(org=org, key=field).first()
            if contact_field:
                label = contact_field.label
            else:
                ContactField.get_or_create(org, get_flow_user(org), field, label)

        return label

    @classmethod
    def from_json(cls, org, json_obj):
        # they are creating a new field
        label = json_obj.get(cls.LABEL)
        field = json_obj.get(cls.FIELD)
        value = json_obj.get(cls.VALUE)

        if label and label.startswith('[_NEW_]'):
            label = label[7:]

        # create our contact field if necessary
        if not field:
            field = ContactField.make_key(label)

        # look up our label
        label = cls.get_label(org, field, label)

        return cls(json_obj.get(cls.UUID), label, field, value)

    def as_json(self):
        return dict(type=self.TYPE, uuid=self.uuid, label=self.label, field=self.field, value=self.value)

    def execute(self, run, context, actionset_uuid, msg, offline_on=None):
        # evaluate our value
        contact = run.contact
        user = get_flow_user(run.org)
        (value, errors) = Msg.evaluate_template(self.value, context, org=run.flow.org)

        if contact.is_test and errors:  # pragma: needs cover
            ActionLog.warn(run, _("Expression contained errors: %s") % ', '.join(errors))

        value = value.strip()

        if self.field == 'name':
            new_value = value[:128]
            contact.name = new_value
            contact.modified_by = user
            contact.save(update_fields=('name', 'modified_by', 'modified_on'))
            self.logger(run, new_value)

        elif self.field == 'first_name':
            new_value = value[:128]
            contact.set_first_name(new_value)
            contact.modified_by = user
            contact.save(update_fields=('name', 'modified_by', 'modified_on'))
            self.logger(run, new_value)

        elif self.field in ContactURN.CONTEXT_KEYS_TO_SCHEME.keys():
            new_value = value[:128]

            # add in our new urn number
            scheme = ContactURN.CONTEXT_KEYS_TO_SCHEME[self.field]

            # trim off '@' for twitter handles
            if self.field == 'twitter':  # pragma: needs cover
                if len(new_value) > 0:
                    if new_value[0] == '@':
                        new_value = new_value[1:]

            # only valid urns get added, sorry
            new_urn = None
            if new_value:
                new_urn = URN.normalize(URN.from_parts(scheme, new_value))
                if not URN.validate(new_urn, contact.org.get_country_code()):
                    new_urn = False
                    if contact.is_test:
                        ActionLog.warn(run, _('Contact not updated, invalid connection for contact (%s:%s)' % (scheme, new_value)))
            else:
                if contact.is_test:
                    ActionLog.warn(run, _('Contact not updated, missing connection for contact'))

            if new_urn:
                urns = [six.text_type(urn) for urn in contact.urns.all()]
                urns += [new_urn]

                # don't really update URNs on test contacts
                if contact.is_test:
                    ActionLog.info(run, _("Added %s as @contact.%s - skipped in simulator" % (new_value, scheme)))
                else:
                    contact.update_urns(user, urns)

        else:
            new_value = value[:Value.MAX_VALUE_LEN]
            contact.set_field(user, self.field, new_value)
            self.logger(run, new_value)

        return []

    def logger(self, run, new_value):  # pragma: needs cover
        # only log for test contact
        if not run.contact.is_test:
            return False

        label = SaveToContactAction.get_label(run.flow.org, self.field, self.label)
        log_txt = _("Updated %s to '%s'") % (label, new_value)

        log = ActionLog.create(run, log_txt)

        return log


class SetChannelAction(Action):
    """
    Action which sets the preferred channel to use for this Contact. If the contact has no URNs that match
    the Channel being set then this is a no-op.
    """
    TYPE = 'channel'
    CHANNEL = 'channel'
    NAME = 'name'

    def __init__(self, uuid, channel):
        super(SetChannelAction, self).__init__(uuid)

        self.channel = channel

    @classmethod
    def from_json(cls, org, json_obj):
        channel_uuid = json_obj.get(SetChannelAction.CHANNEL)

        if channel_uuid:
            channel = Channel.objects.filter(org=org, is_active=True, uuid=channel_uuid).first()
        else:  # pragma: needs cover
            channel = None
        return cls(json_obj.get(cls.UUID), channel)

    def as_json(self):
        channel_uuid = self.channel.uuid if self.channel else None
        channel_name = "%s: %s" % (self.channel.get_channel_type_display(), self.channel.get_address_display()) if self.channel else None
        return dict(type=self.TYPE, uuid=self.uuid, channel=channel_uuid, name=channel_name)

    def execute(self, run, context, actionset_uuid, msg, offline_on=None):
        # if we found the channel to set
        if self.channel:

            # don't set preferred channel for test contacts
            if not run.contact.is_test:
                run.contact.set_preferred_channel(self.channel)

            self.log(run, _("Updated preferred channel to %s") % self.channel.name)
            return []
        else:
            self.log(run, _("Channel not found, no action taken"))
            return []

    def log(self, run, text):  # pragma: no cover
        if run.contact.is_test:
            ActionLog.create(run, text)


class SendAction(VariableContactAction):
    """
    Action which sends a message to a specified set of contacts and groups.
    """
    TYPE = 'send'
    MESSAGE = 'msg'
    MEDIA = 'media'

    def __init__(self, uuid, msg, groups, contacts, variables, media=None):
        super(SendAction, self).__init__(uuid, groups, contacts, variables)

        self.msg = msg
        self.media = media if media else {}

    @classmethod
    def from_json(cls, org, json_obj):
        groups = VariableContactAction.parse_groups(org, json_obj)
        contacts = VariableContactAction.parse_contacts(org, json_obj)
        variables = VariableContactAction.parse_variables(org, json_obj)

        return cls(json_obj.get(cls.UUID), json_obj.get(cls.MESSAGE), groups, contacts, variables,
                   json_obj.get(cls.MEDIA, None))

    def as_json(self):
        contact_ids = [dict(uuid=_.uuid) for _ in self.contacts]
        group_ids = [dict(uuid=_.uuid, name=_.name) for _ in self.groups]
        variables = [dict(id=_) for _ in self.variables]
        return dict(type=self.TYPE, uuid=self.uuid, msg=self.msg,
                    contacts=contact_ids, groups=group_ids, variables=variables,
                    media=self.media)

    def execute(self, run, context, actionset_uuid, msg, offline_on=None):
        if self.msg or self.media:
            flow = run.flow
            (groups, contacts) = self.build_groups_and_contacts(run, msg)

            # create our broadcast and send it
            if not run.contact.is_test:
                # no-op if neither text nor media are defined in the flow base language
                if not (self.msg.get(flow.base_language) or self.media.get(flow.base_language)):
                    return list()

                recipients = groups + contacts

                broadcast = Broadcast.create(flow.org, flow.modified_by, self.msg, recipients,
                                             media=self.media, base_language=flow.base_language)
                broadcast.send(trigger_send=False, expressions_context=context)
                return list(broadcast.get_messages())

            else:
                unique_contacts = set()
                for contact in contacts:
                    unique_contacts.add(contact.pk)

                for group in groups:
                    for contact in group.contacts.all():
                        unique_contacts.add(contact.pk)

                text = run.flow.get_localized_text(self.msg, run.contact)
                (message, errors) = Msg.evaluate_template(text, context, org=run.flow.org, partial_vars=True)

                self.logger(run, message, len(unique_contacts))

            return []
        else:  # pragma: no cover
            return []

    def logger(self, run, text, contact_count):
        log_txt = _n("Sending '%(msg)s' to %(count)d contact",
                     "Sending '%(msg)s' to %(count)d contacts",
                     contact_count) % dict(msg=text, count=contact_count)
        log = ActionLog.create(run, log_txt)
        return log


class Rule(object):

    def __init__(self, uuid, category, destination, destination_type, test, label=None):
        self.uuid = uuid
        self.category = category
        self.destination = destination
        self.destination_type = destination_type
        self.test = test
        self.label = label

    def get_category_name(self, flow_lang, contact_lang=None):
        if not self.category:  # pragma: needs cover
            if isinstance(self.test, BetweenTest):
                return "%s-%s" % (self.test.min, self.test.max)

        # return the category name for the flow language version
        if isinstance(self.category, dict):
            category = None
            if contact_lang:
                category = self.category.get(contact_lang)

            if not category and flow_lang:
                category = self.category.get(flow_lang)

            if not category:  # pragma: needs cover
                category = list(self.category.values())[0]

            return category

        return self.category  # pragma: needs cover

    def matches(self, run, sms, context, text):
        return self.test.evaluate(run, sms, context, text)

    def as_json(self):
        return dict(uuid=self.uuid,
                    category=self.category,
                    destination=self.destination,
                    destination_type=self.destination_type,
                    test=self.test.as_json(),
                    label=self.label)

    @classmethod
    def from_json_array(cls, org, json):
        rules = []
        for rule in json:
            category = rule.get('category', None)

            if isinstance(category, dict):
                # prune all of our translations to 36
                for k, v in category.items():
                    if isinstance(v, six.string_types):
                        category[k] = v[:36]
            elif category:
                category = category[:36]

            destination = rule.get('destination', None)
            destination_type = None

            # determine our destination type, if its not set its an action set
            if destination:
                destination_type = rule.get('destination_type', FlowStep.TYPE_ACTION_SET)

            rules.append(Rule(rule.get('uuid'),
                              category,
                              destination,
                              destination_type,
                              Test.from_json(org, rule['test']),
                              rule.get('label')))

        return rules


class Test(object):
    TYPE = 'type'
    __test_mapping = None

    @classmethod
    def from_json(cls, org, json_dict):
        if not cls.__test_mapping:
            cls.__test_mapping = {
                AirtimeStatusTest.TYPE: AirtimeStatusTest,
                AndTest.TYPE: AndTest,
                BetweenTest.TYPE: BetweenTest,
                ContainsAnyTest.TYPE: ContainsAnyTest,
                ContainsOnlyPhraseTest.TYPE: ContainsOnlyPhraseTest,
                ContainsPhraseTest.TYPE: ContainsPhraseTest,
                ContainsTest.TYPE: ContainsTest,
                DateAfterTest.TYPE: DateAfterTest,
                DateBeforeTest.TYPE: DateBeforeTest,
                DateEqualTest.TYPE: DateEqualTest,
                EqTest.TYPE: EqTest,
                FalseTest.TYPE: FalseTest,
                GtTest.TYPE: GtTest,
                GteTest.TYPE: GteTest,
                DateTest.TYPE: DateTest,
                HasDistrictTest.TYPE: HasDistrictTest,
                HasEmailTest.TYPE: HasEmailTest,
                HasStateTest.TYPE: HasStateTest,
                HasWardTest.TYPE: HasWardTest,
                InGroupTest.TYPE: InGroupTest,
                InterruptTest.TYPE: InterruptTest,
                LtTest.TYPE: LtTest,
                LteTest.TYPE: LteTest,
                NotEmptyTest.TYPE: NotEmptyTest,
                NumberTest.TYPE: NumberTest,
                OrTest.TYPE: OrTest,
                PhoneTest.TYPE: PhoneTest,
                RegexTest.TYPE: RegexTest,
                StartsWithTest.TYPE: StartsWithTest,
                SubflowTest.TYPE: SubflowTest,
                TimeoutTest.TYPE: TimeoutTest,
                TrueTest.TYPE: TrueTest,
                WebhookStatusTest.TYPE: WebhookStatusTest,
            }

        type = json_dict.get(cls.TYPE, None)
        if not type:  # pragma: no cover
            raise FlowException("Test definition missing 'type' field: %s", json_dict)

        if type not in cls.__test_mapping:  # pragma: no cover
            raise FlowException("Unknown type: '%s' in definition: %s" % (type, json_dict))

        return cls.__test_mapping[type].from_json(org, json_dict)

    @classmethod
    def from_json_array(cls, org, json):
        tests = []
        for inner in json:
            tests.append(Test.from_json(org, inner))

        return tests

    def evaluate(self, run, sms, context, text):  # pragma: no cover
        """
        Where the work happens, subclasses need to be able to evalute their Test
        according to their definition given the passed in message. Tests do not have
        side effects.
        """
        raise FlowException("Subclasses must implement evaluate, returning a tuple containing 1 or 0 and the value tested")


class WebhookStatusTest(Test):
    """
    {op: 'webhook', status: 'success' }
    """
    TYPE = 'webhook_status'
    STATUS = 'status'

    STATUS_SUCCESS = 'success'
    STATUS_FAILURE = 'failure'

    def __init__(self, status):
        self.status = status

    @classmethod
    def from_json(cls, org, json):
        return WebhookStatusTest(json.get('status'))

    def as_json(self):  # pragma: needs cover
        return dict(type=WebhookStatusTest.TYPE, status=self.status)

    def evaluate(self, run, sms, context, text):
        # we treat any 20* return code as successful
        success = 200 <= int(text) < 300

        if success and self.status == WebhookStatusTest.STATUS_SUCCESS:
            return 1, text
        elif not success and self.status == WebhookStatusTest.STATUS_FAILURE:
            return 1, text
        else:
            return 0, None


class AirtimeStatusTest(Test):
    """
    {op: 'airtime_status'}
    """
    TYPE = 'airtime_status'
    EXIT = 'exit_status'

    STATUS_SUCCESS = 'success'
    STATUS_FAILED = 'failed'

    STATUS_MAP = {STATUS_SUCCESS: AirtimeTransfer.SUCCESS,
                  STATUS_FAILED: AirtimeTransfer.FAILED}

    def __init__(self, exit_status):
        self.exit_status = exit_status

    @classmethod
    def from_json(cls, org, json):
        return AirtimeStatusTest(json.get('exit_status'))

    def as_json(self):  # pragma: needs cover
        return dict(type=AirtimeStatusTest.TYPE, exit_status=self.exit_status)

    def evaluate(self, run, sms, context, text):
        status = text
        if status and AirtimeStatusTest.STATUS_MAP[self.exit_status] == status:
            return 1, status
        return 0, None


class InGroupTest(Test):
    """
    { op: "in_group" }
    """
    TYPE = 'in_group'
    NAME = 'name'
    UUID = 'uuid'
    TEST = 'test'

    def __init__(self, group):
        self.group = group

    @classmethod
    def from_json(cls, org, json):
        group = json.get(InGroupTest.TEST)
        name = group.get(InGroupTest.NAME)
        uuid = group.get(InGroupTest.UUID)
        return InGroupTest(ContactGroup.get_or_create(org, org.created_by, name, uuid))

    def as_json(self):
        group = ContactGroup.get_or_create(self.group.org, self.group.org.created_by, self.group.name, self.group.uuid)
        return dict(type=InGroupTest.TYPE, test=dict(name=group.name, uuid=group.uuid))

    def evaluate(self, run, sms, context, text):
        if run.contact.user_groups.filter(id=self.group.id).first():
            return 1, self.group.name
        return 0, None


class SubflowTest(Test):
    """
    { op: "subflow" }
    """
    TYPE = 'subflow'
    EXIT = 'exit_type'

    TYPE_COMPLETED = 'completed'
    TYPE_EXPIRED = 'expired'

    EXIT_MAP = {TYPE_COMPLETED: FlowRun.EXIT_TYPE_COMPLETED,
                TYPE_EXPIRED: FlowRun.EXIT_TYPE_EXPIRED}

    def __init__(self, exit_type):
        self.exit_type = exit_type

    @classmethod
    def from_json(cls, org, json):
        return SubflowTest(json.get(SubflowTest.EXIT))

    def as_json(self):  # pragma: needs cover
        return dict(type=SubflowTest.TYPE, exit_type=self.exit_type)

    def evaluate(self, run, sms, context, text):
        # lookup the subflow run
        subflow_run = FlowRun.objects.filter(parent=run).order_by('-created_on').first()

        if subflow_run and SubflowTest.EXIT_MAP[self.exit_type] == subflow_run.exit_type:
            return 1, text
        return 0, None


class TimeoutTest(Test):
    """
    { op: "timeout", minutes: 60 }
    """
    TYPE = 'timeout'
    MINUTES = 'minutes'

    def __init__(self, minutes):
        self.minutes = minutes

    @classmethod
    def from_json(cls, org, json):
        return TimeoutTest(int(json.get(TimeoutTest.MINUTES)))

    def as_json(self):  # pragma: needs cover
        return {'type': TimeoutTest.TYPE, TimeoutTest.MINUTES: self.minutes}

    def evaluate(self, run, sms, context, text):
        if run.timeout_on < timezone.now():
            return 1, None
        else:  # pragma: needs cover
            return 0, None


class TrueTest(Test):
    """
    { op: "true" }
    """
    TYPE = 'true'

    def __init__(self):
        pass

    @classmethod
    def from_json(cls, org, json):
        return TrueTest()

    def as_json(self):
        return dict(type=TrueTest.TYPE)

    def evaluate(self, run, sms, context, text):
        return 1, text


class FalseTest(Test):
    """
    { op: "false" }
    """
    TYPE = 'false'

    def __init__(self):
        pass

    @classmethod
    def from_json(cls, org, json):
        return FalseTest()

    def as_json(self):
        return dict(type=FalseTest.TYPE)

    def evaluate(self, run, sms, context, text):
        return 0, None


class AndTest(Test):
    """
    { op: "and",  "tests": [ ... ] }
    """
    TESTS = 'tests'
    TYPE = 'and'

    def __init__(self, tests):
        self.tests = tests

    @classmethod
    def from_json(cls, org, json):
        return AndTest(Test.from_json_array(org, json[cls.TESTS]))

    def as_json(self):
        return dict(type=AndTest.TYPE, tests=[_.as_json() for _ in self.tests])

    def evaluate(self, run, sms, context, text):  # pragma: needs cover
        matches = []
        for test in self.tests:
            (result, value) = test.evaluate(run, sms, context, text)
            if result:
                matches.append(value)
            else:
                return 0, None

        # all came out true, we are true
        return 1, " ".join(matches)


class OrTest(Test):
    """
    { op: "or",  "tests": [ ... ] }
    """
    TESTS = 'tests'
    TYPE = 'or'

    def __init__(self, tests):
        self.tests = tests

    @classmethod
    def from_json(cls, org, json):
        return OrTest(Test.from_json_array(org, json[cls.TESTS]))

    def as_json(self):
        return dict(type=OrTest.TYPE, tests=[_.as_json() for _ in self.tests])

    def evaluate(self, run, sms, context, text):  # pragma: needs cover
        for test in self.tests:
            (result, value) = test.evaluate(run, sms, context, text)
            if result:
                return result, value

        return 0, None


class NotEmptyTest(Test):
    """
    { op: "not_empty" }
    """

    TYPE = 'not_empty'

    def __init__(self):  # pragma: needs cover
        pass

    @classmethod
    def from_json(cls, org, json):  # pragma: needs cover
        return NotEmptyTest()

    def as_json(self):  # pragma: needs cover
        return dict(type=NotEmptyTest.TYPE)

    def evaluate(self, run, sms, context, text):  # pragma: needs cover
        if text and len(text.strip()):
            return 1, text.strip()
        return 0, None


class ContainsTest(Test):
    """
    { op: "contains", "test": "red" }
    """
    TEST = 'test'
    TYPE = 'contains'

    def __init__(self, test):
        self.test = test

    @classmethod
    def from_json(cls, org, json):
        return cls(json[cls.TEST])

    def as_json(self):
        json = dict(type=ContainsTest.TYPE, test=self.test)
        return json

    def test_in_words(self, test, words, raw_words):
        matches = []
        for index, word in enumerate(words):
            if word == test:
                matches.append(index)
                continue

        return matches

    def evaluate(self, run, sms, context, text):
        # substitute any variables
        test = run.flow.get_localized_text(self.test, run.contact)
        test, errors = Msg.evaluate_template(test, context, org=run.flow.org)

        # tokenize our test
        tests = tokenize(test.lower())

        # tokenize our sms
        words = tokenize(text.lower())
        raw_words = tokenize(text)

        tests = [elt for elt in tests if elt != '']
        words = [elt for elt in words if elt != '']
        raw_words = [elt for elt in raw_words if elt != '']

        # run through each of our tests
        matches = set()
        matched_tests = 0
        for test in tests:
            match = self.test_in_words(test, words, raw_words)
            if match:
                matched_tests += 1
                matches.update(match)

        # we are a match only if every test matches
        if matched_tests == len(tests):
            matches = sorted(list(matches))
            matched_words = " ".join([raw_words[idx] for idx in matches])
            return len(tests), matched_words
        else:
            return 0, None


class HasEmailTest(Test):
    """
    { op: "has_email" }
    """
    TYPE = 'has_email'

    def __init__(self):
        pass

    @classmethod
    def from_json(cls, org, json):
        return cls()

    def as_json(self):
        return dict(type=self.TYPE)

    def evaluate(self, run, sms, context, text):
        # split on whitespace
        words = text.split()
        for word in words:
            word = word.strip(',.;:|()[]"\'<>?&*/\\')
            if is_valid_address(word):
                return 1, word

        return 0, None


class ContainsAnyTest(ContainsTest):
    """
    { op: "contains_any", "test": "red" }
    """
    TEST = 'test'
    TYPE = 'contains_any'

    def as_json(self):
        return dict(type=ContainsAnyTest.TYPE, test=self.test)

    def evaluate(self, run, sms, context, text):
        # substitute any variables
        test = run.flow.get_localized_text(self.test, run.contact)
        test, errors = Msg.evaluate_template(test, context, org=run.flow.org)

        # tokenize our test
        tests = tokenize(test.lower())

        # tokenize our sms
        words = tokenize(text.lower())
        raw_words = tokenize(text)

        tests = [elt for elt in tests if elt != '']
        words = [elt for elt in words if elt != '']
        raw_words = [elt for elt in raw_words if elt != '']

        # run through each of our tests
        matches = set()
        for test in tests:
            match = self.test_in_words(test, words, raw_words)
            if match:
                matches.update(match)

        # we are a match if at least one test matches
        if matches:
            matches = sorted(list(matches))
            matched_words = " ".join([raw_words[idx] for idx in matches])
            return 1, matched_words
        else:
            return 0, None


class ContainsOnlyPhraseTest(ContainsTest):
    """
    { op: "contains_only_phrase", "test": "red" }
    """
    TEST = 'test'
    TYPE = 'contains_only_phrase'

    def as_json(self):
        return dict(type=ContainsOnlyPhraseTest.TYPE, test=self.test)

    def evaluate(self, run, sms, context, text):
        # substitute any variables
        test = run.flow.get_localized_text(self.test, run.contact)
        test, errors = Msg.evaluate_template(test, context, org=run.flow.org)

        # tokenize our test
        tests = tokenize(test.lower())

        # tokenize our sms
        words = tokenize(text.lower())
        raw_words = tokenize(text)

        # they are the same? then we matched
        if tests == words:
            return 1, " ".join(raw_words)
        else:
            return 0, None


class ContainsPhraseTest(ContainsTest):
    """
    { op: "contains_phrase", "test": "red" }
    """
    TEST = 'test'
    TYPE = 'contains_phrase'

    def as_json(self):
        return dict(type=ContainsPhraseTest.TYPE, test=self.test)

    def evaluate(self, run, sms, context, text):
        # substitute any variables
        test = run.flow.get_localized_text(self.test, run.contact)
        test, errors = Msg.evaluate_template(test, context, org=run.flow.org)

        # tokenize our test
        tests = tokenize(test.lower())
        if not tests:
            return True, ""

        # tokenize our sms
        words = tokenize(text.lower())
        raw_words = tokenize(text)

        # look for the phrase
        test_idx = 0
        matches = []
        for i in range(len(words)):
            if tests[test_idx] == words[i]:
                matches.append(raw_words[i])
                test_idx += 1
                if test_idx == len(tests):
                    break
            else:
                matches = []
                test_idx = 0

        # we found the phrase
        if test_idx == len(tests):
            matched_words = " ".join(matches)
            return 1, matched_words
        else:
            return 0, None


class StartsWithTest(Test):
    """
    { op: "starts", "test": "red" }
    """
    TEST = 'test'
    TYPE = 'starts'

    def __init__(self, test):
        self.test = test

    @classmethod
    def from_json(cls, org, json):
        return cls(json[cls.TEST])

    def as_json(self):  # pragma: needs cover
        return dict(type=StartsWithTest.TYPE, test=self.test)

    def evaluate(self, run, sms, context, text):
        # substitute any variables in our test
        test = run.flow.get_localized_text(self.test, run.contact)
        test, errors = Msg.evaluate_template(test, context, org=run.flow.org)

        # strip leading and trailing whitespace
        text = text.strip()

        # see whether we start with our test
        if text.lower().find(test.lower()) == 0:
            return 1, text[:len(test)]
        else:
            return 0, None


class HasStateTest(Test):
    TYPE = 'state'

    def __init__(self):
        pass

    @classmethod
    def from_json(cls, org, json):
        return cls()

    def as_json(self):
        return dict(type=self.TYPE)

    def evaluate(self, run, sms, context, text):
        org = run.flow.org

        # if they removed their country since adding the rule
        if not org.country:
            return 0, None

        state = org.parse_location(text, AdminBoundary.LEVEL_STATE)
        if state:
            return 1, state[0]

        return 0, None


class HasDistrictTest(Test):
    TYPE = 'district'
    TEST = 'test'

    def __init__(self, state=None):
        self.state = state

    @classmethod
    def from_json(cls, org, json):
        return cls(json[cls.TEST])

    def as_json(self):
        return dict(type=self.TYPE, test=self.state)

    def evaluate(self, run, sms, context, text):

        # if they removed their country since adding the rule
        org = run.flow.org
        if not org.country:
            return 0, None

        # evaluate our district in case it has a replacement variable
        state, errors = Msg.evaluate_template(self.state, context, org=run.flow.org)

        parent = org.parse_location(state, AdminBoundary.LEVEL_STATE)
        if parent:
            district = org.parse_location(text, AdminBoundary.LEVEL_DISTRICT, parent[0])
            if district:
                return 1, district[0]
        district = org.parse_location(text, AdminBoundary.LEVEL_DISTRICT)

        # parse location when state contraint is not provided or available
        if (errors or not state) and len(district) == 1:
            return 1, district[0]

        return 0, None


class HasWardTest(Test):
    TYPE = 'ward'
    STATE = 'state'
    DISTRICT = 'district'

    def __init__(self, state=None, district=None):
        self.state = state
        self.district = district

    @classmethod
    def from_json(cls, org, json):
        return cls(json[cls.STATE], json[cls.DISTRICT])

    def as_json(self):
        return dict(type=self.TYPE, state=self.state, district=self.district)

    def evaluate(self, run, sms, context, text):
        # if they removed their country since adding the rule
        org = run.flow.org
        if not org.country:  # pragma: needs cover
            return 0, None
        district = None

        # evaluate our district in case it has a replacement variable
        district_name, missing_district = Msg.evaluate_template(self.district, context, org=run.flow.org)
        state_name, missing_state = Msg.evaluate_template(self.state, context, org=run.flow.org)
        if (district_name and state_name) and (len(missing_district) == 0 and len(missing_state) == 0):
            state = org.parse_location(state_name, AdminBoundary.LEVEL_STATE)
            if state:
                district = org.parse_location(district_name, AdminBoundary.LEVEL_DISTRICT, state[0])
                if district:
                    ward = org.parse_location(text, AdminBoundary.LEVEL_WARD, district[0])
                    if ward:
                        return 1, ward[0]

        # parse location when district contraint is not provided or available
        ward = org.parse_location(text, AdminBoundary.LEVEL_WARD)
        if len(ward) == 1 and district is None:
            return 1, ward[0]

        return 0, None


class DateTest(Test):
    """
    Base class for those tests that check relative dates
    """
    TEST = None
    TYPE = 'date'

    def __init__(self, test=None):
        self.test = test

    @classmethod
    def from_json(cls, org, json):
        if cls.TEST:
            return cls(json[cls.TEST])
        else:
            return cls()

    def as_json(self):
        if self.test:
            return dict(type=self.TYPE, test=self.test)
        else:
            return dict(type=self.TYPE)

    def evaluate_date_test(self, date_message, date_test):
        return date_message is not None

    def evaluate(self, run, sms, context, text):
        org = run.flow.org
        day_first = org.get_dayfirst()
        tz = org.timezone

        test, errors = Msg.evaluate_template(self.test, context, org=org)
        if not errors:
            date_message = str_to_datetime(text, tz=tz, dayfirst=day_first)
            date_test = str_to_datetime(test, tz=tz, dayfirst=day_first)

            if self.evaluate_date_test(date_message, date_test):
                return 1, date_message.astimezone(tz)

        return 0, None


class DateEqualTest(DateTest):
    TEST = 'test'
    TYPE = 'date_equal'

    def evaluate_date_test(self, date_message, date_test):
        return date_message and date_test and date_message.date() == date_test.date()


class DateAfterTest(DateTest):
    TEST = 'test'
    TYPE = 'date_after'

    def evaluate_date_test(self, date_message, date_test):
        return date_message and date_test and date_message >= date_test


class DateBeforeTest(DateTest):
    TEST = 'test'
    TYPE = 'date_before'

    def evaluate_date_test(self, date_message, date_test):
        return date_message and date_test and date_message <= date_test


class NumericTest(Test):
    """
    Base class for those tests that do numeric tests.
    """
    TEST = 'test'
    TYPE = ''

    @classmethod
    def convert_to_decimal(cls, word):
        # common substitutions
        original_word = word
        word = word.replace('l', '1').replace('o', '0').replace('O', '0')

        try:
            return (word, Decimal(word))
        except Exception as e:
            # we only try this hard if we haven't already substituted characters
            if original_word == word:
                # does this start with a number?  just use that part if so
                match = regex.match(r"^[$£€]?([\d,][\d,\.]*([\.,]\d+)?)\D*$", word, regex.UNICODE | regex.V0)

                if match:
                    return (match.group(1), Decimal(match.group(1)))
                else:
                    raise e
            else:
                raise e

    # test every word in the message against our test
    def evaluate(self, run, sms, context, text):
        text = text.replace(',', '')
        for word in regex.split(r"\s+", text, flags=regex.UNICODE | regex.V0):
            try:
                (word, decimal) = NumericTest.convert_to_decimal(word)
                if self.evaluate_numeric_test(run, context, decimal):
                    return 1, decimal
            except Exception:  # pragma: needs cover
                pass
        return 0, None


class BetweenTest(NumericTest):
    """
    Test whether we are between two numbers (inclusive)
    """
    MIN = 'min'
    MAX = 'max'
    TYPE = 'between'

    def __init__(self, min_val, max_val):
        self.min = min_val
        self.max = max_val

    @classmethod
    def from_json(cls, org, json):
        return cls(json[cls.MIN], json[cls.MAX])

    def as_json(self):
        return dict(type=self.TYPE, min=self.min, max=self.max)

    def evaluate_numeric_test(self, run, context, decimal_value):
        min_val, min_errors = Msg.evaluate_template(self.min, context, org=run.flow.org)
        max_val, max_errors = Msg.evaluate_template(self.max, context, org=run.flow.org)

        if not min_errors and not max_errors:
            try:
                return Decimal(min_val) <= decimal_value <= Decimal(max_val)
            except Exception:  # pragma: needs cover
                pass

        return False  # pragma: needs cover


class NumberTest(NumericTest):
    """
    Tests that there is any number in the string.
    """
    TYPE = 'number'

    def __init__(self):
        pass

    @classmethod
    def from_json(cls, org, json):
        return cls()

    def as_json(self):  # pragma: needs cover
        return dict(type=self.TYPE)

    def evaluate_numeric_test(self, run, context, decimal_value):
        return True


class SimpleNumericTest(NumericTest):
    """
    Base class for those tests that do a numeric test with a single value
    """
    TEST = 'test'
    TYPE = ''

    def __init__(self, test):
        self.test = test

    @classmethod
    def from_json(cls, org, json):
        return cls(json[cls.TEST])

    def as_json(self):
        return dict(type=self.TYPE, test=self.test)

    def evaluate_numeric_test(self, message_numeric, test_numeric):  # pragma: no cover
        raise FlowException("Evaluate numeric test needs to be defined by subclass")

    # test every word in the message against our test
    def evaluate(self, run, sms, context, text):
        test, errors = Msg.evaluate_template(str(self.test), context, org=run.flow.org)

        text = text.replace(',', '')
        for word in regex.split(r"\s+", text, flags=regex.UNICODE | regex.V0):
            try:
                (word, decimal) = NumericTest.convert_to_decimal(word)
                if self.evaluate_numeric_test(decimal, Decimal(test)):
                    return 1, decimal
            except Exception:
                pass
        return 0, None


class GtTest(SimpleNumericTest):
    TEST = 'test'
    TYPE = 'gt'

    def evaluate_numeric_test(self, message_numeric, test_numeric):
        return message_numeric > test_numeric


class GteTest(SimpleNumericTest):
    TEST = 'test'
    TYPE = 'gte'

    def evaluate_numeric_test(self, message_numeric, test_numeric):
        return message_numeric >= test_numeric


class LtTest(SimpleNumericTest):
    TEST = 'test'
    TYPE = 'lt'

    def evaluate_numeric_test(self, message_numeric, test_numeric):
        return message_numeric < test_numeric


class LteTest(SimpleNumericTest):
    TEST = 'test'
    TYPE = 'lte'

    def evaluate_numeric_test(self, message_numeric, test_numeric):  # pragma: needs cover
        return message_numeric <= test_numeric


class EqTest(SimpleNumericTest):
    TEST = 'test'
    TYPE = 'eq'

    def evaluate_numeric_test(self, message_numeric, test_numeric):
        return message_numeric == test_numeric


class PhoneTest(Test):
    """
    Test for whether a response contains a phone number
    """
    TYPE = 'phone'

    def __init__(self):
        pass

    @classmethod
    def from_json(cls, org, json):
        return cls()

    def as_json(self):  # pragma: needs cover
        return dict(type=self.TYPE)

    def evaluate(self, run, sms, context, text):
        org = run.flow.org

        # try to find a phone number in the text we have been sent
        country_code = org.get_country_code()
        if not country_code:  # pragma: needs cover
            country_code = 'US'

        number = None
        matches = phonenumbers.PhoneNumberMatcher(text, country_code)

        # try it as an international number if we failed
        if not matches.has_next():  # pragma: needs cover
            matches = phonenumbers.PhoneNumberMatcher('+' + text, country_code)

        for match in matches:
            number = phonenumbers.format_number(match.number, phonenumbers.PhoneNumberFormat.E164)

        return number, number


class RegexTest(Test):  # pragma: needs cover
    """
    Test for whether a response matches a regular expression
    """
    TEST = 'test'
    TYPE = 'regex'

    def __init__(self, test):
        self.test = test

    @classmethod
    def from_json(cls, org, json):
        return cls(json[cls.TEST])

    def as_json(self):
        return dict(type=self.TYPE, test=self.test)

    def evaluate(self, run, sms, context, text):
        try:
            test = run.flow.get_localized_text(self.test, run.contact)

            # check whether we match
            rexp = regex.compile(test, regex.UNICODE | regex.IGNORECASE | regex.MULTILINE | regex.V0)
            match = rexp.search(text)

            # if so, $0 will be what we return
            if match:
                return_match = match.group(0)

                # build up a dictionary that contains indexed group matches
                group_dict = {}
                for idx in range(rexp.groups + 1):
                    group_dict[str(idx)] = match.group(idx)

                # set it on run@extra
                run.update_fields(group_dict)

                # return all matched values
                return True, return_match

        except Exception:
            import traceback
            traceback.print_exc()

        return False, None


class InterruptTest(Test):
    """
    Test if it's an interrupt status message
    """
    TYPE = "interrupted_status"

    def __init__(self):
        pass

    @classmethod
    def from_json(cls, org, json):
        return cls()

    def as_json(self):
        return dict(type=self.TYPE)

    def evaluate(self, run, msg, context, text):
        return (True, self.TYPE) if run.connection and run.connection.status == ChannelSession.INTERRUPTED else (False, None)
