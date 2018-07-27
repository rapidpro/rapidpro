# -*- coding: utf-8 -*-
from __future__ import absolute_import, division, print_function, unicode_literals

import csv
import gc
import six
import time
import os

from datetime import datetime, timedelta
from django.core.files import File
from django.core.files.temp import NamedTemporaryFile
from django.db import models
from django.utils import timezone
from django.utils.translation import ugettext_lazy as _
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.write_only import WriteOnlyCell
from temba.assets.models import BaseAssetStore, get_asset_store
from . import analytics
from .models import TembaModel
from .email import send_template_email
from .text import clean_string


class BaseExportAssetStore(BaseAssetStore):
    def is_asset_ready(self, asset):
        return asset.status == BaseExportTask.STATUS_COMPLETE


class BaseExportTask(TembaModel):
    """
    Base class for export task models, i.e. contacts, messages and flow results
    """
    analytics_key = None
    asset_type = None

    MAX_EXCEL_ROWS = 1048576
    MAX_EXCEL_COLS = 16384

    WIDTH_SMALL = 15
    WIDTH_MEDIUM = 20
    WIDTH_LARGE = 100

    STATUS_PENDING = 'P'
    STATUS_PROCESSING = 'O'
    STATUS_COMPLETE = 'C'
    STATUS_FAILED = 'F'
    STATUS_CHOICES = ((STATUS_PENDING, _("Pending")),
                      (STATUS_PROCESSING, _("Processing")),
                      (STATUS_COMPLETE, _("Complete")),
                      (STATUS_FAILED, _("Failed")))

    org = models.ForeignKey('orgs.Org', related_name='%(class)ss', help_text=_("The organization of the user."))

    status = models.CharField(max_length=1, default=STATUS_PENDING, choices=STATUS_CHOICES)

    def perform(self):
        """
        Performs the actual export. If export generation throws an exception it's caught here and the task is marked
        as failed.
        """
        try:
            self.update_status(self.STATUS_PROCESSING)

            start = time.time()

            temp_file, extension = self.write_export()

            get_asset_store(model=self.__class__).save(self.id, File(temp_file), extension)

            branding = self.org.get_branding()

            # notify user who requested this export
            send_template_email(self.created_by.username, self.email_subject, self.email_template,
                                self.get_email_context(branding), branding)

            # remove temporary file on PY3
            if six.PY3:  # pragma: no cover
                if hasattr(temp_file, 'delete'):
                    if temp_file.delete is False:
                        os.unlink(temp_file.name)
                else:
                    os.unlink(temp_file.name)

        except Exception:
            import traceback
            traceback.print_exc()
            self.update_status(self.STATUS_FAILED)
        else:
            self.update_status(self.STATUS_COMPLETE)
        finally:
            elapsed = time.time() - start
            print("Completed %s in %.1f seconds" % (self.analytics_key, elapsed))
            analytics.track(self.created_by.username, 'temba.%s_latency' % self.analytics_key, properties=dict(value=elapsed))

            gc.collect()  # force garbage collection

    def write_export(self):  # pragma: no cover
        """
        Should return a file handle for a temporary file and the file extension
        """
        pass

    def update_status(self, status):
        self.status = status
        self.save(update_fields=('status',))

    @classmethod
    def get_recent_unfinished(cls, org):
        """
        Checks for unfinished exports created in the last 24 hours for this org, and returns the most recent
        """
        return cls.objects.filter(
            org=org, created_on__gt=timezone.now() - timedelta(hours=24),
            status__in=(cls.STATUS_PENDING, cls.STATUS_PROCESSING)
        ).order_by('-created_on').first()

    def append_row(self, sheet, values):
        row = []
        for value in values:
            cell = WriteOnlyCell(sheet, value=self.prepare_value(value))
            row.append(cell)
        sheet.append(row)

    def prepare_value(self, value):
        if value is None:
            return ''
        elif isinstance(value, six.string_types):
            if value.startswith('='):  # escape = so value isn't mistaken for a formula
                value = '\'' + value
            return clean_string(value)
        elif isinstance(value, datetime):
            return value.astimezone(self.org.timezone).replace(microsecond=0, tzinfo=None)
        else:
            return clean_string(six.text_type(value))

    def set_sheet_column_widths(self, sheet, widths):
        for index, width in enumerate(widths):
            sheet.column_dimensions[get_column_letter(index + 1)].width = widths[index]

    def get_email_context(self, branding):
        asset_store = get_asset_store(model=self.__class__)

        return {'link': branding['link'] + asset_store.get_asset_url(self.id)}

    class Meta:
        abstract = True


class TableExporter(object):
    """
    Class that abstracts out writing a table of data to a CSV or Excel file. This only works for exports that
    have a single sheet (as CSV's don't have sheets) but takes care of writing to a CSV in the case
    where there are more than 16384 columns, which Excel doesn't support.

    When writing to an Excel sheet, this also takes care of creating different sheets every 1048576
    rows, as again, Excel file only support that many per sheet.
    """
    def __init__(self, task, sheet_name, columns):
        self.task = task
        self.columns = columns
        self.is_csv = len(self.columns) > BaseExportTask.MAX_EXCEL_COLS
        self.sheet_name = sheet_name

        self.current_sheet = 0
        self.current_row = 0

        if six.PY2:
            self.file = NamedTemporaryFile(delete=True, suffix='.xlsx', mode='wb+')
        else:  # pragma: no cover
            self.file = NamedTemporaryFile(delete=False, suffix='.xlsx', mode='wt+')

        # if this is a csv file, create our csv writer and write our header
        if self.is_csv:
            self.writer = csv.writer(self.file, quoting=csv.QUOTE_ALL)
            if six.PY2:
                self.writer.writerow([s.encode('utf-8') for s in columns])
            else:  # pragma: no cover
                self.writer.writerow(columns)

        # otherwise, just open a workbook, initializing the first sheet
        else:
            self.workbook = Workbook(write_only=True)
            self.sheet_number = 0
            self._add_sheet()

    def _add_sheet(self):
        self.sheet_number += 1

        # add our sheet
        self.sheet = self.workbook.create_sheet(u"%s %d" % (self.sheet_name, self.sheet_number))

        self.task.append_row(self.sheet, self.columns)

        self.sheet_row = 2

    def write_row(self, values):
        """
        Writes the passed in row to our exporter, taking care of creating new sheets if necessary
        """
        if self.is_csv:
            if six.PY2:
                self.writer.writerow([s.encode('utf-8') for s in values])
            else:  # pragma: no cover
                self.writer.writerow(values)

        else:
            # time for a new sheet? do it
            if self.sheet_row > BaseExportTask.MAX_EXCEL_ROWS:
                self._add_sheet()

            self.task.append_row(self.sheet, values)

            self.sheet_row += 1

    def save_file(self):
        """
        Saves our data to a file, returning the file saved to and the extension
        """
        gc.collect()  # force garbage collection

        if six.PY3:  # pragma: no cover
            self.file.close()
            self.file = open(self.file.name, 'rb+')

        if not self.is_csv:
            print("Writing Excel workbook...")
            self.workbook.save(self.file)

        self.file.flush()
        return self.file, ('csv' if self.is_csv else 'xlsx')
