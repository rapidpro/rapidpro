import logging
from datetime import date, datetime, timedelta, timezone as tzone
from decimal import Decimal
from itertools import chain
from pathlib import Path
from typing import Any

import iso8601
import phonenumbers
import regex
from django_redis import get_redis_connection
from openpyxl import load_workbook
from smartmin.models import SmartModel

from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db import models, transaction
from django.db.models import Count, F, Max, Q, Sum, Value
from django.db.models.functions import Concat, Lower
from django.utils import timezone
from django.utils.translation import gettext_lazy as _

from temba import mailroom
from temba.channels.models import Channel
from temba.locations.models import AdminBoundary
from temba.mailroom import ContactSpec, modifiers, queue_populate_dynamic_group
from temba.orgs.models import DependencyMixin, Export, ExportType, Org, OrgRole, User
from temba.utils import chunk_list, format_number, on_transaction_commit
from temba.utils.export import MultiSheetExporter
from temba.utils.models import JSONField, LegacyUUIDMixin, SquashableModel, TembaModel, delete_in_batches
from temba.utils.text import unsnakify
from temba.utils.urns import ParsedURN, parse_number, parse_urn
from temba.utils.uuid import uuid4

logger = logging.getLogger(__name__)


class URN:
    """
    Support class for URN strings. We differ from the strict definition of a URN (https://tools.ietf.org/html/rfc2141)
    in that:
        * We only supports URNs with scheme and path parts (no netloc, query, params or fragment)
        * Path component can be any non-blank unicode string
        * No hex escaping in URN path
    """

    DELETED_SCHEME = "deleted"
    DISCORD_SCHEME = "discord"
    EMAIL_SCHEME = "mailto"
    EXTERNAL_SCHEME = "ext"
    FACEBOOK_SCHEME = "facebook"
    FCM_SCHEME = "fcm"
    FRESHCHAT_SCHEME = "freshchat"
    INSTAGRAM_SCHEME = "instagram"
    JIOCHAT_SCHEME = "jiochat"
    LINE_SCHEME = "line"
    ROCKETCHAT_SCHEME = "rocketchat"
    SLACK_SCHEME = "slack"
    TELEGRAM_SCHEME = "telegram"
    TEL_SCHEME = "tel"
    TWITTERID_SCHEME = "twitterid"
    TWITTER_SCHEME = "twitter"
    VIBER_SCHEME = "viber"
    VK_SCHEME = "vk"
    WEBCHAT_SCHEME = "webchat"
    WECHAT_SCHEME = "wechat"
    WHATSAPP_SCHEME = "whatsapp"

    SCHEME_CHOICES = (
        (TEL_SCHEME, _("Phone Number")),
        (DISCORD_SCHEME, _("Discord Identifier")),
        (EMAIL_SCHEME, _("Email Address")),
        (EXTERNAL_SCHEME, _("External Identifier")),
        (FACEBOOK_SCHEME, _("Facebook Identifier")),
        (FCM_SCHEME, _("Firebase Cloud Messaging Identifier")),
        (FRESHCHAT_SCHEME, _("Freshchat Identifier")),
        (INSTAGRAM_SCHEME, _("Instagram Identifier")),
        (JIOCHAT_SCHEME, _("JioChat Identifier")),
        (LINE_SCHEME, _("LINE Identifier")),
        (ROCKETCHAT_SCHEME, _("RocketChat Identifier")),
        (SLACK_SCHEME, _("Slack Identifier")),
        (TELEGRAM_SCHEME, _("Telegram Identifier")),
        (TWITTERID_SCHEME, _("Twitter ID")),
        (TWITTER_SCHEME, _("Twitter Handle")),
        (VIBER_SCHEME, _("Viber Identifier")),
        (VK_SCHEME, _("VK Identifier")),
        (WECHAT_SCHEME, _("WeChat Identifier")),
        (WEBCHAT_SCHEME, _("Webchat Identifier")),
        (WHATSAPP_SCHEME, _("WhatsApp Identifier")),
    )

    VALID_SCHEMES = {s[0] for s in SCHEME_CHOICES}

    FACEBOOK_PATH_REF_PREFIX = "ref:"

    def __init__(self):  # pragma: no cover
        raise ValueError("Class shouldn't be instantiated")

    @classmethod
    def from_parts(cls, scheme, path, query=None, display=None):
        """
        Formats a URN scheme and path as single URN string, e.g. tel:+250783835665
        """
        if not scheme or (scheme not in cls.VALID_SCHEMES and scheme != cls.DELETED_SCHEME):
            raise ValueError("Invalid scheme component: '%s'" % scheme)

        if not path:
            raise ValueError("Invalid path component: '%s'" % path)

        return str(ParsedURN(scheme, path, query=query, fragment=display))

    @classmethod
    def to_parts(cls, urn):
        """
        Parses a URN string (e.g. tel:+250783835665) into a tuple of scheme and path
        """
        try:
            parsed = parse_urn(urn)
        except ValueError:
            raise ValueError("URN strings must contain scheme and path components")

        if parsed.scheme not in cls.VALID_SCHEMES and parsed.scheme != cls.DELETED_SCHEME:
            raise ValueError("URN contains an invalid scheme component: '%s'" % parsed.scheme)
        return parsed.scheme, parsed.path, parsed.query or None, parsed.fragment or None

    @classmethod
    def format(cls, urn, international=False, formatted=True):
        """
        formats this URN as a human friendly string
        """
        scheme, path, query, display = cls.to_parts(urn)

        if scheme in [cls.TEL_SCHEME, cls.WHATSAPP_SCHEME] and formatted:
            try:
                # whatsapp scheme is E164 without a leading +, add it so parsing works
                if scheme == cls.WHATSAPP_SCHEME:
                    path = "+" + path

                if path and path[0] == "+":
                    phone_format = phonenumbers.PhoneNumberFormat.NATIONAL
                    if international:
                        phone_format = phonenumbers.PhoneNumberFormat.INTERNATIONAL
                    return phonenumbers.format_number(phonenumbers.parse(path, None), phone_format)
            except phonenumbers.NumberParseException:  # pragma: no cover
                pass

        if display:
            return display

        return path

    @classmethod
    def validate(cls, urn, country_code=None):
        """
        Validates a normalized URN
        """
        try:
            scheme, path, query, display = cls.to_parts(urn)
        except ValueError:
            return False

        if scheme == cls.TEL_SCHEME:
            try:
                parse_number(path, country_code)
                return True
            except ValueError:
                return False

        # validate twitter URNs look like handles
        elif scheme == cls.TWITTER_SCHEME:
            return regex.match(r"^[a-zA-Z0-9_]{1,15}$", path, regex.V0)

        # validate path is a number and display is a handle if present
        elif scheme == cls.TWITTERID_SCHEME:
            valid = path.isdigit()
            if valid and display:
                valid = regex.match(r"^[a-zA-Z0-9_]{1,15}$", display, regex.V0)

            return valid

        elif scheme == cls.EMAIL_SCHEME:
            try:
                validate_email(path)
                return True
            except ValidationError:
                return False

        # facebook use integer ids or temp ref ids
        elif scheme in [cls.FACEBOOK_SCHEME]:
            # we don't validate facebook refs since they come from the outside
            if URN.is_path_fb_ref(path):
                return True

            # otherwise, this should be an int
            else:
                try:
                    int(path)
                    return True
                except ValueError:
                    return False

        # telegram, whatsapp and instagram use integer ids
        elif scheme in [cls.TELEGRAM_SCHEME, cls.WHATSAPP_SCHEME, cls.INSTAGRAM_SCHEME]:
            return regex.match(r"^[0-9]+$", path, regex.V0)

        # validate Viber URNS look right (this is a guess)
        elif scheme == cls.VIBER_SCHEME:  # pragma: needs cover
            return regex.match(r"^[a-zA-Z0-9_=+/]{1,24}$", path, regex.V0)

        # validate Freshchat URNS look right (this is a guess)
        elif scheme == cls.FRESHCHAT_SCHEME:  # pragma: needs cover
            return regex.match(
                r"^[0-9a-fA-F]{8}\-[0-9a-fA-F]{4}\-[0-9a-fA-F]{4}\-[0-9a-fA-F]{4}\-[0-9a-fA-F]{12}/[0-9a-fA-F]{8}\-[0-9a-fA-F]{4}\-[0-9a-fA-F]{4}\-[0-9a-fA-F]{4}\-[0-9a-fA-F]{12}$",
                path,
                regex.V0,
            )
        # Discord IDs are snowflakes, which are int64s internally
        elif scheme == cls.DISCORD_SCHEME:
            try:
                int(path)
                return True
            except ValueError:
                return False

        # anything goes for external schemes
        return True

    @classmethod
    def normalize(cls, urn, country_code=None):
        """
        Normalizes the path of a URN string. Should be called anytime looking for a URN match.
        """
        scheme, path, query, display = cls.to_parts(urn)

        country_code = str(country_code) if country_code else ""
        norm_path = str(path).strip()

        if scheme == cls.TEL_SCHEME:
            norm_path = cls.normalize_number(norm_path, country_code)
        elif scheme == cls.TWITTER_SCHEME:
            norm_path = norm_path.lower()
            if norm_path[0:1] == "@":  # strip @ prefix if provided
                norm_path = norm_path[1:]
            norm_path = norm_path.lower()  # Twitter handles are case-insensitive, so we always store as lowercase

        elif scheme == cls.TWITTERID_SCHEME:
            if display:
                display = str(display).strip().lower()
                if display and display[0] == "@":
                    display = display[1:]

        elif scheme == cls.EMAIL_SCHEME:
            norm_path = norm_path.lower()

        return cls.from_parts(scheme, norm_path, query, display)

    @classmethod
    def normalize_number(cls, number: str, country_code: str):
        """
        Normalizes the passed in number, they should be only digits, some backends prepend + and
        maybe crazy users put in dashes or parentheses in the console.
        """

        number = number.strip()
        normalized = number.lower()

        # if the number ends with e11, then that is Excel corrupting it, remove it
        if normalized.endswith("e+11") or normalized.endswith("e+12"):
            normalized = normalized[0:-4].replace(".", "")

        # remove non alphanumeric characters
        normalized = regex.sub(r"[^0-9a-z]", "", normalized, regex.V0)

        parse_as = normalized

        # if we started with + prefix, or we have a sufficiently long number that doesn't start with 0, add + prefix
        if number.startswith("+") or (len(normalized) >= 11 and not normalized.startswith("0")):
            parse_as = "+" + normalized

        try:
            formatted = parse_number(parse_as, country_code)
        except ValueError:
            # if it's not a possible number, just return what we have minus the +
            return normalized

        return formatted

    @classmethod
    def identity(cls, urn):
        scheme, path, query, display = URN.to_parts(urn)
        return URN.from_parts(scheme, path)

    @classmethod
    def is_path_fb_ref(cls, path):
        return path.startswith(cls.FACEBOOK_PATH_REF_PREFIX)

    @classmethod
    def from_tel(cls, path):
        return cls.from_parts(cls.TEL_SCHEME, path)

    @classmethod
    def from_discord(cls, path):
        return cls.from_parts(cls.DISCORD_SCHEME, path)


class UserContactFieldsQuerySet(models.QuerySet):
    pass


class UserContactFieldsManager(models.Manager):
    def get_queryset(self):
        return UserContactFieldsQuerySet(self.model, using=self._db).filter(is_system=False)


class ContactField(TembaModel, DependencyMixin):
    """
    A custom user field for contacts.
    """

    MAX_KEY_LEN = 36
    MAX_NAME_LEN = 36

    TYPE_TEXT = "T"
    TYPE_NUMBER = "N"
    TYPE_DATETIME = "D"
    TYPE_STATE = "S"
    TYPE_DISTRICT = "I"
    TYPE_WARD = "W"

    TYPE_CHOICES_BASIC = (
        (TYPE_TEXT, _("Text")),
        (TYPE_NUMBER, _("Number")),
        (TYPE_DATETIME, _("Date & Time")),
    )
    TYPE_CHOICES_LOCATIONS = (
        (TYPE_STATE, _("State")),
        (TYPE_DISTRICT, _("District")),
        (TYPE_WARD, _("Ward")),
    )
    TYPE_CHOICES = TYPE_CHOICES_BASIC + TYPE_CHOICES_LOCATIONS

    ACCESS_NONE = "N"
    ACCESS_VIEW = "V"
    ACCESS_EDIT = "E"
    ACCESS_CHOICES = ((ACCESS_NONE, _("Hidden")), (ACCESS_VIEW, _("View")), (ACCESS_EDIT, "Edit"))

    ENGINE_TYPES = {
        TYPE_TEXT: "text",
        TYPE_NUMBER: "number",
        TYPE_DATETIME: "datetime",
        TYPE_STATE: "state",
        TYPE_DISTRICT: "district",
        TYPE_WARD: "ward",
    }

    # system fields that all workspaces get
    SYSTEM_FIELDS = (
        # used for campaign events
        {"key": "created_on", "name": "Created On", "value_type": TYPE_DATETIME, "is_proxy": True},
        {"key": "last_seen_on", "name": "Last Seen On", "value_type": TYPE_DATETIME, "is_proxy": True},
    )

    # can't create custom contact fields with these keys
    RESERVED_KEYS = {"has", "is", "fields", "urns", "created_on", "last_seen_on"}

    org = models.ForeignKey(Org, on_delete=models.PROTECT, related_name="fields")

    key = models.CharField(max_length=MAX_KEY_LEN)
    name = models.CharField(max_length=MAX_NAME_LEN)
    value_type = models.CharField(choices=TYPE_CHOICES, max_length=1, default=TYPE_TEXT)
    is_proxy = models.BooleanField(default=False)  # field is just a proxy for a contact property

    # how field is displayed in the UI
    show_in_table = models.BooleanField(default=False)
    priority = models.PositiveIntegerField(default=0)
    agent_access = models.CharField(max_length=1, choices=ACCESS_CHOICES, default=ACCESS_VIEW)

    # model managers
    objects = models.Manager()
    user_fields = UserContactFieldsManager()

    org_limit_key = Org.LIMIT_FIELDS
    soft_dependent_types = {"flow", "campaign_event"}

    @classmethod
    def create_system_fields(cls, org):
        assert not org.fields.filter(is_system=True).exists(), "org already has system fields"

        for spec in cls.SYSTEM_FIELDS:
            org.fields.create(
                is_system=True,
                key=spec["key"],
                name=spec["name"],
                value_type=spec["value_type"],
                show_in_table=False,
                is_proxy=spec["is_proxy"],
                created_by=org.created_by,
                modified_by=org.modified_by,
            )

    @classmethod
    def create(
        cls, org, user, name: str, value_type: str = TYPE_TEXT, featured: bool = False, agent_access: str = ACCESS_VIEW
    ):
        """
        Creates a new non-system field based on the given name
        """
        assert cls.is_valid_name(name), f"{name} is not a valid field name"

        key = cls.make_key(name)

        assert cls.is_valid_key(key), f"{key} is not a valid field key"
        assert not org.fields.filter(is_active=True, key=key).exists()  # TODO replace with db constraint
        assert not org.fields.filter(is_active=True, name__iexact=name.lower()).exists()

        return cls.objects.create(
            org=org,
            key=key,
            name=name,
            value_type=value_type,
            is_system=False,
            show_in_table=featured,
            agent_access=agent_access,
            created_by=user,
            modified_by=user,
        )

    @classmethod
    def make_key(cls, name: str) -> str:
        """
        Generates a key from a name. There is no guarantee that the key is valid so should be checked with is_valid_key
        """
        key = regex.sub(r"([^a-z0-9]+)", " ", name.lower(), regex.V0)
        return regex.sub(r"([^a-z0-9]+)", "_", key.strip(), regex.V0)

    @classmethod
    def is_valid_key(cls, key: str) -> bool:
        if not regex.match(r"^[a-z][a-z0-9_]*$", key, regex.V0):
            return False
        if key in cls.RESERVED_KEYS or len(key) > cls.MAX_KEY_LEN:
            return False
        return True

    @classmethod
    def is_valid_name(cls, name: str) -> bool:
        return name == name.strip() and regex.match(r"^[A-Za-z0-9\- ]+$", name) and len(name) <= cls.MAX_NAME_LEN

    @classmethod
    def get_or_create(cls, org, user, key: str, name: str = None, value_type=None):
        """
        Gets the existing non-system contact field or creates a new field if it doesn't exist. This is method is only
        used for imports and may ignore or modify the given name to ensure validity and uniqueness.
        """

        existing = org.fields.filter(is_system=False, is_active=True, key=key).first()

        if existing:
            changed = False

            if name and existing.name != name and cls.is_valid_name(name):
                existing.name = cls.get_unique_name(org, name, ignore=existing)
                changed = True

            # update our type if we were given one
            if value_type and existing.value_type != value_type:
                # no changing away from datetime if we have campaign events
                is_date = existing.value_type == ContactField.TYPE_DATETIME
                if is_date and existing.campaign_events.filter(is_active=True).exists():
                    raise ValueError(f"Cannot change type for field '{key}' while it is used in campaigns.")

                existing.value_type = value_type
                changed = True

            if changed:
                existing.modified_by = user
                existing.save(update_fields=("name", "value_type", "modified_on", "modified_by"))

            return existing

        if not ContactField.is_valid_key(key):
            raise ValueError(f"'{key}' is not valid contact field key")

        # generate a name if we don't have one or the given one isn't valid
        if not name or not cls.is_valid_name(name):
            name = unsnakify(key)

        return org.fields.create(
            key=key,
            name=cls.get_unique_name(org, name),  # make unique if necessary
            is_system=False,
            value_type=value_type or cls.TYPE_TEXT,
            created_by=user,
            modified_by=user,
        )

    @classmethod
    def get_fields(cls, org: Org, viewable_by=None):
        """
        Gets the fields for the given org
        """

        fields = org.fields.filter(is_active=True, is_proxy=False)

        if viewable_by and org.get_user_role(viewable_by) == OrgRole.AGENT:
            fields = fields.exclude(agent_access=cls.ACCESS_NONE)

        return fields

    @classmethod
    def import_fields(cls, org, user, field_defs: list):
        """
        Import fields from a list of exported fields
        """

        db_types = {value: key for key, value in cls.ENGINE_TYPES.items()}

        for field_def in field_defs:
            field_key = field_def.get("key")
            field_name = field_def.get("name")
            field_type = field_def.get("type")
            cls.get_or_create(org, user, key=field_key, name=field_name, value_type=db_types[field_type])

    def as_export_def(self):
        return {"key": self.key, "name": self.name, "type": self.ENGINE_TYPES[self.value_type]}

    def get_dependents(self):
        dependents = super().get_dependents()
        dependents["group"] = self.dependent_groups.filter(is_active=True)
        dependents["campaign_event"] = self.campaign_events.filter(is_active=True)
        return dependents

    def get_access(self, user) -> str:
        return self.agent_access if self.org.get_user_role(user) == OrgRole.AGENT else self.ACCESS_EDIT

    def get_attrs(self):
        return {"icon": "info" if self.is_proxy else "fields"}

    def release(self, user):
        assert not (self.is_system and self.org.is_active), "can't release system fields"

        super().release(user)

        for event in self.campaign_events.all():
            event.release(user)

        self.name = self._deleted_name()
        self.is_active = False
        self.modified_by = user
        self.save(update_fields=("name", "is_active", "modified_on", "modified_by"))


class Contact(LegacyUUIDMixin, SmartModel):
    """
    A contact represents an individual with which we can communicate and collect data
    """

    STATUS_ACTIVE = "A"  # is active in flows, campaigns etc
    STATUS_BLOCKED = "B"  # was blocked by a user and their message will always be ignored
    STATUS_STOPPED = "S"  # opted out and their messages will be ignored until they message in again
    STATUS_ARCHIVED = "V"  # user intends to delete them
    STATUS_CHOICES = (
        (STATUS_ACTIVE, "Active"),
        (STATUS_BLOCKED, "Blocked"),
        (STATUS_STOPPED, "Stopped"),
        (STATUS_ARCHIVED, "Archived"),
    )
    ENGINE_STATUSES = {
        STATUS_ACTIVE: "active",
        STATUS_BLOCKED: "blocked",
        STATUS_STOPPED: "stopped",
        STATUS_ARCHIVED: "archived",
    }

    org = models.ForeignKey(Org, on_delete=models.PROTECT, related_name="contacts")

    name = models.CharField(verbose_name=_("Name"), max_length=128, blank=True, null=True)

    language = models.CharField(
        max_length=3,
        verbose_name=_("Language"),
        null=True,
        blank=True,
    )

    # custom field values for this contact, keyed by field UUID
    fields = JSONField(null=True)

    status = models.CharField(max_length=1, choices=STATUS_CHOICES, default=STATUS_ACTIVE)
    current_flow = models.ForeignKey("flows.Flow", on_delete=models.PROTECT, null=True, db_index=False)
    ticket_count = models.IntegerField(default=0)
    last_seen_on = models.DateTimeField(null=True)

    created_by = models.ForeignKey(
        settings.AUTH_USER_MODEL, on_delete=models.PROTECT, null=True, db_index=False, related_name="+"
    )
    modified_by = models.ForeignKey(
        settings.AUTH_USER_MODEL, on_delete=models.PROTECT, null=True, db_index=False, related_name="+"
    )

    # maximum number of contacts to release without using a background task
    BULK_RELEASE_IMMEDIATELY_LIMIT = 50

    @classmethod
    def create(
        cls,
        org,
        user,
        *,
        name: str,
        language: str,
        status: str,
        urns: list[str],
        fields: dict[ContactField, str],
        groups: list,
    ):
        engine_status = cls.ENGINE_STATUSES[status]
        fields_by_key = {f.key: v for f, v in fields.items()}
        group_uuids = [g.uuid for g in groups]

        return mailroom.get_client().contact_create(
            org,
            user,
            ContactSpec(
                name=name, language=language, status=engine_status, urns=urns, fields=fields_by_key, groups=group_uuids
            ),
        )

    @property
    def anon_display(self):
        """
        The displayable identifier used in place of URNs for anonymous orgs
        """
        return f"{self.id:010}"

    @classmethod
    def get_status_counts(cls, org) -> dict:
        """
        Returns the counts for each contact status for the given org
        """
        groups = org.groups.filter(group_type__in=ContactGroup.CONTACT_STATUS_TYPES)
        return {g.group_type: count for g, count in ContactGroupCount.get_totals(groups).items()}

    def get_scheduled_broadcasts(self):
        from temba.msgs.models import SystemLabel

        return (
            SystemLabel.get_queryset(self.org, SystemLabel.TYPE_SCHEDULED)
            .filter(schedule__next_fire__gte=timezone.now())
            .filter(Q(contacts__in=[self]) | Q(groups__in=self.groups.all()))
            .select_related("org", "schedule")
        )

    def get_scheduled_triggers(self):
        from temba.triggers.models import Trigger

        return (
            self.org.triggers.filter(
                trigger_type=Trigger.TYPE_SCHEDULE, schedule__next_fire__gte=timezone.now(), is_archived=False
            )
            .filter(Q(contacts__in=[self]) | Q(groups__in=self.groups.all()))
            .exclude(exclude_groups__in=self.groups.all())
            .select_related("schedule")
        )

    def get_scheduled(self, *, reverse: bool = False) -> list:
        """
        Gets this contact's upcoming scheduled events
        """
        from temba.campaigns.models import CampaignEvent

        fires = self.campaign_fires.filter(
            event__is_active=True, event__campaign__is_archived=False, scheduled__gte=timezone.now()
        ).select_related("event", "event__flow", "event__campaign")

        merged = []
        for fire in fires:
            obj = {
                "type": "campaign_event",
                "scheduled": fire.scheduled.isoformat(),
                "repeat_period": None,
                "campaign": fire.event.campaign.as_export_ref(),
            }
            if fire.event.event_type == CampaignEvent.TYPE_FLOW:
                obj["flow"] = fire.event.flow.as_export_ref()
            else:
                obj["message"] = fire.event.get_message(contact=self)

            merged.append(obj)

        for broadcast in self.get_scheduled_broadcasts():
            merged.append(
                {
                    "type": "scheduled_broadcast",
                    "scheduled": broadcast.schedule.next_fire.isoformat(),
                    "repeat_period": broadcast.schedule.repeat_period,
                    "message": broadcast.get_translation()["text"],
                }
            )

        for trigger in self.get_scheduled_triggers():
            merged.append(
                {
                    "type": "scheduled_trigger",
                    "scheduled": trigger.schedule.next_fire.isoformat(),
                    "repeat_period": trigger.schedule.repeat_period,
                    "flow": trigger.flow.as_export_ref(),
                }
            )

        return sorted(merged, key=lambda k: k["scheduled"], reverse=reverse)

    def get_history(self, after: datetime, before: datetime, include_event_types: set, ticket, limit: int) -> list:
        """
        Gets this contact's history of messages, calls, runs etc in the given time window
        """
        from temba.flows.models import FlowExit
        from temba.ivr.models import Call
        from temba.mailroom.events import get_event_time
        from temba.msgs.models import Msg
        from temba.tickets.models import TicketEvent

        msgs = (
            self.msgs.filter(created_on__gte=after, created_on__lt=before)
            .exclude(status=Msg.STATUS_PENDING)
            .order_by("-created_on", "-id")
            .select_related("channel", "contact_urn", "broadcast", "optin")[:limit]
        )

        # get all runs start started or ended in this period
        runs = (
            self.runs.filter(
                Q(created_on__gte=after, created_on__lt=before)
                | Q(exited_on__isnull=False, exited_on__gte=after, exited_on__lt=before)
            )
            .exclude(flow__is_system=True)
            .order_by("-created_on")
            .select_related("flow")[:limit]
        )
        started_runs = [r for r in runs if after <= r.created_on < before]
        exited_runs = [FlowExit(r) for r in runs if r.exited_on and after <= r.exited_on < before]

        channel_events = (
            self.channel_events.filter(created_on__gte=after, created_on__lt=before)
            .order_by("-created_on")
            .select_related("channel", "optin")[:limit]
        )

        campaign_events = (
            self.campaign_fires.filter(fired__gte=after, fired__lt=before)
            .exclude(fired=None)
            .order_by("-fired")
            .select_related("event__campaign", "event__relative_to")[:limit]
        )

        calls = (
            Call.objects.filter(contact=self, created_on__gte=after, created_on__lt=before)
            .exclude(status__in=[Call.STATUS_PENDING, Call.STATUS_WIRED])
            .order_by("-created_on")
            .select_related("channel")[:limit]
        )

        ticket_events = (
            self.ticket_events.filter(created_on__gte=after, created_on__lt=before)
            .select_related("ticket__topic", "assignee", "created_by")
            .order_by("-created_on")
        )

        if ticket:
            # if we have a ticket this is for the ticket UI, so we want *all* events for *only* that ticket
            ticket_events = ticket_events.filter(ticket=ticket)
        else:
            # if not then this for the contact read page so only show ticket opened/closed/reopened events
            ticket_events = ticket_events.filter(
                event_type__in=[TicketEvent.TYPE_OPENED, TicketEvent.TYPE_CLOSED, TicketEvent.TYPE_REOPENED]
            )

        ticket_events = ticket_events[:limit]

        transfers = self.airtime_transfers.filter(created_on__gte=after, created_on__lt=before).order_by("-created_on")[
            :limit
        ]

        session_events = self.get_session_events(after, before, include_event_types)

        # chain all items together, sort by their event time, and slice
        items = chain(
            msgs,
            started_runs,
            exited_runs,
            ticket_events,
            channel_events,
            campaign_events,
            calls,
            transfers,
            session_events,
        )

        # sort and slice
        return sorted(items, key=get_event_time, reverse=True)[:limit]

    def get_session_events(self, after: datetime, before: datetime, types: set) -> list:
        """
        Extracts events from this contacts sessions that overlap with the given time window
        """

        # limit to 100 sessions at a time to prevent melting when a contact has a lot of sessions
        sessions = self.sessions.filter(
            Q(created_on__gte=after, created_on__lt=before) | Q(ended_on__gte=after, ended_on__lt=before)
        ).order_by("-created_on")[:100]
        events = []
        for session in sessions:
            for run in session.output_json.get("runs", []):
                for event in run.get("events", []):
                    event["session_uuid"] = str(session.uuid)
                    event_time = iso8601.parse_date(event["created_on"])
                    if event["type"] in types and after <= event_time < before:
                        events.append(event)

        return events

    def get_field_serialized(self, field) -> str:
        """
        Given the passed in contact field object, returns the value (as a string) for this contact or None.
        """

        value_dict = self.fields.get(str(field.uuid)) if self.fields else None
        if not value_dict:
            return

        engine_type = ContactField.ENGINE_TYPES[field.value_type]

        if field.value_type == ContactField.TYPE_NUMBER:
            dec_value = value_dict.get(engine_type, value_dict.get("decimal"))
            return format_number(Decimal(dec_value)) if dec_value is not None else None

        return value_dict.get(engine_type)

    def get_field_value(self, field):
        """
        Given the passed in contact field object, returns the value (as a string, decimal, datetime, AdminBoundary)
        for this contact or None.
        """

        if field.is_proxy:  # i.e. created_on, last_seen_on
            return getattr(self, field.key)
        else:
            string_value = self.get_field_serialized(field)
            if string_value is None:
                return None

            if field.value_type == ContactField.TYPE_TEXT:
                return string_value
            elif field.value_type == ContactField.TYPE_DATETIME:
                return iso8601.parse_date(string_value)
            elif field.value_type == ContactField.TYPE_NUMBER:
                return Decimal(string_value)
            elif field.value_type in [ContactField.TYPE_STATE, ContactField.TYPE_DISTRICT, ContactField.TYPE_WARD]:
                return AdminBoundary.get_by_path(self.org, string_value)

    def get_field_display(self, field):
        """
        Returns the display value for the passed in field, or empty string if None
        """
        value = self.get_field_value(field)
        if value is None:
            return ""

        if field.value_type == ContactField.TYPE_DATETIME:
            return self.org.format_datetime(value)
        elif field.value_type == ContactField.TYPE_NUMBER:
            return format_number(value)
        elif (
            field.value_type in [ContactField.TYPE_STATE, ContactField.TYPE_DISTRICT, ContactField.TYPE_WARD] and value
        ):
            return value.name
        else:
            return str(value)

    def update(self, name: str, language: str) -> list[modifiers.Modifier]:
        """
        Updates attributes of this contact
        """
        mods = []
        if (self.name or "") != (name or ""):
            mods.append(modifiers.Name(name or ""))

        if (self.language or "") != (language or ""):
            mods.append(modifiers.Language(language or ""))

        return mods

    def update_fields(self, values: dict[ContactField, str]) -> list[modifiers.Modifier]:
        """
        Updates custom field values of this contact
        """
        mods = []

        for field, value in values.items():
            field_ref = modifiers.FieldRef(key=field.key, name=field.name)
            mods.append(modifiers.Field(field=field_ref, value=value))

        return mods

    def update_static_groups(self, groups) -> list[modifiers.Modifier]:
        """
        Updates the static groups for this contact to match the provided list
        """
        assert all([g for g in groups if g.group_type == ContactGroup.TYPE_MANUAL]), "can only update manual groups"

        current = self.groups.filter(group_type=ContactGroup.TYPE_MANUAL)

        # figure out our diffs, what groups need to be added or removed
        to_remove = [g for g in current if g not in groups]
        to_add = [g for g in groups if g not in current]

        def refs(gs):
            return [modifiers.GroupRef(uuid=str(g.uuid), name=g.name) for g in gs]

        mods = []

        if to_remove:
            mods.append(modifiers.Groups(groups=refs(to_remove), modification="remove"))
        if to_add:
            mods.append(modifiers.Groups(groups=refs(to_add), modification="add"))

        return mods

    def update_urns(self, urns: list[str]) -> list[modifiers.Modifier]:
        return [modifiers.URNs(urns=urns, modification="set")]

    def modify(self, user, mods: list[modifiers.Modifier], refresh=True):
        self.bulk_modify(user, [self], mods)
        if refresh:
            self.refresh_from_db()

    @classmethod
    def bulk_modify(cls, user, contacts, mods: list[modifiers.Modifier]):
        if not contacts:
            return

        org = contacts[0].org
        client = mailroom.get_client()
        try:
            response = client.contact_modify(org, user, contacts, mods)
        except mailroom.RequestException as e:
            logger.error(f"Contact update failed: {str(e)}", exc_info=True)
            raise e

        def modified(contact):
            c = response.get("modified", {}).get(contact.id, {}) or response.get(contact.id, {})
            return len(c.get("events", [])) > 0

        return [c.id for c in contacts if modified(c)]

    @classmethod
    def bulk_change_status(cls, user, contacts, status):
        cls.bulk_modify(user, contacts, [modifiers.Status(status=status)])

    @classmethod
    def bulk_change_group(cls, user, contacts, group, add: bool):
        mod = modifiers.Groups(
            groups=[modifiers.GroupRef(uuid=str(group.uuid), name=group.name)], modification="add" if add else "remove"
        )
        cls.bulk_modify(user, contacts, mods=[mod])

    @classmethod
    def apply_action_block(cls, user, contacts):
        cls.bulk_change_status(user, contacts, modifiers.Status.BLOCKED)

    @classmethod
    def apply_action_archive(cls, user, contacts):
        cls.bulk_change_status(user, contacts, modifiers.Status.ARCHIVED)

    @classmethod
    def apply_action_restore(cls, user, contacts):
        cls.bulk_change_status(user, contacts, modifiers.Status.ACTIVE)

    @classmethod
    def apply_action_label(cls, user, contacts, group):  # pragma: no cover
        cls.bulk_change_group(user, contacts, group, add=True)

    @classmethod
    def apply_action_unlabel(cls, user, contacts, group):  # pragma: no cover
        cls.bulk_change_group(user, contacts, group, add=False)

    @classmethod
    def apply_action_delete(cls, user, contacts):
        if len(contacts) <= cls.BULK_RELEASE_IMMEDIATELY_LIMIT:
            for contact in contacts:
                contact.release(user)
        else:
            from .tasks import release_contacts

            on_transaction_commit(lambda: release_contacts.delay(user.id, [c.id for c in contacts]))

    def set_note(self, user, text):
        """
        Adds a note to this contact, prunes old ones if necessary
        """
        self.notes.create(text=text, created_by=user)

        # remove all notes except the last 5
        notes = self.notes.order_by("-id").values_list("id", flat=True)[5:]
        self.notes.filter(id__in=notes).delete()

    def open_ticket(self, user, *, topic, assignee, note: str):
        """
        Opens a new ticket for this contact.
        """
        mod = modifiers.Ticket(
            topic=modifiers.TopicRef(uuid=str(topic.uuid), name=topic.name),
            assignee=modifiers.UserRef(email=assignee.email, name=assignee.name) if assignee else None,
            note=note,
        )
        self.modify(user, [mod], refresh=False)
        return self.tickets.order_by("id").last()

    def interrupt(self, user) -> bool:
        """
        Interrupts this contact's current flow
        """
        if self.current_flow:
            return mailroom.get_client().contact_interrupt(self.org, user, self) > 0

        return False

    def block(self, user):
        """
        Blocks this contact removing it from all non-smart groups
        """

        Contact.bulk_change_status(user, [self], modifiers.Status.BLOCKED)
        self.refresh_from_db()

    def stop(self, user):
        """
        Marks this contact has stopped, removing them from all groups.
        """

        Contact.bulk_change_status(user, [self], modifiers.Status.STOPPED)
        self.refresh_from_db()

    def archive(self, user):
        """
        Blocks this contact removing it from all non-smart groups
        """

        Contact.bulk_change_status(user, [self], modifiers.Status.ARCHIVED)
        self.refresh_from_db()

    def restore(self, user):
        """
        Restores a contact to active, re-adding them to any smart groups they belong to
        """

        Contact.bulk_change_status(user, [self], modifiers.Status.ACTIVE)
        self.refresh_from_db()

    def release(self, user, *, immediately=False, deindex=True):
        """
        Releases this contact. Note that we clear all identifying data but don't hard delete the contact because we need
        to expose deleted contacts over the API to allow external systems to know that contacts have been deleted.
        """
        from .tasks import full_release_contact

        # do de-indexing first so if it fails for some reason, we don't go through with the delete
        if deindex:
            mailroom.get_client().contact_deindex(self.org, [self])

        with transaction.atomic():
            # prep our urns for deletion so our old path creates a new urn
            for urn in self.urns.all():
                path = str(uuid4())
                urn.identity = f"{URN.DELETED_SCHEME}:{path}"
                urn.path = path
                urn.scheme = URN.DELETED_SCHEME
                urn.channel = None
                urn.save(update_fields=("identity", "path", "scheme", "channel"))

            # remove from non-db trigger groups
            for group in self.get_groups():
                group.contacts.remove(self)

            # delete any unfired campaign event fires
            self.campaign_fires.filter(fired=None).delete()

            # remove from scheduled broadcasts
            for bc in self.addressed_broadcasts.exclude(schedule=None):
                bc.contacts.remove(self)

            # now deactivate the contact itself
            self.is_active = False
            self.name = None
            self.fields = None
            self.modified_by = user
            self.save(update_fields=("name", "is_active", "fields", "modified_by", "modified_on"))

        # the hard work of removing everything this contact owns can be given to a celery task
        if immediately:
            self._full_release()
        else:
            on_transaction_commit(lambda: full_release_contact.delay(self.id))

    def _full_release(self):
        """
        Deletes everything owned by this contact
        """

        from temba.msgs.models import Msg

        assert not self.is_active, "can't fully release a contact which hasn't been released"

        with transaction.atomic():
            # release our tickets
            for ticket in self.tickets.all():
                ticket.delete()

            # delete our messages in batches
            while True:
                msg_batch = list(self.msgs.all()[:1000])
                if not msg_batch:
                    break
                Msg.bulk_delete(msg_batch)

            # any urns currently owned by us
            for urn in self.urns.all():
                # release any messages attached with each urn, these could include messages that began life
                # on a different contact
                for msg in urn.msgs.all():
                    msg.delete()

                # same thing goes for calls
                for call in urn.calls.all():
                    call.release()

                urn.release()

            # release our channel events
            delete_in_batches(self.channel_events.all())

            for run in self.runs.all():
                run.delete(interrupt=False)  # don't try interrupting sessions that are about to be deleted

            for session in self.sessions.all():
                session.delete()

            for call in self.calls.all():  # pragma: needs cover
                call.release()

            # and any event fire history
            self.campaign_fires.all().delete()

            # take us out of broadcast addressed contacts
            for broadcast in self.addressed_broadcasts.all():
                broadcast.contacts.remove(self)

    @classmethod
    def bulk_urn_cache_initialize(cls, contacts, *, using="default"):
        """
        Initializes the URN caches on the given contacts.
        """
        if not contacts:
            return

        contact_map = dict()
        for contact in contacts:
            contact_map[contact.id] = contact
            # initialize URN list cache
            setattr(contact, "_urns_cache", list())

        # cache all URN values (a priority ordered list on each contact)
        urns = (
            ContactURN.objects.filter(contact__in=contact_map.keys())
            .using(using)
            .order_by("contact", "-priority", "id")
        )
        for urn in urns:
            contact = contact_map[urn.contact_id]
            urn.org = contact.org
            getattr(contact, "_urns_cache").append(urn)

    @classmethod
    def bulk_inspect(self, contacts) -> dict:
        """
        Fetches additional information about the given contacts from mailroom
        """
        if not contacts:
            return {}

        return mailroom.get_client().contact_inspect(contacts[0].org, contacts)

    def get_groups(self, *, manual_only=False):
        """
        Gets the groups that this contact is a member of, excluding the status groups.
        """
        types = (ContactGroup.TYPE_MANUAL,) if manual_only else (ContactGroup.TYPE_MANUAL, ContactGroup.TYPE_SMART)
        return self.groups.filter(group_type__in=types, is_active=True)

    def get_urns(self):
        """
        Gets all URNs ordered by priority
        """
        cache_attr = "_urns_cache"
        if hasattr(self, cache_attr):
            return getattr(self, cache_attr)

        urns = self.urns.order_by("-priority", "pk").select_related("org")
        setattr(self, cache_attr, urns)
        return urns

    def get_urn(self, schemes=None):
        """
        Gets the highest priority matching URN for this contact. Schemes may be a single scheme or a set/list/tuple
        """
        if isinstance(schemes, str):
            schemes = (schemes,)

        urns = self.get_urns()

        if schemes is not None:
            for urn in urns:
                if urn.scheme in schemes:
                    return urn
            return None
        else:
            # otherwise return highest priority of any scheme
            return urns[0] if urns else None

    def get_display(self, org=None, formatted=True):
        """
        Gets a displayable name or URN for the contact. If available, org can be provided to avoid having to fetch it
        again based on the contact.
        """
        if not org:
            org = self.org

        if self.name:
            return self.name
        elif org.is_anon:
            return self.anon_display

        return self.get_urn_display(org=org, formatted=formatted)

    def get_urn_display(self, org=None, scheme=None, formatted=True, international=False):
        """
        Gets a displayable URN for the contact. If available, org can be provided to avoid having to fetch it again
        based on the contact.
        """
        if not org:
            org = self.org

        urn = self.get_urn(scheme)

        if not urn:
            return ""

        if org.is_anon:
            return ContactURN.ANON_MASK

        return urn.get_display(org=org, formatted=formatted, international=international) if urn else ""

    def __str__(self):
        return self.get_display()

    def __repr__(self):  # pragma: no cover
        return f'<Contact: id={self.id} name="{self.name}">'

    class Meta:
        indexes = [
            # for API endpoint access
            models.Index(name="contacts_by_org", fields=("org", "-modified_on", "-id"), condition=Q(is_active=True)),
            models.Index(
                name="contacts_by_org_deleted", fields=("org", "-modified_on", "-id"), condition=Q(is_active=False)
            ),
            # for getting the last modified_on during smart group population
            models.Index(name="contacts_contact_org_modified", fields=["org", "-modified_on"]),
            # for indexing modified contacts
            models.Index(name="contacts_modified", fields=("modified_on",)),
        ]
        constraints = [
            models.CheckConstraint(check=Q(status__in=("A", "B", "S", "V")), name="contact_status_valid"),
        ]


class ContactURN(models.Model):
    """
    A Universal Resource Name used to uniquely identify contacts, e.g. tel:+1234567890 or twitter:example
    """

    # schemes that support "new conversation" triggers
    SCHEMES_SUPPORTING_NEW_CONVERSATION = {
        URN.FACEBOOK_SCHEME,
        URN.VIBER_SCHEME,
        URN.TELEGRAM_SCHEME,
        URN.INSTAGRAM_SCHEME,
        URN.WEBCHAT_SCHEME,
    }
    SCHEMES_SUPPORTING_REFERRALS = {URN.FACEBOOK_SCHEME}  # schemes that support "referral" triggers
    SCHEMES_SUPPORTING_OPTINS = {URN.FACEBOOK_SCHEME}  # schemes that support opt-in/opt-out triggers

    # mailroom sets priorites like 1000, 999, ...
    PRIORITY_HIGHEST = 1000

    ANON_MASK = "*" * 8  # Returned instead of URN values for anon orgs
    ANON_MASK_HTML = "•" * 8  # Pretty HTML version of anon mask

    org = models.ForeignKey(Org, related_name="urns", on_delete=models.PROTECT)
    contact = models.ForeignKey(Contact, on_delete=models.PROTECT, null=True, related_name="urns")

    # the scheme and path which together should be unique
    identity = models.CharField(max_length=255)

    # individual parts of the URN
    scheme = models.CharField(max_length=128)
    path = models.CharField(max_length=255)
    display = models.CharField(max_length=255, null=True)

    priority = models.IntegerField(default=PRIORITY_HIGHEST)

    # the channel affinity of this URN
    channel = models.ForeignKey(Channel, related_name="urns", on_delete=models.PROTECT, null=True)

    # auth tokens - usage is channel specific, e.g. every FCM URN has its own token, FB channels have per opt-in tokens
    auth_tokens = models.JSONField(null=True)

    def release(self):
        delete_in_batches(self.channel_events.all())

        self.delete()

    def ensure_number_normalization(self, country_code):
        """
        Tries to normalize our phone number from a possible 10 digit (0788 383 383) to a 12 digit number
        with country code (+250788383383) using the country we now know about the channel.
        """
        number = self.path

        if number and not number[0] == "+" and country_code:
            norm_number = URN.normalize_number(number, country_code)

            # don't trounce existing contacts with that country code already
            norm_urn = URN.from_tel(norm_number)
            if not ContactURN.objects.filter(identity=norm_urn, org_id=self.org_id).exclude(id=self.id):
                self.identity = norm_urn
                self.path = norm_number
                self.save(update_fields=["identity", "path"])

        return self

    def get_display(self, org=None, international: bool = False, formatted: bool = True) -> str:
        """
        Gets a representation of the URN for display, e.g. tel:+12345678901 becomes +1 234 567-8901
        """
        if (org or self.org).is_anon:
            return self.ANON_MASK

        return URN.format(self.urn, international=international, formatted=formatted)

    @property
    def urn(self) -> str:
        """
        Returns a full representation of this contact URN as a string
        """
        return URN.from_parts(self.scheme, self.path, display=self.display)

    def __str__(self):  # pragma: no cover
        return self.urn

    class Meta:
        unique_together = ("identity", "org")
        ordering = ("-priority", "id")
        constraints = [
            models.CheckConstraint(check=~(Q(scheme="") | Q(path="")), name="non_empty_scheme_and_path"),
            models.CheckConstraint(
                check=Q(identity=Concat(F("scheme"), Value(":"), F("path"))), name="identity_matches_scheme_and_path"
            ),
        ]


class ContactGroup(LegacyUUIDMixin, TembaModel, DependencyMixin):
    """
    A group of contacts whose membership can be manual or query based
    """

    TYPE_DB_ACTIVE = "A"  # maintained by db trigger on status=A
    TYPE_DB_BLOCKED = "B"  # maintained by db trigger on status=B
    TYPE_DB_STOPPED = "S"  # maintained by db trigger on status=S
    TYPE_DB_ARCHIVED = "V"  # maintained by db trigger on status=V
    TYPE_MANUAL = "M"  # manual membership changes
    TYPE_SMART = "Q"  # maintained by engine using query
    TYPE_CHOICES = (
        (TYPE_DB_ACTIVE, "Active"),
        (TYPE_DB_BLOCKED, "Blocked"),
        (TYPE_DB_STOPPED, "Stopped"),
        (TYPE_DB_ARCHIVED, "Archived"),
        (TYPE_MANUAL, "Manual"),
        (TYPE_SMART, "Smart"),
    )

    CONTACT_STATUS_TYPES = [TYPE_DB_ACTIVE, TYPE_DB_BLOCKED, TYPE_DB_STOPPED, TYPE_DB_ARCHIVED]

    STATUS_INITIALIZING = "I"  # group has been created but not yet (re)evaluated
    STATUS_EVALUATING = "V"  # a task is currently (re)evaluating this group
    STATUS_READY = "R"  # group is ready for use
    STATUS_CHOICES = (
        (STATUS_INITIALIZING, _("Initializing")),
        (STATUS_EVALUATING, _("Evaluating")),
        (STATUS_READY, _("Ready")),
    )

    org = models.ForeignKey(Org, on_delete=models.PROTECT, related_name="groups")
    group_type = models.CharField(max_length=1, choices=TYPE_CHOICES, default=TYPE_MANUAL)
    status = models.CharField(max_length=1, choices=STATUS_CHOICES, default=STATUS_INITIALIZING)
    contacts = models.ManyToManyField(Contact, related_name="groups")

    # fields used by smart groups
    query = models.TextField(null=True)
    query_fields = models.ManyToManyField(ContactField, related_name="dependent_groups")

    org_limit_key = Org.LIMIT_GROUPS
    soft_dependent_types = {"flow", "trigger"}

    @classmethod
    def create_system_groups(cls, org):
        """
        Creates our system groups for the given organization so that we can keep track of counts etc..
        """

        assert not org.groups.filter(is_system=True).exists(), "org already has system groups"

        org.groups.create(
            name="Active",
            group_type=ContactGroup.TYPE_DB_ACTIVE,
            is_system=True,
            status=cls.STATUS_READY,
            created_by=org.created_by,
            modified_by=org.modified_by,
        )
        org.groups.create(
            name="Blocked",
            group_type=ContactGroup.TYPE_DB_BLOCKED,
            is_system=True,
            status=cls.STATUS_READY,
            created_by=org.created_by,
            modified_by=org.modified_by,
        )
        org.groups.create(
            name="Stopped",
            group_type=ContactGroup.TYPE_DB_STOPPED,
            is_system=True,
            status=cls.STATUS_READY,
            created_by=org.created_by,
            modified_by=org.modified_by,
        )
        org.groups.create(
            name="Archived",
            group_type=ContactGroup.TYPE_DB_ARCHIVED,
            is_system=True,
            status=cls.STATUS_READY,
            created_by=org.created_by,
            modified_by=org.modified_by,
        )
        org.groups.create(
            name="Open Tickets",
            group_type=ContactGroup.TYPE_SMART,
            is_system=True,
            query="tickets > 0",
            status=cls.STATUS_READY,  # since this group will always be empty for new orgs
            created_by=org.created_by,
            modified_by=org.modified_by,
        )

    @classmethod
    def get_groups(cls, org: Org, *, manual_only=False, ready_only=False):
        """
        Gets the groups (excluding db trigger based status groups) for the given org
        """
        types = (cls.TYPE_MANUAL,) if manual_only else (cls.TYPE_MANUAL, cls.TYPE_SMART)
        groups = cls.objects.filter(org=org, group_type__in=types, is_active=True)

        if ready_only:
            groups = groups.filter(status=cls.STATUS_READY)

        return groups

    @classmethod
    def get_group_by_name(cls, org, name):
        """
        Returns the user group with the passed in name
        """
        return cls.get_groups(org).filter(name__iexact=name).first()

    @classmethod
    def get_or_create(cls, org, user, name, query=None, uuid=None, parsed_query=None):
        existing = None

        if uuid:
            existing = org.groups.filter(uuid=uuid, is_active=True).first()

        if not existing and name:
            existing = cls.get_group_by_name(org, name)

        if existing:
            return existing

        assert name, "can't create group without a name"

        if query:
            return cls.create_smart(org, user, name, query, parsed_query=parsed_query)
        else:
            return cls.create_manual(org, user, name)

    @classmethod
    def create_manual(cls, org, user, name, *, status=STATUS_READY):
        """
        Creates a manual group whose members will be manually added and removed
        """
        return cls._create(org, user, name, status=status)

    @classmethod
    def create_smart(cls, org, user, name, query, evaluate=True, parsed_query=None):
        """
        Creates a smart group with the given query, e.g. gender=M
        """
        assert query, "query can't be empty for a smart group"

        group = cls._create(org, user, name, ContactGroup.STATUS_INITIALIZING, query=query)
        group.update_query(query=query, reevaluate=evaluate, parsed=parsed_query)
        return group

    @classmethod
    def _create(cls, org, user, name, status, query=None):
        assert cls.is_valid_name(name), f"'{name}' is not a valid group name"

        # look for name collision and append count if necessary
        name = cls.get_unique_name(org, base_name=name)

        return cls.objects.create(
            org=org,
            name=name,
            group_type=cls.TYPE_SMART if query else cls.TYPE_MANUAL,
            is_system=False,
            query=query,
            status=status,
            created_by=user,
            modified_by=user,
        )

    @property
    def icon(self) -> str:
        return "group_smart" if self.group_type == self.TYPE_SMART else "group"

    def get_attrs(self):
        return {"icon": self.icon}

    def update_query(self, query, reevaluate=True, parsed=None):
        """
        Updates the query for a smart group
        """

        assert self.group_type == self.TYPE_SMART, "can only update queries on smart groups"
        assert self.status != self.STATUS_EVALUATING, "group is already re-evaluating"

        try:
            if not parsed:
                parsed = mailroom.get_client().contact_parse_query(self.org, query)

            if not parsed.metadata.allow_as_group:
                raise ValueError(f"Cannot use query '{query}' as a smart group")

            self.query = parsed.query
            self.status = ContactGroup.STATUS_INITIALIZING
            self.save(update_fields=("query", "status"))

            self.query_fields.clear()

            # build our list of the fields we are dependent on
            field_keys = [f["key"] for f in parsed.metadata.fields]
            field_ids = []
            for c in self.org.fields.filter(is_active=True, key__in=field_keys).only("id"):
                field_ids.append(c.id)

            # and add them as dependencies
            self.query_fields.add(*field_ids)

        except mailroom.QueryValidationException as e:
            raise ValueError(str(e))

        # start background task to re-evaluate who belongs in this group
        if reevaluate:
            on_transaction_commit(lambda: queue_populate_dynamic_group(self))

    def get_member_count(self):
        """
        Returns the number of contacts in the group
        """
        return ContactGroupCount.get_totals([self])[self]

    def get_dependents(self):
        dependents = super().get_dependents()
        dependents["campaign"] = self.campaigns.filter(is_active=True)
        dependents["trigger"] = self.triggers.filter(is_active=True)
        return dependents

    def release(self, user, immediate: bool = False):
        """
        Releases this group, removing all contacts and marking as inactive
        """

        assert not (self.is_system and self.org.is_active), "can't release system groups"

        from .tasks import release_group_task

        # delete all triggers for this group
        for trigger in self.triggers.filter(is_active=True):
            trigger.release(user)

        super().release(user)

        self.name = self._deleted_name()
        self.is_active = False
        self.modified_by = user
        self.save(update_fields=("name", "is_active", "modified_by", "modified_on"))

        if immediate:
            self._full_release()
        else:
            # do the hard work of actually clearing out contacts etc in a background task
            on_transaction_commit(lambda: release_group_task.delay(self.id))

    def _full_release(self):
        # detach from contact imports associated with this group
        ContactImport.objects.filter(group=self).update(group=None)

        # remove from any scheduled broadcasts
        for bc in self.addressed_broadcasts.exclude(schedule=None):
            bc.groups.remove(self)

        # delete all counts for this group
        self.counts.all().delete()

        # delete the m2m related rows in batches, updating the contacts' modified_on as we go
        ContactGroupContacts = self.contacts.through
        memberships = ContactGroupContacts.objects.filter(contactgroup_id=self.id)

        for batch in chunk_list(memberships, 100):
            ContactGroupContacts.objects.filter(id__in=[m.id for m in batch]).delete()
            Contact.objects.filter(id__in=[m.contact_id for m in batch]).update(modified_on=timezone.now())

    @property
    def is_smart(self):
        return self.query is not None

    @property
    def triggers(self):
        from temba.triggers.models import Trigger

        return Trigger.objects.filter(Q(groups=self) | Q(exclude_groups=self))

    @classmethod
    def import_groups(cls, org, user, group_defs, dependency_mapping):
        """
        Import groups from a list of exported groups
        """

        for group_def in group_defs:
            group_uuid = group_def.get("uuid")
            group_name = group_def.get("name")
            group_query = group_def.get("query")

            parsed_query = None
            if group_query:
                parsed_query = mailroom.get_client().contact_parse_query(org, group_query, parse_only=True)
                for field_ref in parsed_query.metadata.fields:
                    ContactField.get_or_create(org, user, key=field_ref["key"])

            group = ContactGroup.get_or_create(
                org, user, cls.clean_name(group_name), group_query, uuid=group_uuid, parsed_query=parsed_query
            )

            dependency_mapping[group_uuid] = str(group.uuid)

    def as_export_def(self):
        return {"uuid": str(self.uuid), "name": self.name, "query": self.query}

    class Meta:
        verbose_name = _("Group")
        verbose_name_plural = _("Groups")

        constraints = [models.UniqueConstraint("org", Lower("name"), name="unique_contact_group_names")]


class ContactNote(models.Model):
    """
    Note attached to a contact, with last 5 versions kept for history.
    """

    MAX_LENGTH = 10_000

    contact = models.ForeignKey(Contact, on_delete=models.PROTECT, related_name="notes")
    text = models.TextField(max_length=MAX_LENGTH, blank=True)
    created_on = models.DateTimeField(default=timezone.now)
    created_by = models.ForeignKey(User, on_delete=models.PROTECT, related_name="contact_notes")


class ContactGroupCount(SquashableModel):
    """
    Maintains counts of contact groups. These are calculated via triggers on the database and squashed
    by a recurring task.
    """

    squash_over = ("group_id",)

    group = models.ForeignKey(ContactGroup, on_delete=models.PROTECT, related_name="counts", db_index=True)
    count = models.IntegerField(default=0)

    @classmethod
    def get_squash_query(cls, distinct_set):
        sql = """
        WITH deleted as (
            DELETE FROM %(table)s WHERE "group_id" = %%s RETURNING "count"
        )
        INSERT INTO %(table)s("group_id", "count", "is_squashed")
        VALUES (%%s, GREATEST(0, (SELECT SUM("count") FROM deleted)), TRUE);
        """ % {
            "table": cls._meta.db_table
        }

        return sql, (distinct_set.group_id,) * 2

    @classmethod
    def get_totals(cls, groups) -> dict:
        """
        Gets total counts for all the given groups
        """
        counts = cls.objects.filter(group__in=groups)
        counts = counts.values("group").order_by("group").annotate(count_sum=Sum("count"))
        counts_by_group_id = {c["group"]: c["count_sum"] for c in counts}
        return {g: counts_by_group_id.get(g.id, 0) for g in groups}

    @classmethod
    def populate_for_group(cls, group):
        # remove old ones
        ContactGroupCount.objects.filter(group=group).delete()

        # calculate our count for the group
        count = group.contacts.all().count()

        # insert updated count, returning it
        return ContactGroupCount.objects.create(group=group, count=count)

    class Meta:
        indexes = [
            models.Index(fields=("group",), condition=Q(is_squashed=False), name="contactgroupcounts_unsquashed")
        ]


class ContactExport(ExportType):
    """
    Export of contacts
    """

    slug = "contact"
    name = _("Contacts")
    download_prefix = "contacts"
    download_template = "contacts/export_download.html"

    @classmethod
    def create(cls, org, user, group=None, search=None, with_groups=()):
        export = Export.objects.create(
            org=org,
            export_type=cls.slug,
            config={
                "group_id": group.id if group else None,
                "search": search,
                "with_groups": [g.id for g in with_groups],
            },
            created_by=user,
        )
        return export

    def get_export_fields_and_schemes(self, export):
        fields = [
            dict(label="Contact UUID", key="uuid", field=None, urn_scheme=None),
            dict(label="Name", key="name", field=None, urn_scheme=None),
            dict(label="Language", key="language", field=None, urn_scheme=None),
            dict(label="Status", key="status", field=None, urn_scheme=None),
            dict(label="Created On", key="created_on", field=None, urn_scheme=None),
            dict(label="Last Seen On", key="last_seen_on", field=None, urn_scheme=None),
        ]

        # anon orgs also get an ID column that is just the PK
        if export.org.is_anon:
            fields = [
                dict(label="ID", key="id", field=None, urn_scheme=None),
                dict(label="Scheme", key="scheme", field=None, urn_scheme=None),
            ] + fields

        scheme_counts = dict()
        if not export.org.is_anon:
            org_urns = export.org.urns.using("readonly")
            schemes_in_use = sorted(list(org_urns.order_by().values_list("scheme", flat=True).distinct()))
            scheme_contact_max = {}

            # for each scheme used by this org, calculate the max number of URNs owned by a single contact
            for scheme in schemes_in_use:
                scheme_contact_max[scheme] = (
                    org_urns.filter(scheme=scheme)
                    .exclude(contact=None)
                    .values("contact")
                    .annotate(count=Count("contact"))
                    .aggregate(Max("count"))["count__max"]
                )

            for scheme in schemes_in_use:
                contact_max = scheme_contact_max.get(scheme) or 0
                for i in range(contact_max):
                    fields.append(
                        dict(
                            label=f"URN:{scheme.capitalize()}",
                            key=None,
                            field=None,
                            urn_scheme=scheme,
                            position=i,
                        )
                    )

        contact_fields_list = (
            ContactField.get_fields(export.org).using("readonly").select_related("org").order_by("-priority", "pk")
        )
        for contact_field in contact_fields_list:
            fields.append(
                dict(
                    field=contact_field,
                    label="Field:%s" % contact_field.name,
                    key=contact_field.key,
                    urn_scheme=None,
                )
            )

        group_fields = []
        for group in export.get_contact_groups():
            group_fields.append(dict(label="Group:%s" % group.name, key=None, group_id=group.id, group=group))

        return fields, scheme_counts, group_fields

    def get_group(self, export):
        group_id = export.config.get("group_id")
        if group_id:
            return export.org.groups.filter(id=group_id).get()
        else:
            return export.org.active_contacts_group

    def write(self, export):
        fields, scheme_counts, group_fields = self.get_export_fields_and_schemes(export)
        group = self.get_group(export)
        search = export.config.get("search")

        include_group_memberships = bool(len(group_fields) > 0)

        if search:
            contact_ids = mailroom.get_client().contact_export(export.org, group, query=search)
        else:
            contact_ids = group.contacts.using("readonly").order_by("id").values_list("id", flat=True)

        # create our exporter
        exporter = MultiSheetExporter(
            "Contact", [f["label"] for f in fields] + [g["label"] for g in group_fields], export.org.timezone
        )

        num_records = 0

        # write out contacts in batches to limit memory usage
        for batch_ids in chunk_list(contact_ids, 1000):
            # fetch all the contacts for our batch
            batch_contacts = (
                Contact.objects.filter(id__in=batch_ids).prefetch_related("org", "groups").using("readonly")
            )

            # to maintain our sort, we need to lookup by id, create a map of our id->contact to aid in that
            contact_by_id = {c.id: c for c in batch_contacts}

            Contact.bulk_urn_cache_initialize(batch_contacts, using="readonly")

            for contact_id in batch_ids:
                contact = contact_by_id[contact_id]

                values = []
                for field in fields:
                    values.append(self.get_field_value(export.org, field, contact=contact))

                group_values = []
                if include_group_memberships:
                    contact_groups_ids = [g.id for g in contact.groups.all()]
                    for col in range(len(group_fields)):
                        field = group_fields[col]
                        group_values.append(field["group_id"] in contact_groups_ids)

                # write this contact's values
                exporter.write_row(values + group_values)
                num_records += 1

            # keep bumping our modified_on to show we're alive
            if timezone.now() - export.modified_on > timedelta(minutes=3):  # pragma: no cover
                export.modified_on = timezone.now()
                export.save(update_fields=("modified_on",))

        return *exporter.save_file(), num_records

    def get_field_value(self, org, field: dict, contact: Contact):
        if field["key"] == "name":
            return contact.name
        elif field["key"] == "uuid":
            return contact.uuid
        elif field["key"] == "language":
            return contact.language
        elif field["key"] == "status":
            return contact.get_status_display()
        elif field["key"] == "created_on":
            return contact.created_on
        elif field["key"] == "last_seen_on":
            return contact.last_seen_on
        elif field["key"] == "id":
            return str(contact.id)
        elif field["key"] == "scheme":
            contact_urns = contact.get_urns()
            return contact_urns[0].scheme if contact_urns else ""
        elif field["urn_scheme"] is not None:
            contact_urns = contact.get_urns()
            scheme_urns = []
            for urn in contact_urns:
                if urn.scheme == field["urn_scheme"]:
                    scheme_urns.append(urn)
            position = field["position"]
            if len(scheme_urns) > position:
                urn_obj = scheme_urns[position]
                return urn_obj.get_display(org=org, formatted=False) if urn_obj else ""
            else:
                return ""
        else:
            return contact.get_field_display(field["field"])

    def get_download_context(self, export) -> dict:
        return {"group": self.get_group(export)}


def get_import_upload_path(instance: Any, filename: str):
    ext = Path(filename).suffix.lower()
    return f"orgs/{instance.org_id}/contact_imports/{uuid4()}{ext}"


class ContactImport(SmartModel):
    MAX_RECORDS = 25_000
    BATCH_SIZE = 100
    EXPLICIT_CLEAR = "--"

    # how many sequential URNs triggers flagging
    SEQUENTIAL_URNS_THRESHOLD = 250

    STATUS_PENDING = "P"
    STATUS_PROCESSING = "O"
    STATUS_COMPLETE = "C"
    STATUS_FAILED = "F"
    STATUS_CHOICES = (
        (STATUS_PENDING, "Pending"),
        (STATUS_PROCESSING, "Processing"),
        (STATUS_COMPLETE, "Complete"),
        (STATUS_FAILED, "Failed"),
    )

    MAPPING_IGNORE = {"type": "ignore"}

    org = models.ForeignKey(Org, on_delete=models.PROTECT, related_name="contact_imports")
    file = models.FileField(upload_to=get_import_upload_path)
    original_filename = models.TextField()
    mappings = models.JSONField()
    num_records = models.IntegerField()
    group_name = models.CharField(null=True, max_length=ContactGroup.MAX_NAME_LEN)
    group = models.ForeignKey(ContactGroup, on_delete=models.PROTECT, null=True, related_name="imports")
    started_on = models.DateTimeField(null=True)
    status = models.CharField(max_length=1, default=STATUS_PENDING, choices=STATUS_CHOICES)
    finished_on = models.DateTimeField(null=True)

    @classmethod
    def try_to_parse(cls, org: Org, file, filename: str) -> tuple[list, int]:
        """
        Tries to parse the given file stream as an import. If successful it returns the automatic column mappings and
        total number of records. Otherwise raises a ValidationError.
        """

        try:
            workbook = load_workbook(filename=file, read_only=True, data_only=True)
        except Exception:
            raise ValidationError(_("Import file appears to be corrupted."))
        ws = workbook.active

        # see https://openpyxl.readthedocs.io/en/latest/optimized.html#worksheet-dimensions but even with this we need
        # to ignore empty columns after the last column with data
        ws.reset_dimensions()

        data = ws.iter_rows()

        try:
            header_row = next(data)
        except StopIteration:
            raise ValidationError(_("Import file appears to be empty."))

        headers = [h.value for h in header_row]
        headers = [str(h).strip() if h else "" for h in headers]

        # ignore empty header columns after the last column with data
        max_col = 0
        for h, header in enumerate(headers):
            if header:
                max_col = h
        headers = headers[: max_col + 1]

        if any([h == "" for h in headers]):
            raise ValidationError(_("Import file contains an empty header."))

        mappings = cls._auto_mappings(org, headers)

        # iterate over rest of the rows to do row-level validation
        seen_uuids = set()
        seen_urns = set()
        num_records = 0
        row_num = 1

        while True:
            row_num += 1

            try:
                raw_row = next(data)
            except StopIteration:
                break

            row = cls._parse_row(raw_row, len(mappings))
            uuid, urns = cls._extract_uuid_and_urns(row, mappings)
            if uuid:
                if uuid in seen_uuids:
                    raise ValidationError(
                        _("Import file contains duplicated contact UUID '%(uuid)s' on row %(row)s."),
                        params={"uuid": uuid, "row": row_num},
                    )
                seen_uuids.add(uuid)
            for urn in urns:
                if urn in seen_urns:
                    raise ValidationError(
                        _("Import file contains duplicated contact URN '%(urn)s' on row %(row)s."),
                        params={"urn": urn, "row": row_num},
                    )
                seen_urns.add(urn)

            if uuid or urns:  # if we have a UUID or URN on this row it's an importable record
                num_records += 1

            # check if we exceed record limit
            if num_records > ContactImport.MAX_RECORDS:
                raise ValidationError(
                    _("Import files can contain a maximum of %(max)d records."),
                    params={"max": ContactImport.MAX_RECORDS},
                )

        if num_records == 0:
            raise ValidationError(_("Import file doesn't contain any records."))

        file.seek(0)  # seek back to beginning so subsequent reads work

        return mappings, num_records

    @staticmethod
    def _extract_uuid_and_urns(row, mappings) -> tuple[str, list[str]]:
        """
        Extracts any UUIDs and URNs from the given row so they can be checked for uniqueness
        """
        uuid = ""
        urns = []
        for value, item in zip(row, mappings):
            mapping = item["mapping"]
            if mapping["type"] == "attribute" and mapping["name"] == "uuid":
                uuid = value.lower()
            elif mapping["type"] == "scheme" and value:
                urn = URN.from_parts(mapping["scheme"], value)
                try:
                    urn = URN.normalize(urn)
                except ValueError:
                    pass
                urns.append(urn)
        return uuid, urns

    @classmethod
    def _auto_mappings(cls, org: Org, headers: list[str]) -> list:
        """
        Automatic mappings for the given list of headers - users can customize these later
        """

        fields_by_key = {}
        fields_by_name = {}
        for f in org.fields.filter(is_system=False, is_active=True):
            fields_by_key[f.key] = f
            fields_by_name[f.name.lower()] = f

        mappings = []

        for header in headers:
            header_prefix, header_name = cls._parse_header(header)
            mapping = ContactImport.MAPPING_IGNORE

            if header_prefix == "":
                attribute = header_name.lower()
                attribute = attribute.removeprefix("contact ")  # header "contact uuid" -> "uuid" etc

                if attribute in ("uuid", "name", "language", "status"):
                    mapping = {"type": "attribute", "name": attribute}
            elif header_prefix == "urn" and header_name:
                mapping = {"type": "scheme", "scheme": header_name.lower()}
            elif header_prefix == "field" and header_name:
                field_key = ContactField.make_key(header_name)

                # try to match by field name, then by key
                field = fields_by_name.get(header_name.lower())
                if not field:
                    field = fields_by_key.get(field_key)

                if field:
                    mapping = {"type": "field", "key": field.key, "name": field.name}
                else:
                    # can be created or selected in next step
                    mapping = {
                        "type": "new_field",
                        "key": field_key,
                        "name": unsnakify(header_name),
                        "value_type": "T",
                    }

            mappings.append({"header": header, "mapping": mapping})

        cls._validate_mappings(mappings)
        return mappings

    @staticmethod
    def _validate_mappings(mappings: list):
        non_ignored_mappings = []

        has_uuid, has_urn = False, False
        for item in mappings:
            header, mapping = item["header"], item["mapping"]

            if mapping["type"] == "attribute" and mapping["name"] == "uuid":
                has_uuid = True
            elif mapping["type"] == "scheme":
                has_urn = True
                if mapping["scheme"] not in URN.VALID_SCHEMES:
                    raise ValidationError(_("Header '%(header)s' is not a valid URN type."), params={"header": header})
            elif mapping["type"] == "new_field":
                if not ContactField.is_valid_key(mapping["key"]):
                    raise ValidationError(
                        _("Header '%(header)s' is not a valid field name."), params={"header": header}
                    )

            if mapping != ContactImport.MAPPING_IGNORE:
                non_ignored_mappings.append(mapping)

        if not (has_uuid or has_urn):
            raise ValidationError(_("Import files must contain either UUID or a URN header."))

        if has_uuid and len(non_ignored_mappings) == 1:
            raise ValidationError(_("Import files must contain columns besides UUID."))

    def start_async(self):
        from .tasks import import_contacts_task

        on_transaction_commit(lambda: import_contacts_task.delay(self.id))

    def delete(self):
        # delete our source import file
        self.file.delete()

        # delete any batches associated with this import
        ContactImportBatch.objects.filter(contact_import=self).delete()

        # delete any notifications attached this import
        self.notifications.all().delete()

        # then ourselves
        super().delete()

    def start(self):
        """
        Starts this import, creating batches to be handled by mailroom
        """

        assert self.status == self.STATUS_PENDING, "trying to start an already started import"

        # mark us as processing to prevent double starting
        self.status = self.STATUS_PROCESSING
        self.started_on = timezone.now()
        self.save(update_fields=("status", "started_on"))

        # create new contact fields as necessary
        for item in self.mappings:
            mapping = item["mapping"]
            if mapping["type"] == "new_field":
                ContactField.create(self.org, self.created_by, mapping["name"], value_type=mapping["value_type"])

        # if user wants contacts added to a new group, create it
        if self.group_name and not self.group:
            self.group = ContactGroup.create_manual(self.org, self.created_by, name=self.group_name)
            self.save(update_fields=("group",))

        # parse each row, creating batch tasks for mailroom
        workbook = load_workbook(filename=self.file, read_only=True, data_only=True)
        ws = workbook.active
        ws.reset_dimensions()  # see https://openpyxl.readthedocs.io/en/latest/optimized.html#worksheet-dimensions
        data = ws.iter_rows(min_row=2)

        urns = []
        batches = []

        for batch_specs, batch_start, batch_end in self._batches_generator(data):
            batches.append(self.batches.create(specs=batch_specs, record_start=batch_start, record_end=batch_end))

            for spec in batch_specs:
                urns.extend(spec.get("urns", []))

        # set redis key which mailroom batch tasks can decrement to know when import has completed
        r = get_redis_connection()
        r.set(f"contact_import_batches_remaining:{self.id}", len(batches), ex=24 * 60 * 60)

        # start each batch...
        for batch in batches:
            batch.import_async()

        # flag org if the set of imported URNs looks suspicious
        if not self.org.is_verified and self._detect_spamminess(urns):
            self.org.flag()

    def _batches_generator(self, row_iter):
        """
        Generator which takes an iterable of raw rows and returns tuples of 1. a batches of specs, 2. the record index
        at which the batch starts, 3. the record number at which the batch ends
        """
        record = 0
        batch_specs = []
        batch_start = record
        row = 1  # 1-based rows like Excel uses

        for raw_row in row_iter:
            row_data = self._parse_row(raw_row, len(self.mappings), tz=self.org.timezone)
            spec = self._row_to_spec(row_data)
            row += 1
            if spec:
                spec["_import_row"] = row
                batch_specs.append(spec)
                record += 1

            if len(batch_specs) == ContactImport.BATCH_SIZE:
                yield batch_specs, batch_start, record
                batch_specs = []
                batch_start = record

        if batch_specs:
            yield batch_specs, batch_start, record

    def get_info(self):
        """
        Gets info about this import by merging info from its batches
        """

        num_created = 0
        num_updated = 0
        num_errored = 0
        errors = []
        batches = self.batches.values("num_created", "num_updated", "num_errored", "errors")

        for batch in batches:
            num_created += batch["num_created"]
            num_updated += batch["num_updated"]
            num_errored += batch["num_errored"]
            errors.extend(batch["errors"])

        # sort errors by record #
        errors = sorted(errors, key=lambda e: e["record"])

        if self.finished_on:
            time_taken = self.finished_on - self.started_on
        elif self.started_on:
            time_taken = timezone.now() - self.started_on
        else:
            time_taken = timedelta(seconds=0)

        return {
            "status": self.status,
            "num_created": num_created,
            "num_updated": num_updated,
            "num_errored": num_errored,
            "errors": errors,
            "time_taken": int(time_taken.total_seconds()),
        }

    @staticmethod
    def _parse_header(header: str) -> tuple[str, str]:
        """
        Parses a header like "Field: Foo" into ("field", "Foo")
        """
        parts = header.split(":", maxsplit=1)
        parts = [p.strip() for p in parts]
        prefix, name = (parts[0], parts[1]) if len(parts) >= 2 else ("", parts[0])
        return prefix.lower(), name

    def _row_to_spec(self, row: list[str]) -> dict:
        """
        Convert a record (dict of headers to values) to a contact spec
        """

        spec = {}
        if self.group_id:
            spec["groups"] = [str(self.group.uuid)]

        for value, item in zip(row, self.mappings):
            mapping = item["mapping"]

            if not value:  # blank values interpreted as leaving values unchanged
                continue
            if value == ContactImport.EXPLICIT_CLEAR:
                value = ""

            if mapping["type"] == "attribute":
                attribute = mapping["name"]
                if attribute in ("uuid", "language", "status"):
                    value = value.lower()
                spec[attribute] = value
            elif mapping["type"] == "scheme":
                scheme = mapping["scheme"]
                if value:
                    if "urns" not in spec:
                        spec["urns"] = []
                    urn = URN.from_parts(scheme, value)
                    try:
                        urn = URN.normalize(urn, country_code=self.org.default_country_code)
                    except ValueError:
                        pass
                    spec["urns"].append(urn)

            elif mapping["type"] in ("field", "new_field"):
                if "fields" not in spec:
                    spec["fields"] = {}
                key = mapping["key"]
                spec["fields"][key] = value

        # Make sure the row has a UUID or URNs
        if not spec.get("uuid", "") and not spec.get("urns", []):
            return {}

        return spec

    @classmethod
    def _parse_row(cls, row: list[str], size: int, tz=None) -> list[str]:
        """
        Parses the raw values in the given row, returning a new list with the given size
        """
        parsed = []
        for i in range(size):
            parsed.append(cls._parse_value(row[i].value, tz=tz) if i < len(row) else "")
        return parsed

    @staticmethod
    def _parse_value(value: Any, tz=None) -> str:
        """
        Parses a record value into a string that can be serialized and understood by mailroom
        """

        if isinstance(value, datetime):
            # make naive datetime timezone-aware
            if not value.tzinfo and tz:
                value = value.replace(tzinfo=tz) if tz else value.replace(tzinfo=tzone.utc)

            return value.isoformat()
        elif isinstance(value, date):
            return value.isoformat()
        else:
            return str(value).strip() if value is not None else ""

    @classmethod
    def _detect_spamminess(cls, urns: list[str]) -> bool:
        """
        Takes the list of URNs that have been imported and tries to detect spamming
        """

        # extract all numerical URN paths
        numerical_paths = []
        for urn in urns:
            scheme, path, query, display = URN.to_parts(urn)
            try:
                numerical_paths.append(int(path))
            except ValueError:
                pass

        if len(numerical_paths) < cls.SEQUENTIAL_URNS_THRESHOLD:
            return False

        numerical_paths = sorted(numerical_paths)
        last_path = numerical_paths[0]
        num_sequential = 1
        for path in numerical_paths[1:]:
            if path == last_path + 1:
                num_sequential += 1
            last_path = path

            if num_sequential >= cls.SEQUENTIAL_URNS_THRESHOLD:
                return True

        return False

    def get_default_group_name(self):
        name = Path(self.original_filename).stem.title()
        name = name.replace("_", " ").replace("-", " ").strip()  # convert _- to spaces
        name = regex.sub(r"[^\w\s]", "", name)  # remove any non-word or non-space chars

        if len(name) >= ContactGroup.MAX_NAME_LEN - 10:  # truncate
            name = name[: ContactGroup.MAX_NAME_LEN - 10]
        elif len(name) < 4:  # default if too short
            name = "Import"

        return ContactGroup.get_unique_name(self.org, name)


class ContactImportBatch(models.Model):
    """
    A batch of contact records to be handled by mailroom
    """

    STATUS_CHOICES = (
        (ContactImport.STATUS_PENDING, "Pending"),
        (ContactImport.STATUS_PROCESSING, "Processing"),
        (ContactImport.STATUS_COMPLETE, "Complete"),
        (ContactImport.STATUS_FAILED, "Failed"),
    )

    contact_import = models.ForeignKey(ContactImport, on_delete=models.PROTECT, related_name="batches")
    status = models.CharField(max_length=1, default=ContactImport.STATUS_PENDING, choices=STATUS_CHOICES)
    specs = models.JSONField()

    # the range of records from the entire import contained in this batch
    record_start = models.IntegerField()
    record_end = models.IntegerField()

    # results written by mailroom after processing this batch
    num_created = models.IntegerField(default=0)
    num_updated = models.IntegerField(default=0)
    num_errored = models.IntegerField(default=0)
    errors = models.JSONField(default=list)
    finished_on = models.DateTimeField(null=True)

    def import_async(self):
        mailroom.queue_contact_import_batch(self)
