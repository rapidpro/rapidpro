# -*- coding: utf-8 -*-
from __future__ import absolute_import, division, print_function, unicode_literals

import json
import logging
import pytz
import regex
import requests
import six
import time
import traceback
import uuid

from datetime import datetime, timedelta
from django.conf import settings
from django.contrib.auth.models import User
from django.contrib.postgres.fields import ArrayField
from django.core.files.temp import NamedTemporaryFile
from django.db import models, transaction
from django.db.models import Count, Prefetch, Sum
from django.db.models.functions import Upper
from django.utils import timezone
from django.utils.html import escape
from django.utils.translation import ugettext, ugettext_lazy as _
from django_redis import get_redis_connection
from temba_expressions.evaluator import EvaluationContext, DateStyle
from temba.channels.courier import push_courier_msgs
from temba.assets.models import register_asset_store
from temba.contacts.models import Contact, ContactGroup, ContactURN, URN
from temba.channels.models import Channel, ChannelEvent
from temba.orgs.models import Org, TopUp, Language
from temba.schedules.models import Schedule
from temba.utils import analytics, chunk_list, on_transaction_commit, dict_to_json, get_anonymous_user
from temba.utils.dates import get_datetime_format, datetime_to_str, datetime_to_s
from temba.utils.export import BaseExportTask, BaseExportAssetStore
from temba.utils.expressions import evaluate_template
from temba.utils.http import http_headers
from temba.utils.models import SquashableModel, TembaModel, TranslatableField, JSONAsTextField
from temba.utils.queues import DEFAULT_PRIORITY, push_task, LOW_PRIORITY, HIGH_PRIORITY
from temba.utils.text import clean_string
from .handler import MessageHandler

logger = logging.getLogger(__name__)
__message_handlers = None

MSG_QUEUE = 'msgs'
SEND_MSG_TASK = 'send_msg_task'

HANDLER_QUEUE = 'handler'
HANDLE_EVENT_TASK = 'handle_event_task'
MSG_EVENT = 'msg'
FIRE_EVENT = 'fire'
TIMEOUT_EVENT = 'timeout'

BATCH_SIZE = 500

INITIALIZING = 'I'
PENDING = 'P'
QUEUED = 'Q'
WIRED = 'W'
SENT = 'S'
DELIVERED = 'D'
HANDLED = 'H'
ERRORED = 'E'
FAILED = 'F'
RESENT = 'R'

INCOMING = 'I'
OUTGOING = 'O'

INBOX = 'I'
FLOW = 'F'
IVR = 'V'
USSD = 'U'

MSG_SENT_KEY = 'msgs_sent_%y_%m_%d'

# status codes used for both messages and broadcasts (single char constant, human readable, API readable)
STATUS_CONFIG = (
    # special state for flows used to hold off sending the message until the flow is ready to receive a response
    (INITIALIZING, _("Initializing"), 'initializing'),

    (PENDING, _("Pending"), 'pending'),        # initial state for all messages

    # valid only for outgoing messages
    (QUEUED, _("Queued"), 'queued'),
    (WIRED, _("Wired"), 'wired'),              # message was handed off to the provider and credits were deducted for it
    (SENT, _("Sent"), 'sent'),                 # we have confirmation that a message was sent
    (DELIVERED, _("Delivered"), 'delivered'),

    # valid only for incoming messages
    (HANDLED, _("Handled"), 'handled'),

    (ERRORED, _("Error Sending"), 'errored'),  # there was an error during delivery
    (FAILED, _("Failed Sending"), 'failed'),   # we gave up on sending this message
    (RESENT, _("Resent message"), 'resent'),   # we retried this message
)


def get_message_handlers():
    """
    Initializes all our message handlers
    """
    global __message_handlers
    if not __message_handlers:
        handlers = []
        for handler_class in settings.MESSAGE_HANDLERS:
            try:
                cls = MessageHandler.find(handler_class)
                handlers.append(cls())
            except Exception:  # pragma: no cover
                traceback.print_exc()

        __message_handlers = handlers

    return __message_handlers


def get_unique_recipients(urns, contacts, groups):
    """
    Builds a list of the unique contacts and URNs by merging urns, contacts and groups
    """
    unique_urns = set()
    unique_contacts = set()
    included_by_urn = set()  # contact ids of contacts included by URN

    for urn in urns:
        unique_urns.add(urn)
        included_by_urn.add(urn.contact_id)

    for group in groups:
        for contact in group.contacts.all():
            if contact.id not in included_by_urn:
                unique_contacts.add(contact)

    for contact in contacts:
        if contact.id not in included_by_urn:
            unique_contacts.add(contact)

    return unique_urns, unique_contacts


class UnreachableException(Exception):
    """
    Exception thrown when a message is being sent to a contact that we don't have a sendable URN for
    """
    pass


class BroadcastRecipient(models.Model):
    """
    Through table for broadcast recipients many-to-many
    """
    broadcast = models.ForeignKey('msgs.Broadcast')

    contact = models.ForeignKey(Contact)

    purged_status = models.CharField(null=True, max_length=1,
                                     help_text=_("Used when broadcast is purged to record contact's message's state"))

    class Meta:
        db_table = 'msgs_broadcast_recipients'


@six.python_2_unicode_compatible
class Broadcast(models.Model):
    """
    A broadcast is a message that is sent out to more than one recipient, such
    as a ContactGroup or a list of Contacts. It's nothing more than a way to tie
    messages sent from the same bundle together
    """
    STATUS_CHOICES = [(s[0], s[1]) for s in STATUS_CONFIG]

    MAX_TEXT_LEN = settings.MSG_FIELD_SIZE

    METADATA_QUICK_REPLIES = 'quick_replies'

    org = models.ForeignKey(Org, verbose_name=_("Org"),
                            help_text=_("The org this broadcast is connected to"))

    groups = models.ManyToManyField(ContactGroup, verbose_name=_("Groups"), related_name='addressed_broadcasts',
                                    help_text=_("The groups to send the message to"))

    contacts = models.ManyToManyField(Contact, verbose_name=_("Contacts"), related_name='addressed_broadcasts',
                                      help_text=_("Individual contacts included in this message"))

    urns = models.ManyToManyField(ContactURN, verbose_name=_("URNs"), related_name='addressed_broadcasts',
                                  help_text=_("Individual URNs included in this message"))

    recipients = models.ManyToManyField(Contact, through=BroadcastRecipient, verbose_name=_("Recipients"),
                                        related_name='broadcasts',
                                        help_text=_("The contacts which received this message"))

    recipient_count = models.IntegerField(verbose_name=_("Number of recipients"), null=True,
                                          help_text=_("Number of urns which received this broadcast"))

    channel = models.ForeignKey(Channel, null=True, verbose_name=_("Channel"),
                                help_text=_("Channel to use for message sending"))

    status = models.CharField(max_length=1, verbose_name=_("Status"), choices=STATUS_CHOICES, default=INITIALIZING,
                              help_text=_("The current status for this broadcast"))

    schedule = models.OneToOneField(Schedule, verbose_name=_("Schedule"), null=True,
                                    help_text=_("Our recurring schedule if we have one"), related_name="broadcast")

    parent = models.ForeignKey('Broadcast', verbose_name=_("Parent"), null=True, related_name='children')

    text = TranslatableField(verbose_name=_("Translations"), max_length=MAX_TEXT_LEN,
                             help_text=_("The localized versions of the message text"))

    base_language = models.CharField(max_length=4,
                                     help_text=_('The language used to send this to contacts without a language'))

    is_active = models.BooleanField(default=True, help_text="Whether this broadcast is active")

    created_by = models.ForeignKey(User, related_name="%(app_label)s_%(class)s_creations",
                                   help_text="The user which originally created this item")

    created_on = models.DateTimeField(default=timezone.now, blank=True, editable=False, db_index=True,
                                      help_text=_("When this broadcast was created"))

    modified_by = models.ForeignKey(User, related_name="%(app_label)s_%(class)s_modifications",
                                    help_text="The user which last modified this item")

    modified_on = models.DateTimeField(auto_now=True,
                                       help_text="When this item was last modified")

    purged = models.BooleanField(default=False,
                                 help_text="If the messages for this broadcast have been purged")

    media = TranslatableField(verbose_name=_("Media"), max_length=2048,
                              help_text=_("The localized versions of the media"), null=True)

    send_all = models.BooleanField(default=False,
                                   help_text="Whether this broadcast should send to all URNs for each contact")

    metadata = JSONAsTextField(null=True, help_text=_("The metadata for messages of this broadcast"), default=dict)

    @classmethod
    def create(cls, org, user, text, recipients, base_language=None, channel=None, media=None, send_all=False,
               quick_replies=None, **kwargs):
        # for convenience broadcasts can still be created with single translation and no base_language
        if isinstance(text, six.string_types):
            base_language = org.primary_language.iso_code if org.primary_language else 'base'
            text = {base_language: text}

        if base_language not in text:  # pragma: no cover
            raise ValueError("Base language '%s' doesn't exist in the provided translations dict" % base_language)
        if media and base_language not in media:  # pragma: no cover
            raise ValueError("Base language '%s' doesn't exist in the provided media dict" % base_language)
        if quick_replies:
            for quick_reply in quick_replies:
                if base_language not in quick_reply:
                    raise ValueError("Base language '%s' doesn't exist for one or more of the provided quick replies" % base_language)

        metadata = dict(quick_replies=quick_replies) if quick_replies else {}

        broadcast = cls.objects.create(org=org, channel=channel, send_all=send_all,
                                       base_language=base_language, text=text, media=media,
                                       created_by=user, modified_by=user, metadata=metadata, **kwargs)
        broadcast.update_recipients(recipients)
        return broadcast

    def update_contacts(self, contact_ids):
        """
        Optimization for broadcasts that only contain contacts. Updates our contacts according to the passed in
        queryset or array.
        """
        self.urns.clear()
        self.groups.clear()
        self.contacts.clear()

        # get our through model
        RelatedModel = self.contacts.through

        # clear called automatically by django
        for chunk in chunk_list(contact_ids, 1000):
            bulk_contacts = [RelatedModel(contact_id=id, broadcast_id=self.id) for id in chunk]
            RelatedModel.objects.bulk_create(bulk_contacts)

        self.recipient_count = len(contact_ids)
        self.save(update_fields=('recipient_count',))

        # cache on object for use in subsequent send(..) calls
        delattr(self, '_recipient_cache')

    def update_recipients(self, recipients):
        """
        Updates the recipients which may be contact groups, contacts or contact URNs. Normally you can't update a
        broadcast after it has been created - the exception is scheduled broadcasts which are never really sent (clones
        of them are sent).
        """
        urns = []
        contacts = []
        groups = []

        for recipient in recipients:
            if isinstance(recipient, ContactURN):
                urns.append(recipient)
            elif isinstance(recipient, Contact):
                contacts.append(recipient)
            elif isinstance(recipient, ContactGroup):
                groups.append(recipient)
            else:  # pragma: needs cover
                raise ValueError("Recipient item is not a Contact, ContactURN or ContactGroup")

        self.urns.clear()
        self.urns.add(*urns)
        self.contacts.clear()
        self.contacts.add(*contacts)
        self.groups.clear()
        self.groups.add(*groups)

        urns, contacts = get_unique_recipients(urns, contacts, groups)

        # update the recipient count - the number of messages we intend to send
        self.recipient_count = len(urns) + len(contacts)
        self.save(update_fields=('recipient_count',))

        # cache on object for use in subsequent send(..) calls
        setattr(self, '_recipient_cache', (urns, contacts))

        return urns, contacts

    def has_pending_fire(self):  # pragma: needs cover
        return self.schedule and self.schedule.has_pending_fire()

    def fire(self):
        recipients = list(self.urns.all()) + list(self.contacts.all()) + list(self.groups.all())
        broadcast = Broadcast.create(self.org, self.created_by, self.text, recipients,
                                     media=self.media, base_language=self.base_language,
                                     parent=self)

        broadcast.send(trigger_send=True, expressions_context={})

        return broadcast

    @classmethod
    def get_broadcasts(cls, org, scheduled=False):
        qs = Broadcast.objects.filter(org=org).exclude(contacts__is_test=True)
        return qs.exclude(schedule=None) if scheduled else qs.filter(schedule=None)

    def get_messages(self):
        return self.msgs.exclude(status=RESENT)

    def get_message_count(self):
        return self.get_messages().count()

    def get_message_sending_count(self):  # pragma: needs cover
        return self.get_messages().filter(status__in=[PENDING, QUEUED]).count()

    def get_message_sent_count(self):  # pragma: needs cover
        return self.get_messages().filter(status__in=[SENT, DELIVERED, WIRED]).count()

    def get_message_delivered_count(self):  # pragma: needs cover
        return self.get_messages().filter(status=DELIVERED).count()

    def get_message_failed_count(self):  # pragma: needs cover
        return self.get_messages().filter(status__in=[FAILED, RESENT]).count()

    def get_preferred_languages(self, contact=None, org=None):
        """
        Gets the ordered list of language preferences for the given contact
        """
        org = org or self.org  # org object can be provided to allow caching of org languages
        preferred_languages = []

        # if contact has a language and it's a valid org language, it has priority
        if contact is not None and contact.language and contact.language in org.get_language_codes():
            preferred_languages.append(contact.language)

        if org.primary_language:
            preferred_languages.append(org.primary_language.iso_code)

        preferred_languages.append(self.base_language)

        return preferred_languages

    def get_default_text(self):
        """
        Gets the appropriate display text for the broadcast without a contact
        """
        return self.text[self.base_language]

    def get_translated_text(self, contact, org=None):
        """
        Gets the appropriate translation for the given contact
        """
        preferred_languages = self.get_preferred_languages(contact, org)
        return Language.get_localized_text(self.text, preferred_languages)

    def get_translated_quick_replies(self, contact, org=None):
        """
        Gets the appropriate quick replies translation for the given contact
        """
        preferred_languages = self.get_preferred_languages(contact, org)
        language_metadata = []
        metadata = self.metadata

        for item in metadata.get(self.METADATA_QUICK_REPLIES, []):
            text = Language.get_localized_text(text_translations=item, preferred_languages=preferred_languages)
            language_metadata.append(text)

        return language_metadata

    def get_translated_media(self, contact, org=None):
        """
        Gets the appropriate media for the given contact
        """
        preferred_languages = self.get_preferred_languages(contact, org)
        return Language.get_localized_text(self.media, preferred_languages)

    def send(self, trigger_send=True, expressions_context=None, response_to=None, status=PENDING, msg_type=INBOX,
             created_on=None, partial_recipients=None, run_map=None, high_priority=False):
        """
        Sends this broadcast by creating outgoing messages for each recipient.
        """
        # ignore mock messages
        if response_to and not response_to.id:  # pragma: no cover
            response_to = None

        # cannot ask for sending by us AND specify a created on, blow up in that case
        if trigger_send and created_on:  # pragma: no cover
            raise ValueError("Cannot trigger send and specify a created_on, breaks creating batches")

        if partial_recipients:
            # if flow is being started, it'll provide a batch of unique contacts itself
            urns, contacts = partial_recipients
        elif hasattr(self, '_recipient_cache'):
            # look to see if previous call to update_recipients left a cached value
            urns, contacts = self._recipient_cache
        else:
            # otherwise fetch everything and calculate
            urns, contacts = get_unique_recipients(self.urns.all(), self.contacts.all(), self.groups.all())

        Contact.bulk_cache_initialize(self.org, contacts)
        recipients = list(urns) + list(contacts)

        if self.send_all:
            recipients = list(urns)
            contact_list = list(contacts)
            for contact in contact_list:
                contact_urns = contact.get_urns()
                for c_urn in contact_urns:
                    recipients.append(c_urn)

            recipients = set(recipients)

        RelatedRecipient = Broadcast.recipients.through

        # we batch up our SQL calls to speed up the creation of our SMS objects
        batch = []
        batch_recipients = []
        existing_recipients = set(BroadcastRecipient.objects.filter(broadcast_id=self.id).values_list('contact_id', flat=True))

        # if they didn't pass in a created on, create one ourselves
        if not created_on:
            created_on = timezone.now()

        for recipient in recipients:
            contact = recipient if isinstance(recipient, Contact) else recipient.contact
            contact.org = self.org

            # get the appropriate translation for this contact
            text = self.get_translated_text(contact)

            # get the appropriate quick replies translation for this contact
            quick_replies = self.get_translated_quick_replies(contact)

            media = self.get_translated_media(contact)
            if media:
                media_type, media_url = media.split(':', 1)
                # arbitrary media urls don't have a full content type, so only
                # make uploads into fully qualified urls
                if media_url and len(media_type.split('/')) > 1:
                    media = "%s:https://%s/%s" % (media_type, settings.AWS_BUCKET_DOMAIN, media_url)

            # build our message specific context
            if expressions_context is not None:
                message_context = expressions_context.copy()
                if 'contact' not in message_context:
                    message_context['contact'] = contact.build_expressions_context()
            else:
                message_context = None

            # add in our parent context if the message references @parent
            if run_map:
                run = run_map.get(recipient.pk, None)
                if run and run.flow:
                    # a bit kludgy here, but should avoid most unnecessary context creations.
                    # since this path is an optimization for flow starts, we don't need to
                    # worry about the @child context.
                    if 'parent' in text:
                        if run.parent:
                            run.parent.org = self.org
                            message_context.update(dict(parent=run.parent.build_expressions_context()))

            try:
                msg = Msg.create_outgoing(self.org,
                                          self.created_by,
                                          recipient,
                                          text,
                                          broadcast=self,
                                          channel=self.channel,
                                          response_to=response_to,
                                          expressions_context=message_context,
                                          status=status,
                                          msg_type=msg_type,
                                          high_priority=high_priority,
                                          insert_object=False,
                                          attachments=[media] if media else None,
                                          created_on=created_on,
                                          quick_replies=quick_replies)

            except UnreachableException:
                # there was no way to reach this contact, do not create a message
                msg = None

            # only add it to our batch if it was legit
            if msg:
                batch.append(msg)

                # if this isn't an existing recipient, add it as one
                if msg.contact_id not in existing_recipients:
                    existing_recipients.add(msg.contact_id)
                    batch_recipients.append(RelatedRecipient(contact_id=msg.contact_id, broadcast_id=self.id))

            # we commit our messages in batches
            if len(batch) >= BATCH_SIZE:
                Msg.objects.bulk_create(batch)
                RelatedRecipient.objects.bulk_create(batch_recipients)

                # send any messages
                if trigger_send:
                    self.org.trigger_send(Msg.objects.filter(broadcast=self, created_on=created_on).select_related('contact', 'contact_urn', 'channel'))

                    # increment our created on so we can load our next batch
                    created_on = created_on + timedelta(seconds=1)

                batch = []
                batch_recipients = []

        # commit any remaining objects
        if batch:
            Msg.objects.bulk_create(batch)
            RelatedRecipient.objects.bulk_create(batch_recipients)

            if trigger_send:
                self.org.trigger_send(Msg.objects.filter(broadcast=self, created_on=created_on).select_related('contact', 'contact_urn', 'channel'))

        # for large batches, status is handled externally
        # we do this as with the high concurrency of sending we can run into postgresl deadlocks
        # (this could be our fault, or could be: http://www.postgresql.org/message-id/20140731233051.GN17765@andrew-ThinkPad-X230)
        if not partial_recipients:
            self.status = QUEUED if len(recipients) > 0 else SENT
            self.save(update_fields=('status',))

    def update(self):
        """
        Check the status of our messages and update ours accordingly
        """
        # build a map from status to the count for that status
        statuses = self.get_messages().values('status').order_by('status').annotate(count=Count('status'))
        total = 0
        status_map = dict()
        for status in statuses:
            status_map[status['status']] = status['count']
            total += status['count']

        # if errored msgs are greater than the half of all msgs
        if status_map.get(ERRORED, 0) > total // 2:
            self.status = ERRORED

        # if there are more than half failed, show failed
        elif status_map.get(FAILED, 0) > total // 2:
            self.status = FAILED

        # if there are any in Q, we are Q
        elif status_map.get(QUEUED, 0) or status_map.get(PENDING, 0):
            self.status = QUEUED

        # at this point we are either sent or delivered

        # if there are any messages that are only in a sent state
        elif status_map.get(SENT, 0) or status_map.get(WIRED, 0):
            self.status = SENT

        # otherwise, all messages delivered
        elif status_map.get(DELIVERED, 0) == total:
            self.status = DELIVERED

        self.save(update_fields=['status'])

    def __str__(self):
        return "%s (%s)" % (self.org.name, self.pk)


class Attachment(object):
    """
    Represents a message attachment stored as type:url
    """
    def __init__(self, content_type, url):
        self.content_type = content_type
        self.url = url

    @classmethod
    def parse(cls, s):
        return cls(*s.split(':', 1))

    @classmethod
    def parse_all(cls, attachments):
        return [cls.parse(s) for s in attachments] if attachments else []

    def as_json(self):
        return {'content_type': self.content_type, 'url': self.url}

    def __eq__(self, other):
        return self.content_type == other.content_type and self.url == other.url


@six.python_2_unicode_compatible
class Msg(models.Model):
    """
    Messages are the main building blocks of a RapidPro application. Channels send and receive
    these, Triggers and Flows handle them when appropriate.

    Messages are either inbound or outbound and can have varying states depending on their
    direction. Generally an outbound message will go through the following states:

      INITIALIZING > QUEUED > WIRED > SENT > DELIVERED

    If things go wrong, they can be put into an ERRORED state where they can be retried. Once
    we've given up then they can be put in the FAILED state.

    Inbound messages are much simpler. They start as PENDING and the can be picked up by Triggers
    or Flows where they would get set to the HANDLED state once they've been dealt with.
    """
    STATUS_CHOICES = [(s[0], s[1]) for s in STATUS_CONFIG]

    VISIBILITY_VISIBLE = 'V'
    VISIBILITY_ARCHIVED = 'A'
    VISIBILITY_DELETED = 'D'

    # single char flag, human readable name, API readable name
    VISIBILITY_CONFIG = ((VISIBILITY_VISIBLE, _("Visible"), 'visible'),
                         (VISIBILITY_ARCHIVED, _("Archived"), 'archived'),
                         (VISIBILITY_DELETED, _("Deleted"), 'deleted'))

    VISIBILITY_CHOICES = [(s[0], s[1]) for s in VISIBILITY_CONFIG]

    DIRECTION_CHOICES = ((INCOMING, _("Incoming")),
                         (OUTGOING, _("Outgoing")))

    MSG_TYPES = ((INBOX, _("Inbox Message")),
                 (FLOW, _("Flow Message")),
                 (IVR, _("IVR Message")),
                 (USSD, _("USSD Message")))

    MEDIA_GPS = 'geo'
    MEDIA_IMAGE = 'image'
    MEDIA_VIDEO = 'video'
    MEDIA_AUDIO = 'audio'

    MEDIA_TYPES = [MEDIA_AUDIO, MEDIA_GPS, MEDIA_IMAGE, MEDIA_VIDEO]

    CONTACT_HANDLING_QUEUE = 'ch:%d'

    MAX_TEXT_LEN = settings.MSG_FIELD_SIZE

    uuid = models.UUIDField(null=True, default=uuid.uuid4,
                            help_text=_("The UUID for this message"))

    org = models.ForeignKey(Org, related_name='msgs', verbose_name=_("Org"),
                            help_text=_("The org this message is connected to"))

    channel = models.ForeignKey(Channel, null=True,
                                related_name='msgs', verbose_name=_("Channel"),
                                help_text=_("The channel object that this message is associated with"))

    contact = models.ForeignKey(Contact,
                                related_name='msgs', verbose_name=_("Contact"),
                                help_text=_("The contact this message is communicating with"),
                                db_index=False)

    contact_urn = models.ForeignKey(ContactURN, null=True,
                                    related_name='msgs', verbose_name=_("Contact URN"),
                                    help_text=_("The URN this message is communicating with"))

    broadcast = models.ForeignKey(Broadcast, null=True, blank=True,
                                  related_name='msgs', verbose_name=_("Broadcast"),
                                  help_text=_("If this message was sent to more than one recipient"))

    text = models.TextField(verbose_name=_("Text"),
                            help_text=_("The actual message content that was sent"))

    high_priority = models.NullBooleanField(help_text=_("Give this message higher priority than other messages"))

    created_on = models.DateTimeField(verbose_name=_("Created On"), db_index=True,
                                      help_text=_("When this message was created"))

    modified_on = models.DateTimeField(null=True, blank=True, verbose_name=_("Modified On"), auto_now=True,
                                       help_text=_("When this message was last modified"))

    sent_on = models.DateTimeField(null=True, blank=True, verbose_name=_("Sent On"),
                                   help_text=_("When this message was sent to the endpoint"))

    queued_on = models.DateTimeField(null=True, blank=True, verbose_name=_("Queued On"),
                                     help_text=_("When this message was queued to be sent or handled."))

    direction = models.CharField(max_length=1, choices=DIRECTION_CHOICES, verbose_name=_("Direction"),
                                 help_text=_("The direction for this message, either incoming or outgoing"))

    status = models.CharField(max_length=1, choices=STATUS_CHOICES, default='P', verbose_name=_("Status"), db_index=True,
                              help_text=_("The current status for this message"))

    response_to = models.ForeignKey('Msg', null=True, blank=True, related_name='responses',
                                    verbose_name=_("Response To"), db_index=False,
                                    help_text=_("The message that this message is in reply to"))

    labels = models.ManyToManyField('Label', related_name='msgs', verbose_name=_("Labels"),
                                    help_text=_("Any labels on this message"))

    visibility = models.CharField(max_length=1, choices=VISIBILITY_CHOICES, default=VISIBILITY_VISIBLE,
                                  verbose_name=_("Visibility"),
                                  help_text=_("The current visibility of this message, either visible, archived or deleted"))

    msg_type = models.CharField(max_length=1, choices=MSG_TYPES, null=True, verbose_name=_("Message Type"),
                                help_text=_('The type of this message'))

    msg_count = models.IntegerField(default=1, verbose_name=_("Message Count"),
                                    help_text=_("The number of messages that were used to send this message, calculated on Twilio channels"))

    error_count = models.IntegerField(default=0, verbose_name=_("Error Count"),
                                      help_text=_("The number of times this message has errored"))

    next_attempt = models.DateTimeField(auto_now_add=True, verbose_name=_("Next Attempt"),
                                        help_text=_("When we should next attempt to deliver this message"))

    external_id = models.CharField(max_length=255, null=True, blank=True, verbose_name=_("External ID"),
                                   help_text=_("External id used for integrating with callbacks from other APIs"))

    topup = models.ForeignKey(TopUp, null=True, blank=True, related_name='msgs', on_delete=models.SET_NULL,
                              help_text="The topup that this message was deducted from")

    attachments = ArrayField(models.URLField(max_length=2048), null=True,
                             help_text=_("The media attachments on this message if any"))

    connection = models.ForeignKey('channels.ChannelSession', null=True,
                                   help_text=_("The session this message was a part of if any"))

    metadata = JSONAsTextField(null=True, help_text=_("The metadata for this msg"), default=dict)

    @classmethod
    def send_messages(cls, all_msgs):
        """
        Adds the passed in messages to our sending queue, this will also update the status of the message to
        queued.
        :return:
        """
        rapid_batches = []
        courier_batches = []

        # we send in chunks of 1,000 to help with contention
        for msgs in chunk_list(all_msgs, 1000):
            # build our id list
            msg_ids = set([m.id for m in msgs])

            with transaction.atomic():
                queued_on = timezone.now()
                courier_msgs = []
                task_msgs = []

                task_priority = None
                last_contact = None
                last_channel = None

                # update them to queued
                send_messages = Msg.objects.filter(id__in=msg_ids)\
                                           .exclude(channel__channel_type=Channel.TYPE_ANDROID)\
                                           .exclude(msg_type=IVR)\
                                           .exclude(topup=None)\
                                           .exclude(contact__is_test=True)
                send_messages.update(status=QUEUED, queued_on=queued_on, modified_on=queued_on)

                # now push each onto our queue
                for msg in msgs:
                    if (msg.msg_type != IVR and msg.channel and msg.channel.channel_type != Channel.TYPE_ANDROID) and msg.topup and not msg.contact.is_test:
                        if msg.channel.channel_type in settings.COURIER_CHANNELS and msg.uuid:
                            courier_msgs.append(msg)
                            continue

                        # if this is a different contact than our last, and we have msgs, queue the task
                        if task_msgs and last_contact != msg.contact_id:
                            # if no priority was set, default to DEFAULT
                            task_priority = DEFAULT_PRIORITY if task_priority is None else task_priority
                            rapid_batches.append(dict(org=task_msgs[0]['org'], msgs=task_msgs, priority=task_priority))
                            task_msgs = []
                            task_priority = None

                        # serialize the model to a dictionary
                        msg.queued_on = queued_on
                        task = msg.as_task_json()

                        # only be low priority if no priority has been set for this task group
                        if not msg.high_priority and task_priority is None:
                            task_priority = LOW_PRIORITY

                        task_msgs.append(task)
                        last_contact = msg.contact_id

                if task_msgs:
                    task_priority = DEFAULT_PRIORITY if task_priority is None else task_priority
                    rapid_batches.append(dict(org=task_msgs[0]['org'], msgs=task_msgs, priority=task_priority))
                    task_msgs = []

                # ok, now push our courier msgs
                last_contact = None
                last_channel = None
                for msg in courier_msgs:
                    if task_msgs and (last_contact != msg.contact_id or last_channel != msg.channel_id):
                        courier_batches.append(dict(channel=task_msgs[0].channel, msgs=task_msgs,
                                                    high_priority=task_msgs[0].high_priority))
                        task_msgs = []

                    last_contact = msg.contact_id
                    last_channel = msg.channel_id
                    task_msgs.append(msg)

                # push any remaining courier msgs
                if task_msgs:
                    courier_batches.append(dict(channel=task_msgs[0].channel, msgs=task_msgs,
                                                high_priority=task_msgs[0].high_priority))

        # send our batches
        on_transaction_commit(lambda: cls._send_rapid_msg_batches(rapid_batches))
        on_transaction_commit(lambda: cls._send_courier_msg_batches(courier_batches))

    @classmethod
    def _send_rapid_msg_batches(cls, batches):
        for batch in batches:
            push_task(batch['org'], MSG_QUEUE, SEND_MSG_TASK, batch['msgs'], priority=batch['priority'])

    @classmethod
    def _send_courier_msg_batches(cls, batches):
        for batch in batches:
            push_courier_msgs(batch['channel'], batch['msgs'], batch['high_priority'])

    @classmethod
    def process_message(cls, msg):
        """
        Processes a message, running it through all our handlers
        """
        from temba.orgs.models import CHATBASE_TYPE_USER

        handlers = get_message_handlers()

        if msg.contact.is_blocked:
            msg.visibility = Msg.VISIBILITY_ARCHIVED
            msg.save(update_fields=['visibility', 'modified_on'])
        else:
            for handler in handlers:
                try:
                    start = None
                    if settings.DEBUG:  # pragma: no cover
                        start = time.time()

                    handled = handler.handle(msg)

                    if start:  # pragma: no cover
                        print("[%0.2f] %s for %d" % (time.time() - start, handler.name, msg.pk or 0))

                    if handled:
                        break
                except Exception as e:  # pragma: no cover
                    import traceback
                    traceback.print_exc()
                    logger.exception("Error in message handling: %s" % e)

        cls.mark_handled(msg)

        # chatbase parameters to track logs
        chatbase_not_handled = msg.msg_type != FLOW

        # Sending data to Chatbase API
        if not msg.contact.is_test:
            (chatbase_api_key, chatbase_version) = msg.org.get_chatbase_credentials()
            if chatbase_api_key:
                cls.send_chatbase_log(chatbase_api_key, chatbase_version, msg.channel.name, msg.text, msg.contact.id,
                                      CHATBASE_TYPE_USER, chatbase_not_handled)

        # record our handling latency for this object
        if msg.queued_on:
            analytics.gauge('temba.handling_latency', (msg.modified_on - msg.queued_on).total_seconds())

        # this is the latency from when the message was received at the channel, which may be different than
        # above if people above us are queueing (or just because clocks are out of sync)
        analytics.gauge('temba.channel_handling_latency', (msg.modified_on - msg.created_on).total_seconds())

    @classmethod
    def send_chatbase_log(cls, chatbase_api_key, chatbase_version, channel_name, text, contact_id, log_type,
                          not_handled=True):
        """
        Send messages logs in batch to Chatbase
        """
        if not settings.SEND_CHATBASE:
            raise Exception("!! Skipping Chatbase request, SEND_CHATBASE set to False")

        message = {
            'type': log_type,
            'user_id': contact_id,
            'platform': channel_name,
            'message': text,
            'time_stamp': int(time.time()),
            'api_key': chatbase_api_key
        }

        if chatbase_version:
            message['version'] = chatbase_version

        if log_type == 'user' and not_handled:
            message['not_handled'] = not_handled

        headers = http_headers(extra={'Content-Type': 'application/json'})

        requests.post(settings.CHATBASE_API_URL, data=json.dumps(message), headers=headers)

    @classmethod
    def get_messages(cls, org, is_archived=False, direction=None, msg_type=None):
        messages = Msg.objects.filter(org=org)

        if is_archived:  # pragma: needs cover
            messages = messages.filter(visibility=Msg.VISIBILITY_ARCHIVED)
        else:
            messages = messages.filter(visibility=Msg.VISIBILITY_VISIBLE)

        if direction:  # pragma: needs cover
            messages = messages.filter(direction=direction)

        if msg_type:  # pragma: needs cover
            messages = messages.filter(msg_type=msg_type)

        return messages.filter(contact__is_test=False)

    @classmethod
    def fail_old_messages(cls):  # pragma: needs cover
        """
        Looks for any errored or queued messages more than a week old and fails them. Messages that old would
        probably be confusing to go out.
        """
        one_week_ago = timezone.now() - timedelta(days=7)
        failed_messages = Msg.objects.filter(created_on__lte=one_week_ago, direction=OUTGOING,
                                             status__in=[QUEUED, PENDING, ERRORED])

        failed_broadcasts = list(failed_messages.order_by('broadcast').values('broadcast').distinct())

        # fail our messages
        failed_messages.update(status='F', modified_on=timezone.now())

        # and update all related broadcast statuses
        for broadcast in Broadcast.objects.filter(id__in=[b['broadcast'] for b in failed_broadcasts]):
            broadcast.update()

    @classmethod
    def mark_handled(cls, msg):
        """
        Marks an incoming message as HANDLED
        """
        update_fields = ['status', 'modified_on']

        # if flows or IVR haven't claimed this message, then it's going to the inbox
        if not msg.msg_type:
            msg.msg_type = INBOX
            update_fields.append('msg_type')

        msg.status = HANDLED

        # make sure we don't overwrite any async message changes by only saving specific fields
        msg.save(update_fields=update_fields)

    @classmethod
    def mark_error(cls, r, channel, msg, fatal=False):
        """
        Marks an outgoing message as FAILED or ERRORED
        :param msg: a JSON representation of the message or a Msg object
        """
        msg.error_count += 1
        if msg.error_count >= 3 or fatal:
            if isinstance(msg, Msg):
                msg.status_fail()
            else:
                Msg.objects.select_related('org').get(pk=msg.id).status_fail()

            if channel:
                analytics.gauge('temba.msg_failed_%s' % channel.channel_type.lower())
        else:
            msg.status = ERRORED
            msg.next_attempt = timezone.now() + timedelta(minutes=5 * msg.error_count)

            if isinstance(msg, Msg):
                msg.save(update_fields=('status', 'modified_on', 'next_attempt', 'error_count'))
            else:
                Msg.objects.filter(id=msg.id).update(status=msg.status, next_attempt=msg.next_attempt,
                                                     error_count=msg.error_count, modified_on=msg.modified_on)

            # clear that we tried to send this message (otherwise we'll ignore it when we retry)
            pipe = r.pipeline()
            pipe.srem(timezone.now().strftime(MSG_SENT_KEY), str(msg.id))
            pipe.srem((timezone.now() - timedelta(days=1)).strftime(MSG_SENT_KEY), str(msg.id))
            pipe.execute()

            if channel:
                analytics.gauge('temba.msg_errored_%s' % channel.channel_type.lower())

    @classmethod
    def mark_sent(cls, r, msg, status, external_id=None):
        """
        Marks an outgoing message as WIRED or SENT
        :param msg: a JSON representation of the message
        """
        msg.status = status
        msg.sent_on = timezone.now()
        if external_id:
            msg.external_id = external_id

        # use redis to mark this message sent
        pipe = r.pipeline()
        sent_key = timezone.now().strftime(MSG_SENT_KEY)
        pipe.sadd(sent_key, str(msg.id))
        pipe.expire(sent_key, 86400)
        pipe.execute()

        if external_id:
            Msg.objects.filter(id=msg.id).update(status=status, sent_on=msg.sent_on, external_id=external_id)
        else:
            Msg.objects.filter(id=msg.id).update(status=status, sent_on=msg.sent_on)

    def as_json(self):
        return dict(direction=self.direction,
                    text=self.text,
                    id=self.id,
                    attachments=self.attachments,
                    created_on=self.created_on.strftime('%x %X'),
                    model="msg",
                    metadata=self.metadata)

    def simulator_json(self):
        msg_json = self.as_json()
        msg_json['text'] = escape(self.text).replace('\n', "<br/>")
        return msg_json

    @classmethod
    def get_text_parts(cls, text, max_length=160):
        """
        Breaks our message into 160 character parts
        """
        if len(text) < max_length or max_length <= 0:
            return [text]

        else:
            def next_part(text):
                if len(text) <= max_length:
                    return text, None

                else:
                    # search for a space to split on, up to 140 characters in
                    index = max_length
                    while index > max_length - 20:
                        if text[index] == ' ':
                            break
                        index -= 1

                    # couldn't find a good split, oh well, 160 it is
                    if index == max_length - 20:
                        return text[:max_length], text[max_length:]
                    else:
                        return text[:index], text[index + 1:]

            parts = []
            rest = text
            while rest:
                (part, rest) = next_part(rest)
                parts.append(part)

            return parts

    @classmethod
    def get_sync_commands(cls, msgs):
        """
        Returns the minimal # of broadcast commands for the given Android channel to uniquely represent all the
        messages which are being sent to tel URNs. This will return an array of dicts that look like:
             dict(cmd="mt_bcast", to=[dict(phone=msg.contact.tel, id=msg.pk) for msg in msgs], msg=broadcast.text))
        """
        commands = []
        current_text = None
        contact_id_pairs = []

        for m in msgs.values('id', 'text', 'contact_urn__path').order_by('created_on'):
            if m['text'] != current_text and contact_id_pairs:
                commands.append(dict(cmd='mt_bcast', to=contact_id_pairs, msg=current_text))
                contact_id_pairs = []

            current_text = m['text']
            contact_id_pairs.append(dict(phone=m['contact_urn__path'], id=m['id']))

        if contact_id_pairs:
            commands.append(dict(cmd='mt_bcast', to=contact_id_pairs, msg=current_text))

        return commands

    def get_attachments(self):
        """
        Gets this message's attachments parsed into actual attachment objects
        """
        return Attachment.parse_all(self.attachments)

    def get_last_log(self):
        """
        Gets the last channel log for this message. Performs sorting in Python to ease pre-fetching.
        """
        sorted_logs = None
        if self.channel and self.channel.is_active:
            sorted_logs = sorted(self.channel_logs.all(), key=lambda l: l.created_on, reverse=True)
        return sorted_logs[0] if sorted_logs else None

    def reply(self, text, user, trigger_send=False, expressions_context=None, connection=None, attachments=None, msg_type=None,
              send_all=False, created_on=None, quick_replies=None):

        return self.contact.send(text, user, trigger_send=trigger_send, expressions_context=expressions_context,
                                 response_to=self if self.id else None, connection=connection, attachments=attachments,
                                 msg_type=msg_type or self.msg_type, created_on=created_on, all_urns=send_all,
                                 high_priority=True, quick_replies=quick_replies)

    def update(self, cmd):
        """
        Updates our message according to the provided client command
        """
        from temba.api.models import WebHookEvent
        date = datetime.fromtimestamp(int(cmd['ts']) // 1000).replace(tzinfo=pytz.utc)

        keyword = cmd['cmd']
        handled = False

        if keyword == 'mt_error':
            self.status = ERRORED
            handled = True

        elif keyword == 'mt_fail':
            self.status = FAILED
            handled = True
            WebHookEvent.trigger_sms_event(WebHookEvent.TYPE_SMS_FAIL, self, date)

        elif keyword == 'mt_sent':
            self.status = SENT
            self.sent_on = date
            handled = True
            WebHookEvent.trigger_sms_event(WebHookEvent.TYPE_SMS_SENT, self, date)

        elif keyword == 'mt_dlvd':
            self.status = DELIVERED
            handled = True
            WebHookEvent.trigger_sms_event(WebHookEvent.TYPE_SMS_DELIVERED, self, date)

        self.save()  # first save message status before updating the broadcast status

        # update our broadcast if we have one
        if self.broadcast:
            self.broadcast.update()

        return handled

    def queue_handling(self, new_message=False, new_contact=False):
        """
        Queues this message to be handled by one of our celery queues

        new_message - should be true for messages which were created outside rapidpro
        new_contact - should be true for contacts which were created outside rapidpro
        """
        payload = dict(type=MSG_EVENT, contact_id=self.contact.id, id=self.id, new_message=new_message, new_contact=new_contact)

        # first push our msg on our contact's queue using our created date
        r = get_redis_connection('default')
        queue_time = self.sent_on if self.sent_on else timezone.now()
        r.zadd(Msg.CONTACT_HANDLING_QUEUE % self.contact_id, datetime_to_s(queue_time), dict_to_json(payload))

        # queue up our celery task
        push_task(self.org, HANDLER_QUEUE, HANDLE_EVENT_TASK, payload, priority=HIGH_PRIORITY)

    def handle(self):
        if self.direction == OUTGOING:
            raise ValueError(ugettext("Cannot process an outgoing message."))

        # process Android and test contact messages inline
        if not self.channel or self.channel.channel_type == Channel.TYPE_ANDROID or self.contact.is_test:
            Msg.process_message(self)

        # others do in celery
        else:
            on_transaction_commit(lambda: self.queue_handling())

    def build_expressions_context(self):
        date_format = get_datetime_format(self.org.get_dayfirst())[1]
        value = six.text_type(self)
        attachments = {six.text_type(a): attachment.url for a, attachment in enumerate(self.get_attachments())}

        return {
            '__default__': value,
            'value': value,
            'text': self.text,
            'attachments': attachments,
            'time': datetime_to_str(self.created_on, format=date_format, tz=self.org.timezone)
        }

    def resend(self):
        """
        Resends this message by creating a clone and triggering a send of that clone
        """
        now = timezone.now()
        (topup_id, amount) = self.org.decrement_credit()  # costs 1 credit to resend message

        # see if we should use a new channel
        channel = self.org.get_send_channel(contact_urn=self.contact_urn)

        cloned = Msg.objects.create(org=self.org,
                                    channel=channel,
                                    contact=self.contact,
                                    contact_urn=self.contact_urn,
                                    created_on=now,
                                    modified_on=now,
                                    text=self.text,
                                    response_to=self.response_to,
                                    direction=self.direction,
                                    topup_id=topup_id,
                                    status=PENDING,
                                    broadcast=self.broadcast,
                                    metadata=self.metadata)

        # mark ourselves as resent
        self.status = RESENT
        self.topup = None
        self.save()

        # update our broadcast
        if cloned.broadcast:
            cloned.broadcast.update()

        # send our message
        self.org.trigger_send([cloned])
        return cloned

    def as_task_json(self):
        """
        Used internally to serialize to JSON when queueing messages in Redis
        """
        data = dict(id=self.id, org=self.org_id, channel=self.channel_id, broadcast=self.broadcast_id,
                    text=self.text, urn_path=self.contact_urn.path, urn=six.text_type(self.contact_urn),
                    contact=self.contact_id, contact_urn=self.contact_urn_id,
                    error_count=self.error_count, next_attempt=self.next_attempt,
                    status=self.status, direction=self.direction, attachments=self.attachments,
                    external_id=self.external_id, response_to_id=self.response_to_id,
                    sent_on=self.sent_on, queued_on=self.queued_on,
                    created_on=self.created_on, modified_on=self.modified_on,
                    high_priority=self.high_priority,
                    metadata=self.metadata,
                    connection_id=self.connection_id)

        if self.contact_urn.auth:
            data.update(dict(auth=self.contact_urn.auth))

        (chatbase_api_key, chatbase_version) = self.org.get_chatbase_credentials()
        if chatbase_api_key:
            data.update(dict(chatbase_api_key=chatbase_api_key, chatbase_version=chatbase_version,
                             is_org_connected_to_chatbase=True))

        return data

    def __str__(self):
        if self.attachments:
            parts = ([self.text] if self.text else []) + [a.url for a in self.get_attachments()]
            return "\n".join(parts)
        else:
            return self.text

    @classmethod
    def create_incoming(cls, channel, urn, text, user=None, date=None, org=None, contact=None,
                        status=PENDING, attachments=None, msg_type=None, topup=None, external_id=None, connection=None):

        from temba.api.models import WebHookEvent
        if not org and channel:
            org = channel.org

        if not org:
            raise Exception(_("Can't create an incoming message without an org"))

        if not user:
            user = get_anonymous_user()

        if not date:
            date = timezone.now()  # no date?  set it to now

        contact_urn = None
        if not contact:
            contact, contact_urn = Contact.get_or_create(org, urn, channel, user=user)
        elif urn:
            contact_urn = ContactURN.get_or_create(org, contact, urn, channel=channel)

        # set the preferred channel for this contact
        contact.set_preferred_channel(channel)

        # and update this URN to make sure it is associated with this channel
        if contact_urn:
            contact_urn.update_affinity(channel)

        # we limit our text message length and remove any invalid chars
        if text:
            text = clean_string(text[:cls.MAX_TEXT_LEN])

        existing = Msg.objects.filter(text=text, sent_on=date, contact=contact, direction='I').first()
        if existing:
            return existing

        # costs 1 credit to receive a message
        topup_id = None
        if topup:  # pragma: needs cover
            topup_id = topup.pk
        elif not contact.is_test:
            (topup_id, amount) = org.decrement_credit()

        now = timezone.now()

        msg_args = dict(contact=contact,
                        contact_urn=contact_urn,
                        org=org,
                        channel=channel,
                        text=text,
                        sent_on=date,
                        created_on=now,
                        modified_on=now,
                        queued_on=now,
                        direction=INCOMING,
                        msg_type=msg_type,
                        attachments=attachments,
                        status=status,
                        external_id=external_id,
                        connection=connection)

        if topup_id is not None:
            msg_args['topup_id'] = topup_id

        msg = Msg.objects.create(**msg_args)

        # if this contact is currently stopped, unstop them
        if contact.is_stopped:
            contact.unstop(user)

        if channel:
            analytics.gauge('temba.msg_incoming_%s' % channel.channel_type.lower())

        # ivr messages are handled in handle_call
        if status == PENDING and msg_type != IVR:
            msg.handle()

            # fire an event off for this message
            WebHookEvent.trigger_sms_event(WebHookEvent.TYPE_SMS_RECEIVED, msg, date)

        return msg

    @classmethod
    def evaluate_template(cls, text, context, org=None, url_encode=False, partial_vars=False):
        """
        Given input ```text```, tries to find variables in the format @foo.bar and replace them according to
        the passed in context, contact and org. If some variables are not resolved to values, then the variable
        name will remain (ie, @foo.bar).

        Returns a tuple of the substituted text and whether there were are substitution failures.
        """
        # shortcut for cases where there is no way we would substitute anything as there are no variables
        if not text or text.find('@') < 0:
            return text, []

        # add 'step.contact' if it isn't populated for backwards compatibility
        if 'step' not in context:
            context['step'] = dict()
        if 'contact' not in context['step']:
            context['step']['contact'] = context.get('contact')

        if not org:
            dayfirst = True
            tz = timezone.get_current_timezone()
        else:
            dayfirst = org.get_dayfirst()
            tz = org.timezone

        (format_date, format_time) = get_datetime_format(dayfirst)

        now = timezone.now().astimezone(tz)

        # add date.* constants to context
        context['date'] = {
            '__default__': now.isoformat(),
            'now': now.isoformat(),
            'today': datetime_to_str(timezone.now(), format=format_date, tz=tz),
            'tomorrow': datetime_to_str(timezone.now() + timedelta(days=1), format=format_date, tz=tz),
            'yesterday': datetime_to_str(timezone.now() - timedelta(days=1), format=format_date, tz=tz)
        }

        date_style = DateStyle.DAY_FIRST if dayfirst else DateStyle.MONTH_FIRST
        context = EvaluationContext(context, tz, date_style)

        # returns tuple of output and errors
        return evaluate_template(text, context, url_encode, partial_vars)

    @classmethod
    def create_outgoing(cls, org, user, recipient, text, broadcast=None, channel=None, high_priority=False,
                        created_on=None, response_to=None, expressions_context=None, status=PENDING, insert_object=True,
                        attachments=None, topup_id=None, msg_type=INBOX, connection=None, quick_replies=None):

        if not org or not user:  # pragma: no cover
            raise ValueError("Trying to create outgoing message with no org or user")

        # for IVR messages we need a channel that can call
        if msg_type == IVR:
            role = Channel.ROLE_CALL
        elif msg_type == USSD:
            role = Channel.ROLE_USSD
        else:
            role = Channel.ROLE_SEND

        if status != SENT:
            # if message will be sent, resolve the recipient to a contact and URN
            contact, contact_urn = cls.resolve_recipient(org, user, recipient, channel, role=role)

            if not contact_urn:
                raise UnreachableException("No suitable URN found for contact")

            if not channel:
                if msg_type == IVR:
                    channel = org.get_call_channel()
                elif msg_type == USSD:
                    channel = org.get_ussd_channel(contact_urn=contact_urn)
                else:
                    channel = org.get_send_channel(contact_urn=contact_urn)

                if not channel and not contact.is_test:  # pragma: needs cover
                    raise UnreachableException("No suitable channel available for this org")
        else:
            # if message has already been sent, recipient must be a tuple of contact and URN
            contact, contact_urn = recipient

        # no creation date? set it to now
        if not created_on:
            created_on = timezone.now()

        # evaluate expressions in the text and attachments if a context was provided
        if expressions_context is not None:
            # make sure 'channel' is populated if we have a channel
            if channel and 'channel' not in expressions_context:
                expressions_context['channel'] = channel.build_expressions_context()

            (text, errors) = Msg.evaluate_template(text, expressions_context, org=org)
            if text:
                text = text[:Msg.MAX_TEXT_LEN]

            evaluated_attachments = []
            if attachments:
                for attachment in attachments:
                    (attachment, errors) = Msg.evaluate_template(attachment, expressions_context, org=org)
                    evaluated_attachments.append(attachment)
        else:
            text = text[:Msg.MAX_TEXT_LEN]
            evaluated_attachments = attachments

        # prefer none to empty lists in the database
        if evaluated_attachments is not None and len(evaluated_attachments) == 0:
            evaluated_attachments = None

        # if we are doing a single message, check whether this might be a loop of some kind
        if insert_object:
            # prevent the loop of message while the sending phone is the channel
            # get all messages with same text going to same number
            same_msgs = Msg.objects.filter(contact_urn=contact_urn,
                                           contact__is_test=False,
                                           channel=channel,
                                           attachments=evaluated_attachments,
                                           text=text,
                                           direction=OUTGOING,
                                           created_on__gte=created_on - timedelta(minutes=10))

            # we aren't considered with robo detection on calls
            same_msg_count = same_msgs.exclude(msg_type=IVR).count()

            if same_msg_count >= 10:
                analytics.gauge('temba.msg_loop_caught')
                return None

            # be more aggressive about short codes for duplicate messages
            # we don't want machines talking to each other
            tel = contact.raw_tel()
            if tel and len(tel) < 6:
                same_msg_count = Msg.objects.filter(contact_urn=contact_urn,
                                                    contact__is_test=False,
                                                    channel=channel,
                                                    text=text,
                                                    direction=OUTGOING,
                                                    created_on__gte=created_on - timedelta(hours=24)).count()
                if same_msg_count >= 10:  # pragma: needs cover
                    analytics.gauge('temba.msg_shortcode_loop_caught')
                    return None

        # costs 1 credit to send a message
        if not topup_id and not contact.is_test:
            (topup_id, _) = org.decrement_credit()

        if response_to:
            msg_type = response_to.msg_type

        text = text.strip()

        # track this if we have a channel
        if channel:
            analytics.gauge('temba.msg_outgoing_%s' % channel.channel_type.lower())

        metadata = {}  # init metadata to the same as the default value of the Msg.metadata field
        if quick_replies:
            for counter, reply in enumerate(quick_replies):
                (value, errors) = Msg.evaluate_template(text=reply, context=expressions_context, org=org)
                if value:
                    quick_replies[counter] = value
            metadata = dict(quick_replies=quick_replies)

        msg_args = dict(contact=contact,
                        contact_urn=contact_urn,
                        org=org,
                        channel=channel,
                        text=text,
                        created_on=created_on,
                        modified_on=created_on,
                        direction=OUTGOING,
                        status=status,
                        broadcast=broadcast,
                        response_to=response_to,
                        msg_type=msg_type,
                        high_priority=high_priority,
                        attachments=evaluated_attachments,
                        metadata=metadata,
                        connection=connection)

        if topup_id is not None:
            msg_args['topup_id'] = topup_id

        return Msg.objects.create(**msg_args) if insert_object else Msg(**msg_args)

    @staticmethod
    def resolve_recipient(org, user, recipient, channel, role=Channel.ROLE_SEND):
        """
        Recipient can be a contact, a URN object, or a URN tuple, e.g. ('tel', '123'). Here we resolve the contact and
        contact URN to use for an outgoing message.
        """
        contact = None
        contact_urn = None

        resolved_schemes = set(channel.schemes) if channel else org.get_schemes(role)

        if isinstance(recipient, Contact):
            if recipient.is_test:
                contact = recipient
                contact_urn = contact.urns.all().first()
            else:
                contact = recipient
                contact_urn = contact.get_urn(schemes=resolved_schemes)  # use highest priority URN we can send to
        elif isinstance(recipient, ContactURN):
            if recipient.scheme in resolved_schemes:
                contact = recipient.contact
                contact_urn = recipient
        elif isinstance(recipient, six.string_types):
            scheme, path, display = URN.to_parts(recipient)
            if scheme in resolved_schemes:
                contact, contact_urn = Contact.get_or_create(org, recipient, user=user)
        else:  # pragma: no cover
            raise ValueError("Message recipient must be a Contact, ContactURN or URN string")

        return contact, contact_urn

    def status_fail(self):
        """
        Update the message status to FAILED
        """
        self.status = FAILED
        self.save(update_fields=('status', 'modified_on'))

        Channel.track_status(self.channel, "Failed")

    def status_sent(self):
        """
        Update the message status to SENT
        """
        now = timezone.now()
        self.status = SENT
        self.sent_on = now
        self.save(update_fields=('status', 'sent_on', 'modified_on'))

        Channel.track_status(self.channel, "Sent")

    def status_delivered(self):
        """
        Update the message status to DELIVERED
        """
        self.status = DELIVERED
        if not self.sent_on:
            self.sent_on = timezone.now()
        self.save(update_fields=('status', 'modified_on', 'sent_on'))

        Channel.track_status(self.channel, "Delivered")

    def archive(self):
        """
        Archives this message
        """
        if self.direction != INCOMING or self.contact.is_test:
            raise ValueError("Can only archive incoming non-test messages")

        self.visibility = Msg.VISIBILITY_ARCHIVED
        self.save(update_fields=('visibility', 'modified_on'))

    @classmethod
    def archive_all_for_contacts(cls, contacts):
        """
        Archives all incoming messages for the given contacts
        """
        msgs = Msg.objects.filter(direction=INCOMING, visibility=Msg.VISIBILITY_VISIBLE, contact__in=contacts)
        msg_ids = list(msgs.values_list('pk', flat=True))

        # update modified on in small batches to avoid long table lock, and having too many non-unique values for
        # modified_on which is the primary ordering for the API
        for batch in chunk_list(msg_ids, 100):
            Msg.objects.filter(pk__in=batch).update(visibility=Msg.VISIBILITY_ARCHIVED, modified_on=timezone.now())

    def restore(self):
        """
        Restores (i.e. un-archives) this message
        """
        if self.direction != INCOMING or self.contact.is_test:  # pragma: needs cover
            raise ValueError("Can only restore incoming non-test messages")

        self.visibility = Msg.VISIBILITY_VISIBLE
        self.save(update_fields=('visibility', 'modified_on'))

    def release(self):
        """
        Releases (i.e. deletes) this message
        """
        self.visibility = Msg.VISIBILITY_DELETED
        self.text = ""
        self.save(update_fields=('visibility', 'text', 'modified_on'))

        # remove labels
        self.labels.clear()

    @classmethod
    def apply_action_label(cls, user, msgs, label, add):
        return label.toggle_label(msgs, add)

    @classmethod
    def apply_action_archive(cls, user, msgs):
        changed = []

        for msg in msgs:
            msg.archive()
            changed.append(msg.pk)

        return changed

    @classmethod
    def apply_action_restore(cls, user, msgs):
        changed = []

        for msg in msgs:
            msg.restore()
            changed.append(msg.pk)

        return changed

    @classmethod
    def apply_action_delete(cls, user, msgs):
        changed = []

        for msg in msgs:
            msg.release()
            changed.append(msg.pk)

        return changed

    @classmethod
    def apply_action_resend(cls, user, msgs):
        changed = []

        for msg in msgs:
            msg.resend()
            changed.append(msg.pk)
        return changed


STOP_WORDS = 'a,able,about,across,after,all,almost,also,am,among,an,and,any,are,as,at,be,because,been,but,by,can,' \
             'cannot,could,dear,did,do,does,either,else,ever,every,for,from,get,got,had,has,have,he,her,hers,him,his,' \
             'how,however,i,if,in,into,is,it,its,just,least,let,like,likely,may,me,might,most,must,my,neither,no,nor,' \
             'not,of,off,often,on,only,or,other,our,own,rather,said,say,says,she,should,since,so,some,than,that,the,' \
             'their,them,then,there,these,they,this,tis,to,too,twas,us,wants,was,we,were,what,when,where,which,while,' \
             'who,whom,why,will,with,would,yet,you,your'.split(',')


class SystemLabel(object):
    TYPE_INBOX = 'I'
    TYPE_FLOWS = 'W'
    TYPE_ARCHIVED = 'A'
    TYPE_OUTBOX = 'O'
    TYPE_SENT = 'S'
    TYPE_FAILED = 'X'
    TYPE_SCHEDULED = 'E'
    TYPE_CALLS = 'C'

    TYPE_CHOICES = ((TYPE_INBOX, "Inbox"),
                    (TYPE_FLOWS, "Flows"),
                    (TYPE_ARCHIVED, "Archived"),
                    (TYPE_OUTBOX, "Outbox"),
                    (TYPE_SENT, "Sent"),
                    (TYPE_FAILED, "Failed"),
                    (TYPE_SCHEDULED, "Scheduled"),
                    (TYPE_CALLS, "Calls"))

    @classmethod
    def get_counts(cls, org):
        return SystemLabelCount.get_totals(org)

    @classmethod
    def get_queryset(cls, org, label_type, exclude_test_contacts=True):
        """
        Gets the queryset for the given system label. Any change here needs to be reflected in a change to the db
        trigger used to maintain the label counts.
        """
        # TODO: (Indexing) Sent and Failed require full message history
        if label_type == cls.TYPE_INBOX:
            qs = Msg.objects.filter(direction=INCOMING, visibility=Msg.VISIBILITY_VISIBLE, msg_type=INBOX)
        elif label_type == cls.TYPE_FLOWS:
            qs = Msg.objects.filter(direction=INCOMING, visibility=Msg.VISIBILITY_VISIBLE, msg_type=FLOW)
        elif label_type == cls.TYPE_ARCHIVED:
            qs = Msg.objects.filter(direction=INCOMING, visibility=Msg.VISIBILITY_ARCHIVED)
        elif label_type == cls.TYPE_OUTBOX:
            qs = Msg.objects.filter(direction=OUTGOING, visibility=Msg.VISIBILITY_VISIBLE, status__in=(PENDING, QUEUED))
        elif label_type == cls.TYPE_SENT:
            qs = Msg.objects.filter(direction=OUTGOING, visibility=Msg.VISIBILITY_VISIBLE,
                                    status__in=(WIRED, SENT, DELIVERED))
        elif label_type == cls.TYPE_FAILED:
            qs = Msg.objects.filter(direction=OUTGOING, visibility=Msg.VISIBILITY_VISIBLE, status=FAILED)
        elif label_type == cls.TYPE_SCHEDULED:
            qs = Broadcast.objects.exclude(schedule=None)
        elif label_type == cls.TYPE_CALLS:
            qs = ChannelEvent.objects.filter(event_type__in=ChannelEvent.CALL_TYPES)
        else:  # pragma: needs cover
            raise ValueError("Invalid label type: %s" % label_type)

        qs = qs.filter(org=org)

        if exclude_test_contacts:
            if label_type == cls.TYPE_SCHEDULED:
                qs = qs.exclude(contacts__is_test=True)
            else:
                qs = qs.exclude(contact__is_test=True)

        return qs


class SystemLabelCount(SquashableModel):
    """
    Counts of messages/broadcasts/calls maintained by database level triggers
    """
    SQUASH_OVER = ('org_id', 'label_type')

    org = models.ForeignKey(Org, related_name='system_labels')

    label_type = models.CharField(max_length=1, choices=SystemLabel.TYPE_CHOICES)

    count = models.IntegerField(default=0, help_text=_("Number of items with this system label"))

    @classmethod
    def get_squash_query(cls, distinct_set):
        sql = """
        WITH deleted as (
            DELETE FROM %(table)s WHERE "org_id" = %%s AND "label_type" = %%s RETURNING "count"
        )
        INSERT INTO %(table)s("org_id", "label_type", "count", "is_squashed")
        VALUES (%%s, %%s, GREATEST(0, (SELECT SUM("count") FROM deleted)), TRUE);
        """ % {'table': cls._meta.db_table}

        return sql, (distinct_set.org_id, distinct_set.label_type) * 2

    @classmethod
    def get_totals(cls, org):
        """
        Gets all system label counts by type for the given org
        """
        counts = cls.objects.filter(org=org)
        counts = counts.values_list('label_type').annotate(count_sum=Sum('count'))
        counts_by_type = {c[0]: c[1] for c in counts}

        # for convenience, include all label types
        return {l: counts_by_type.get(l, 0) for l, n in SystemLabel.TYPE_CHOICES}

    class Meta:
        index_together = ('org', 'label_type')


class UserFolderManager(models.Manager):
    def get_queryset(self):
        return super(UserFolderManager, self).get_queryset().filter(label_type=Label.TYPE_FOLDER)


class UserLabelManager(models.Manager):
    def get_queryset(self):
        return super(UserLabelManager, self).get_queryset().filter(label_type=Label.TYPE_LABEL)


@six.python_2_unicode_compatible
class Label(TembaModel):
    """
    Labels represent both user defined labels and folders of labels. User defined labels that can be applied to messages
    much the same way labels or tags apply to messages in web-based email services.
    """
    MAX_NAME_LEN = 64
    MAX_ORG_LABELS = 250
    MAX_ORG_FOLDERS = 250

    TYPE_FOLDER = 'F'
    TYPE_LABEL = 'L'

    TYPE_CHOICES = ((TYPE_FOLDER, "Folder of labels"),
                    (TYPE_LABEL, "Regular label"))

    org = models.ForeignKey(Org)

    name = models.CharField(max_length=MAX_NAME_LEN, verbose_name=_("Name"), help_text=_("The name of this label"))

    folder = models.ForeignKey('Label', verbose_name=_("Folder"), null=True, related_name="children")

    label_type = models.CharField(max_length=1, choices=TYPE_CHOICES, default=TYPE_LABEL, help_text=_("Label type"))

    # define some custom managers to do the filtering of label types for us
    all_objects = models.Manager()
    folder_objects = UserFolderManager()
    label_objects = UserLabelManager()

    @classmethod
    def get_or_create(cls, org, user, name, folder=None):
        name = name.strip()

        if not cls.is_valid_name(name):
            raise ValueError("Invalid label name: %s" % name)

        if folder and not folder.is_folder():  # pragma: needs cover
            raise ValueError("%s is not a label folder" % six.text_type(folder))

        label = cls.label_objects.filter(org=org, name__iexact=name).first()
        if label:
            return label

        return cls.label_objects.create(org=org, name=name, folder=folder, created_by=user, modified_by=user)

    @classmethod
    def get_or_create_folder(cls, org, user, name):
        name = name.strip()

        if not cls.is_valid_name(name):  # pragma: needs cover
            raise ValueError("Invalid folder name: %s" % name)

        folder = cls.folder_objects.filter(org=org, name__iexact=name).first()
        if folder:  # pragma: needs cover
            return folder

        return cls.folder_objects.create(org=org, name=name, label_type=Label.TYPE_FOLDER,
                                         created_by=user, modified_by=user)

    @classmethod
    def get_hierarchy(cls, org):
        """
        Gets labels and folders organized into their hierarchy and with their message counts
        """
        labels_and_folders = list(Label.all_objects.filter(org=org).order_by(Upper('name')))
        label_counts = LabelCount.get_totals([l for l in labels_and_folders if not l.is_folder()])

        folder_nodes = {}
        all_nodes = []
        for obj in labels_and_folders:
            node = {'obj': obj, 'count': label_counts.get(obj), 'children': []}
            all_nodes.append(node)

            if obj.is_folder():
                folder_nodes[obj.id] = node

        top_nodes = []
        for node in all_nodes:
            if node['obj'].folder_id is None:
                top_nodes.append(node)
            else:
                folder_nodes[node['obj'].folder_id]['children'].append(node)

        return top_nodes

    @classmethod
    def is_valid_name(cls, name):
        # don't allow empty strings, blanks, initial or trailing whitespace
        if not name or name.strip() != name:
            return False

        if len(name) > cls.MAX_NAME_LEN:
            return False

        # first character must be a word char
        return regex.match('\w', name[0], flags=regex.UNICODE)

    def filter_messages(self, queryset):
        if self.is_folder():
            return queryset.filter(labels__in=self.children.all()).distinct()

        return queryset.filter(labels=self)

    def get_messages(self):
        # TODO: consider purpose built indexes
        return self.filter_messages(Msg.objects.all())

    def get_visible_count(self):
        """
        Returns the count of visible, non-test message tagged with this label
        """
        if self.is_folder():
            raise ValueError("Message counts are not tracked for user folders")

        return LabelCount.get_totals([self])[self]

    def toggle_label(self, msgs, add):
        """
        Adds or removes this label from the given messages
        """
        if self.is_folder():  # pragma: needs cover
            raise ValueError("Can only assign messages to user labels")

        changed = set()

        for msg in msgs:
            if msg.direction != INCOMING:
                raise ValueError("Can only apply labels to incoming messages")

            if msg.contact.is_test:
                raise ValueError("Cannot apply labels to test messages")

            # if we are adding the label and this message doesnt have it, add it
            if add:
                if not msg.labels.filter(pk=self.pk):
                    msg.labels.add(self)
                    changed.add(msg.pk)

            # otherwise, remove it if not already present
            else:
                if msg.labels.filter(pk=self.pk):
                    msg.labels.remove(self)
                    changed.add(msg.pk)

        # update modified on all our changed msgs
        Msg.objects.filter(id__in=changed).update(modified_on=timezone.now())

        return changed

    def is_folder(self):
        return self.label_type == Label.TYPE_FOLDER

    def release(self):
        self.delete()

    def __str__(self):
        if self.folder:
            return "%s > %s" % (six.text_type(self.folder), self.name)
        return self.name

    class Meta:
        unique_together = ('org', 'name')


class LabelCount(SquashableModel):
    """
    Counts of user labels maintained by database level triggers
    """
    SQUASH_OVER = ('label_id',)

    label = models.ForeignKey(Label, related_name='counts')

    count = models.IntegerField(default=0, help_text=_("Number of items with this system label"))

    @classmethod
    def get_squash_query(cls, distinct_set):
        sql = """
            WITH deleted as (
                DELETE FROM %(table)s WHERE "label_id" = %%s RETURNING "count"
            )
            INSERT INTO %(table)s("label_id", "count", "is_squashed")
            VALUES (%%s, GREATEST(0, (SELECT SUM("count") FROM deleted)), TRUE);
            """ % {'table': cls._meta.db_table}

        return sql, (distinct_set.label_id, distinct_set.label_id)

    @classmethod
    def get_totals(cls, labels):
        """
        Gets total counts for all the given labels
        """
        counts = cls.objects.filter(label__in=labels).values_list('label_id').annotate(count_sum=Sum('count'))
        counts_by_label_id = {c[0]: c[1] for c in counts}
        return {l: counts_by_label_id.get(l.id, 0) for l in labels}


class MsgIterator(six.Iterator):
    """
    Queryset wrapper to chunk queries and reduce in-memory footprint
    """
    def __init__(self, ids, order_by=None, select_related=None, prefetch_related=None, max_obj_num=1000):
        self._ids = ids
        self._order_by = order_by
        self._select_related = select_related
        self._prefetch_related = prefetch_related
        self._generator = self._setup()
        self.max_obj_num = max_obj_num

    def _setup(self):
        for i in six.moves.xrange(0, len(self._ids), self.max_obj_num):
            chunk_queryset = Msg.objects.filter(id__in=self._ids[i:i + self.max_obj_num])

            if self._order_by:
                chunk_queryset = chunk_queryset.order_by(*self._order_by)

            if self._select_related:
                chunk_queryset = chunk_queryset.select_related(*self._select_related)

            if self._prefetch_related:
                chunk_queryset = chunk_queryset.prefetch_related(*self._prefetch_related)

            for obj in chunk_queryset:
                yield obj

    def __iter__(self):
        return self

    def __next__(self):
        return next(self._generator)


class ExportMessagesTask(BaseExportTask):
    """
    Wrapper for handling exports of raw messages. This will export all selected messages in
    an Excel spreadsheet, adding sheets as necessary to fall within the guidelines of Excel 97
    (the library we depend on requires this) which has column and row size limits.

    When the export is done, we store the file on the server and send an e-mail notice with a
    link to download the results.
    """
    analytics_key = 'msg_export'
    email_subject = "Your messages export is ready"
    email_template = 'msgs/email/msg_export_download'

    groups = models.ManyToManyField(ContactGroup)

    label = models.ForeignKey(Label, null=True)

    system_label = models.CharField(null=True, max_length=1)

    start_date = models.DateField(null=True, blank=True, help_text=_("The date for the oldest message to export"))

    end_date = models.DateField(null=True, blank=True, help_text=_("The date for the newest message to export"))

    @classmethod
    def create(cls, org, user, system_label=None, label=None, groups=(), start_date=None, end_date=None):
        if label and system_label:  # pragma: no cover
            raise ValueError("Can't specify both label and system label")

        export = cls.objects.create(org=org, system_label=system_label, label=label,
                                    start_date=start_date, end_date=end_date,
                                    created_by=user, modified_by=user)
        export.groups.add(*groups)
        return export

    def write_export(self):
        from openpyxl import Workbook

        book = Workbook(write_only=True)

        fields = ['Date', 'Contact', 'Contact Type', 'Name', 'Contact UUID', 'Direction', 'Text', 'Labels', "Status"]
        fields_col_width = [self.WIDTH_MEDIUM,  # Date
                            self.WIDTH_MEDIUM,  # Contact
                            self.WIDTH_SMALL,   # Contact Type
                            self.WIDTH_MEDIUM,  # Name
                            self.WIDTH_MEDIUM,  # Contact UUID
                            self.WIDTH_SMALL,   # Direction
                            self.WIDTH_LARGE,   # Text
                            self.WIDTH_MEDIUM,  # Labels
                            self.WIDTH_SMALL]   # Status

        if self.system_label:
            messages = SystemLabel.get_queryset(self.org, self.system_label)
        elif self.label:
            messages = self.label.get_messages()
        else:
            messages = Msg.get_messages(self.org)

        tz = self.org.timezone

        if self.start_date:
            start_date = tz.localize(datetime.combine(self.start_date, datetime.min.time()))
            messages = messages.filter(created_on__gte=start_date)

        if self.end_date:
            end_date = tz.localize(datetime.combine(self.end_date, datetime.max.time()))
            messages = messages.filter(created_on__lte=end_date)

        if self.groups.all():
            messages = messages.filter(contact__all_groups__in=self.groups.all())

        all_message_ids = list(messages.order_by('-created_on').values_list('id', flat=True))

        messages_sheet_number = 1

        current_messages_sheet = book.create_sheet(six.text_type(_("Messages %d" % messages_sheet_number)))

        self.set_sheet_column_widths(current_messages_sheet, fields_col_width)
        self.append_row(current_messages_sheet, fields)

        row = 2
        processed = 0
        start = time.time()

        prefetch = Prefetch('labels', queryset=Label.label_objects.order_by('name'))
        for msg in MsgIterator(all_message_ids,
                               order_by=[''
                                         '-created_on'],
                               select_related=['contact', 'contact_urn'],
                               prefetch_related=[prefetch]):

            if row >= self.MAX_EXCEL_ROWS:  # pragma: needs cover
                messages_sheet_number += 1
                current_messages_sheet = book.create_sheet(six.text_type(_("Messages %d" % messages_sheet_number)))

                self.append_row(current_messages_sheet, fields)
                self.set_sheet_column_widths(current_messages_sheet, fields_col_width)
                row = 2

            # only show URN path if org isn't anon and there is a URN
            if self.org.is_anon:  # pragma: needs cover
                urn_path = msg.contact.anon_identifier
            elif msg.contact_urn:
                urn_path = msg.contact_urn.get_display(org=self.org, formatted=False)
            else:
                urn_path = ''

            urn_scheme = msg.contact_urn.scheme if msg.contact_urn else ''

            self.append_row(current_messages_sheet, [
                msg.created_on,
                urn_path,
                urn_scheme,
                msg.contact.name,
                msg.contact.uuid,
                msg.get_direction_display(),
                msg.text,
                ", ".join(msg_label.name for msg_label in msg.labels.all()),
                msg.get_status_display()
            ])

            row += 1
            processed += 1

            if processed % 10000 == 0:  # pragma: needs cover
                print("Export of %d msgs for %s - %d%% complete in %0.2fs" %
                      (len(all_message_ids), self.org.name, processed * 100 // len(all_message_ids), time.time() - start))

        temp = NamedTemporaryFile(delete=True)
        book.save(temp)
        temp.flush()
        return temp, 'xlsx'


@register_asset_store
class MessageExportAssetStore(BaseExportAssetStore):
    model = ExportMessagesTask
    key = 'message_export'
    directory = 'message_exports'
    permission = 'msgs.msg_export'
    extensions = ('xlsx',)
