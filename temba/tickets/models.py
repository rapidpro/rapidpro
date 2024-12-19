import itertools
import logging
from abc import ABCMeta
from datetime import date

import openpyxl

from django.conf import settings
from django.db import models
from django.db.models import F, Q, Sum, Value
from django.db.models.functions import Cast, Lower
from django.utils import timezone
from django.utils.translation import gettext_lazy as _

from temba import mailroom
from temba.contacts.models import Contact
from temba.orgs.models import DependencyMixin, Export, ExportType, Org, OrgMembership, User
from temba.utils.dates import date_range
from temba.utils.db.functions import SplitPart
from temba.utils.export import MultiSheetExporter
from temba.utils.models import TembaModel
from temba.utils.models.counts import DailyCountModel, DailyTimingModel
from temba.utils.uuid import is_uuid, uuid4

logger = logging.getLogger(__name__)


class Shortcut(TembaModel):
    """
    A canned response available from the ticketing interface.
    """

    org = models.ForeignKey(Org, on_delete=models.PROTECT, related_name="shortcuts")
    text = models.TextField(max_length=10_000)

    @classmethod
    def create(cls, org, user, name: str, text: str):
        assert cls.is_valid_name(name), f"'{name}' is not a valid shortcut name"
        assert not org.shortcuts.filter(name__iexact=name).exists(), f"shortcut with name '{name}' already exists"

        return org.shortcuts.create(name=name, text=text, created_by=user, modified_by=user)

    def release(self, user):
        self.is_active = False
        self.name = self._deleted_name()
        self.modified_by = user
        self.save(update_fields=("name", "is_active", "modified_by", "modified_on"))

    class Meta:
        constraints = [models.UniqueConstraint("org", Lower("name"), name="unique_shortcut_names")]


class Topic(TembaModel, DependencyMixin):
    """
    The topic of a ticket which controls who can access that ticket.
    """

    org = models.ForeignKey(Org, on_delete=models.PROTECT, related_name="topics")
    is_default = models.BooleanField(default=False)

    org_limit_key = Org.LIMIT_TOPICS

    @classmethod
    def create_system(cls, org):
        assert not org.topics.filter(is_default=True).exists(), "org already has default topic"

        org.topics.create(
            name="General",
            is_default=True,
            is_system=True,
            created_by=org.created_by,
            modified_by=org.modified_by,
        )

    @classmethod
    def create(cls, org, user, name: str):
        assert cls.is_valid_name(name), f"'{name}' is not a valid topic name"
        assert not org.topics.filter(name__iexact=name).exists(), f"topic with name '{name}' already exists"

        return org.topics.create(name=name, created_by=user, modified_by=user)

    @classmethod
    def create_from_import_def(cls, org, user, definition: dict):
        return cls.create(org, user, definition["name"])

    @classmethod
    def get_accessible(cls, org, user):
        """
        Gets the topics accessible to the given user in the given org.
        """

        if not user.is_staff:
            membership = org.get_membership(user)
            if membership.team and not membership.team.all_topics:
                return membership.team.topics.all()

        return org.topics.filter(is_active=True)

    def release(self, user):
        assert not (self.is_system and self.org.is_active), "can't release system topics"
        assert not self.tickets.exists(), "can't release topic with tickets"

        super().release(user)

        for team in self.teams.all():
            team.topics.remove(self)

        # delete ticket counts for this topic
        self.org.counts.prefix(
            [f"tickets:{Ticket.STATUS_OPEN}:{self.id}:", f"tickets:{Ticket.STATUS_CLOSED}:{self.id}:"]
        ).delete()

        self.is_active = False
        self.name = self._deleted_name()
        self.modified_by = user
        self.save(update_fields=("name", "is_active", "modified_by", "modified_on"))

    class Meta:
        constraints = [models.UniqueConstraint("org", Lower("name"), name="unique_topic_names")]


class Team(TembaModel):
    """
    Agent users are assigned to a team which controls which topics they can access.
    """

    org = models.ForeignKey(Org, on_delete=models.PROTECT, related_name="teams")
    topics = models.ManyToManyField(Topic, related_name="teams")
    all_topics = models.BooleanField(default=False)
    is_default = models.BooleanField(default=False)

    org_limit_key = Org.LIMIT_TEAMS
    max_topics = 10

    @classmethod
    def create_system(cls, org):
        assert not org.teams.filter(is_default=True).exists(), "org already has default team"

        org.teams.create(
            name="All Topics",
            is_default=True,
            is_system=True,
            all_topics=True,
            created_by=org.created_by,
            modified_by=org.modified_by,
        )

    @classmethod
    def create(cls, org, user, name: str, *, topics=(), all_topics: bool = False):
        assert cls.is_valid_name(name), f"'{name}' is not a valid team name"
        assert not org.teams.filter(name__iexact=name, is_active=True).exists()
        assert not (topics and all_topics), "can't specify topics and all_topics"

        team = org.teams.create(name=name, all_topics=all_topics, created_by=user, modified_by=user)
        team.topics.add(*topics)
        return team

    def get_users(self):
        return self.org.users.filter(orgmembership__team=self)

    def release(self, user):
        assert not (self.is_system and self.org.is_active), "can't release system teams"

        # re-assign agents in this team to the default team
        OrgMembership.objects.filter(org=self.org, team=self).update(team=self.org.default_ticket_team)

        self.name = self._deleted_name()
        self.is_active = False
        self.modified_by = user
        self.save(update_fields=("name", "is_active", "modified_by", "modified_on"))

    class Meta:
        constraints = [models.UniqueConstraint("org", Lower("name"), name="unique_team_names")]


class Ticket(models.Model):
    """
    A ticket represents a period of human interaction with a contact.
    """

    STATUS_OPEN = "O"
    STATUS_CLOSED = "C"
    STATUS_CHOICES = ((STATUS_OPEN, _("Open")), (STATUS_CLOSED, _("Closed")))

    # permission that users need to have a ticket assigned to them
    ASSIGNEE_PERMISSION = "tickets.ticket_assignee"

    MAX_NOTE_LENGTH = 10_000

    uuid = models.UUIDField(unique=True, default=uuid4)
    org = models.ForeignKey(Org, on_delete=models.PROTECT, related_name="tickets", db_index=False)  # indexed below
    contact = models.ForeignKey(Contact, on_delete=models.PROTECT, related_name="tickets", db_index=False)
    topic = models.ForeignKey(Topic, on_delete=models.PROTECT, related_name="tickets")

    # the status of this ticket and who it's currently assigned to
    status = models.CharField(max_length=1, choices=STATUS_CHOICES)
    assignee = models.ForeignKey(User, on_delete=models.PROTECT, null=True, related_name="assigned_tickets")

    opened_on = models.DateTimeField(default=timezone.now)
    opened_in = models.ForeignKey("flows.Flow", null=True, on_delete=models.PROTECT, related_name="opened_tickets")
    opened_by = models.ForeignKey(User, null=True, on_delete=models.PROTECT, related_name="opened_tickets")

    # when this ticket was first replied to, closed, modified
    replied_on = models.DateTimeField(null=True)
    closed_on = models.DateTimeField(null=True)
    modified_on = models.DateTimeField(default=timezone.now)

    # when this ticket last had activity which includes messages being sent and received, and is used for ordering
    last_activity_on = models.DateTimeField(default=timezone.now)

    def assign(self, user: User, *, assignee: User):
        self.bulk_assign(self.org, user, [self], assignee=assignee)

    def add_note(self, user: User, *, note: str):
        self.bulk_add_note(self.org, user, [self], note=note)

    @classmethod
    def bulk_assign(cls, org, user: User, tickets: list, assignee: User):
        return mailroom.get_client().ticket_assign(org, user, tickets, assignee)

    @classmethod
    def bulk_add_note(cls, org, user: User, tickets: list, note: str):
        return mailroom.get_client().ticket_add_note(org, user, tickets, note)

    @classmethod
    def bulk_change_topic(cls, org, user: User, tickets: list, topic: Topic):
        return mailroom.get_client().ticket_change_topic(org, user, tickets, topic)

    @classmethod
    def bulk_close(cls, org, user, tickets, *, force: bool = False):
        return mailroom.get_client().ticket_close(org, user, tickets, force=force)

    @classmethod
    def bulk_reopen(cls, org, user, tickets):
        return mailroom.get_client().ticket_reopen(org, user, tickets)

    @classmethod
    def get_allowed_assignees(cls, org):
        return org.get_users(with_perm=cls.ASSIGNEE_PERMISSION)

    @classmethod
    def get_assignee_count(cls, org, user, topics, status: str) -> int:
        """
        Gets the count of tickets assigned to the given user across the given topics and status.
        """
        return org.counts.filter(scope__in=[f"tickets:{status}:{t.id}:{user.id if user else 0}" for t in topics]).sum()

    @classmethod
    def get_status_count(cls, org, topics, status: str) -> int:
        """
        Gets the count across the given topics and status.
        """
        return org.counts.prefix([f"tickets:{status}:{t.id}:" for t in topics]).sum()

    @classmethod
    def get_topic_counts(cls, org, topics, status: str) -> dict[Topic, int]:
        """
        Gets the count for each topic and the given status.
        """

        # count scopes are stored as 'tickets:<status>:<topic-id>:<assignee-id>' so get all counts with the prefix
        # 'tickets:<status>:' and group by topic-id extracted as the 3rd split part.
        counts = (
            org.counts.prefix(f"tickets:{status}:")
            .annotate(topic_id=Cast(SplitPart(F("scope"), Value(":"), Value(3)), output_field=models.IntegerField()))
            .values_list("topic_id")
            .annotate(count_sum=Sum("count"))
        )
        by_topic_id = {c[0]: c[1] for c in counts}
        return {t: by_topic_id.get(t.id, 0) for t in topics}

    def delete(self):
        self.events.all().delete()

        super().delete()

    def __str__(self):
        return f"Ticket[uuid={self.uuid}, topic={self.topic.name}]"

    class Meta:
        indexes = [
            # used by the All folder
            models.Index(name="tickets_org_status", fields=["org", "status", "-last_activity_on", "-id"]),
            # used by the Unassigned and Mine folders
            models.Index(
                name="tickets_org_assignee_status",
                fields=["org", "assignee", "status", "-last_activity_on", "-id"],
            ),
            # used by message handling to find open tickets for contact
            models.Index(name="tickets_contact_open", fields=["contact", "-opened_on"], condition=Q(status="O")),
            # used by API tickets endpoint hence the ordering, and general fetching by org or contact
            models.Index(name="tickets_api_by_org", fields=["org", "-modified_on", "-id"]),
            models.Index(name="tickets_api_by_contact", fields=["contact", "-modified_on", "-id"]),
        ]


class TicketEvent(models.Model):
    """
    Models the history of a ticket.
    """

    TYPE_OPENED = "O"
    TYPE_ASSIGNED = "A"
    TYPE_NOTE_ADDED = "N"
    TYPE_TOPIC_CHANGED = "T"
    TYPE_CLOSED = "C"
    TYPE_REOPENED = "R"
    TYPE_CHOICES = (
        (TYPE_OPENED, "Opened"),
        (TYPE_ASSIGNED, "Assigned"),
        (TYPE_NOTE_ADDED, "Note Added"),
        (TYPE_TOPIC_CHANGED, "Topic Changed"),
        (TYPE_CLOSED, "Closed"),
        (TYPE_REOPENED, "Reopened"),
    )

    org = models.ForeignKey(Org, on_delete=models.PROTECT, related_name="ticket_events")
    ticket = models.ForeignKey(Ticket, on_delete=models.PROTECT, related_name="events")
    contact = models.ForeignKey(Contact, on_delete=models.PROTECT, related_name="ticket_events")
    event_type = models.CharField(max_length=1, choices=TYPE_CHOICES)
    note = models.TextField(null=True, max_length=Ticket.MAX_NOTE_LENGTH)
    topic = models.ForeignKey(Topic, on_delete=models.PROTECT, null=True, related_name="ticket_events")
    assignee = models.ForeignKey(User, on_delete=models.PROTECT, null=True, related_name="ticket_assignee_events")

    created_by = models.ForeignKey(
        settings.AUTH_USER_MODEL, on_delete=models.PROTECT, null=True, related_name="ticket_events"
    )
    created_on = models.DateTimeField(default=timezone.now)

    class Meta:
        indexes = [
            # used for contact history
            models.Index(name="ticketevents_contact_created", fields=["contact", "created_on"])
        ]


class TicketFolder(metaclass=ABCMeta):
    slug = None
    name = None
    icon = None
    verbose_name = None

    def get_icon(self, count) -> str:
        return self.icon

    def get_queryset(self, org, user, *, ordered: bool):
        qs = org.tickets.all()

        if not user.is_staff:
            membership = org.get_membership(user)
            if membership.team and not membership.team.all_topics:
                qs = qs.filter(topic__in=list(membership.team.topics.all()))

        if ordered:
            qs = qs.order_by("-last_activity_on", "-id")

        return qs.select_related("topic", "assignee").prefetch_related("contact")

    @classmethod
    def from_slug(cls, org, user, slug_or_uuid: str):
        if is_uuid(slug_or_uuid):
            topic = Topic.get_accessible(org, user).filter(uuid=slug_or_uuid).first()
            if topic:
                return TopicFolder(topic)

        return FOLDERS.get(slug_or_uuid, None)

    @classmethod
    def all(cls):
        return FOLDERS


class MineFolder(TicketFolder):
    """
    Tickets assigned to the current user.
    """

    slug = "mine"
    name = _("My Tickets")
    icon = "tickets_mine"

    def get_icon(self, count) -> str:
        return self.icon if count else "tickets_mine_done"

    def get_queryset(self, org, user, *, ordered: bool):
        return super().get_queryset(org, user, ordered=ordered).filter(assignee=user)


class UnassignedFolder(TicketFolder):
    """
    Tickets not assigned to any user.
    """

    slug = "unassigned"
    name = _("Unassigned")
    verbose_name = _("Unassigned Tickets")
    icon = "tickets_unassigned"

    def get_queryset(self, org, user, *, ordered: bool):
        return super().get_queryset(org, user, ordered=ordered).filter(assignee=None)


class AllFolder(TicketFolder):
    """
    All tickets the user can access.
    """

    slug = "all"
    name = _("All")
    verbose_name = _("All Tickets")
    icon = "tickets_all"


FOLDERS = {f.slug: f() for f in TicketFolder.__subclasses__()}


class TopicFolder(TicketFolder):
    """
    Wraps a topic so we can use it like a folder.
    """

    def __init__(self, topic: Topic):
        self.slug = topic.uuid
        self.name = topic.name
        self.topic = topic

    def get_queryset(self, org, user, *, ordered: bool):
        return super().get_queryset(org, user, ordered=ordered).filter(topic=self.topic)


class TicketDailyCount(DailyCountModel):
    """
    Ticket activity daily counts by who did it and when. Mailroom writes these.
    """

    TYPE_OPENING = "O"
    TYPE_ASSIGNMENT = "A"  # includes tickets opened with assignment but excludes re-assignments
    TYPE_REPLY = "R"

    @classmethod
    def get_by_org(cls, org, count_type: str, since=None, until=None):
        return cls._get_count_set(count_type, {f"o:{org.id}": org}, since, until)

    @classmethod
    def get_by_teams(cls, teams, count_type: str, since=None, until=None):
        return cls._get_count_set(count_type, {f"t:{t.id}": t for t in teams}, since, until)

    @classmethod
    def get_by_users(cls, org, users, count_type: str, since=None, until=None):
        return cls._get_count_set(count_type, {f"o:{org.id}:u:{u.id}": u for u in users}, since, until)

    class Meta:
        indexes = [
            models.Index(name="tickets_dailycount_type_scope", fields=("count_type", "scope", "day")),
            models.Index(
                name="tickets_dailycount_unsquashed",
                fields=("count_type", "scope", "day"),
                condition=Q(is_squashed=False),
            ),
        ]


class TicketDailyTiming(DailyTimingModel):
    """
    Ticket activity daily timings. Mailroom writes these.
    """

    TYPE_FIRST_REPLY = "R"
    TYPE_LAST_CLOSE = "C"

    @classmethod
    def get_by_org(cls, org, count_type: str, since=None, until=None):
        return cls._get_count_set(count_type, {f"o:{org.id}": org}, since, until)

    class Meta:
        indexes = [
            models.Index(name="tickets_dailytiming_type_scope", fields=("count_type", "scope", "day")),
            models.Index(
                name="tickets_dailytiming_unsquashed",
                fields=("count_type", "scope", "day"),
                condition=Q(is_squashed=False),
            ),
        ]


def export_ticket_stats(org: Org, since: date, until: date) -> openpyxl.Workbook:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Tickets"
    sheet.merge_cells("A1:A2")
    sheet.cell(row=1, column=2, value="Workspace")
    sheet.merge_cells("B1:D1")
    sheet.cell(row=2, column=2, value="Opened")
    sheet.cell(row=2, column=3, value="Replies")
    sheet.cell(row=2, column=4, value="Reply Time (Secs)")

    users = list(org.users.order_by("email"))

    user_col = 5
    for user in users:
        cell = sheet.cell(row=1, column=user_col, value=str(user))
        cell.hyperlink = f"mailto:{user.email}"
        cell.style = "Hyperlink"
        sheet.merge_cells(start_row=1, start_column=user_col, end_row=1, end_column=user_col + 1)

        sheet.cell(row=2, column=user_col, value="Assigned")
        sheet.cell(row=2, column=user_col + 1, value="Replies")
        user_col += 2

    def by_day(cs: list) -> dict:
        return {c[0]: c[1] for c in cs}

    org_openings = by_day(TicketDailyCount.get_by_org(org, TicketDailyCount.TYPE_OPENING, since, until).day_totals())
    org_replies = by_day(TicketDailyCount.get_by_org(org, TicketDailyCount.TYPE_REPLY, since, until).day_totals())
    org_avg_reply_time = by_day(
        TicketDailyTiming.get_by_org(org, TicketDailyTiming.TYPE_FIRST_REPLY, since, until).day_averages(rounded=True)
    )

    user_assignments = {}
    user_replies = {}
    for user in users:
        user_assignments[user] = by_day(
            TicketDailyCount.get_by_users(org, [user], TicketDailyCount.TYPE_ASSIGNMENT, since, until).day_totals()
        )
        user_replies[user] = by_day(
            TicketDailyCount.get_by_users(org, [user], TicketDailyCount.TYPE_REPLY, since, until).day_totals()
        )

    day_row = 3
    for day in date_range(since, until):
        sheet.cell(row=day_row, column=1, value=day)
        sheet.cell(row=day_row, column=2, value=org_openings.get(day, 0))
        sheet.cell(row=day_row, column=3, value=org_replies.get(day, 0))
        sheet.cell(row=day_row, column=4, value=org_avg_reply_time.get(day, ""))

        user_col = 5
        for user in users:
            sheet.cell(row=day_row, column=user_col, value=user_assignments[user].get(day, 0))
            sheet.cell(row=day_row, column=user_col + 1, value=user_replies[user].get(day, 0))
            user_col += 2

        day_row += 1

    return workbook


class TicketExport(ExportType):
    """
    Export of tickets
    """

    slug = "ticket"
    name = _("Tickets")
    download_prefix = "tickets"

    @classmethod
    def create(cls, org, user, start_date, end_date, with_fields=(), with_groups=()):
        return Export.objects.create(
            org=org,
            export_type=cls.slug,
            start_date=start_date,
            end_date=end_date,
            config={"with_fields": [f.id for f in with_fields], "with_groups": [g.id for g in with_groups]},
            created_by=user,
        )

    def write(self, export):
        headers = ["UUID", "Opened On", "Closed On", "Topic", "Assigned To"] + export.get_contact_headers()
        start_date, end_date = export.get_date_range()

        # get the ticket ids, filtered and ordered by opened on
        ticket_ids = (
            Ticket.objects.filter(org=export.org, opened_on__gte=start_date, opened_on__lte=end_date)
            .order_by("opened_on")
            .values_list("id", flat=True)
            .using("readonly")
        )

        exporter = MultiSheetExporter("Tickets", headers, export.org.timezone)
        num_records = 0

        # add tickets to the export in batches of 1k to limit memory usage
        for batch_ids in itertools.batched(ticket_ids, 1000):
            tickets = (
                Ticket.objects.filter(id__in=batch_ids)
                .order_by("opened_on")
                .prefetch_related("org", "contact", "contact__org", "contact__groups", "assignee", "topic")
                .using("readonly")
            )

            Contact.bulk_urn_cache_initialize([t.contact for t in tickets], using="readonly")

            for ticket in tickets:
                values = [
                    str(ticket.uuid),
                    ticket.opened_on,
                    ticket.closed_on,
                    ticket.topic.name,
                    ticket.assignee.email if ticket.assignee else None,
                ]
                values += export.get_contact_columns(ticket.contact)

                exporter.write_row(values)

            num_records += len(tickets)

            export.modified_on = timezone.now()
            export.save(update_fields=("modified_on",))

        return *exporter.save_file(), num_records
