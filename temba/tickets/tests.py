from datetime import date, datetime, timedelta, timezone as tzone
from unittest.mock import call, patch

from openpyxl import load_workbook

from django.core.files.storage import default_storage
from django.test.utils import override_settings
from django.urls import reverse
from django.utils import timezone

from temba.contacts.models import Contact, ContactField, ContactURN
from temba.orgs.models import Export, Invitation, Org, OrgMembership, OrgRole
from temba.orgs.tasks import squash_item_counts
from temba.tests import CRUDLTestMixin, TembaTest, matchers, mock_mailroom
from temba.utils.dates import datetime_to_timestamp
from temba.utils.uuid import uuid4

from .models import (
    Shortcut,
    Team,
    Ticket,
    TicketDailyCount,
    TicketDailyTiming,
    TicketEvent,
    TicketExport,
    Topic,
    export_ticket_stats,
)
from .tasks import squash_ticket_counts


class TicketTest(TembaTest):
    @mock_mailroom
    def test_model(self, mr_mocks):
        topic = Topic.create(self.org, self.admin, "Sales")
        contact = self.create_contact("Bob", urns=["twitter:bobby"])

        ticket = Ticket.objects.create(
            org=self.org,
            contact=contact,
            topic=self.org.default_ticket_topic,
            status="O",
        )

        self.assertEqual(f"Ticket[uuid={ticket.uuid}, topic=General]", str(ticket))

        # test bulk assignment
        Ticket.bulk_assign(self.org, self.admin, [ticket], self.agent)

        # test bulk un-assignment
        Ticket.bulk_assign(self.org, self.admin, [ticket], None)

        self.assertEqual(
            [
                call(self.org, self.admin, [ticket], self.agent),
                call(self.org, self.admin, [ticket], None),
            ],
            mr_mocks.calls["ticket_assign"],
        )

        # test bulk adding a note
        Ticket.bulk_add_note(self.org, self.admin, [ticket], "please handle")

        self.assertEqual([call(self.org, self.admin, [ticket], "please handle")], mr_mocks.calls["ticket_add_note"])

        # test bulk changing topic
        Ticket.bulk_change_topic(self.org, self.admin, [ticket], topic)

        self.assertEqual([call(self.org, self.admin, [ticket], topic)], mr_mocks.calls["ticket_change_topic"])

        # test bulk closing
        Ticket.bulk_close(self.org, self.admin, [ticket], force=True)

        self.assertEqual([call(self.org, self.admin, [ticket], force=True)], mr_mocks.calls["ticket_close"])

        # test bulk re-opening
        Ticket.bulk_reopen(self.org, self.admin, [ticket])

        self.assertEqual([call(self.org, self.admin, [ticket])], mr_mocks.calls["ticket_reopen"])

    def test_allowed_assignees(self):
        self.assertEqual({self.admin, self.editor, self.agent}, set(Ticket.get_allowed_assignees(self.org)))
        self.assertEqual({self.admin2}, set(Ticket.get_allowed_assignees(self.org2)))

    @mock_mailroom
    def test_counts(self, mr_mocks):
        general = self.org.default_ticket_topic
        cats = Topic.create(self.org, self.admin, "Cats")

        contact1 = self.create_contact("Bob", urns=["twitter:bobby"])
        contact2 = self.create_contact("Jim", urns=["twitter:jimmy"])

        org2_general = self.org2.default_ticket_topic
        org2_contact = self.create_contact("Bob", urns=["twitter:bobby"], org=self.org2)

        t1 = self.create_ticket(contact1, topic=general)
        t2 = self.create_ticket(contact2, topic=general)
        t3 = self.create_ticket(contact1, topic=general)
        t4 = self.create_ticket(contact2, topic=cats)
        t5 = self.create_ticket(contact1, topic=cats)
        t6 = self.create_ticket(org2_contact, topic=org2_general)

        def assert_counts(
            org, *, assignee_open: dict, assignee_closed: dict, topic_open: dict, topic_closed: dict, contacts: dict
        ):
            all_topics = org.topics.filter(is_active=True)
            assignees = [None] + list(Ticket.get_allowed_assignees(org))

            self.assertEqual(
                assignee_open, {u: Ticket.get_assignee_count(org, u, all_topics, Ticket.STATUS_OPEN) for u in assignees}
            )
            self.assertEqual(
                assignee_closed,
                {u: Ticket.get_assignee_count(org, u, all_topics, Ticket.STATUS_CLOSED) for u in assignees},
            )

            self.assertEqual(sum(assignee_open.values()), Ticket.get_status_count(org, all_topics, Ticket.STATUS_OPEN))
            self.assertEqual(
                sum(assignee_closed.values()), Ticket.get_status_count(org, all_topics, Ticket.STATUS_CLOSED)
            )

            self.assertEqual(topic_open, Ticket.get_topic_counts(org, list(org.topics.all()), Ticket.STATUS_OPEN))
            self.assertEqual(topic_closed, Ticket.get_topic_counts(org, list(org.topics.all()), Ticket.STATUS_CLOSED))

            self.assertEqual(contacts, {c: Contact.objects.get(id=c.id).ticket_count for c in contacts})

        # t1:O/None/General t2:O/None/General t3:O/None/General t4:O/None/Cats t5:O/None/Cats t6:O/None/General
        assert_counts(
            self.org,
            assignee_open={None: 5, self.agent: 0, self.editor: 0, self.admin: 0},
            assignee_closed={None: 0, self.agent: 0, self.editor: 0, self.admin: 0},
            topic_open={general: 3, cats: 2},
            topic_closed={general: 0, cats: 0},
            contacts={contact1: 3, contact2: 2},
        )
        assert_counts(
            self.org2,
            assignee_open={None: 1, self.admin2: 0},
            assignee_closed={None: 0, self.admin2: 0},
            topic_open={org2_general: 1},
            topic_closed={org2_general: 0},
            contacts={org2_contact: 1},
        )

        Ticket.bulk_assign(self.org, self.admin, [t1, t2], assignee=self.agent)
        Ticket.bulk_assign(self.org, self.admin, [t3], assignee=self.editor)
        Ticket.bulk_assign(self.org2, self.admin2, [t6], assignee=self.admin2)

        # t1:O/Agent/General t2:O/Agent/General t3:O/Editor/General t4:O/None/Cats t5:O/None/Cats t6:O/Admin2/General
        assert_counts(
            self.org,
            assignee_open={None: 2, self.agent: 2, self.editor: 1, self.admin: 0},
            assignee_closed={None: 0, self.agent: 0, self.editor: 0, self.admin: 0},
            topic_open={general: 3, cats: 2},
            topic_closed={general: 0, cats: 0},
            contacts={contact1: 3, contact2: 2},
        )
        assert_counts(
            self.org2,
            assignee_open={None: 0, self.admin2: 1},
            assignee_closed={None: 0, self.admin2: 0},
            topic_open={org2_general: 1},
            topic_closed={org2_general: 0},
            contacts={org2_contact: 1},
        )

        Ticket.bulk_close(self.org, self.admin, [t1, t4])
        Ticket.bulk_close(self.org2, self.admin2, [t6])

        # t1:C/Agent/General t2:O/Agent/General t3:O/Editor/General t4:C/None/Cats t5:O/None/Cats t6:C/Admin2/General
        assert_counts(
            self.org,
            assignee_open={None: 1, self.agent: 1, self.editor: 1, self.admin: 0},
            assignee_closed={None: 1, self.agent: 1, self.editor: 0, self.admin: 0},
            topic_open={general: 2, cats: 1},
            topic_closed={general: 1, cats: 1},
            contacts={contact1: 2, contact2: 1},
        )
        assert_counts(
            self.org2,
            assignee_open={None: 0, self.admin2: 0},
            assignee_closed={None: 0, self.admin2: 1},
            topic_open={org2_general: 0},
            topic_closed={org2_general: 1},
            contacts={org2_contact: 0},
        )

        Ticket.bulk_assign(self.org, self.admin, [t1, t5], assignee=self.admin)

        # t1:C/Admin/General t2:O/Agent/General t3:O/Editor/General t4:C/None/Cats t5:O/Admin/Cats t6:C/Admin2/General
        assert_counts(
            self.org,
            assignee_open={None: 0, self.agent: 1, self.editor: 1, self.admin: 1},
            assignee_closed={None: 1, self.agent: 0, self.editor: 0, self.admin: 1},
            topic_open={general: 2, cats: 1},
            topic_closed={general: 1, cats: 1},
            contacts={contact1: 2, contact2: 1},
        )

        Ticket.bulk_reopen(self.org, self.admin, [t4])
        Ticket.bulk_change_topic(self.org, self.admin, [t1], cats)

        # t1:C/Admin/General t2:O/Agent/General t3:O/Editor/General t4:O/None/Cats t5:O/Admin/Cats t6:C/Admin2/General
        assert_counts(
            self.org,
            assignee_open={None: 1, self.agent: 1, self.editor: 1, self.admin: 1},
            assignee_closed={None: 0, self.agent: 0, self.editor: 0, self.admin: 1},
            topic_open={general: 2, cats: 2},
            topic_closed={general: 0, cats: 1},
            contacts={contact1: 2, contact2: 2},
        )

        squash_item_counts()  # shouldn't change counts

        assert_counts(
            self.org,
            assignee_open={None: 1, self.agent: 1, self.editor: 1, self.admin: 1},
            assignee_closed={None: 0, self.agent: 0, self.editor: 0, self.admin: 1},
            topic_open={general: 2, cats: 2},
            topic_closed={general: 0, cats: 1},
            contacts={contact1: 2, contact2: 2},
        )

        TicketEvent.objects.all().delete()
        t1.delete()
        t2.delete()
        t6.delete()

        # t3:O/Editor/General t4:O/None/Cats t5:O/Admin/Cats
        assert_counts(
            self.org,
            assignee_open={None: 1, self.agent: 0, self.editor: 1, self.admin: 1},
            assignee_closed={None: 0, self.agent: 0, self.editor: 0, self.admin: 0},
            topic_open={general: 1, cats: 2},
            topic_closed={general: 0, cats: 0},
            contacts={contact1: 2, contact2: 1},
        )
        assert_counts(
            self.org2,
            assignee_open={None: 0, self.admin2: 0},
            assignee_closed={None: 0, self.admin2: 0},
            topic_open={org2_general: 0},
            topic_closed={org2_general: 0},
            contacts={org2_contact: 0},
        )

        squash_item_counts()

        # check count model raw values are consistent
        self.assertEqual(
            {
                f"tickets:O:{general.id}:{self.editor.id}": 1,
                f"tickets:O:{cats.id}:0": 1,
                f"tickets:O:{cats.id}:{self.admin.id}": 1,
            },
            {c["scope"]: c["count"] for c in self.org.counts.order_by("scope").values("scope", "count")},
        )


class ShortcutCRUDLTest(TembaTest, CRUDLTestMixin):
    def test_create(self):
        create_url = reverse("tickets.shortcut_create")

        self.assertRequestDisallowed(create_url, [None, self.agent, self.user])

        self.assertCreateFetch(create_url, [self.editor, self.admin], form_fields=("name", "text"))

        # try to create with empty values
        self.assertCreateSubmit(
            create_url,
            self.admin,
            {"name": "", "text": ""},
            form_errors={"name": "This field is required.", "text": "This field is required."},
        )

        # try to create with name that is already taken
        Shortcut.create(self.org, self.admin, "Reboot", "Try switching it off and on again")

        self.assertCreateSubmit(
            create_url,
            self.admin,
            {"name": "reboot", "text": "Have you tried..."},
            form_errors={"name": "Shortcut with this name already exists."},
        )

        # try to create with name that has invalid characters
        self.assertCreateSubmit(
            create_url,
            self.admin,
            {"name": "\\reboot", "text": "x"},
            form_errors={"name": "Cannot contain the character: \\"},
        )

        # try to create with name that is too long
        self.assertCreateSubmit(
            create_url,
            self.admin,
            {"name": "X" * 65, "text": "x"},
            form_errors={"name": "Ensure this value has at most 64 characters (it has 65)."},
        )

        self.assertCreateSubmit(
            create_url,
            self.admin,
            {"name": "Not Interested", "text": "We're not interested"},
            new_obj_query=Shortcut.objects.filter(name="Not Interested", text="We're not interested", is_system=False),
            success_status=302,
        )

    def test_update(self):
        shortcut = Shortcut.create(self.org, self.admin, "Planes", "Planes are...")
        Shortcut.create(self.org, self.admin, "Trains", "Trains are...")

        update_url = reverse("tickets.shortcut_update", args=[shortcut.id])

        self.assertRequestDisallowed(update_url, [None, self.user, self.agent, self.admin2])

        self.assertUpdateFetch(update_url, [self.editor, self.admin], form_fields=["name", "text"])

        # names must be unique (case-insensitive)
        self.assertUpdateSubmit(
            update_url,
            self.admin,
            {"name": "trains", "text": "Trains are..."},
            form_errors={"name": "Shortcut with this name already exists."},
            object_unchanged=shortcut,
        )

        self.assertUpdateSubmit(update_url, self.admin, {"name": "Cars", "text": "Cars are..."}, success_status=302)

        shortcut.refresh_from_db()
        self.assertEqual(shortcut.name, "Cars")
        self.assertEqual(shortcut.text, "Cars are...")

    def test_delete(self):
        shortcut1 = Shortcut.create(self.org, self.admin, "Planes", "Planes are...")
        shortcut2 = Shortcut.create(self.org, self.admin, "Trains", "Trains are...")

        delete_url = reverse("tickets.shortcut_delete", args=[shortcut1.id])

        self.assertRequestDisallowed(delete_url, [None, self.user, self.agent, self.admin2])

        response = self.assertDeleteFetch(delete_url, [self.editor, self.admin])
        self.assertContains(response, "You are about to delete")

        # submit to delete it
        response = self.assertDeleteSubmit(delete_url, self.admin, object_deactivated=shortcut1, success_status=302)

        # other shortcut unaffected
        shortcut2.refresh_from_db()
        self.assertTrue(shortcut2.is_active)

    def test_list(self):
        shortcut1 = Shortcut.create(self.org, self.admin, "Planes", "Planes are...")
        shortcut2 = Shortcut.create(self.org, self.admin, "Trains", "Trains are...")
        Shortcut.create(self.org2, self.admin, "Cars", "Other org")

        list_url = reverse("tickets.shortcut_list")

        self.assertRequestDisallowed(list_url, [None, self.agent])

        self.assertListFetch(list_url, [self.editor, self.admin], context_objects=[shortcut1, shortcut2])


class TopicCRUDLTest(TembaTest, CRUDLTestMixin):
    @override_settings(ORG_LIMIT_DEFAULTS={"topics": 2})
    def test_create(self):
        create_url = reverse("tickets.topic_create")

        self.assertRequestDisallowed(create_url, [None, self.agent, self.user])

        self.assertCreateFetch(create_url, [self.editor, self.admin], form_fields=("name",))

        # try to create with empty name
        self.assertCreateSubmit(
            create_url,
            self.admin,
            {"name": ""},
            form_errors={"name": "This field is required."},
        )

        # try to create with name that is too long
        self.assertCreateSubmit(
            create_url,
            self.admin,
            {"name": "X" * 65},
            form_errors={"name": "Ensure this value has at most 64 characters (it has 65)."},
        )

        self.assertCreateSubmit(
            create_url,
            self.admin,
            {"name": "Sales"},
            new_obj_query=Topic.objects.filter(name="Sales", is_system=False),
            success_status=302,
        )

        # try again with same name
        self.assertCreateSubmit(
            create_url,
            self.admin,
            {"name": "sales"},
            form_errors={"name": "Topic with this name already exists."},
        )

        self.assertCreateSubmit(
            create_url,
            self.admin,
            {"name": "Support"},
            new_obj_query=Topic.objects.filter(name="Support", is_system=False),
            success_status=302,
        )

        # try to create another now that we've reached the limit
        self.assertCreateSubmit(
            create_url,
            self.admin,
            {"name": "Training"},
            form_errors={
                "__all__": "This workspace has reached its limit of 2 topics. You must delete existing ones before you can create new ones."
            },
        )

    def test_update(self):
        topic = Topic.create(self.org, self.admin, "Hot Topic")

        update_url = reverse("tickets.topic_update", args=[topic.id])

        self.assertRequestDisallowed(update_url, [None, self.user, self.agent, self.admin2])

        self.assertUpdateFetch(update_url, [self.editor, self.admin], form_fields=["name"])

        # names must be unique (case-insensitive)
        self.assertUpdateSubmit(
            update_url,
            self.admin,
            {"name": "general"},
            form_errors={"name": "Topic with this name already exists."},
            object_unchanged=topic,
        )

        self.assertUpdateSubmit(update_url, self.admin, {"name": "Boring"}, success_status=302)

        topic.refresh_from_db()
        self.assertEqual(topic.name, "Boring")

        # can't edit a system topic
        self.assertRequestDisallowed(
            reverse("tickets.topic_update", args=[self.org.default_ticket_topic.id]), [self.admin]
        )

    def test_delete(self):
        topic1 = Topic.create(self.org, self.admin, "Planes")
        topic2 = Topic.create(self.org, self.admin, "Trains")
        ticket = self.create_ticket(self.create_contact("Bob", urns=["twitter:bobby"]), topic=topic1)

        delete_url = reverse("tickets.topic_delete", args=[topic1.id])

        self.assertRequestDisallowed(delete_url, [None, self.user, self.agent, self.admin2])

        # deleting blocked for topic with tickets
        response = self.assertDeleteFetch(delete_url, [self.editor, self.admin])
        self.assertContains(response, "Sorry, the <b>Planes</b> topic can't be deleted")

        ticket.topic = topic2
        ticket.save(update_fields=("topic",))

        # try again...
        response = self.assertDeleteFetch(delete_url, [self.editor, self.admin])
        self.assertContains(response, "You are about to delete the <b>Planes</b> topic")

        response = self.assertDeleteSubmit(delete_url, self.admin, object_deactivated=topic1, success_status=302)

        # other topic unafected
        topic2.refresh_from_db()
        self.assertTrue(topic2.is_active)

        # we should have been redirected to the default topic
        self.assertEqual(f"/ticket/{self.org.default_ticket_topic.uuid}/open/", response.url)


class TeamCRUDLTest(TembaTest, CRUDLTestMixin):
    @override_settings(ORG_LIMIT_DEFAULTS={"teams": 1})
    def test_create(self):
        create_url = reverse("tickets.team_create")

        # nobody can access if new orgs feature not enabled
        response = self.requestView(create_url, self.admin)
        self.assertRedirect(response, reverse("orgs.org_workspace"))

        self.org.features = [Org.FEATURE_TEAMS]
        self.org.save(update_fields=("features",))

        self.assertRequestDisallowed(create_url, [None, self.agent, self.user, self.editor])

        self.assertCreateFetch(create_url, [self.admin], form_fields=("name", "topics"))

        sales = Topic.create(self.org, self.admin, "Sales")
        for n in range(Team.max_topics + 1):
            Topic.create(self.org, self.admin, f"Topic {n}")

        # try to create with empty values
        self.assertCreateSubmit(
            create_url,
            self.admin,
            {"name": "", "topics": []},
            form_errors={"name": "This field is required."},
        )

        self.assertCreateSubmit(
            create_url,
            self.admin,
            {"name": "all topics", "topics": []},
            form_errors={"name": "Team with this name already exists."},
        )

        # try to create with name that has invalid characters
        self.assertCreateSubmit(
            create_url,
            self.admin,
            {"name": "\\ministry", "topics": []},
            form_errors={"name": "Cannot contain the character: \\"},
        )

        # try to create with name that is too long
        self.assertCreateSubmit(
            create_url,
            self.admin,
            {"name": "X" * 65, "topics": []},
            form_errors={"name": "Ensure this value has at most 64 characters (it has 65)."},
        )

        # try to create with too many topics
        self.assertCreateSubmit(
            create_url,
            self.admin,
            {"name": "Everything", "topics": [t.id for t in self.org.topics.all()]},
            form_errors={"topics": "Teams can have at most 10 topics."},
        )

        self.assertCreateSubmit(
            create_url,
            self.admin,
            {"name": "Sales", "topics": [sales.id]},
            new_obj_query=Team.objects.filter(name="Sales", is_system=False),
            success_status=302,
        )

        team = Team.objects.get(name="Sales")
        self.assertEqual({sales}, set(team.topics.all()))

        # try to create another now that we've reached the limit
        self.assertCreateSubmit(
            create_url,
            self.admin,
            {"name": "Training", "topics": [sales.id]},
            form_errors={
                "__all__": "This workspace has reached its limit of 1 teams. You must delete existing ones before you can create new ones."
            },
        )

    def test_update(self):
        sales = Topic.create(self.org, self.admin, "Sales")
        marketing = Topic.create(self.org, self.admin, "Marketing")
        team = Team.create(self.org, self.admin, "Sales", topics=[sales])

        update_url = reverse("tickets.team_update", args=[team.id])

        self.assertRequestDisallowed(update_url, [None, self.user, self.agent, self.editor, self.admin2])

        self.assertUpdateFetch(update_url, [self.admin], form_fields=["name", "topics"])

        # names must be unique (case-insensitive)
        self.assertUpdateSubmit(
            update_url,
            self.admin,
            {"name": "all topics"},
            form_errors={"name": "Team with this name already exists."},
            object_unchanged=team,
        )

        self.assertUpdateSubmit(
            update_url, self.admin, {"name": "Marketing", "topics": [marketing.id]}, success_status=302
        )

        team.refresh_from_db()
        self.assertEqual(team.name, "Marketing")
        self.assertEqual({marketing}, set(team.topics.all()))

        # can't edit a system team
        self.assertRequestDisallowed(
            reverse("tickets.team_update", args=[self.org.default_ticket_team.id]), [self.admin]
        )

    def test_delete(self):
        sales = Topic.create(self.org, self.admin, "Sales")
        team1 = Team.create(self.org, self.admin, "Sales", topics=[sales])
        team2 = Team.create(self.org, self.admin, "Other", topics=[sales])
        self.org.add_user(self.agent, OrgRole.AGENT, team=team1)
        invite = Invitation.create(self.org, self.admin, "newagent@textit.com", OrgRole.AGENT, team=team1)

        delete_url = reverse("tickets.team_delete", args=[team1.id])

        self.assertRequestDisallowed(delete_url, [None, self.user, self.agent, self.editor, self.admin2])

        # deleting blocked for team with agents
        response = self.assertDeleteFetch(delete_url, [self.admin])
        self.assertContains(response, "Sorry, the <b>Sales</b> team can't be deleted while it still has agents")

        self.org.add_user(self.agent, OrgRole.AGENT, team=team2)

        # deleting blocked for team with pending invitations
        response = self.assertDeleteFetch(delete_url, [self.admin])
        self.assertContains(
            response, "Sorry, the <b>Sales</b> team can't be deleted while it still has pending invitations"
        )

        invite.release()

        # try again...
        response = self.assertDeleteFetch(delete_url, [self.admin])
        self.assertContains(response, "You are about to delete the <b>Sales</b> team")

        response = self.assertDeleteSubmit(delete_url, self.admin, object_deactivated=team1, success_status=302)

        # other team unafected
        team2.refresh_from_db()
        self.assertTrue(team2.is_active)

        # we should have been redirected to the team list
        self.assertEqual("/team/", response.url)

    def test_list(self):
        sales = Topic.create(self.org, self.admin, "Sales")
        team1 = Team.create(self.org, self.admin, "Sales", topics=[sales])
        team2 = Team.create(self.org, self.admin, "Other", topics=[sales])
        Team.create(self.org2, self.admin2, "Cars", topics=[])

        list_url = reverse("tickets.team_list")

        # nobody can access if new orgs feature not enabled
        response = self.requestView(list_url, self.admin)
        self.assertRedirect(response, reverse("orgs.org_workspace"))

        self.org.features = [Org.FEATURE_TEAMS]
        self.org.save(update_fields=("features",))

        self.assertRequestDisallowed(list_url, [None, self.agent, self.editor])

        self.assertListFetch(list_url, [self.admin], context_objects=[self.org.default_ticket_team, team2, team1])


class TicketCRUDLTest(TembaTest, CRUDLTestMixin):
    def setUp(self):
        super().setUp()

        self.contact = self.create_contact("Bob", urns=["twitter:bobby"])
        self.sales = Topic.create(self.org, self.admin, "Sales")
        self.support = Topic.create(self.org, self.admin, "Support")

        # create other agent users in teams with limited topic access
        self.agent2 = self.create_user("agent2@textit.com")
        sales_only = Team.create(self.org, self.admin, "Sales", topics=[self.sales])
        self.org.add_user(self.agent2, OrgRole.AGENT, team=sales_only)

        self.agent3 = self.create_user("agent3@textit.com")
        support_only = Team.create(self.org, self.admin, "Support", topics=[self.support])
        self.org.add_user(self.agent3, OrgRole.AGENT, team=support_only)

    def test_list(self):
        list_url = reverse("tickets.ticket_list")

        ticket = self.create_ticket(self.contact, assignee=self.admin, topic=self.support)

        # just a placeholder view for frontend components
        self.assertRequestDisallowed(list_url, [None])
        self.assertListFetch(
            list_url, [self.user, self.editor, self.admin, self.agent, self.agent2, self.agent3], context_objects=[]
        )

        # link to our ticket within the All folder
        deep_link = f"{list_url}all/open/{ticket.uuid}/"

        response = self.assertListFetch(
            deep_link, [self.user, self.editor, self.admin, self.agent, self.agent3], context_objects=[]
        )
        self.assertEqual("All", response.context["title"])
        self.assertEqual("all", response.context["folder"])
        self.assertEqual("open", response.context["status"])

        # our ticket exists on the first page, so it'll get flagged to be focused
        self.assertEqual(str(ticket.uuid), response.context["nextUUID"])

        # we have a specific ticket so we should show context menu for it
        self.assertContentMenu(deep_link, self.admin, ["Edit", "Add Note", "Start Flow"])

        with self.assertNumQueries(11):
            self.client.get(deep_link)

        # try same request but for agent that can't see this ticket
        response = self.assertListFetch(deep_link, [self.agent2], context_objects=[])
        self.assertEqual("All", response.context["title"])
        self.assertEqual("all", response.context["folder"])
        self.assertEqual("open", response.context["status"])
        self.assertNotIn("nextUUID", response.context)

        # can also link to our ticket within the Support topic
        deep_link = f"{list_url}{self.support.uuid}/open/{ticket.uuid}/"

        self.assertRequestDisallowed(deep_link, [self.agent2])  # doesn't have access to that topic

        response = self.assertListFetch(
            deep_link, [self.user, self.editor, self.admin, self.agent, self.agent3], context_objects=[]
        )
        self.assertEqual("Support", response.context["title"])
        self.assertEqual(str(self.support.uuid), response.context["folder"])
        self.assertEqual("open", response.context["status"])

        # try to link to our ticket but with mismatched topic
        deep_link = f"{list_url}{self.sales.uuid}/closed/{str(ticket.uuid)}/"

        # redirected to All
        response = self.assertListFetch(deep_link, [self.agent], context_objects=[])
        self.assertEqual("all", response.context["folder"])
        self.assertEqual("open", response.context["status"])
        self.assertEqual(str(ticket.uuid), response.context["uuid"])

        # try to link to our ticket but with mismatched status
        deep_link = f"{list_url}all/closed/{ticket.uuid}/"

        # now our ticket is listed as the uuid and we were redirected to All folder with Open status
        response = self.assertListFetch(deep_link, [self.agent], context_objects=[])
        self.assertEqual("all", response.context["folder"])
        self.assertEqual("open", response.context["status"])
        self.assertEqual(str(ticket.uuid), response.context["uuid"])

        # and again we have a specific ticket so we should show context menu for it
        self.assertContentMenu(deep_link, self.admin, ["Edit", "Add Note", "Start Flow"])

        # non-existent topic should give a 404
        bad_topic_link = f"{list_url}{uuid4()}/open/{ticket.uuid}/"
        response = self.requestView(bad_topic_link, self.agent)
        self.assertEqual(404, response.status_code)

        response = self.client.get(
            list_url,
            content_type="application/json",
            HTTP_X_TEMBA_REFERER_PATH=f"/tickets/mine/open/{ticket.uuid}",
        )
        self.assertEqual(("tickets", "mine", "open", str(ticket.uuid)), response.context["temba_referer"])

        # contacts in a flow don't get a start flow option
        flow = self.create_flow("Test")
        self.contact.current_flow = flow
        self.contact.save()
        deep_link = f"{list_url}all/open/{str(ticket.uuid)}/"
        self.assertContentMenu(deep_link, self.admin, ["Edit", "Add Note"])

        # closed our tickets don't get extra menu options
        ticket.status = Ticket.STATUS_CLOSED
        ticket.save(update_fields=("status",))
        deep_link = f"{list_url}all/closed/{str(ticket.uuid)}/"
        self.assertContentMenu(deep_link, self.admin, [])

    def test_update(self):
        ticket = self.create_ticket(self.contact, assignee=self.admin)

        update_url = reverse("tickets.ticket_update", args=[ticket.uuid])

        self.assertRequestDisallowed(update_url, [None, self.user, self.admin2])
        self.assertUpdateFetch(update_url, [self.agent, self.editor, self.admin], form_fields=["topic"])

        user_topic = Topic.objects.create(org=self.org, name="Hot Topic", created_by=self.admin, modified_by=self.admin)

        # edit successfully
        self.assertUpdateSubmit(update_url, self.admin, {"topic": user_topic.id}, success_status=302)

        ticket.refresh_from_db()
        self.assertEqual(user_topic, ticket.topic)

    def test_menu(self):
        menu_url = reverse("tickets.ticket_menu")

        self.create_ticket(self.contact, assignee=self.admin)
        self.create_ticket(self.contact, assignee=self.admin, topic=self.sales)
        self.create_ticket(self.contact, assignee=None)
        self.create_ticket(self.contact, closed_on=timezone.now())

        self.assertRequestDisallowed(menu_url, [None])
        self.assertPageMenu(
            menu_url,
            self.admin,
            [
                "My Tickets (2)",
                "Unassigned (1)",
                "All (3)",
                "Shortcuts (0)",
                "Export",
                "New Topic",
                "General (2)",
                "Sales (1)",
                "Support (0)",
            ],
        )
        self.assertPageMenu(
            menu_url,
            self.agent,
            ["My Tickets (0)", "Unassigned (1)", "All (3)", "General (2)", "Sales (1)", "Support (0)"],
        )
        self.assertPageMenu(menu_url, self.agent2, ["My Tickets (0)", "Unassigned (0)", "All (1)", "Sales (1)"])
        self.assertPageMenu(menu_url, self.agent3, ["My Tickets (0)", "Unassigned (0)", "All (0)", "Support (0)"])

    @mock_mailroom
    def test_folder(self, mr_mocks):
        self.login(self.admin)

        user_topic = Topic.objects.create(org=self.org, name="Hot Topic", created_by=self.admin, modified_by=self.admin)

        contact1 = self.create_contact("Joe", phone="123", last_seen_on=timezone.now())
        contact2 = self.create_contact("Frank", phone="124", last_seen_on=timezone.now())
        contact3 = self.create_contact("Anne", phone="125", last_seen_on=timezone.now())
        self.create_contact("Mary No tickets", phone="126", last_seen_on=timezone.now())
        self.create_contact("Mr Other Org", phone="126", last_seen_on=timezone.now(), org=self.org2)
        topic = Topic.objects.filter(org=self.org, is_system=True).first()

        open_url = reverse("tickets.ticket_folder", kwargs={"folder": "all", "status": "open"})
        closed_url = reverse("tickets.ticket_folder", kwargs={"folder": "all", "status": "closed"})
        mine_url = reverse("tickets.ticket_folder", kwargs={"folder": "mine", "status": "open"})
        unassigned_url = reverse("tickets.ticket_folder", kwargs={"folder": "unassigned", "status": "open"})
        system_topic_url = reverse("tickets.ticket_folder", kwargs={"folder": topic.uuid, "status": "open"})
        user_topic_url = reverse("tickets.ticket_folder", kwargs={"folder": user_topic.uuid, "status": "open"})
        bad_topic_url = reverse("tickets.ticket_folder", kwargs={"folder": uuid4(), "status": "open"})

        def assert_tickets(resp, tickets: list):
            actual_tickets = [t["ticket"]["uuid"] for t in resp.json()["results"]]
            expected_tickets = [str(t.uuid) for t in tickets]
            self.assertEqual(expected_tickets, actual_tickets)

        # system topic has no menu options
        self.assertContentMenu(system_topic_url, self.admin, [])

        # user topic gets edit too
        self.assertContentMenu(user_topic_url, self.admin, ["Edit", "Delete"])

        # no tickets yet so no contacts returned
        response = self.client.get(open_url)
        assert_tickets(response, [])

        # contact 1 has two open tickets and some messages
        c1_t1 = self.create_ticket(contact1)
        # assign it
        c1_t1.assign(self.admin, assignee=self.admin)
        c1_t2 = self.create_ticket(contact1)
        self.create_incoming_msg(contact1, "I have an issue")
        self.create_outgoing_msg(contact1, "We can help", created_by=self.admin)

        # contact 2 has an open ticket and a closed ticket
        c2_t1 = self.create_ticket(contact2)
        c2_t2 = self.create_ticket(contact2, closed_on=timezone.now())

        self.create_incoming_msg(contact2, "Anyone there?")
        self.create_incoming_msg(contact2, "Hello?")

        # contact 3 has two closed tickets
        c3_t1 = self.create_ticket(contact3, closed_on=timezone.now())
        c3_t2 = self.create_ticket(contact3, closed_on=timezone.now())

        self.create_outgoing_msg(contact3, "Yes", created_by=self.agent)

        # fetching open folder returns all open tickets
        with self.assertNumQueries(12):
            response = self.client.get(open_url)

        assert_tickets(response, [c2_t1, c1_t2, c1_t1])

        joes_open_tickets = contact1.tickets.filter(status="O").order_by("-opened_on")

        expected_json = {
            "results": [
                {
                    "uuid": str(contact2.uuid),
                    "name": "Frank",
                    "last_seen_on": matchers.ISODate(),
                    "last_msg": {
                        "text": "Hello?",
                        "direction": "I",
                        "type": "T",
                        "created_on": matchers.ISODate(),
                        "sender": None,
                        "attachments": [],
                    },
                    "ticket": {
                        "uuid": str(contact2.tickets.filter(status="O").first().uuid),
                        "assignee": None,
                        "topic": {"uuid": matchers.UUID4String(), "name": "General"},
                        "last_activity_on": matchers.ISODate(),
                        "closed_on": None,
                    },
                },
                {
                    "uuid": str(contact1.uuid),
                    "name": "Joe",
                    "last_seen_on": matchers.ISODate(),
                    "last_msg": {
                        "text": "We can help",
                        "direction": "O",
                        "type": "T",
                        "created_on": matchers.ISODate(),
                        "sender": {"id": self.admin.id, "email": "admin@textit.com"},
                        "attachments": [],
                    },
                    "ticket": {
                        "uuid": str(joes_open_tickets[0].uuid),
                        "assignee": None,
                        "topic": {"uuid": matchers.UUID4String(), "name": "General"},
                        "last_activity_on": matchers.ISODate(),
                        "closed_on": None,
                    },
                },
                {
                    "uuid": str(contact1.uuid),
                    "name": "Joe",
                    "last_seen_on": matchers.ISODate(),
                    "last_msg": {
                        "text": "We can help",
                        "direction": "O",
                        "type": "T",
                        "created_on": matchers.ISODate(),
                        "sender": {"id": self.admin.id, "email": "admin@textit.com"},
                        "attachments": [],
                    },
                    "ticket": {
                        "uuid": str(joes_open_tickets[1].uuid),
                        "assignee": {
                            "id": self.admin.id,
                            "first_name": "Andy",
                            "last_name": "",
                            "email": "admin@textit.com",
                        },
                        "topic": {"uuid": matchers.UUID4String(), "name": "General"},
                        "last_activity_on": matchers.ISODate(),
                        "closed_on": None,
                    },
                },
            ]
        }
        self.assertEqual(expected_json, response.json())

        # test before and after windowing
        response = self.client.get(f"{open_url}?before={datetime_to_timestamp(c2_t1.last_activity_on)}")
        self.assertEqual(2, len(response.json()["results"]))

        response = self.client.get(f"{open_url}?after={datetime_to_timestamp(c1_t2.last_activity_on)}")
        self.assertEqual(1, len(response.json()["results"]))

        # the two unassigned tickets
        response = self.client.get(unassigned_url)
        assert_tickets(response, [c2_t1, c1_t2])

        # one assigned ticket for mine
        response = self.client.get(mine_url)
        assert_tickets(response, [c1_t1])

        # three tickets for our general topic
        response = self.client.get(system_topic_url)
        assert_tickets(response, [c2_t1, c1_t2, c1_t1])

        # bad topic should be a 404
        response = self.client.get(bad_topic_url)
        self.assertEqual(response.status_code, 404)

        # fetching closed folder returns all closed tickets
        response = self.client.get(closed_url)
        assert_tickets(response, [c3_t2, c3_t1, c2_t2])
        self.assertEqual(
            {
                "uuid": str(contact3.uuid),
                "name": "Anne",
                "last_seen_on": matchers.ISODate(),
                "last_msg": {
                    "text": "Yes",
                    "direction": "O",
                    "type": "T",
                    "created_on": matchers.ISODate(),
                    "sender": {"id": self.agent.id, "email": "agent@textit.com"},
                    "attachments": [],
                },
                "ticket": {
                    "uuid": str(c3_t2.uuid),
                    "assignee": None,
                    "topic": {"uuid": matchers.UUID4String(), "name": "General"},
                    "last_activity_on": matchers.ISODate(),
                    "closed_on": matchers.ISODate(),
                },
            },
            response.json()["results"][0],
        )

        # deep linking to a single ticket returns just that ticket
        response = self.client.get(f"{open_url}{str(c1_t1.uuid)}")
        assert_tickets(response, [c1_t1])

        # make sure when paging we get a next url
        with patch("temba.tickets.views.TicketCRUDL.Folder.paginate_by", 1):
            response = self.client.get(open_url + "?_format=json")
            self.assertIsNotNone(response.json()["next"])

        # requesting my tickets as servicing staff should return empty list
        response = self.requestView(mine_url, self.customer_support, choose_org=self.org)
        assert_tickets(response, [])

        response = self.requestView(unassigned_url, self.customer_support, choose_org=self.org)
        assert_tickets(response, [c2_t1, c1_t2])

    @mock_mailroom
    def test_note(self, mr_mocks):
        ticket = self.create_ticket(self.contact)

        update_url = reverse("tickets.ticket_note", args=[ticket.uuid])

        self.assertRequestDisallowed(update_url, [None, self.user, self.admin2])
        self.assertUpdateFetch(update_url, [self.agent, self.editor, self.admin], form_fields=["note"])

        self.assertUpdateSubmit(
            update_url,
            self.admin,
            {"note": ""},
            form_errors={"note": "This field is required."},
            object_unchanged=ticket,
        )

        self.assertUpdateSubmit(
            update_url, self.admin, {"note": "I have a bad feeling about this."}, success_status=200
        )

        self.assertEqual(1, ticket.events.filter(event_type=TicketEvent.TYPE_NOTE_ADDED).count())

    def test_export_stats(self):
        export_url = reverse("tickets.ticket_export_stats")

        self.login(self.admin)

        response = self.client.get(export_url)
        self.assertEqual(200, response.status_code)
        self.assertEqual("application/ms-excel", response["Content-Type"])
        self.assertEqual(
            f"attachment; filename=ticket-stats-{timezone.now().strftime('%Y-%m-%d')}.xlsx",
            response["Content-Disposition"],
        )

    @mock_mailroom
    def test_export(self, mr_mocks):
        export_url = reverse("tickets.ticket_export")

        self.assertRequestDisallowed(export_url, [None, self.agent])
        response = self.assertUpdateFetch(
            export_url,
            [self.user, self.editor, self.admin],
            form_fields=("start_date", "end_date", "with_fields", "with_groups"),
        )
        self.assertNotContains(response, "already an export in progress")

        # create a dummy export task so that we won't be able to export
        blocking_export = TicketExport.create(
            self.org, self.admin, start_date=date.today() - timedelta(days=7), end_date=date.today()
        )

        response = self.client.get(export_url)
        self.assertContains(response, "already an export in progress")

        # check we can't submit in case a user opens the form and whilst another user is starting an export
        response = self.client.post(export_url, {"start_date": "2022-06-28", "end_date": "2022-09-28"})
        self.assertContains(response, "already an export in progress")
        self.assertEqual(1, Export.objects.count())

        # mark that one as finished so it's no longer a blocker
        blocking_export.status = Export.STATUS_COMPLETE
        blocking_export.save(update_fields=("status",))

        # try to submit with no values
        response = self.client.post(export_url, {})
        self.assertFormError(response.context["form"], "start_date", "This field is required.")
        self.assertFormError(response.context["form"], "end_date", "This field is required.")

        # try to submit with start date in future
        response = self.client.post(export_url, {"start_date": "2200-01-01", "end_date": "2022-09-28"})
        self.assertFormError(response.context["form"], None, "Start date can't be in the future.")

        # try to submit with start date > end date
        response = self.client.post(export_url, {"start_date": "2022-09-01", "end_date": "2022-03-01"})
        self.assertFormError(response.context["form"], None, "End date can't be before start date.")

        # try to submit with too many fields or groups
        too_many_fields = [self.create_field(f"Field {i}", f"field{i}") for i in range(11)]
        too_many_groups = [self.create_group(f"Group {i}", contacts=[]) for i in range(11)]

        response = self.client.post(
            export_url,
            {
                "start_date": "2022-06-28",
                "end_date": "2022-09-28",
                "with_fields": [cf.id for cf in too_many_fields],
                "with_groups": [cg.id for cg in too_many_groups],
            },
        )
        self.assertFormError(response.context["form"], "with_fields", "You can only include up to 10 fields.")
        self.assertFormError(response.context["form"], "with_groups", "You can only include up to 10 groups.")

        testers = self.create_group("Testers", contacts=[])
        gender = self.create_field("gender", "Gender")

        response = self.client.post(
            export_url,
            {
                "start_date": "2022-06-28",
                "end_date": "2022-09-28",
                "with_groups": [testers.id],
                "with_fields": [gender.id],
            },
        )
        self.assertEqual(200, response.status_code)

        export = Export.objects.exclude(id=blocking_export.id).get()
        self.assertEqual("ticket", export.export_type)
        self.assertEqual(date(2022, 6, 28), export.start_date)
        self.assertEqual(date(2022, 9, 28), export.end_date)
        self.assertEqual(
            {"with_groups": [testers.id], "with_fields": [gender.id]},
            export.config,
        )


class TicketExportTest(TembaTest):
    def _export(self, start_date: date, end_date: date, with_fields=(), with_groups=()):
        export = TicketExport.create(
            self.org,
            self.admin,
            start_date=start_date,
            end_date=end_date,
            with_fields=with_fields,
            with_groups=with_groups,
        )
        export.perform()

        workbook = load_workbook(filename=default_storage.open(f"orgs/{self.org.id}/ticket_exports/{export.uuid}.xlsx"))
        return workbook.worksheets, export

    def test_export_empty(self):
        # check results of sheet in workbook (no Contact ID column)
        sheets, export = self._export(start_date=date.today() - timedelta(days=7), end_date=date.today())
        self.assertExcelSheet(
            sheets[0],
            [
                [
                    "UUID",
                    "Opened On",
                    "Closed On",
                    "Topic",
                    "Assigned To",
                    "Contact UUID",
                    "Contact Name",
                    "URN Scheme",
                    "URN Value",
                ]
            ],
            tz=self.org.timezone,
        )

        with self.anonymous(self.org):
            # anon org doesn't see URN value column
            sheets, export = self._export(start_date=date.today() - timedelta(days=7), end_date=date.today())
            self.assertExcelSheet(
                sheets[0],
                [
                    [
                        "UUID",
                        "Opened On",
                        "Closed On",
                        "Topic",
                        "Assigned To",
                        "Contact UUID",
                        "Contact Name",
                        "URN Scheme",
                        "Anon Value",
                    ]
                ],
                tz=self.org.timezone,
            )

    def test_export(self):
        gender = self.create_field("gender", "Gender")
        age = self.create_field("age", "Age", value_type=ContactField.TYPE_NUMBER)

        # messages can't be older than org
        self.org.created_on = datetime(2016, 1, 2, 10, tzinfo=tzone.utc)
        self.org.save(update_fields=("created_on",))

        topic = Topic.create(self.org, self.admin, "AFC Richmond")
        assignee = self.admin
        today = timezone.now().astimezone(self.org.timezone).date()

        # create a contact with no urns
        nate = self.create_contact("Nathan Shelley", fields={"gender": "Male"})

        # create a contact with one urn
        jamie = self.create_contact(
            "Jamie Tartt", urns=["twitter:jamietarttshark"], fields={"gender": "Male", "age": 25}
        )

        # create a contact with multiple urns that have different max priority
        roy = self.create_contact(
            "Roy Kent", urns=["tel:+12345678900", "twitter:roykent"], fields={"gender": "Male", "age": 41}
        )

        # create a contact with multiple urns that have the same max priority
        sam = self.create_contact(
            "Sam Obisanya", urns=["twitter:nigerianprince", "tel:+9876543210"], fields={"gender": "Male", "age": 22}
        )
        sam.urns.update(priority=50)

        testers = self.create_group("Testers", contacts=[nate, roy])

        # create an open ticket for nate, opened 30 days ago
        ticket1 = self.create_ticket(
            nate, topic=topic, assignee=assignee, opened_on=timezone.now() - timedelta(days=30)
        )
        # create an open ticket for jamie, opened 25 days ago
        ticket2 = self.create_ticket(
            jamie, topic=topic, assignee=assignee, opened_on=timezone.now() - timedelta(days=25)
        )

        # create a closed ticket for roy, opened yesterday
        ticket3 = self.create_ticket(
            roy, topic=topic, assignee=assignee, opened_on=timezone.now() - timedelta(days=1), closed_on=timezone.now()
        )
        # create a closed ticket for sam, opened today
        ticket4 = self.create_ticket(
            sam, topic=topic, assignee=assignee, opened_on=timezone.now(), closed_on=timezone.now()
        )

        # create a ticket on another org for rebecca
        self.create_ticket(self.create_contact("Rebecca", urns=["twitter:rwaddingham"], org=self.org2))

        # check requesting export for last 90 days
        with self.mockReadOnly(assert_models={Ticket, ContactURN}):
            with self.assertNumQueries(17):
                sheets, export = self._export(start_date=today - timedelta(days=90), end_date=today)

        expected_headers = [
            "UUID",
            "Opened On",
            "Closed On",
            "Topic",
            "Assigned To",
            "Contact UUID",
            "Contact Name",
            "URN Scheme",
            "URN Value",
        ]

        self.assertExcelSheet(
            sheets[0],
            rows=[
                expected_headers,
                [
                    ticket1.uuid,
                    ticket1.opened_on,
                    "",
                    ticket1.topic.name,
                    ticket1.assignee.email,
                    ticket1.contact.uuid,
                    "Nathan Shelley",
                    "",
                    "",
                ],
                [
                    ticket2.uuid,
                    ticket2.opened_on,
                    "",
                    ticket2.topic.name,
                    ticket2.assignee.email,
                    ticket2.contact.uuid,
                    "Jamie Tartt",
                    "twitter",
                    "jamietarttshark",
                ],
                [
                    ticket3.uuid,
                    ticket3.opened_on,
                    ticket3.closed_on,
                    ticket3.topic.name,
                    ticket3.assignee.email,
                    ticket3.contact.uuid,
                    "Roy Kent",
                    "tel",
                    "+12345678900",
                ],
                [
                    ticket4.uuid,
                    ticket4.opened_on,
                    ticket4.closed_on,
                    ticket4.topic.name,
                    ticket4.assignee.email,
                    ticket4.contact.uuid,
                    "Sam Obisanya",
                    "twitter",
                    "nigerianprince",
                ],
            ],
            tz=self.org.timezone,
        )

        # check requesting export for last 7 days
        with self.mockReadOnly(assert_models={Ticket, ContactURN}):
            sheets, export = self._export(start_date=today - timedelta(days=7), end_date=today)

        self.assertExcelSheet(
            sheets[0],
            rows=[
                expected_headers,
                [
                    ticket3.uuid,
                    ticket3.opened_on,
                    ticket3.closed_on,
                    ticket3.topic.name,
                    ticket3.assignee.email,
                    ticket3.contact.uuid,
                    "Roy Kent",
                    "tel",
                    "+12345678900",
                ],
                [
                    ticket4.uuid,
                    ticket4.opened_on,
                    ticket4.closed_on,
                    ticket4.topic.name,
                    ticket4.assignee.email,
                    ticket4.contact.uuid,
                    "Sam Obisanya",
                    "twitter",
                    "nigerianprince",
                ],
            ],
            tz=self.org.timezone,
        )

        # check requesting with contact fields and groups
        with self.mockReadOnly(assert_models={Ticket, ContactURN}):
            sheets, export = self._export(
                start_date=today - timedelta(days=7), end_date=today, with_fields=(age, gender), with_groups=(testers,)
            )

        self.assertExcelSheet(
            sheets[0],
            rows=[
                expected_headers + ["Field:Age", "Field:Gender", "Group:Testers"],
                [
                    ticket3.uuid,
                    ticket3.opened_on,
                    ticket3.closed_on,
                    ticket3.topic.name,
                    ticket3.assignee.email,
                    ticket3.contact.uuid,
                    "Roy Kent",
                    "tel",
                    "+12345678900",
                    "41",
                    "Male",
                    True,
                ],
                [
                    ticket4.uuid,
                    ticket4.opened_on,
                    ticket4.closed_on,
                    ticket4.topic.name,
                    ticket4.assignee.email,
                    ticket4.contact.uuid,
                    "Sam Obisanya",
                    "twitter",
                    "nigerianprince",
                    "22",
                    "Male",
                    False,
                ],
            ],
            tz=self.org.timezone,
        )

        with self.anonymous(self.org):
            with self.mockReadOnly(assert_models={Ticket, ContactURN}):
                sheets, export = self._export(start_date=today - timedelta(days=90), end_date=today)
            self.assertExcelSheet(
                sheets[0],
                [
                    [
                        "UUID",
                        "Opened On",
                        "Closed On",
                        "Topic",
                        "Assigned To",
                        "Contact UUID",
                        "Contact Name",
                        "URN Scheme",
                        "Anon Value",
                    ],
                    [
                        ticket1.uuid,
                        ticket1.opened_on,
                        "",
                        ticket1.topic.name,
                        ticket1.assignee.email,
                        ticket1.contact.uuid,
                        "Nathan Shelley",
                        "",
                        ticket1.contact.anon_display,
                    ],
                    [
                        ticket2.uuid,
                        ticket2.opened_on,
                        "",
                        ticket2.topic.name,
                        ticket2.assignee.email,
                        ticket2.contact.uuid,
                        "Jamie Tartt",
                        "twitter",
                        ticket2.contact.anon_display,
                    ],
                    [
                        ticket3.uuid,
                        ticket3.opened_on,
                        ticket3.closed_on,
                        ticket3.topic.name,
                        ticket3.assignee.email,
                        ticket3.contact.uuid,
                        "Roy Kent",
                        "tel",
                        ticket3.contact.anon_display,
                    ],
                    [
                        ticket4.uuid,
                        ticket4.opened_on,
                        ticket4.closed_on,
                        ticket4.topic.name,
                        ticket4.assignee.email,
                        ticket4.contact.uuid,
                        "Sam Obisanya",
                        "twitter",
                        ticket4.contact.anon_display,
                    ],
                ],
                tz=self.org.timezone,
            )


class TopicTest(TembaTest):
    def test_create(self):
        topic1 = Topic.create(self.org, self.admin, "Sales")

        self.assertEqual("Sales", topic1.name)
        self.assertEqual("Sales", str(topic1))
        self.assertEqual(f'<Topic: id={topic1.id} name="Sales">', repr(topic1))

        # try to create with invalid name
        with self.assertRaises(AssertionError):
            Topic.create(self.org, self.admin, '"Support"')

        # try to create with name that already exists
        with self.assertRaises(AssertionError):
            Topic.create(self.org, self.admin, "Sales")

    @override_settings(ORG_LIMIT_DEFAULTS={"topics": 3})
    def test_import(self):
        def _import(definition, preview=False):
            return Topic.import_def(self.org, self.admin, definition, preview=preview)

        # preview import as dependency ref from flow inspection
        topic1, result = _import({"uuid": "0c81be38-8481-4a20-92ca-67e9a5617e77", "name": "Sales"}, preview=True)
        self.assertIsNone(topic1)
        self.assertEqual(Topic.ImportResult.CREATED, result)
        self.assertEqual(0, Topic.objects.filter(name="Sales").count())

        # import as dependency ref from flow inspection
        topic1, result = _import({"uuid": "0c81be38-8481-4a20-92ca-67e9a5617e77", "name": "Sales"})
        self.assertNotEqual("0c81be38-8481-4a20-92ca-67e9a5617e77", str(topic1.uuid))  # UUIDs never trusted
        self.assertEqual("Sales", topic1.name)
        self.assertEqual(Topic.ImportResult.CREATED, result)

        # preview import same definition again
        topic2, result = _import({"uuid": "0c81be38-8481-4a20-92ca-67e9a5617e77", "name": "Sales"}, preview=True)
        self.assertEqual(topic1, topic2)
        self.assertEqual(Topic.ImportResult.MATCHED, result)

        # import same definition again
        topic2, result = _import({"uuid": "0c81be38-8481-4a20-92ca-67e9a5617e77", "name": "Sales"})
        self.assertEqual(topic1, topic2)
        self.assertEqual("Sales", topic2.name)
        self.assertEqual(Topic.ImportResult.MATCHED, result)

        # import different UUID but same name
        topic3, result = _import({"uuid": "89a2265b-0caf-478f-837c-187fc8c32b46", "name": "Sales"})
        self.assertEqual(topic2, topic3)
        self.assertEqual(Topic.ImportResult.MATCHED, result)

        topic4 = Topic.create(self.org, self.admin, "Support")

        # import with UUID of existing thing (i.e. importing an export from this workspace)
        topic5, result = _import({"uuid": str(topic4.uuid), "name": "Support"})
        self.assertEqual(topic4, topic5)
        self.assertEqual(Topic.ImportResult.MATCHED, result)

        # preview import with UUID of existing thing with different name
        topic6, result = _import({"uuid": str(topic4.uuid), "name": "Help"}, preview=True)
        self.assertEqual(topic5, topic6)
        self.assertEqual("Support", topic6.name)  # not actually updated
        self.assertEqual(Topic.ImportResult.UPDATED, result)

        # import with UUID of existing thing with different name
        topic6, result = _import({"uuid": str(topic4.uuid), "name": "Help"})
        self.assertEqual(topic5, topic6)
        self.assertEqual("Help", topic6.name)  # updated
        self.assertEqual(Topic.ImportResult.UPDATED, result)

        # import with UUID of existing thing and name that conflicts with another existing thing
        topic7, result = _import({"uuid": str(topic4.uuid), "name": "Sales"})
        self.assertEqual(topic6, topic7)
        self.assertEqual("Sales 2", topic7.name)  # updated with suffix to make it unique
        self.assertEqual(Topic.ImportResult.UPDATED, result)

        # import definition of default topic from other workspace
        topic8, result = _import({"uuid": "bfacf01f-50d5-4236-9faa-7673bb4a9520", "name": "General"})
        self.assertEqual(self.org.default_ticket_topic, topic8)
        self.assertEqual(Topic.ImportResult.MATCHED, result)

        # import definition of default topic from this workspace
        topic9, result = _import({"uuid": str(self.org.default_ticket_topic.uuid), "name": "General"})
        self.assertEqual(self.org.default_ticket_topic, topic9)
        self.assertEqual(Topic.ImportResult.MATCHED, result)

        # import definition of default topic from this workspace... but with different name
        topic10, result = _import({"uuid": str(self.org.default_ticket_topic.uuid), "name": "Default"})
        self.assertEqual(self.org.default_ticket_topic, topic10)
        self.assertEqual("General", topic10.name)  # unchanged
        self.assertEqual(Topic.ImportResult.MATCHED, result)

        # import definition with name that can be cleaned and then matches existing
        topic11, result = _import({"uuid": "e694bad8-9cca-4efd-9f07-cb13248ed5e8", "name": " Sales\0 "})
        self.assertEqual("Sales", topic11.name)
        self.assertEqual(Topic.ImportResult.MATCHED, result)

        # import definition with name that can be cleaned and created new
        topic12, result = _import({"uuid": "c537ad58-ab2e-4b3a-8677-2766a2d14efe", "name": ' "Testing" '})
        self.assertEqual("'Testing'", topic12.name)
        self.assertEqual(Topic.ImportResult.CREATED, result)

        # try to import with name that can't be cleaned to something valid
        topic13, result = _import({"uuid": "c537ad58-ab2e-4b3a-8677-2766a2d14efe", "name": "  "})
        self.assertIsNone(topic13)
        self.assertEqual(Topic.ImportResult.IGNORED_INVALID, result)

        # import with UUID of existing thing and invalid name which will be ignored
        topic14, result = _import({"uuid": str(topic4.uuid), "name": "  "})
        self.assertEqual(topic4, topic14)
        self.assertEqual(Topic.ImportResult.MATCHED, result)

        # try to import new now that we've reached org limit
        topic15, result = _import({"uuid": "bef5f64c-0ad5-4ee0-9c9f-b3f471ec3b0c", "name": "Yet More"})
        self.assertIsNone(topic15)
        self.assertEqual(Topic.ImportResult.IGNORED_LIMIT_REACHED, result)

    def test_get_accessible(self):
        topic1 = Topic.create(self.org, self.admin, "Sales")
        topic2 = Topic.create(self.org, self.admin, "Support")
        team1 = Team.create(self.org, self.admin, "Sales & Support", topics=[topic1, topic2])
        team2 = Team.create(self.org, self.admin, "Nothing", topics=[])
        agent2 = self.create_user("agent2@textit.com")
        self.org.add_user(agent2, OrgRole.AGENT, team=team1)
        agent3 = self.create_user("agent3@textit.com")
        self.org.add_user(agent3, OrgRole.AGENT, team=team2)

        self.assertEqual(
            {self.org.default_ticket_topic, topic1, topic2}, set(Topic.get_accessible(self.org, self.admin))
        )
        self.assertEqual(
            {self.org.default_ticket_topic, topic1, topic2}, set(Topic.get_accessible(self.org, self.agent))
        )
        self.assertEqual({topic1, topic2}, set(Topic.get_accessible(self.org, agent2)))
        self.assertEqual(set(), set(Topic.get_accessible(self.org, agent3)))
        self.assertEqual(
            {self.org.default_ticket_topic, topic1, topic2}, set(Topic.get_accessible(self.org, self.customer_support))
        )

    def test_release(self):
        topic1 = Topic.create(self.org, self.admin, "Sales")
        topic2 = Topic.create(self.org, self.admin, "Support")
        flow = self.create_flow("Test")
        flow.topic_dependencies.add(topic1)
        team = Team.create(self.org, self.admin, "Sales & Support", topics=[topic1, topic2])
        ticket = self.create_ticket(self.create_contact("Ann"), topic=topic1)
        self.create_ticket(self.create_contact("Bob"), topic=topic2)

        # can't release a topic with tickets
        with self.assertRaises(AssertionError):
            topic1.release(self.admin)

        ticket.delete()

        topic1.release(self.admin)

        self.assertFalse(topic1.is_active)
        self.assertTrue(topic1.name.startswith("deleted-"))

        # topic should be removed from team
        self.assertEqual({topic2}, set(team.topics.all()))

        # counts should be deleted
        self.assertEqual(0, self.org.counts.filter(scope__startswith=f"tickets:O:{topic1.id}:").count())
        self.assertEqual(1, self.org.counts.filter(scope__startswith=f"tickets:O:{topic2.id}:").count())

        # flow should be flagged as having issues
        flow.refresh_from_db()
        self.assertTrue(flow.has_issues)

        # can't release system topic
        with self.assertRaises(AssertionError):
            self.org.default_ticket_topic.release(self.admin)

        # can't release a topic with tickets
        ticket = self.create_ticket(self.create_contact("Bob"), topic=topic1)
        with self.assertRaises(AssertionError):
            topic1.release(self.admin)


class TeamTest(TembaTest):
    def test_create(self):
        sales = Topic.create(self.org, self.admin, "Sales")
        support = Topic.create(self.org, self.admin, "Support")
        team1 = Team.create(self.org, self.admin, "Sales & Support", topics=[sales, support])
        agent2 = self.create_user("tickets@textit.com")
        self.org.add_user(self.agent, OrgRole.AGENT, team=team1)
        self.org.add_user(agent2, OrgRole.AGENT, team=team1)

        self.assertEqual("Sales & Support", team1.name)
        self.assertEqual("Sales & Support", str(team1))
        self.assertEqual(f'<Team: id={team1.id} name="Sales & Support">', repr(team1))
        self.assertEqual({self.agent, agent2}, set(team1.get_users()))
        self.assertEqual({sales, support}, set(team1.topics.all()))
        self.assertFalse(team1.all_topics)

        # create an unrestricted team
        team2 = Team.create(self.org, self.admin, "Any Topic", all_topics=True)
        self.assertEqual(set(), set(team2.topics.all()))
        self.assertTrue(team2.all_topics)

        # try to create with invalid name
        with self.assertRaises(AssertionError):
            Team.create(self.org, self.admin, '"Support"')

        # try to create with name that already exists
        with self.assertRaises(AssertionError):
            Team.create(self.org, self.admin, "Sales & Support")

    def test_release(self):
        team1 = Team.create(self.org, self.admin, "Sales")
        self.org.add_user(self.agent, OrgRole.AGENT, team=team1)

        team1.release(self.admin)

        self.assertFalse(team1.is_active)
        self.assertTrue(team1.name.startswith("deleted-"))
        self.assertEqual(0, team1.get_users().count())

        # check agent was re-assigned to default team
        self.assertEqual({self.agent}, set(self.org.default_ticket_team.get_users()))

        # can't release system team
        with self.assertRaises(AssertionError):
            self.org.default_ticket_team.release(self.admin)


class TicketDailyCountTest(TembaTest):
    def test_model(self):
        sales = Team.create(self.org, self.admin, "Sales")
        self.org.add_user(self.agent, OrgRole.AGENT, team=sales)
        self.org.add_user(self.editor, OrgRole.AGENT, team=sales)

        self._record_opening(self.org, date(2022, 4, 30))
        self._record_opening(self.org, date(2022, 5, 3))
        self._record_assignment(self.org, self.admin, date(2022, 5, 3))
        self._record_reply(self.org, self.admin, date(2022, 5, 3))

        self._record_reply(self.org, self.editor, date(2022, 5, 4))
        self._record_reply(self.org, self.agent, date(2022, 5, 4))

        self._record_reply(self.org, self.admin, date(2022, 5, 5))
        self._record_reply(self.org, self.admin, date(2022, 5, 5))
        self._record_opening(self.org, date(2022, 5, 5))
        self._record_reply(self.org, self.agent, date(2022, 5, 5))

        def assert_counts():
            # openings tracked at org scope
            self.assertEqual(3, TicketDailyCount.get_by_org(self.org, TicketDailyCount.TYPE_OPENING).total())
            self.assertEqual(
                2, TicketDailyCount.get_by_org(self.org, TicketDailyCount.TYPE_OPENING, since=date(2022, 5, 1)).total()
            )
            self.assertEqual(
                1, TicketDailyCount.get_by_org(self.org, TicketDailyCount.TYPE_OPENING, until=date(2022, 5, 1)).total()
            )
            self.assertEqual(0, TicketDailyCount.get_by_org(self.org2, TicketDailyCount.TYPE_OPENING).total())
            self.assertEqual(
                [(date(2022, 4, 30), 1), (date(2022, 5, 3), 1), (date(2022, 5, 5), 1)],
                TicketDailyCount.get_by_org(self.org, TicketDailyCount.TYPE_OPENING).day_totals(),
            )
            self.assertEqual(
                [(4, 1), (5, 2)], TicketDailyCount.get_by_org(self.org, TicketDailyCount.TYPE_OPENING).month_totals()
            )

            # assignments tracked at org+user scope
            self.assertEqual(
                1, TicketDailyCount.get_by_users(self.org, [self.admin], TicketDailyCount.TYPE_ASSIGNMENT).total()
            )
            self.assertEqual(
                0, TicketDailyCount.get_by_users(self.org, [self.agent], TicketDailyCount.TYPE_ASSIGNMENT).total()
            )
            self.assertEqual(
                {self.admin: 1, self.agent: 0},
                TicketDailyCount.get_by_users(
                    self.org, [self.admin, self.agent], TicketDailyCount.TYPE_ASSIGNMENT
                ).scope_totals(),
            )
            self.assertEqual(
                [(date(2022, 5, 3), 1)],
                TicketDailyCount.get_by_users(self.org, [self.admin], TicketDailyCount.TYPE_ASSIGNMENT).day_totals(),
            )

            # replies tracked at org scope, team scope and user-in-org scope
            self.assertEqual(6, TicketDailyCount.get_by_org(self.org, TicketDailyCount.TYPE_REPLY).total())
            self.assertEqual(0, TicketDailyCount.get_by_org(self.org2, TicketDailyCount.TYPE_REPLY).total())
            self.assertEqual(3, TicketDailyCount.get_by_teams([sales], TicketDailyCount.TYPE_REPLY).total())
            self.assertEqual(
                3, TicketDailyCount.get_by_users(self.org, [self.admin], TicketDailyCount.TYPE_REPLY).total()
            )
            self.assertEqual(
                1, TicketDailyCount.get_by_users(self.org, [self.editor], TicketDailyCount.TYPE_REPLY).total()
            )
            self.assertEqual(
                2, TicketDailyCount.get_by_users(self.org, [self.agent], TicketDailyCount.TYPE_REPLY).total()
            )

        assert_counts()
        self.assertEqual(19, TicketDailyCount.objects.count())

        TicketDailyCount.squash()

        assert_counts()
        self.assertEqual(14, TicketDailyCount.objects.count())

        workbook = export_ticket_stats(self.org, date(2022, 4, 30), date(2022, 5, 6))
        self.assertEqual(["Tickets"], workbook.sheetnames)
        self.assertExcelRow(
            workbook.active, 1, ["", "Opened", "Replies", "Reply Time (Secs)"] + ["Assigned", "Replies"] * 4
        )
        self.assertExcelRow(workbook.active, 2, [date(2022, 4, 30), 1, 0, "", 0, 0, 0, 0, 0, 0, 0, 0])
        self.assertExcelRow(workbook.active, 3, [date(2022, 5, 1), 0, 0, "", 0, 0, 0, 0, 0, 0, 0, 0])
        self.assertExcelRow(workbook.active, 4, [date(2022, 5, 2), 0, 0, "", 0, 0, 0, 0, 0, 0, 0, 0])
        self.assertExcelRow(workbook.active, 5, [date(2022, 5, 3), 1, 1, "", 1, 1, 0, 0, 0, 0, 0, 0])
        self.assertExcelRow(workbook.active, 6, [date(2022, 5, 4), 0, 2, "", 0, 0, 0, 1, 0, 1, 0, 0])
        self.assertExcelRow(workbook.active, 7, [date(2022, 5, 5), 1, 3, "", 0, 2, 0, 1, 0, 0, 0, 0])

    def _record_opening(self, org, d: date):
        TicketDailyCount.objects.create(count_type=TicketDailyCount.TYPE_OPENING, scope=f"o:{org.id}", day=d, count=1)

    def _record_assignment(self, org, user, d: date):
        TicketDailyCount.objects.create(
            count_type=TicketDailyCount.TYPE_ASSIGNMENT, scope=f"o:{org.id}:u:{user.id}", day=d, count=1
        )

    def _record_reply(self, org, user, d: date):
        TicketDailyCount.objects.create(count_type=TicketDailyCount.TYPE_REPLY, scope=f"o:{org.id}", day=d, count=1)

        team = OrgMembership.objects.get(org=org, user=user).team
        if team:
            TicketDailyCount.objects.create(
                count_type=TicketDailyCount.TYPE_REPLY, scope=f"t:{team.id}", day=d, count=1
            )
        TicketDailyCount.objects.create(
            count_type=TicketDailyCount.TYPE_REPLY, scope=f"o:{org.id}:u:{user.id}", day=d, count=1
        )


class TicketDailyTimingTest(TembaTest):
    def test_model(self):
        self._record_first_reply(self.org, date(2022, 4, 30), 60)
        self._record_first_reply(self.org, date(2022, 5, 1), 60)
        self._record_first_reply(self.org, date(2022, 5, 1), 120)
        self._record_first_reply(self.org, date(2022, 5, 1), 180)
        self._record_first_reply(self.org, date(2022, 5, 2), 11)
        self._record_first_reply(self.org, date(2022, 5, 2), 70)
        self._record_last_close(self.org, date(2022, 5, 1), 100)
        self._record_last_close(self.org, date(2022, 5, 1), 100, undo=True)
        self._record_last_close(self.org, date(2022, 5, 1), 200)
        self._record_last_close(self.org, date(2022, 5, 1), 300)
        self._record_last_close(self.org, date(2022, 5, 2), 100)

        def assert_timings():
            self.assertEqual(6, TicketDailyTiming.get_by_org(self.org, TicketDailyTiming.TYPE_FIRST_REPLY).total())
            self.assertEqual(
                [(date(2022, 4, 30), 1), (date(2022, 5, 1), 3), (date(2022, 5, 2), 2)],
                TicketDailyTiming.get_by_org(self.org, TicketDailyTiming.TYPE_FIRST_REPLY).day_totals(),
            )
            self.assertEqual(
                [(date(2022, 4, 30), 60.0), (date(2022, 5, 1), 120.0), (date(2022, 5, 2), 40.5)],
                TicketDailyTiming.get_by_org(self.org, TicketDailyTiming.TYPE_FIRST_REPLY).day_averages(rounded=False),
            )

            self.assertEqual(3, TicketDailyTiming.get_by_org(self.org, TicketDailyTiming.TYPE_LAST_CLOSE).total())
            self.assertEqual(
                [(date(2022, 5, 1), 2), (date(2022, 5, 2), 1)],
                TicketDailyTiming.get_by_org(self.org, TicketDailyTiming.TYPE_LAST_CLOSE).day_totals(),
            )
            self.assertEqual(
                [(date(2022, 5, 1), 250.0), (date(2022, 5, 2), 100.0)],
                TicketDailyTiming.get_by_org(self.org, TicketDailyTiming.TYPE_LAST_CLOSE).day_averages(),
            )

        assert_timings()

        squash_ticket_counts()

        assert_timings()

        workbook = export_ticket_stats(self.org, date(2022, 4, 30), date(2022, 5, 4))
        self.assertEqual(["Tickets"], workbook.sheetnames)
        self.assertExcelRow(
            workbook.active, 1, ["", "Opened", "Replies", "Reply Time (Secs)"] + ["Assigned", "Replies"] * 4
        )
        self.assertExcelRow(workbook.active, 2, [date(2022, 4, 30), 0, 0, 60, 0, 0, 0, 0, 0, 0, 0, 0])
        self.assertExcelRow(workbook.active, 3, [date(2022, 5, 1), 0, 0, 120, 0, 0, 0, 0, 0, 0, 0, 0])
        self.assertExcelRow(workbook.active, 4, [date(2022, 5, 2), 0, 0, 40, 0, 0, 0, 0, 0, 0, 0, 0])
        self.assertExcelRow(workbook.active, 5, [date(2022, 5, 3), 0, 0, "", 0, 0, 0, 0, 0, 0, 0, 0])

    def _record_first_reply(self, org, d: date, seconds: int):
        TicketDailyTiming.objects.create(
            count_type=TicketDailyTiming.TYPE_FIRST_REPLY, scope=f"o:{org.id}", day=d, count=1, seconds=seconds
        )

    def _record_last_close(self, org, d: date, seconds: int, undo: bool = False):
        count, seconds = (-1, -seconds) if undo else (1, seconds)

        TicketDailyTiming.objects.create(
            count_type=TicketDailyTiming.TYPE_LAST_CLOSE, scope=f"o:{org.id}", day=d, count=count, seconds=seconds
        )
